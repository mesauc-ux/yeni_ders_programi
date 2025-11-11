from flask import Flask, render_template_string, request, jsonify, send_file, make_response
from datetime import datetime, timedelta
import io
import sqlite3
import json
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from weasyprint import HTML

app = Flask(__name__)

# SQLite veritabanı bağlantısı
def get_db():
    conn = sqlite3.connect('/home/mesauc/mysite/ders_programi.db')
    conn.row_factory = sqlite3.Row
    return conn

# Veritabanı tablolarını oluştur
def init_db():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            surname TEXT NOT NULL,
            branch TEXT NOT NULL,
            schedule TEXT NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            surname TEXT NOT NULL,
            class TEXT NOT NULL,
            restrictions TEXT,
            priorities TEXT,
            manual_lessons TEXT
        )
    ''')

    # Mevcut tabloya yeni kolonları ekle (eğer yoksa)
    try:
        cursor.execute('ALTER TABLE students ADD COLUMN priorities TEXT')
    except:
        pass

    try:
        cursor.execute('ALTER TABLE students ADD COLUMN manual_lessons TEXT')
    except:
        pass

    try:
        cursor.execute('ALTER TABLE students ADD COLUMN teacher_blocks TEXT')
    except:
        pass  # Kolon zaten varsa hata verme

    # 🆕 ÖĞRETMEN TABLOSUNA BLOKLAMA KOLONU EKLE
    try:
        cursor.execute('ALTER TABLE teachers ADD COLUMN blocked_slots TEXT')
    except:
        pass  # Kolon zaten varsa hata verme

    # 🆕 GEÇMİŞ PROGRAMLAR TABLOSU
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS saved_schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            schedule_data TEXT NOT NULL,
            teachers_snapshot TEXT,
            students_snapshot TEXT
        )
    ''')

    # 🆕 BAŞLANGIÇ TARİHİ KOLONU EKLE
    try:
        cursor.execute('ALTER TABLE saved_schedules ADD COLUMN start_date TEXT')
    except:
        pass  # Kolon zaten varsa hata verme

    # 📚 SINIF DERSLERİ TABLOSU
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS class_lessons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT NOT NULL,
            teacher_id INTEGER NOT NULL,
            day TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            weeks TEXT NOT NULL,
            is_group INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id)
        )
    ''')

    # 🆕 Mevcut tabloya is_group field ekle (eğer yoksa)
    try:
        cursor.execute("ALTER TABLE class_lessons ADD COLUMN is_group INTEGER DEFAULT 0")
        print("✅ class_lessons tablosuna is_group field eklendi")
    except:
        print("ℹ️ is_group field zaten mevcut")

    # 🆕 Mevcut tabloya is_forced field ekle (eğer yoksa)
    try:
        cursor.execute("ALTER TABLE class_lessons ADD COLUMN is_forced INTEGER DEFAULT 0")
        print("✅ class_lessons tablosuna is_forced field eklendi")
    except:
        print("ℹ️ is_forced field zaten mevcut")


    conn.commit()
    conn.close()

# Uygulama başlatıldığında veritabanını oluştur
init_db()

schedule_data = None

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Özel Ders Programı Yönetim Sistemi</title>
    <style>
        /* ⚙️ TEMEL STİLLER - HER ZAMAN AKTİF */
        * { margin: 0; padding: 0; box-sizing: border-box; }

        body {
            font-family: 'Inter', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }

        /* 🎨 GÖRSEL İYİLEŞTİRMELER - SADECE EKRANDA */
        @media screen {

        /* 🎨 MODERN RENK PALETİ VE DEĞİŞKENLER */
        :root {
            --primary-color: #667eea;
            --primary-dark: #5568d3;
            --secondary-color: #764ba2;
            --accent-color: #f093fb;
            --success-color: #10b981;
            --warning-color: #f59e0b;
            --danger-color: #ef4444;
            --info-color: #3b82f6;
            --dark-bg: #1e293b;
            --light-bg: #f8fafc;
            --text-primary: #1f2937;
            --text-secondary: #6b7280;
            --border-color: #e5e7eb;
            --shadow-sm: 0 2px 4px rgba(0,0,0,0.1);
            --shadow-md: 0 4px 15px rgba(0,0,0,0.15);
            --shadow-lg: 0 10px 30px rgba(0,0,0,0.2);
            --shadow-xl: 0 20px 50px rgba(0,0,0,0.25);
            --transition-fast: all 0.2s ease;
            --transition-medium: all 0.3s ease;
            --transition-slow: all 0.5s ease;
            --border-radius-sm: 8px;
            --border-radius-md: 12px;
            --border-radius-lg: 20px;
        }

        /* 🌙 DARK MODE DEĞİŞKENLERİ */
        body.dark-mode {
            --primary-color: #818cf8;
            --primary-dark: #6366f1;
            --dark-bg: #0f172a;
            --light-bg: #1e293b;
            --text-primary: #f1f5f9;
            --text-secondary: #94a3b8;
            --border-color: #334155;
        }

        /* 🌙 DARK MODE STİLLERİ */
        body.dark-mode {
            background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        }

        body.dark-mode .container {
            background: #1e293b;
            color: #f1f5f9;
        }

        body.dark-mode .header {
            background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
        }

        body.dark-mode .list-section {
            background: #0f172a;
        }

        body.dark-mode .data-table {
            background: #1e293b;
        }

        body.dark-mode .data-table thead {
            background: linear-gradient(135deg, #334155 0%, #1e293b 100%);
        }

        body.dark-mode .data-table tbody tr:hover {
            background: linear-gradient(90deg, #334155 0%, #475569 100%);
        }

        body.dark-mode .data-table tbody tr:nth-child(even) {
            background: #0f172a;
        }

        body.dark-mode .data-table tbody tr:nth-child(even):hover {
            background: linear-gradient(90deg, #334155 0%, #475569 100%);
        }

        body.dark-mode .data-table td {
            border-bottom-color: #334155;
            color: #e2e8f0;
        }

        body.dark-mode .modal-content {
            background: #1e293b;
            color: #f1f5f9;
        }

        body.dark-mode .form-group input,
        body.dark-mode .form-group select {
            background: #0f172a;
            color: #f1f5f9;
            border-color: #334155;
        }

        /* 🔍 ARAMA KUTUSU DARK MODE */
        body.dark-mode #teacherSearchInput,
        body.dark-mode #studentSearchInput {
            background: #0f172a;
            color: #f1f5f9;
            border-color: #334155;
        }

        body.dark-mode #teacherSearchInput:focus,
        body.dark-mode #studentSearchInput:focus {
            border-color: #818cf8;
            box-shadow: 0 0 0 3px rgba(129,140,248,0.2);
        }

        body.dark-mode #teacherSearchCount,
        body.dark-mode #studentSearchCount {
            color: #94a3b8;
        }

        /* 🎨 ARAMA İKONLARI DARK MODE */
        body.dark-mode .accordion-content .fa-chalkboard-teacher,
        body.dark-mode .accordion-content .fa-user-graduate {
            color: #818cf8 !important;
        }

        body.dark-mode .day-group,
        body.dark-mode .restriction-group {
            background: #0f172a;
            border-color: #334155;
        }

        body.dark-mode .lesson-slot {
            background: #1e293b;
        }

        body.dark-mode .stat-card {
            background: #1e293b;
            color: #f1f5f9;
            border-left-color: #818cf8;
        }

        body.dark-mode .stat-value {
            color: #f1f5f9;
        }

        body.dark-mode #weeklyScheduleSection > div {
            background: #1e293b;
        }

        body.dark-mode .detail-modal-content {
            background: #1e293b;
            color: #f1f5f9;
        }

        body.dark-mode .detail-section {
            background: #0f172a;
            border-left-color: #818cf8;
        }

        body.dark-mode .detail-item {
            background: #1e293b;
            color: #e2e8f0;
        }

        /* 🌙 DARK MODE TOGGLE BUTON HOVER */
        #darkModeToggle:hover {
            background: rgba(255,255,255,0.3);
            transform: scale(1.05);
        }

        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Inter', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--secondary-color) 100%);
            background-attachment: fixed;
            min-height: 100vh;
            padding: 20px;
            transition: var(--transition-medium);
            position: relative;
            overflow-x: hidden;
        }

        /* ✨ ARKA PLAN DESENI */
        body::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-image:
                radial-gradient(circle at 20% 50%, rgba(255, 255, 255, 0.1) 0%, transparent 50%),
                radial-gradient(circle at 80% 80%, rgba(255, 255, 255, 0.1) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: var(--border-radius-lg);
            box-shadow: var(--shadow-xl);
            overflow: hidden;
            position: relative;
            z-index: 1;
            animation: fadeInUp 0.6s ease;
        }

        /* 📊 HAFTALIK PROGRAM BÖLÜMÜ - SCROLL YOK */
        #weeklyScheduleSection {
            overflow-x: hidden;  /* Scroll bar yok */
            overflow-y: visible;
        }

        /* 🎬 FADE IN ANİMASYONU */
        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }
        .header h1 { font-size: 2.5em; margin-bottom: 10px; }
        .header p { font-size: 1.1em; opacity: 0.9; }
        .main-content { padding: 30px; }
        .button-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }
        .main-btn {
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--secondary-color) 100%);
            color: white;
            border: none;
            padding: 20px;
            border-radius: var(--border-radius-md);
            font-size: 1.1em;
            font-weight: 700;
            cursor: pointer;
            transition: var(--transition-medium);
            box-shadow: var(--shadow-md);
            position: relative;
            overflow: hidden;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        /* ✨ BUTON HOVER EFEKTİ */
        .main-btn:hover {
            transform: translateY(-5px) scale(1.02);
            box-shadow: 0 15px 35px rgba(102, 126, 234, 0.5);
        }

        /* 💫 BUTON RIPPLE EFEKTİ */
        .main-btn::before {
            content: '';
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.3);
            transform: translate(-50%, -50%);
            transition: width 0.6s, height 0.6s;
        }

        .main-btn:active::before {
            width: 300px;
            height: 300px;
        }

        /* 🎯 BUTON AKTİF DURUMU */
        .main-btn:active {
            transform: translateY(-2px) scale(0.98);
        }



        /* 🎨 BUTON İKONLARI */
        .main-btn i {
            margin-right: 8px;
            font-size: 1.1em;
        }

        .modal {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.7);
            z-index: 1000;
            overflow-y: auto;
        }
        .modal-content {
            background: white;
            max-width: 850px;
            margin: 50px auto;
            border-radius: 20px;
            padding: 30px;
            position: relative;
            max-height: 90vh;
            overflow-y: auto;
        }
        .close-btn {
            position: absolute;
            top: 15px;
            right: 20px;
            font-size: 30px;
            cursor: pointer;
            color: #999;
        }
        .close-btn:hover { color: #333; }
        .form-group { margin-bottom: 20px; }
        .form-group label {
            display: block;
            margin-bottom: 8px;
            font-weight: bold;
            color: #333;
        }
        .form-group input, .form-group select {
            width: 100%;
            padding: 12px;
            border: 2px solid #ddd;
            border-radius: 10px;
            font-size: 1em;
        }
        .form-group input:focus, .form-group select:focus {
            outline: none;
            border-color: #667eea;
        }
        .day-group {
            border: 2px solid #667eea;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            background: #f9f9f9;
        }
        .day-group-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
        }
        .remove-day-btn {
            background: #f44336;
            color: white;
            border: none;
            padding: 8px 15px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: bold;
        }
        .lesson-slot {
            display: grid;
            grid-template-columns: 120px 100px 120px auto;
            gap: 10px;
            margin-bottom: 15px;
            align-items: center;
            background: white;
            padding: 10px;
            border-radius: 8px;
        }
        .lesson-number {
            font-weight: bold;
            color: #667eea;
        }
        .lesson-slot input[type="time"] {
            padding: 8px;
            border: 2px solid #ddd;
            border-radius: 6px;
        }
        .lesson-slot select {
            padding: 8px;
            border: 2px solid #ddd;
            border-radius: 6px;
            background: #fff;
        }
        .restriction-group {
            border: 2px solid #ff9800;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 15px;
            background: #fff3e0;
        }
        .restriction-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }
        .restriction-title {
            font-weight: bold;
            color: #ff9800;
        }
        .remove-restriction-btn {
            background: #f44336;
            color: white;
            border: none;
            padding: 6px 12px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: bold;
            font-size: 0.9em;
        }
        .restriction-row {
            display: grid;
            grid-template-columns: 140px 120px 1fr;
            gap: 10px;
            margin-bottom: 10px;
            align-items: center;
        }
        .restriction-row select {
            padding: 8px;
            border: 2px solid #ddd;
            border-radius: 6px;
        }
        .lesson-checkboxes {
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
        }
        .lesson-checkbox {
            display: flex;
            align-items: center;
            gap: 5px;
        }
        .lesson-checkbox input[type="checkbox"] {
            width: auto;
        }
        .add-day-btn {
            background: #4CAF50;
            color: white;
            border: none;
            padding: 12px 20px;
            border-radius: 10px;
            cursor: pointer;
            font-weight: bold;
            width: 100%;
            margin-bottom: 20px;
        }
        .submit-btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 15px 30px;
            border-radius: 10px;
            font-size: 1.1em;
            font-weight: bold;
            cursor: pointer;
            width: 100%;
            margin-top: 20px;
        }
        .list-section {
            background: #f9f9f9;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
        }

        .results-container {
            margin-top: 30px;
        }
        .section-title {
            font-size: 1.8em;
            font-weight: bold;
            margin: 30px 0 20px 0;
            color: #333;
            border-bottom: 3px solid #667eea;
            padding-bottom: 10px;
        }

        .teacher-distribution {
            background: white;
            border-radius: 15px;
            padding: 25px;
            margin-bottom: 30px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }

        .branch-distribution {
            background: white;
            border-radius: 15px;
            padding: 25px;
            margin-bottom: 30px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }
        .branch-dist-table {
            width: 100%;
            border-collapse: collapse;
        }
        .branch-dist-table th {
            background: #dc2626;
            color: white;
            padding: 15px;
            text-align: center;
            font-weight: bold;
            border: 1px solid #b91c1c;
        }
        .branch-dist-table td {
            padding: 12px;
            text-align: center;
            border: 1px solid #e5e7eb;
        }
        .branch-dist-table .student-name {
            text-align: left;
            font-weight: 600;
        }
        .branch-cell {
            background: #fef3c7;
            font-size: 0.9em;
        }

        .success-message, .error-message {
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
            display: none;
        }
        .success-message { background: #4CAF50; color: white; }
        .error-message { background: #f44336; color: white; }
        .lessons-container {
            margin-bottom: 10px;
        }
        .remove-lesson-btn:hover {
            background: #da190b !important;
        }

        /* 🎴 ACCORDION KARTLARI - MODERN TASARIM */
        .accordion-header {
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--secondary-color) 100%);
            color: white;
            padding: 18px 25px;
            border-radius: var(--border-radius-md);
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-size: 1.3em;
            font-weight: 700;
            margin-bottom: 15px;
            transition: var(--transition-medium);
            user-select: none;
            box-shadow: var(--shadow-md);
            position: relative;
            overflow: hidden;
        }

        /* ✨ HOVER ANIMASYONU */
        .accordion-header:hover {
            transform: translateY(-4px) scale(1.01);
            box-shadow: 0 8px 25px rgba(102, 126, 234, 0.5);
        }

        /* 🆕 SINIF DERSLERİ ÖZEL YEŞİL RENK */
        #classLessonsSection .accordion-header {
            background: linear-gradient(135deg, #10b981 0%, #059669 100%);
            user-select: none;
            -webkit-user-select: none;
            -moz-user-select: none;
            -ms-user-select: none;
        }

        #classLessonsSection .accordion-header:hover {
            box-shadow: 0 8px 25px rgba(16, 185, 129, 0.5);
        }

        /* 💫 PARLAMA EFEKTİ */
        .accordion-header::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
            transition: left 0.5s;
        }

        .accordion-header:hover::before {
            left: 100%;
        }

        /* 🎯 İKON ARALIĞI */
        .accordion-header i {
            margin-right: 12px;
            font-size: 1.1em;
        }
        .accordion-arrow {
            transition: transform 0.3s;
            font-size: 1.3em;
        }
        .accordion-arrow.open {
            transform: rotate(180deg);
        }
        .accordion-content {
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.3s ease-out;
            opacity: 0;
        }
        .accordion-content.open {
            max-height: 600px; /* 🆕 Scroll bar için limit */
            overflow-y: auto; /* 🆕 Scroll bar */
            transition: max-height 0.5s ease-in;
            opacity: 1;
        }

        /* Accordion içindeki liste container'larına scroll ekle */
        #restrictionGroups,
        #manualLessonGroups,
        #studentTeacherBlockGroups {
            max-height: 700px;
            overflow-y: auto;
            padding-right: 10px; /* Scroll bar için boşluk */
        }

        /* Her hafta için priority listleri */
        #week1PriorityList,
        #week2PriorityList,
        #week3PriorityList,
        #week4PriorityList {
            max-height: 500px;
            overflow-y: auto;
            padding-right: 10px;
        }

        /* Özel scroll bar tasarımı (Chrome, Edge, Safari) */
        #restrictionGroups::-webkit-scrollbar,
        #manualLessonGroups::-webkit-scrollbar,
        #studentTeacherBlockGroups::-webkit-scrollbar,
        #week1PriorityList::-webkit-scrollbar,
        #week2PriorityList::-webkit-scrollbar,
        #week3PriorityList::-webkit-scrollbar,
        #week4PriorityList::-webkit-scrollbar,
        #todayLessonsWidget::-webkit-scrollbar {
            width: 8px;
        }

        #restrictionGroups::-webkit-scrollbar-track,
        #manualLessonGroups::-webkit-scrollbar-track,
        #studentTeacherBlockGroups::-webkit-scrollbar-track,
        #week1PriorityList::-webkit-scrollbar-track,
        #week2PriorityList::-webkit-scrollbar-track,
        #week3PriorityList::-webkit-scrollbar-track,
        #week4PriorityList::-webkit-scrollbar-track,
        #todayLessonsWidget::-webkit-scrollbar-track {
            background: #f1f1f1;
            border-radius: 10px;
        }

        #restrictionGroups::-webkit-scrollbar-thumb,
        #manualLessonGroups::-webkit-scrollbar-thumb,
        #studentTeacherBlockGroups::-webkit-scrollbar-thumb,
        #week1PriorityList::-webkit-scrollbar-thumb,
        #week2PriorityList::-webkit-scrollbar-thumb,
        #week3PriorityList::-webkit-scrollbar-thumb,
        #week4PriorityList::-webkit-scrollbar-thumb,
        #todayLessonsWidget::-webkit-scrollbar-thumb {
            background: linear-gradient(135deg, #10b981 0%, #059669 100%);
            border-radius: 10px;
        }

        #restrictionGroups::-webkit-scrollbar-thumb:hover,
        #manualLessonGroups::-webkit-scrollbar-thumb:hover,
        #studentTeacherBlockGroups::-webkit-scrollbar-thumb:hover,
        #week1PriorityList::-webkit-scrollbar-thumb:hover,
        #week2PriorityList::-webkit-scrollbar-thumb:hover,
        #week3PriorityList::-webkit-scrollbar-thumb:hover,
        #week4PriorityList::-webkit-scrollbar-thumb:hover,
        #todayLessonsWidget::-webkit-scrollbar-thumb:hover {
            background: linear-gradient(135deg, #059669 0%, #047857 100%);
        }

                /* 🆕 ÖĞRETMEN MODALI İÇİN SCROLL */
        #dayGroups,
        #teacherBlockGroups {
            max-height: 600px;
            overflow-y: auto;
            padding-right: 10px;
        }

        #dayGroups::-webkit-scrollbar,
        #teacherBlockGroups::-webkit-scrollbar {
            width: 8px;
        }

        #dayGroups::-webkit-scrollbar-track,
        #teacherBlockGroups::-webkit-scrollbar-track {
            background: #f1f1f1;
            border-radius: 10px;
        }

        #dayGroups::-webkit-scrollbar-thumb,
        #teacherBlockGroups::-webkit-scrollbar-thumb {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            border-radius: 10px;
        }

        #dayGroups::-webkit-scrollbar-thumb:hover,
        #teacherBlockGroups::-webkit-scrollbar-thumb:hover {
            background: linear-gradient(135deg, #5568d3 0%, #6a3d91 100%);
        }

        /* 📊 MODERN TABLO TASARIMI */
        .data-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
            background: white;
            border-radius: var(--border-radius-md);
            overflow: hidden;
            box-shadow: var(--shadow-md);
        }

        .data-table thead {
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--secondary-color) 100%);
        }

        .data-table th {
            color: white;
            padding: 18px 20px;
            text-align: left;
            font-weight: 700;
            font-size: 0.95em;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        .data-table td {
            padding: 15px 20px;
            border-bottom: 1px solid var(--border-color);
            font-size: 0.95em;
        }

        .data-table tr:last-child td {
            border-bottom: none;
        }

        .data-table tbody tr {
            transition: var(--transition-fast);
        }

        .data-table tbody tr:hover {
            background: linear-gradient(90deg, #f9fafb 0%, #f3f4f6 100%);
            transform: scale(1.01);
        }

        /* 🎨 ZEBRA STRİPE */
        .data-table tbody tr:nth-child(even) {
            background: #fafbfc;
        }

        .data-table tbody tr:nth-child(even):hover {
            background: linear-gradient(90deg, #f3f4f6 0%, #e5e7eb 100%);
        }
        .action-buttons {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }

        .view-btn, .edit-btn-small, .delete-btn-small {
            border: none;
            padding: 8px 16px;
            border-radius: var(--border-radius-sm);
            cursor: pointer;
            font-size: 0.85em;
            font-weight: 700;
            transition: var(--transition-fast);
            text-transform: uppercase;
            letter-spacing: 0.3px;
            box-shadow: var(--shadow-sm);
        }

        .view-btn, .edit-btn-small, .delete-btn-small {
            position: relative;
            overflow: hidden;
        }

        .view-btn::before, .edit-btn-small::before, .delete-btn-small::before {
            content: '';
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.5);
            transform: translate(-50%, -50%);
            transition: width 0.4s, height 0.4s;
        }

        .view-btn:active::before, .edit-btn-small:active::before, .delete-btn-small:active::before {
            width: 200px;
            height: 200px;
        }
        .view-btn {
            background: #10b981;
            color: white;
        }
        .view-btn:hover {
            background: #059669;
        }
        .edit-btn-small {
            background: #3b82f6;
            color: white;
        }
        .edit-btn-small:hover {
            background: #2563eb;
        }
        .delete-btn-small {
            background: #ef4444;
            color: white;
        }
        .delete-btn-small:hover {
            background: #dc2626;
        }

                /* 📊 PROGRESS BAR SİSTEMİ */
        .progress-container {
            width: 100%;
            background: #e5e7eb;
            border-radius: 50px;
            overflow: hidden;
            height: 24px;
            box-shadow: inset 0 2px 4px rgba(0,0,0,0.1);
            position: relative;
        }

        .progress-bar {
            height: 100%;
            background: linear-gradient(90deg, var(--primary-color) 0%, var(--secondary-color) 100%);
            border-radius: 50px;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: 700;
            font-size: 0.85em;
            transition: width 0.6s ease;
            position: relative;
            overflow: hidden;
        }

        /* ✨ PROGRESS BAR PARLAMA EFEKTİ */
        .progress-bar::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 50%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.3), transparent);
            animation: progressShine 2s infinite;
        }

        @keyframes progressShine {
            0% { left: -100%; }
            100% { left: 200%; }
        }

        /* 🎨 PROGRESS BAR RENKLERİ */
        .progress-bar.low {
            background: linear-gradient(90deg, #ef4444 0%, #dc2626 100%);
        }

        .progress-bar.medium {
            background: linear-gradient(90deg, #f59e0b 0%, #d97706 100%);
        }

        .progress-bar.high {
            background: linear-gradient(90deg, #10b981 0%, #059669 100%);
        }

        /* 🏷️ BADGE SİSTEMİ */
        .badge {
            display: inline-block;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 0.8em;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            box-shadow: var(--shadow-sm);
        }

        .badge-success {
            background: linear-gradient(135deg, #10b981 0%, #059669 100%);
            color: white;
        }

        .badge-warning {
            background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);
            color: white;
        }

        .badge-danger {
            background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
            color: white;
        }

        .badge-info {
            background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
            color: white;
        }

        .badge-primary {
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--secondary-color) 100%);
            color: white;
        }

        /* 📈 STAT CARD (Mini İstatistik Kartı) */
        .stat-card {
            background: white;
            border-radius: var(--border-radius-md);
            padding: 20px;
            box-shadow: var(--shadow-md);
            text-align: center;
            transition: var(--transition-medium);
            border-left: 4px solid var(--primary-color);
        }

        .stat-card:hover {
            transform: translateY(-5px);
            box-shadow: var(--shadow-lg);
        }

        .stat-icon {
            font-size: 2.5em;
            margin-bottom: 10px;
        }

        .stat-value {
            font-size: 2em;
            font-weight: 700;
            color: var(--text-primary);
            margin-bottom: 5px;
        }

        .stat-label {
            font-size: 0.9em;
            color: var(--text-secondary);
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        /* Detay Modal */
        .detail-modal {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.7);
            z-index: 2000;
            justify-content: center;
            align-items: center;
        }
        .detail-modal-content {
            background: white;
            padding: 30px;
            border-radius: 15px;
            max-width: 700px;
            max-height: 80vh;
            overflow-y: auto;
            position: relative;
        }
        .detail-close {
            position: absolute;
            top: 15px;
            right: 20px;
            font-size: 30px;
            cursor: pointer;
            color: #999;
        }
        .detail-close:hover {
            color: #333;
        }

        .detail-section {
            margin-bottom: 20px;
            padding: 15px;
            background: #f9fafb;
            border-radius: 10px;
            border-left: 4px solid #667eea;
        }
        .detail-section h4 {
            color: #667eea;
            margin-bottom: 10px;
            font-size: 1.1em;
        }
        .detail-item {
            padding: 8px;
            margin: 5px 0;
            background: white;
            border-radius: 5px;
            font-size: 0.95em;
        }

        /* Haftalık Program Butonları */
        #weeklyScheduleSection button[onclick*="changeWeek"] {
            min-width: 140px !important;  /* Önceki/Sonraki hafta butonları */
        }

        #weeklyScheduleSection button[onclick*="printWeeklyTable"] {
            min-width: 140px !important;  /* Yazdır butonu */
        }

        #weeklyScheduleSection button[onclick*="exportWeeklyToPDF"] {
            min-width: 140px !important;  /* PDF İndir butonu */
        }

        #weeklyScheduleSection div[style*="min-width: 150px"] {
            min-width: 150px !important;  /* Hafta numarası kutusu */
        }

        } /* @media screen SONU */

        /* 📝 EKRAN TABLOSU - SADECE YAZI BOYUTU */
        #weeklyScheduleSection table tbody tr td {
            font-size: 12px !important;
            padding: 4px 2px !important;
        }

        /* ═══════════════════════════════════════════════════════════════
           📱 MOBİL UYUMLULUK - GÜÇLÜ VERSİYON
           ═══════════════════════════════════════════════════════════════ */

        @media screen and (max-width: 768px) {
            /* GENEL */
            body {
                padding: 5px !important;
            }

            .container {
                padding: 10px !important;
                border-radius: 0 !important;
                margin: 0 !important;
            }
            .main-content {
                padding: 5px !important;
            }

            /* BAŞLIK */
            .header {
                padding: 15px 10px !important;
            }

            .header h1 {
                font-size: 1.3em !important;
                line-height: 1.2 !important;
            }

            .header p {
                font-size: 0.85em !important;
            }

            #darkModeToggle {
                font-size: 0.75em !important;
                padding: 6px 10px !important;
                top: 10px !important;
                right: 10px !important;
            }

            /* ANA BUTONLAR */
            .button-grid {
                display: flex !important;
                flex-direction: column !important;
                gap: 10px !important;
            }

            .main-btn {
                width: 100% !important;
                max-width: 100% !important;
                padding: 14px !important;
                font-size: 0.9em !important;
                margin: 0 !important;
            }

            /* ACCORDION */
            .accordion-header {
                font-size: 0.95em !important;
                padding: 12px 14px !important;
                user-select: none !important;
                -webkit-user-select: none !important;
                -moz-user-select: none !important;
                -ms-user-select: none !important;
            }

            /* TABLOLAR - ACCORDION İÇİNDE SCROLL */
            .list-section {
                overflow: visible !important;
                margin: 0 !important;
                padding: 0 !important;
            }

            .accordion-content {
                overflow: visible !important;
            }

            .accordion-content.open {
                overflow-x: auto !important;
                overflow-y: visible !important;
                -webkit-overflow-scrolling: touch !important;
                padding: 15px !important;
                margin: 0 -15px !important;
            }

            .accordion-content.open > .data-table {
                min-width: 900px !important;
                width: 900px !important;
                display: table !important;
            }

            .accordion-content.open .data-table {
                display: table !important;
            }

            .accordion-content.open .data-table thead {
                display: table-header-group !important;
            }

            .accordion-content.open .data-table tbody {
                display: table-row-group !important;
            }

            .accordion-content.open .data-table tr {
                display: table-row !important;
            }

            .accordion-content.open .data-table th,
            .accordion-content.open .data-table td {
                display: table-cell !important;
                white-space: nowrap !important;
                padding: 8px 6px !important;
            }

            .accordion-content.open .data-table th:nth-child(1),
            .accordion-content.open .data-table td:nth-child(1) {
                min-width: 150px !important;
            }

            .accordion-content.open .data-table th:nth-child(2),
            .accordion-content.open .data-table td:nth-child(2) {
                min-width: 120px !important;
            }

            .accordion-content.open .data-table th:nth-child(3),
            .accordion-content.open .data-table td:nth-child(3) {
                min-width: 100px !important;
            }

            .accordion-content.open .data-table th:last-child,
            .accordion-content.open .data-table td:last-child {
                min-width: 200px !important;
            }

            /* İşlemler başlığını göster */
            .accordion-content.open .data-table thead {
                display: table-header-group !important;
            }

            .accordion-content.open .data-table thead tr {
                display: table-row !important;
            }

            .accordion-content.open .data-table thead th {
                display: table-cell !important;
                background: #667eea !important;
                color: white !important;
                font-weight: bold !important;
                text-transform: uppercase !important;
                padding: 10px 8px !important;
                font-size: 0.75em !important;
            }

            /* Dark mode'da başlık */
            body.dark-mode .accordion-content.open .data-table thead th {
                background: #5568d3 !important;
            }

            .accordion-content.open .action-buttons {
                display: flex !important;
                flex-direction: column !important;
                gap: 5px !important;
            }

            .accordion-content.open .view-btn,
            .accordion-content.open .edit-btn-small,
            .accordion-content.open .delete-btn-small {
                display: block !important;
                width: 100% !important;
                padding: 6px 10px !important;
                font-size: 0.75em !important;
            }

            .accordion-content.open::before {
                content: '← Kaydırarak tüm sütunları görün →';
                display: block;
                text-align: center;
                padding: 8px;
                background: linear-gradient(90deg, transparent, rgba(102, 126, 234, 0.15), transparent);
                color: #667eea;
                font-size: 0.75em;
                font-weight: bold;
                margin: -15px -15px 10px -15px;
                border-radius: 10px 10px 0 0;
            }

            /* Scroll bar tam çözüm */
            .accordion-content.open {
                position: relative !important;
                z-index: 1 !important;
            }

            .accordion-content.open::-webkit-scrollbar {
                height: 14px !important;
                cursor: pointer !important;
                z-index: 999 !important;
            }

            .accordion-content.open::-webkit-scrollbar-track {
                background: #e2e8f0 !important;
                border-radius: 8px !important;
                border: 1px solid #cbd5e1 !important;
                cursor: pointer !important;
            }

            .accordion-content.open::-webkit-scrollbar-thumb {
                background: linear-gradient(180deg, #667eea 0%, #764ba2 100%) !important;
                border-radius: 8px !important;
                border: 2px solid #e2e8f0 !important;
                cursor: grab !important;
                min-width: 50px !important;
            }

            .accordion-content.open::-webkit-scrollbar-thumb:hover {
                background: linear-gradient(180deg, #5568d3 0%, #6a3d91 100%) !important;
                cursor: grab !important;
            }

            .accordion-content.open::-webkit-scrollbar-thumb:active {
                cursor: grabbing !important;
                background: linear-gradient(180deg, #4c5bc4 0%, #5e347f 100%) !important;
            }

            /* Scroll okları */
            .accordion-content.open::-webkit-scrollbar-button {
                width: 20px !important;
                height: 14px !important;
                background: #667eea !important;
                cursor: pointer !important;
                border: none !important;
            }

            .accordion-content.open::-webkit-scrollbar-button:hover {
                background: #5568d3 !important;
            }

            .accordion-content.open::-webkit-scrollbar-button:active {
                background: #4c5bc4 !important;
            }

            /* Sol ok */
            .accordion-content.open::-webkit-scrollbar-button:horizontal:decrement {
                background: #667eea !important;
                border-radius: 8px 0 0 8px !important;
            }

            .accordion-content.open::-webkit-scrollbar-button:horizontal:decrement::after {
                content: '◀' !important;
                color: white !important;
                font-size: 10px !important;
                position: absolute !important;
                left: 5px !important;
            }

            /* Sağ ok */
            .accordion-content.open::-webkit-scrollbar-button:horizontal:increment {
                background: #667eea !important;
                border-radius: 0 8px 8px 0 !important;
            }

            .accordion-content.open::-webkit-scrollbar-button:horizontal:increment::after {
                content: '▶' !important;
                color: white !important;
                font-size: 10px !important;
                position: absolute !important;
                right: 5px !important;
            }

            /* Tablo genişliğini artır */
            .accordion-content.open > .data-table {
                min-width: 1400px !important;
                width: 1400px !important;
            }

            /* İŞLEM BUTONLARI - ZORUNLU GÖRÜNÜR */
            .data-table td:last-child {
                min-width: 180px !important;
                padding: 10px !important;
            }

            .action-buttons {
                display: flex !important;
                flex-direction: column !important;
                gap: 6px !important;
                width: 100% !important;
            }

            .view-btn,
            .edit-btn-small,
            .delete-btn-small {
                width: 100% !important;
                padding: 8px 12px !important;
                font-size: 0.8em !important;
                display: block !important;
                visibility: visible !important;
                opacity: 1 !important;
                box-sizing: border-box !important;
            }

            /* MODAL */
            .modal-content,
            .detail-modal-content {
                width: 95% !important;
                max-width: 95% !important;
                margin: 10px auto !important;
                padding: 15px !important;
            }

            /* FORM */
            .form-group input,
            .form-group select {
                font-size: 16px !important;
                padding: 10px !important;
            }

            .lesson-slot {
                display: flex !important;
                flex-direction: column !important;
                gap: 8px !important;
            }

           /* ═══════════════════════════════════════════════════
               HAFTALIK PROGRAM - DOĞRU ÇÖZÜM
               ═══════════════════════════════════════════════════ */

            /* Ana konteyner */
            #weeklyScheduleSection {
                width: 100% !important;
                max-width: 100% !important;
                padding: 0 !important;
                margin: 0 !important;
                box-sizing: border-box !important;
            }

            /* Beyaz arka planlı iç konteyner - padding kaldır */
            #weeklyScheduleSection > div {
                padding: 10px !important;
                width: 100% !important;
                max-width: 100% !important;
                margin: 0 !important;
                box-sizing: border-box !important;
            }

            /* Üst kısım: Başlık + Butonlar (yan yana -> alt alta) */
            #weeklyScheduleSection > div > div:first-child {
                display: flex !important;
                flex-direction: column !important;
                gap: 15px !important;
                margin-bottom: 20px !important;
                align-items: stretch !important;
            }

            /* Başlık */
            #weeklyScheduleSection > div > div:first-child > h2 {
                font-size: 1.1em !important;
                margin: 0 !important;
                padding: 0 !important;
                text-align: center !important;
            }

            /* Buton grubu konteyneri (asıl butonların olduğu div) */
            #weeklyScheduleSection > div > div:first-child > div {
                display: flex !important;
                flex-direction: column !important;
                gap: 10px !important;
                width: 100% !important;
                max-width: 100% !important;
                padding: 0 !important;
                margin: 0 !important;
            }

            /* Tüm butonlar */
            #weeklyScheduleSection > div > div:first-child > div > button {
                display: block !important;
                width: 100% !important;
                max-width: 100% !important;
                padding: 14px 12px !important;
                font-size: 0.9em !important;
                min-height: 50px !important;
                height: auto !important;
                white-space: normal !important;
                word-wrap: break-word !important;
                line-height: 1.4 !important;
                margin: 0 !important;
                box-sizing: border-box !important;
                text-align: center !important;
                min-width: 0 !important;
            }

            /* TABLO SCROLL */
            #weeklyScheduleTable {
                overflow-x: auto !important;
                -webkit-overflow-scrolling: touch !important;
                margin: 0 -10px !important;
                padding: 0 10px !important;
            }

            #weeklyScheduleSection table {
                min-width: 900px !important;
                font-size: 0.6em !important;
            }
            /* TABLO KAYDIRMA DÜZELTMESİ - HER YERDEN KAYDIRILABİLİR */
            #weeklyScheduleSection table,
            #weeklyScheduleSection table caption,
            #weeklyScheduleSection table thead,
            #weeklyScheduleSection table tbody,
            #weeklyScheduleSection table tr,
            #weeklyScheduleSection table th,
            #weeklyScheduleSection table td {
                touch-action: pan-x pan-y !important;
                pointer-events: auto !important;
            }

            /* Scroll container tam genişlik */
            #weeklyScheduleTable {
                width: 100% !important;
                position: relative !important;
            }

            /* GRAFİKLER */
            #statsCardsSection {
                padding: 10px !important;
            }

            #statsCardsSection canvas {
                max-width: 100% !important;
                height: auto !important;
            }

            #statsCardsSection > div[style*="grid"] {
                display: block !important;
            }

            #statsCardsSection > div[style*="grid"] > div {
                width: 100% !important;
                margin-bottom: 15px !important;
            }

            /* 4 HAFTALIK PROGRAM */
            #resultsSection h2 {
                font-size: 1.2em !important;
            }

            #resultsSection > div[style*="grid-template-columns: repeat(4"] {
                display: block !important;
            }

            #resultsSection > div[style*="grid-template-columns: repeat(4"] > div {
                width: 100% !important;
                max-width: 100% !important;
                margin-bottom: 15px !important;
            }

            /* ÖĞRENCİ/BRANŞ KARTLARI */
            .student-card,
            .branch-card {
                padding: 12px !important;
            }

            .student-card > div:first-child,
            .branch-card > div:first-child {
                flex-direction: column !important;
                align-items: flex-start !important;
                gap: 8px !important;
            }

            /* FİLTRE */
            #studentSearchInput,
            #branchSearchInput,
            #studentDropdownBtn,
            #branchDropdownBtn {
                width: 100% !important;
                max-width: 100% !important;
                margin-bottom: 8px !important;
            }

            #studentDropdownMenu,
            #branchDropdownMenu {
                width: 100% !important;
                left: 0 !important;
                right: 0 !important;
            }
            /* İşlemler kolonunu her zaman görünür yap */
            .data-table th:last-child,
            .data-table td:last-child {
                position: sticky !important;
                right: 0 !important;
                background: white !important;
                box-shadow: -2px 0 5px rgba(0,0,0,0.1) !important;
                z-index: 5 !important;
            }

            /* Dark mode'da */
            body.dark-mode .data-table th:last-child,
            body.dark-mode .data-table td:last-child {
                background: #1e293b !important;
            }

            /* ÇAKIŞMA DASHBOARD */
            #conflictDashboardModal .modal-content {
                width: 95% !important;
                padding: 12px !important;
            }

            #conflictStats > div:first-child {
                grid-template-columns: repeat(2, 1fr) !important;
                gap: 8px !important;
            }

            .conflict-actions {
                flex-direction: column !important;
            }

            .conflict-btn {
                width: 100% !important;
            }
        }

        /* ÇOK KÜÇÜK EKRANLAR */
        @media screen and (max-width: 480px) {
            .header h1 {
                font-size: 1.1em !important;
            }

            .data-table {
                min-width: 700px !important;
                font-size: 0.65em !important;
            }

            #weeklyScheduleSection table {
                min-width: 800px !important;
                font-size: 0.55em !important;
            }

            #conflictStats > div:first-child {
                grid-template-columns: 1fr !important;
            }
        }

        /* ═══════════════════════════════════════════════════════════════
           🎬 HAFTALIK PROGRAM GEÇİŞ ANİMASYONLARI - ADIM 1
           ═══════════════════════════════════════════════════════════════ */

        /* Haftalık program tablosu için geçiş hazırlığı */
        #weeklyScheduleTable {
            transition: opacity 0.4s ease, transform 0.4s ease;
        }

        /* Soldaki hafta numarası için pulse animasyonu */
        @keyframes pulseScale {
            0%, 100% {
                transform: scale(1);
            }
            50% {
                transform: scale(1.05);
            }
        }

        /* Loading (yükleniyor) yazısı için */
        @keyframes spin {
            to { transform: rotate(360deg); }
        }

        /* Satırların tek tek belirmesi için */
        @keyframes slideInRow {
            from {
                opacity: 0;
                transform: translateX(-20px);
            }
            to {
                opacity: 1;
                transform: translateX(0);
            }
        }

        /* 🎨 ÖĞRENCİ ARAMA VURGULAMA */
        .student-highlight {
            background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%) !important;
            animation: highlightPulse 1.5s ease-in-out infinite;
            border: 2px solid #f59e0b !important;
            font-weight: 700 !important;
            box-shadow: 0 0 15px rgba(245, 158, 11, 0.4) !important;
        }

        @keyframes highlightPulse {
            0%, 100% {
                box-shadow: 0 0 15px rgba(245, 158, 11, 0.4);
            }
            50% {
                box-shadow: 0 0 25px rgba(245, 158, 11, 0.7);
            }
        }

        /* 🎨 ÖĞRETMEN ARAMA VURGULAMA */
        .teacher-highlight {
            background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%) !important;
            animation: teacherHighlightPulse 1.5s ease-in-out infinite;
            border: 2px solid #3b82f6 !important;
            font-weight: 700 !important;
            box-shadow: 0 0 15px rgba(59, 130, 246, 0.4) !important;
        }

        @keyframes teacherHighlightPulse {
            0%, 100% {
                box-shadow: 0 0 15px rgba(59, 130, 246, 0.4);
            }
            50% {
                box-shadow: 0 0 25px rgba(59, 130, 246, 0.7);
            }
        }

        /* ═══════════════════════════════════════════════════════════════
           🎯 SÜRÜKLE-BIRAK SİSTEMİ
           ═══════════════════════════════════════════════════════════════ */

        /* Sürüklenebilir hücreler */
        .draggable-cell {
            cursor: move;
            transition: all 0.2s ease;
            position: relative;
        }

        .draggable-cell:hover {
            background: linear-gradient(135deg, #FFE0B2 0%, #FFCC80 100%) !important;
            transform: scale(1.03);
            box-shadow: 0 4px 15px rgba(255, 152, 0, 0.4);
            transition: all 0.3s ease;
            z-index: 10;
            cursor: move;
        }

        /* Sürüklenirken */
        .dragging {
            opacity: 0.5;
            transform: scale(0.95);
            cursor: grabbing !important;
        }

        /* Drop zone (hedef) */
        .drag-over {
            background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%) !important;
            border: 2px dashed #3b82f6 !important;
            transform: scale(1.05);
        }

        /* Boş slot indicator */
        .empty-slot {
            background: #E3F2FD !important;  /* Açık mavi - dolu slotlarla aynı */
            border: 1px solid #e5e7eb !important;  /* İnce border */
        }

        .empty-slot:hover {
            background: linear-gradient(135deg, #E3F2FD 0%, #BBDEFB 100%) !important;
            border-color: #64B5F6 !important;
            cursor: pointer;
        }

        /* ═══════════════════════════════════════════════════════════════
           🎨 ONAY POPUP MODAL
           ═══════════════════════════════════════════════════════════════ */

        .swap-modal {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.7);
            z-index: 10000;
            justify-content: center;
            align-items: center;
            animation: fadeIn 0.2s ease;
        }

        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }

        .swap-modal-content {
            background: white;
            border-radius: 20px;
            padding: 40px;
            max-width: 500px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            animation: slideUp 0.3s ease;
            position: relative;
        }

        @keyframes slideUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        .swap-modal-header {
            display: flex;
            align-items: center;
            gap: 15px;
            margin-bottom: 25px;
        }

        .swap-modal-icon {
            font-size: 3em;
            animation: bounce 0.5s ease infinite alternate;
        }

        @keyframes bounce {
            from { transform: translateY(0); }
            to { transform: translateY(-10px); }
        }

        .swap-modal-title {
            font-size: 1.5em;
            font-weight: bold;
            color: #1f2937;
        }

        .swap-modal-body {
            background: #f9fafb;
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 25px;
            border-left: 4px solid #f59e0b;
        }

        /* Modern Scroll Bar Stili */
        .swap-modal-body::-webkit-scrollbar {
            width: 8px;
        }

        .swap-modal-body::-webkit-scrollbar-track {
            background: #f1f5f9;
            border-radius: 10px;
        }

        .swap-modal-body::-webkit-scrollbar-thumb {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            border-radius: 10px;
        }

        .swap-modal-body::-webkit-scrollbar-thumb:hover {
            background: linear-gradient(135deg, #5568d3 0%, #653a8a 100%);
        }

        .swap-info {
            display: flex;
            align-items: center;
            gap: 15px;
            margin: 10px 0;
            padding: 12px;
            background: white;
            border-radius: 8px;
            font-weight: 600;
            color: #1f2937;
        }

        .swap-arrow {
            font-size: 2em;
            color: #3b82f6;
            font-weight: bold;
        }

        .swap-modal-buttons {
            display: flex;
            gap: 15px;
        }

        .swap-btn {
            flex: 1;
            padding: 15px 25px;
            border: none;
            border-radius: 12px;
            font-size: 1.1em;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s ease;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        .swap-btn-confirm {
            background: linear-gradient(135deg, #10b981 0%, #059669 100%);
            color: white;
            box-shadow: 0 4px 15px rgba(16, 185, 129, 0.3);
        }

        .swap-btn-confirm:hover {
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(16, 185, 129, 0.5);
        }

        .swap-btn-cancel {
            background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
            color: white;
            box-shadow: 0 4px 15px rgba(239, 68, 68, 0.3);
        }

        .swap-btn-cancel:hover {
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(239, 68, 68, 0.5);
        }

        /* Dark mode uyumluluğu */
        body.dark-mode .swap-modal-content {
            background: #1e293b;
            color: #f1f5f9;
        }

        body.dark-mode .swap-modal-title {
            color: #f1f5f9;
        }

        body.dark-mode .swap-modal-body {
            background: #0f172a;
        }

        /* Dark Mode Scroll Bar */
        body.dark-mode .swap-modal-body::-webkit-scrollbar-track {
            background: #1e293b;
        }

        body.dark-mode .swap-modal-body::-webkit-scrollbar-thumb {
            background: linear-gradient(135deg, #818cf8 0%, #a78bfa 100%);
        }

        body.dark-mode .swap-modal-body::-webkit-scrollbar-thumb:hover {
            background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
        }

        body.dark-mode .swap-info {
            background: #1e293b;
            color: #e2e8f0;
        }

        /* ============================================
           📺 SADECE TABLO İÇİN SCALE
           Sayfa normal, sadece tablo küçülür
           Print/PDF'de scale YOK
           ============================================ */
        @media screen {
            #weeklyPrintTable {
                transform: scale(0.887);
                transform-origin: top left;
                margin-bottom: 20px;
            }
        }

        @media print {
            #weeklyPrintTable {
                transform: none !important;
            }
        }

        /* YAZDIRMA STİLLERİ */
        @media print {
            /* HER ŞEYİ GİZLE */
            .header,
            .main-content > *:not(#weeklyScheduleSection),
            .button-grid,
            .list-section,
            .modal,
            #resultsSection,
            button {
                display: none !important;
            }

            /* 🔍 ARAMA KUTUSUNU GİZLE - GÜÇLÜ YÖNTEM */
            #searchBoxContainer,
            #teacherSearchBoxContainer {
                display: none !important;
                visibility: hidden !important;
                height: 0 !important;
                overflow: hidden !important;
            }

            html, body {
                background: white !important;
                margin: 0 !important;
                padding: 0 !important;
                width: 100% !important;
                height: 100% !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }

            .container {
                box-shadow: none !important;
                border-radius: 0 !important;
                padding: 0 !important;
                margin: 0 !important;
                background: white !important;
            }

            /* HAFTALIK BÖLÜM - DAHA GENİŞ */
            #weeklyScheduleSection {
                display: block !important;
                margin: 0 !important;
                padding: 5mm 6mm !important;  /* ← Dar kenar boşluğu */
                width: 100% !important;
                box-sizing: border-box !important;
            }

            /* İÇ CONTAINER */
            #weeklyScheduleSection > div {
                padding: 0 !important;
                background: white !important;
                border-radius: 0 !important;
            }

            /* KONTROL BUTONLARINI GİZLE */
            #weeklyScheduleSection > div > div:first-child {
                display: none !important;
            }

            /* TABLO CONTAINER */
            #weeklyScheduleTable {
                overflow: visible !important;
                width: 100% !important;
            }

            /* TABLO - YÜKSEK VE GENİŞ */
            #weeklyScheduleSection table {
                width: 100% !important;
                table-layout: fixed !important;
                font-size: 8px !important;  /* ← Biraz büyüttük */
                margin: 0 !important;
                border: 1px solid #e5e7eb !important;
                border-collapse: separate !important;
                border-spacing: 0 !important;
                border-radius: 12px !important;
                overflow: hidden !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
                color-adjust: exact !important;
            }

            /* CAPTION - RENKLER KORUNSUN */
            #weeklyScheduleSection table caption {
                font-size: 14px !important;
                padding: 8px !important;
                border-radius: 12px 12px 0 0 !important;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
                color: white !important;
                font-weight: bold !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }

            /* BAŞLIK SATIRI - RENKLER */
            #weeklyScheduleSection table thead th {
                font-size: 9px !important;
                padding: 6px 3px !important;
                line-height: 1.2 !important;
                white-space: normal !important;
                word-wrap: break-word !important;
                border-right: 1px solid rgba(255,255,255,0.2) !important;
                background: linear-gradient(135deg, #4472C4 0%, #5B9BD5 100%) !important;
                color: white !important;
                font-weight: bold !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }

            /* SON KOLON */
            #weeklyScheduleSection table thead th:last-child {
                border-right: 1px solid #5B9BD5 !important;
            }

            /* TBODY SAĞ KENAR */
            #weeklyScheduleSection table tbody tr td:last-child {
                border-right: 1px solid #e5e7eb !important;
            }

            /* GÜN BAŞLIKLARI - MOR RENK */
            #weeklyScheduleSection table tbody td[colspan] {
                font-size: 8px !important;          /* ← 8px'ten 7px'e düşür */
                padding: 2px !important;            /* ← 3px'ten 2px'e düşür */
                line-height: 1.1 !important;        /* ← 1.2'den 1.1'e düşür */
                border-right: 1px solid #7E57C2 !important;
                background: #9575CD !important;
                color: white !important;
                font-weight: bold !important;
                text-align: center !important;
                text-transform: uppercase !important;  /* ← BU SATIRIA EKLE */
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
                color-adjust: exact !important;
            }

            /* SAAT HÜCRELERİ - MAVİ ARKA PLAN */
            #weeklyScheduleSection table tbody tr td:first-child {
                font-size: 8px !important;          /* ← 9px'ten 7px'e düşür */
                padding: 4px !important;            /* ← 4px'ten 2px'e düşür */
                line-height: 1.1 !important;        /* ← EKLE */
                white-space: nowrap !important;
                background: #E3F2FD !important;
                font-weight: 600 !important;
                color: #1565C0 !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }

            /* ÖĞRENCİ HÜCRELERİ */
            #weeklyScheduleSection table tbody tr td {
                font-size: 8px !important;          /* ← 8px'ten 7px'e düşür */
                padding: 4px 2px !important;        /* ← 4px 2px'ten 2px 2px'e */
                line-height: 1.1 !important;        /* ← 1.3'ten 1.1'e düşür */
                white-space: nowrap !important;
                overflow: hidden !important;
                text-overflow: ellipsis !important;
                border: 1px solid #e5e7eb !important;
            }

            /* SON SATIR - ALT BORDER-RADIUS */
            #weeklyScheduleSection table tbody tr:last-child td {
                border-bottom: 1px solid #e5e7eb !important;
            }

            #weeklyScheduleSection table tbody tr:last-child td:first-child {
                border-bottom-left-radius: 11px !important;
                background: #E3F2FD !important;
                -webkit-print-color-adjust: exact !important;
            }

            #weeklyScheduleSection table tbody tr:last-child td:last-child {
                border-bottom-right-radius: 11px !important;
            }

            /* KOLON GENİŞLİKLERİ - SAAT KOLONU DAR */
            #weeklyScheduleSection table colgroup col:first-child {
                width: 70px !important;  /* ← Daha dar */
            }

            #weeklyScheduleSection table colgroup col {
                width: auto !important;
            }

            /* TÜM ELEMENTLERDE RENK KORUMA */
            #weeklyScheduleSection *,
            #weeklyScheduleSection table *,
            table *,
            thead *,
            tbody *,
            tr *,
            td *,
            th * {
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
                color-adjust: exact !important;
            }

            /* SAYFA AYARLARI */
            @page {
                size: A4 landscape;
                margin: 5mm !important;  /* ← Minimal margin */
            }

            /* SAYFA KIRIMLARI ENGELLE */
            #weeklyScheduleSection,
            #weeklyScheduleTable,
            #weeklyScheduleTable table,
            table,
            thead,
            tbody {
                page-break-inside: avoid !important;
                page-break-after: avoid !important;
                page-break-before: avoid !important;
                break-inside: avoid-page !important;
            }

            /* HER SATIR */
            table tbody tr {
                page-break-inside: avoid !important;
                break-inside: avoid !important;
            }
        }

        /* 🆕 ÇAKIŞMA DASHBOARD STİLLERİ */
        .conflict-item {
            background: white;
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 15px;
            border-left: 5px solid #ef4444;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            transition: all 0.3s;
        }

        .conflict-item:hover {
            transform: translateX(5px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }

        .conflict-item.severity-critical {
            border-left-color: #dc2626;
            background: linear-gradient(90deg, #fee2e2 0%, #ffffff 100%);
        }

        .conflict-item.severity-high {
            border-left-color: #f59e0b;
            background: linear-gradient(90deg, #fef3c7 0%, #ffffff 100%);
        }

        .conflict-item.severity-medium {
            border-left-color: #3b82f6;
            background: linear-gradient(90deg, #dbeafe 0%, #ffffff 100%);
        }

        .conflict-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }

        .conflict-type-badge {
            display: inline-block;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 0.85em;
            font-weight: bold;
            text-transform: uppercase;
        }

        .conflict-type-badge.student {
            background: #ef4444;
            color: white;
        }

        .conflict-type-badge.teacher {
            background: #dc2626;
            color: white;
        }

        .conflict-type-badge.restriction {
            background: #f59e0b;
            color: white;
        }

        .conflict-type-badge.teacher_block {
            background: #8b5cf6;
            color: white;
        }

        .conflict-details {
            color: #4b5563;
            line-height: 1.8;
            margin-top: 10px;
        }

        .conflict-actions {
            display: flex;
            gap: 10px;
            margin-top: 15px;
        }

        .conflict-btn {
            padding: 8px 16px;
            border: none;
            border-radius: 8px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
            font-size: 0.9em;
        }

        .conflict-btn-fix {
            background: #10b981;
            color: white;
        }

        .conflict-btn-fix:hover {
            background: #059669;
        }

        .conflict-btn-timeline {
            background: #667eea;
            color: white;
        }

        .conflict-btn-timeline:hover {
            background: #5568d3;
        }

        .conflict-btn-suggest {
            background: #3b82f6;
            color: white;
        }

        .conflict-btn-suggest:hover {
            background: #2563eb;
        }

        /* 🆕 MODAL İÇİ BÜYÜK BUTONLAR İÇİN HOVER EFEKTİ */
        #conflictStats button {
            transition: all 0.3s ease;
            position: relative;
            overflow: hidden;
        }

        #conflictStats button:hover {
            transform: translateY(-3px);
            box-shadow: 0 8px 20px rgba(0,0,0,0.3);
        }

        #conflictStats button:active {
            transform: translateY(-1px);
        }

        /* 💫 RIPPLE EFEKTİ */
        #conflictStats button::before {
            content: '';
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.3);
            transform: translate(-50%, -50%);
            transition: width 0.5s, height 0.5s;
        }

        #conflictStats button:active::before {
            width: 300px;
            height: 300px;
        }

        /* 💥 HAFTALIK PROGRAM BUTONLARI İÇİN RIPPLE EFEKTİ */
        #prevWeekBtn::before, #nextWeekBtn::before,
        button[onclick="printWeeklyTable()"]::before,
        button[onclick="exportWeeklyToPDF()"]::before,
        button[onclick="exportAllWeeksToPDF()"]::before {
            content: '';
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.5);
            transform: translate(-50%, -50%);
            transition: width 0.6s ease, height 0.6s ease;
            pointer-events: none;
        }

        #prevWeekBtn:active::before, #nextWeekBtn:active::before,
        button[onclick="printWeeklyTable()"]:active::before,
        button[onclick="exportWeeklyToPDF()"]:active::before,
        button[onclick="exportAllWeeksToPDF()"]:active::before {
            width: 300px;
            height: 300px;
        }

        /* 💥 UYGULA VE SIFIRLA BUTONLARI İÇİN RIPPLE EFEKTİ */
        button[onclick="applyStudentFilter()"]::before,
        button[onclick="applyBranchFilter()"]::before,
        button[onclick="resetStudentFilter()"]::before,
        button[onclick="resetBranchFilter()"]::before {
            content: '';
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.5);
            transform: translate(-50%, -50%);
            transition: width 0.6s ease, height 0.6s ease;
            pointer-events: none;
        }

        button[onclick="applyStudentFilter()"]:active::before,
        button[onclick="applyBranchFilter()"]:active::before,
        button[onclick="resetStudentFilter()"]:active::before,
        button[onclick="resetBranchFilter()"]:active::before {
            width: 300px;
            height: 300px;
        }

    </style>

    <!-- PDF Export Kütüphaneleri -->
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/1.5.3/jspdf.min.js"></script>

<!-- 🎨 GÖRSEL İYİLEŞTİRME KÜTÜPHANELERİ -->
<!-- Font Awesome İkonlar -->
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">

<!-- Chart.js Grafikler -->
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.js"></script>

<!-- Google Fonts - Daha Modern Yazı Tipi -->
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">

</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Özel Ders Programı Yönetim Sistemi</h1>
            <p>Modern ve Akıllı Ders Programı Oluşturma Platformu</p>

            <!-- 🌙 DARK MODE TOGGLE -->
            <div style="position: absolute; top: 20px; right: 30px;">
                <button id="darkModeToggle" onclick="toggleDarkMode()" style="background: rgba(255,255,255,0.2); border: 2px solid rgba(255,255,255,0.3); color: white; padding: 12px 20px; border-radius: 50px; cursor: pointer; font-weight: 700; font-size: 1em; transition: all 0.3s; backdrop-filter: blur(10px); display: flex; align-items: center; gap: 8px;">
                    <i class="fas fa-moon" id="darkModeIcon"></i>
                    <span id="darkModeText">Koyu Tema</span>
                </button>
            </div>
        </div>
        <div class="main-content">
            <div class="success-message" id="successMessage"></div>
            <div class="error-message" id="errorMessage"></div>
            <div class="button-grid">
                <button class="main-btn" onclick="openTeacherModal()">
                    <i class="fas fa-chalkboard-teacher"></i> Öğretmen Ekle
                </button>
                <button class="main-btn" onclick="openStudentModal()">
                    <i class="fas fa-user-graduate"></i> Öğrenci Ekle
                </button>
                <button class="main-btn" onclick="openGenerateScheduleModal()">
                    <i class="fas fa-calendar-alt"></i> Yeni Program Oluştur
                </button>
                <button class="main-btn" onclick="openClassLessonModal()" style="background: linear-gradient(135deg, #10b981 0%, #059669 100%);">
                    <i class="fas fa-users-class"></i> Sınıf Dersi Ata
                </button>
                <button class="main-btn" onclick="openSaveScheduleModal()" style="background: linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%);">
                    <i class="fas fa-save"></i> Programı Kaydet
                </button>
                <button class="main-btn" onclick="openSavedSchedulesModal()" style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);">
                    <i class="fas fa-history"></i> Geçmiş Programlar
                </button>
                <button class="main-btn" onclick="exportToExcel()">
                    <i class="fas fa-file-excel"></i> Excel İndir
                </button>
                <button class="main-btn" onclick="exportToHTML()">
                    <i class="fas fa-file-code"></i> HTML İndir
                </button>

                <!-- 🆕 YENİ İHLAL KONTROL BUTONU -->
                <button class="main-btn" onclick="openConflictPanelV2()" style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); position: relative; overflow: visible;">
                    <i class="fas fa-exclamation-triangle"></i> İhlal Kontrolü
                    <span id="conflictBadge" style="display: none; position: absolute; top: -10px; right: -10px; background: #fbbf24; color: #000; padding: 4px 8px; border-radius: 12px; font-size: 0.7em; font-weight: bold; min-width: auto; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.3); z-index: 10; border: 2px solid white; white-space: nowrap;">0</span>
                </button>
            </div>

            <!-- 📚 SINIF DERSLERİ BÖLÜMÜ -->
            <div class="list-section" id="classLessonsSection">
                <div class="accordion-header" onclick="toggleMainClassLessons()">
                    <span>
                        <i class="fas fa-users-class"></i> Sınıf Dersleri
                        <span class="badge badge-info" id="classLessonCount">0</span>
                    </span>
                    <span class="accordion-arrow" id="mainClassLessonsArrow">▼</span>
                </div>
                <div class="accordion-content" id="classLessonsContent">
                    <div id="classLessonsList">
                        <div style="text-align: center; padding: 40px; color: #999;">
                            <i class="fas fa-users-class" style="font-size: 3em; margin-bottom: 15px; opacity: 0.3;"></i>
                            <p style="font-size: 1.1em;">Henüz sınıf dersi eklenmedi.</p>
                            <p style="font-size: 0.9em; margin-top: 8px;">Sınıf dersi eklemek için yukarıdaki <strong>"Sınıf Dersi Ata"</strong> butonuna tıklayın.</p>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Öğretmenler Accordion -->
            <div class="list-section">
                <div class="accordion-header" onclick="toggleAccordion('teachers')">
                    <span>
                        <i class="fas fa-chalkboard-teacher"></i> Kayıtlı Öğretmenler
                        <span class="badge badge-info" id="teacherCount">0</span>
                    </span>
                    <span class="accordion-arrow" id="teachersArrow">▼</span>
                </div>
                <div class="accordion-content" id="teachersContent">
                    <!-- 🔍 ARAMA KUTUSU -->
                    <div style="margin-bottom: 20px; padding: 15px; background: #f8fafc; border-radius: 10px;">
                        <div style="position: relative;">
                            <i class="fas fa-chalkboard-teacher" style="position: absolute; left: 15px; top: 50%; transform: translateY(-50%); color: #667eea; font-size: 1.1em;"></i>
                            <input
                                type="text"
                                id="teacherSearchInput"
                                placeholder="🔍 Öğretmen ara (Ad, Soyad, Branş)..."
                                onkeyup="filterTeachers()"
                                style="width: 100%; padding: 12px 12px 12px 45px; border: 2px solid #e5e7eb; border-radius: 8px; font-size: 1em; transition: all 0.3s;"
                                onfocus="this.style.borderColor='#667eea'; this.style.boxShadow='0 0 0 3px rgba(102,126,234,0.1)'"
                                onblur="this.style.borderColor='#e5e7eb'; this.style.boxShadow='none'"
                            />
                        </div>
                        <div id="teacherSearchCount" style="margin-top: 8px; font-size: 0.9em; color: #6b7280;"></div>
                    </div>
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Ad Soyad</th>
                                <th>Branş</th>
                                <th>Program</th>
                                <th>🚫 Bloklama</th>
                                <th>İşlemler</th>
                            </tr>
                        </thead>
                        <tbody id="teacherTableBody">
                            <tr>
                                <td colspan="5" style="text-align: center; padding: 20px; color: #999;">
                                    Henüz öğretmen eklenmedi.
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- Öğrenciler Accordion -->
            <div class="list-section">
                <div class="accordion-header" onclick="toggleAccordion('students')">
                    <span>
                        <i class="fas fa-user-graduate"></i> Kayıtlı Öğrenciler
                        <span class="badge badge-info" id="studentCount">0</span>
                    </span>
                    <span class="accordion-arrow" id="studentsArrow">▼</span>
                </div>
                <div class="accordion-content" id="studentsContent">
                    <!-- 🔍 ARAMA KUTUSU -->
                    <div style="margin-bottom: 20px; padding: 15px; background: #f8fafc; border-radius: 10px;">
                        <div style="position: relative;">
                            <i class="fas fa-user-graduate" style="position: absolute; left: 15px; top: 50%; transform: translateY(-50%); color: #667eea; font-size: 1.1em;"></i>
                            <input
                                type="text"
                                id="mainStudentSearchInput"
                                placeholder="🔍 Öğrenci ara (Ad, Soyad, Sınıf)..."
                                onkeyup="filterStudentsTable()"
                                style="width: 100%; padding: 12px 12px 12px 45px; border: 2px solid #e5e7eb; border-radius: 8px; font-size: 1em; transition: all 0.3s;"
                                onfocus="this.style.borderColor='#667eea'; this.style.boxShadow='0 0 0 3px rgba(102,126,234,0.1)'"
                                onblur="this.style.borderColor='#e5e7eb'; this.style.boxShadow='none'"
                            />
                        </div>
                        <div id="studentSearchCount" style="margin-top: 8px; font-size: 0.9em; color: #6b7280;"></div>
                    </div>
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Ad Soyad</th>
                                <th>Sınıf</th>
                                <th>Kısıtlama</th>
                                <th>Öncelik</th>
                                <th>Manuel Ders</th>
                                <th>🚫 Öğretmen Engeli</th>
                                <th>İşlemler</th>
                            </tr>
                        </thead>
                        <tbody id="studentTableBody">
                            <tr>
                                <td colspan="6" style="text-align: center; padding: 20px; color: #999;">
                                    Henüz öğrenci eklenmedi.
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- 📅 BUGÜNÜN DERSLERİ WIDGET -->
            <div class="accordion-item" style="margin-bottom: 20px;">
                <div class="accordion-header" onclick="toggleAccordion('todayLessons')"
                     style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); cursor: pointer; display: flex; justify-content: space-between; align-items: center;">
                    <span style="display: flex; align-items: center; gap: 10px;">
                        <i class="fas fa-calendar-day"></i> Günlük Dersler
                        <span class="badge badge-success" id="todayLessonsCount" style="background: rgba(255,255,255,0.3);">0</span>
                    </span>
                    <span class="accordion-arrow" id="todayLessonsArrow">▼</span>
                </div>
                <div class="accordion-content" id="todayLessonsContent">
                    <!-- Gün Seçici Kontroller -->
                    <div id="daySelector" style="display: none; background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); padding: 15px; border-radius: 12px; margin-bottom: 20px; border: 2px solid #10b981;">
                        <div style="display: flex; gap: 10px; align-items: center; justify-content: center; flex-wrap: wrap;">
                            <button id="prevDayBtn" onclick="changeDayView(-1)"
                                style="background: white; color: #10b981; border: 2px solid #10b981; padding: 10px 15px; border-radius: 8px; cursor: pointer; font-weight: 600; transition: all 0.2s; display: flex; align-items: center; gap: 5px;"
                                onmouseover="this.style.background='#10b981'; this.style.color='white'"
                                onmouseout="this.style.background='white'; this.style.color='#10b981'">
                                ◄ Önceki
                            </button>

                            <select id="dayDropdown" onchange="selectDayByDate()"
                                style="padding: 10px 20px; border: 2px solid #10b981; border-radius: 8px; font-size: 1em; font-weight: 600; background: white; color: #065f46; cursor: pointer; min-width: 250px;">
                                <!-- JavaScript ile doldurulacak -->
                            </select>

                            <button id="nextDayBtn" onclick="changeDayView(1)"
                                style="background: white; color: #10b981; border: 2px solid #10b981; padding: 10px 15px; border-radius: 8px; cursor: pointer; font-weight: 600; transition: all 0.2s; display: flex; align-items: center; gap: 5px;"
                                onmouseover="this.style.background='#10b981'; this.style.color='white'"
                                onmouseout="this.style.background='white'; this.style.color='#10b981'">
                                Sonraki ►
                            </button>

                            <button id="todayBtn" onclick="goToToday()"
                                style="background: #10b981; color: white; border: 2px solid #10b981; padding: 10px 20px; border-radius: 8px; cursor: pointer; font-weight: 600; transition: all 0.2s; display: flex; align-items: center; gap: 5px;"
                                onmouseover="this.style.background='#059669'"
                                onmouseout="this.style.background='#10b981'">
                                🔄 Bugün
                            </button>
                        </div>
                    </div>

                    <div id="todayLessonsWidget" style="max-height: 600px; overflow-y: auto; padding-right: 10px;">
                        <!-- JavaScript ile doldurulacak -->
                        <div style="text-align: center; padding: 40px; color: #6b7280;">
                            <i class="fas fa-calendar-day" style="font-size: 3em; margin-bottom: 15px; opacity: 0.3;"></i>
                            <p>Kaydedilmiş program bulunamadı.</p>
                            <p style="font-size: 0.9em; opacity: 0.7;">Lütfen önce bir program oluşturun ve kaydedin.</p>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Detay Modal -->
            <div class="detail-modal" id="detailModal">
                <div class="detail-modal-content">
                    <span class="detail-close" onclick="closeDetailModal()">&times;</span>
                    <div id="detailContent"></div>
                </div>
            </div>

            <!-- ⚠️ BUGÜN BUTONU UYARI MODALI -->
            <div class="detail-modal" id="todayWarningModal">
                <div class="detail-modal-content" style="max-width: 500px;">
                    <span class="detail-close" onclick="closeTodayWarningModal()">&times;</span>
                    <div id="todayWarningContent" style="padding: 20px; text-align: center;">
                        <!-- JavaScript ile doldurulacak -->
                    </div>
                </div>
            </div>

            <!-- 📅 HAFTALIK PROGRAM GÖRÜNTÜLEYICI -->
            <div id="weeklyScheduleSection" style="display: none; margin-top: 30px;">
                <div style="background: white; border-radius: 15px; padding: 30px; box-shadow: 0 4px 20px rgba(0,0,0,0.1);">
                    <!-- Başlık ve Butonlar -->
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; flex-wrap: wrap; gap: 10px;">
                        <h2 style="font-size: 1.8em; color: #667eea; margin: 0;">📅 Ders Programı</h2>
                        <div style="display: flex; gap: 8px; align-items: center; flex-wrap: nowrap;">
                            <!-- Görünüm Seçici -->
                            <button id="tableViewBtn" onclick="switchView('table')"
                                style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border: none; padding: 8px 16px; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 0.9em; transition: all 0.3s; box-shadow: 0 2px 8px rgba(102,126,234,0.3); white-space: nowrap; height: 48px;">
                                📋 Tablo
                            </button>
                            <button id="calendarViewBtn" onclick="switchView('calendar')"
                                style="background: white; color: #6b7280; border: 2px solid #e5e7eb; padding: 8px 16px; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 0.9em; transition: all 0.3s; white-space: nowrap; height: 48px;">
                                📅 Takvim
                            </button>

                            <!-- Ayırıcı -->
                            <div style="width: 1px; height: 35px; background: #e5e7eb; margin: 0 3px;"></div>

                            <!-- Hafta Butonları -->
                            <button id="prevWeekBtn" onclick="changeWeek(-1)"
                                onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 6px 20px rgba(102,126,234,0.5)'"
                                onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 10px rgba(0,0,0,0.1)'"
                                style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border: none; padding: 12px 12px; border-radius: 10px; font-weight: 600; cursor: pointer; font-size: 0.85em; width: 135px; height: 48px; transition: all 0.3s; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">
                                ◄ Önceki Hafta
                            </button>
                            <button id="nextWeekBtn" onclick="changeWeek(1)"
                                onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 6px 20px rgba(102,126,234,0.5)'"
                                onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 10px rgba(0,0,0,0.1)'"
                                style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border: none; padding: 12px 12px; border-radius: 10px; font-weight: 600; cursor: pointer; font-size: 0.85em; width: 135px; height: 48px; transition: all 0.3s; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">
                                Sonraki Hafta ►
                            </button>
                            <button onclick="printWeeklyTable()"
                                onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 6px 20px rgba(76,175,80,0.5)'"
                                onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 10px rgba(0,0,0,0.1)'"
                                style="background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); color: white; border: none; padding: 12px 16px; border-radius: 10px; font-weight: 600; cursor: pointer; font-size: 0.9em; min-width: 100px; height: 48px; white-space: nowrap; transition: all 0.3s;">
                                🖨️ Yazdır
                            </button>
                            <button onclick="window.location.href='/export_weekly_pdf_server/'+currentWeekView"
                                onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 6px 20px rgba(220,38,38,0.5)'"
                                onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 10px rgba(0,0,0,0.1)'"
                                style="background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%); color: white; border: none; padding: 12px 16px; border-radius: 10px; font-weight: 600; cursor: pointer; font-size: 0.9em; min-width: 100px; height: 48px; white-space: nowrap; transition: all 0.3s;">
                                📄 PDF İndir
                            </button>
                            <button onclick="window.location.href='/export_all_weeks_pdf_server'"
                                onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 6px 20px rgba(245,158,11,0.5)'"
                                onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 10px rgba(0,0,0,0.1)'"
                                style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); color: white; border: none; padding: 12px 16px; border-radius: 10px; font-weight: 600; cursor: pointer; font-size: 0.9em; min-width: 120px; height: 48px; white-space: nowrap; transition: all 0.3s;">
                                📊 4 Hafta
                            </button>
                        </div>
                    </div>

                    <!-- 🔍 ARAMA KUTUSU -->
                    <div id="searchBoxContainer" style="background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%); border-radius: 12px; padding: 20px; margin-bottom: 20px; border: 2px solid #0ea5e9;">
                        <div style="display: flex; align-items: center; gap: 15px; flex-wrap: wrap;">
                            <div style="flex: 1; min-width: 250px;">
                                <label style="display: block; font-weight: 600; color: #0c4a6e; margin-bottom: 8px; font-size: 0.95em;">
                                    🔍👨‍🎓 Öğrenci Ara:
                                </label>
                                <input
                                    type="text"
                                    id="studentSearchBox"
                                    placeholder="Öğrenci adı yazın..."
                                    oninput="searchStudentInTable()"
                                    style="width: 100%; padding: 12px 15px; border: 2px solid #0ea5e9; border-radius: 10px; font-size: 1em; transition: all 0.3s;"
                                    onfocus="this.style.borderColor='#0284c7'; this.style.boxShadow='0 0 0 3px rgba(14,165,233,0.1)'"
                                    onblur="this.style.borderColor='#0ea5e9'; this.style.boxShadow='none'"
                                >
                            </div>
                            <button
                                onclick="clearStudentSearch()"
                                style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); color: white; border: none; padding: 12px 24px; border-radius: 10px; font-weight: 600; cursor: pointer; transition: all 0.3s; margin-top: 24px;"
                                onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 4px 12px rgba(239,68,68,0.4)'"
                                onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='none'"
                            >
                                🗑️ Temizle
                            </button>
                        </div>
                        <div id="searchResultInfo" style="margin-top: 12px; font-size: 0.9em; color: #0c4a6e; font-weight: 500;"></div>
                    </div>

                    <!-- 👨‍🏫 ÖĞRETMEN ARAMA KUTUSU -->
                    <div id="teacherSearchBoxContainer" style="background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); border-radius: 12px; padding: 20px; margin-bottom: 20px; border: 2px solid #10b981;">
                        <div style="display: flex; align-items: center; gap: 15px; flex-wrap: wrap;">
                            <div style="flex: 1; min-width: 250px;">
                                <label style="display: block; font-weight: 600; color: #065f46; margin-bottom: 8px; font-size: 0.95em;">
                                    🔍👨‍🏫 Öğretmen Ara:
                                </label>
                                <input
                                    type="text"
                                    id="teacherSearchBox"
                                    placeholder="Öğretmen adı yazın..."
                                    oninput="searchTeacherInTable()"
                                    style="width: 100%; padding: 12px 15px; border: 2px solid #10b981; border-radius: 10px; font-size: 1em; transition: all 0.3s;"
                                    onfocus="this.style.borderColor='#059669'; this.style.boxShadow='0 0 0 3px rgba(16,185,129,0.1)'"
                                    onblur="this.style.borderColor='#10b981'; this.style.boxShadow='none'"
                                >
                            </div>
                            <button
                                onclick="clearTeacherSearch()"
                                style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); color: white; border: none; padding: 12px 24px; border-radius: 10px; font-weight: 600; cursor: pointer; transition: all 0.3s; margin-top: 24px;"
                                onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 4px 12px rgba(239,68,68,0.4)'"
                                onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='none'"
                            >
                                🗑️ Temizle
                            </button>
                        </div>
                        <div id="teacherSearchResultInfo" style="margin-top: 12px; font-size: 0.9em; color: #065f46; font-weight: 500;"></div>
                    </div>

                    <!-- Tablo -->
                    <div id="weeklyScheduleTable" style="overflow-x: hidden;"></div>

                    <!-- TAKVİM GÖRÜNÜMÜ -->
                    <div id="calendarView" style="display: none;">
                        <!-- Ay Başlığı (Butonlar üstte artık) -->
                        <div style="display: flex; justify-content: center; align-items: center; margin-bottom: 30px;">
                            <h3 id="calendarMonthTitle" style="font-size: 1.5em; color: #667eea; margin: 0; text-align: center;">
                                ARALIK 2025
                            </h3>
                        </div>

                        <!-- Takvim Grid -->
                        <div id="calendarGrid" style="background: white; border-radius: 15px; padding: 20px; box-shadow: 0 4px 20px rgba(0,0,0,0.1);">
                            <!-- JavaScript ile doldurulacak -->
                        </div>
                    </div>
                </div>
            </div>

            <!-- GÜN DETAY MODALI -->
            <div id="dayDetailModal" class="detail-modal" style="display: none;">
                <div class="detail-modal-content" style="max-width: 800px;">
                    <span class="detail-close" onclick="closeDayDetailModal()">&times;</span>
                    <div id="dayDetailContent"></div>
                </div>
            </div>

            <!-- 📊 İSTATİSTİK KARTLARI -->
            <div id="statsCardsSection" style="display: none; margin-bottom: 30px;">
                <h2 style="text-align: center; font-size: 1.8em; color: var(--text-primary); margin-bottom: 20px;">
                    <i class="fas fa-chart-line"></i> Genel İstatistikler
                </h2>
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px;">
                    <div class="stat-card">
                        <div class="stat-icon">👨‍🏫</div>
                        <div class="stat-value" id="totalTeachers">0</div>
                        <div class="stat-label">Toplam Öğretmen</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-icon">👨‍🎓</div>
                        <div class="stat-value" id="totalStudents">0</div>
                        <div class="stat-label">Toplam Öğrenci</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-icon">📚</div>
                        <div class="stat-value" id="totalLessons">0</div>
                        <div class="stat-label">Toplam Ders</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-icon">📅</div>
                        <div class="stat-value">4</div>
                        <div class="stat-label">Hafta</div>
                    </div>
                </div>

                <!-- 📊 GRAFİKLER -->
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(400px, 1fr)); gap: 30px;">
                    <!-- Branş Dağılımı Grafiği -->
                    <div style="background: white; border-radius: var(--border-radius-md); padding: 25px; box-shadow: var(--shadow-md);">
                        <h3 style="margin-bottom: 20px; color: var(--text-primary); display: flex; align-items: center; gap: 10px;">
                            <i class="fas fa-chart-pie"></i> Branş Dağılımı
                        </h3>
                        <canvas id="branchChart" style="max-height: 300px;"></canvas>
                    </div>

                    <!-- Haftalık Ders Dağılımı -->
                    <div style="background: white; border-radius: var(--border-radius-md); padding: 25px; box-shadow: var(--shadow-md);">
                        <h3 style="margin-bottom: 20px; color: var(--text-primary); display: flex; align-items: center; gap: 10px;">
                            <i class="fas fa-chart-bar"></i> Haftalık Ders Dağılımı
                        </h3>
                        <canvas id="weeklyChart" style="max-height: 300px;"></canvas>
                    </div>
                </div>
            </div>

            <div id="resultsSection" class="results-container"></div>
        </div>
    </div>

    <div class="modal" id="teacherModal">
        <div class="modal-content">
            <span class="close-btn" onclick="closeTeacherModal()">&times;</span>
            <h2 id="teacherModalTitle">Yeni Öğretmen Ekle</h2>
            <form id="teacherForm" onsubmit="saveTeacher(event)">
                <input type="hidden" id="teacherId">
                <div class="form-group">
                    <label>Ad:</label>
                    <input type="text" id="teacherName" required>
                </div>
                <div class="form-group">
                    <label>Soyad:</label>
                    <input type="text" id="teacherSurname" required>
                </div>
                <div class="form-group">
                    <label>Branş:</label>
                    <select id="teacherBranch" required>
                        <option value="">Seçiniz...</option>
                        <option value="Matematik">Matematik</option>
                        <option value="Matematik-1">Matematik-1</option>
                        <option value="Matematik-2">Matematik-2</option>
                        <option value="Geometri">Geometri</option>
                        <option value="Türkçe">Türkçe</option>
                        <option value="Edebiyat">Edebiyat</option>
                        <option value="İngilizce">İngilizce</option>
                        <option value="Fen Bilgisi">Fen Bilgisi</option>
                        <option value="Fizik">Fizik</option>
                        <option value="Kimya">Kimya</option>
                        <option value="Biyoloji">Biyoloji</option>
                        <option value="Sosyal Bilgiler">Sosyal Bilgiler</option>
                        <option value="Tarih">Tarih</option>
                        <option value="Coğrafya">Coğrafya</option>
                        <option value="Felsefe">Felsefe</option>
                        <option value="Din Kültürü">Din Kültürü</option>
                    </select>
                </div>
                <div class="form-group">
                    <div class="accordion-header" onclick="toggleFormAccordion('teacherSchedule')" style="margin-bottom: 10px; cursor: pointer; background: linear-gradient(135deg, #10b981 0%, #059669 100%);">
                        <span>
                            📅 Ders Günleri ve Saatleri
                            <span class="badge badge-info" id="dayCountBadge" style="display: none; margin-left: 8px; background: rgba(255,255,255,0.3); color: white; font-size: 0.85em;">0</span>
                        </span>
                        <span class="accordion-arrow" id="teacherScheduleArrow">▼</span>
                    </div>
                    <div class="accordion-content" id="teacherScheduleContent" style="padding: 15px; border: 2px solid #10b981; border-radius: 10px; background: #f0fdf4;">
                        <div id="dayGroups"></div>
                        <button type="button" class="add-day-btn" onclick="addDayGroup()">+ Yeni Gün Ekle</button>
                    </div>
                </div>

                <!-- 🆕 SLOT BLOKLAMALARI -->
                <div class="form-group">
                    <div class="accordion-header" onclick="toggleFormAccordion('teacherBlocks')" style="margin-bottom: 10px; cursor: pointer; background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);">
                        <span>
                            🚫 Slot Bloklamaları (İsteğe Bağlı)
                            <span class="badge badge-info" id="blockCountBadge" style="display: none; margin-left: 8px; background: rgba(255,255,255,0.3); color: white; font-size: 0.85em;">0</span>
                        </span>
                        <span class="accordion-arrow" id="teacherBlocksArrow">▼</span>
                    </div>
                    <div class="accordion-content" id="teacherBlocksContent" style="padding: 15px; border: 2px solid #ef4444; border-radius: 10px; background: #fee2e2;">
                        <p style="font-size: 0.9em; color: #991b1b; margin-bottom: 15px; font-weight: 500;">
                            ℹ️ Belirtilen hafta/gün/saatlerde bu öğretmenin slotları program oluştururken kullanılmaz (tatil, toplantı vb. için)
                        </p>
                        <div id="teacherBlockGroups"></div>
                        <button type="button" class="add-day-btn" style="background: #ef4444; margin-top: 10px;" onclick="addTeacherBlock()">+ Bloklama Ekle</button>
                    </div>
                </div>

                <button type="submit" class="submit-btn">Kaydet</button>
            </form>
        </div>
    </div>

    <div class="modal" id="studentModal">
        <div class="modal-content">
            <span class="close-btn" onclick="closeStudentModal()">&times;</span>
            <h2 id="studentModalTitle">Yeni Öğrenci Ekle</h2>
            <form id="studentForm" onsubmit="saveStudent(event)">
                <input type="hidden" id="studentId">
                <div class="form-group">
                    <label>Ad:</label>
                    <input type="text" id="studentName" required>
                </div>
                <div class="form-group">
                    <label>Soyad:</label>
                    <input type="text" id="studentSurname" required>
                </div>
                <div class="form-group">
                    <label>Sınıf:</label>
                    <input type="text" id="studentClass" required placeholder="Örn: 5-A">
                </div>
                <div class="form-group">
                    <div class="accordion-header" onclick="toggleFormAccordion('restrictions')" style="margin-bottom: 10px; cursor: pointer; background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);">
                        <span>
                            🚫 Katılamayacağı Dersler (İsteğe Bağlı)
                            <span class="badge badge-info" id="restrictionCountBadge" style="display: none; margin-left: 8px; background: rgba(255,255,255,0.3); color: white; font-size: 0.85em;">0</span>
                        </span>
                        <span class="accordion-arrow" id="restrictionsArrow">▼</span>
                    </div>
                    <div class="accordion-content" id="restrictionsContent" style="padding: 15px; border: 2px solid #ef4444; border-radius: 10px; background: #fee2e2;">
                        <p style="font-size: 0.9em; color: #991b1b; margin-bottom: 15px; font-weight: 500;">
                            ℹ️ Öğrencinin belirtilen hafta/gün/saatlerde ders alamayacağı durumlar için kullanılır
                        </p>
                        <div id="restrictionGroups"></div>
                        <button type="button" class="add-day-btn" style="background: #ef4444; margin-top: 10px;" onclick="addRestriction()">+ Kısıtlama Ekle</button>
                    </div>
                </div>

                <!-- Haftalık Öncelikler -->
                <div class="form-group">
                    <div class="accordion-header" onclick="toggleFormAccordion('priorities')" style="margin-bottom: 10px; cursor: pointer;">
                        <span>
                            ⭐ Haftalık Ders Öncelikleri (İsteğe Bağlı)
                            <span class="badge badge-info" id="priorityCountBadge" style="display: none; margin-left: 8px; background: rgba(255,255,255,0.3); color: white; font-size: 0.85em;">0</span>
                        </span>
                        <span class="accordion-arrow" id="prioritiesArrow">▼</span>
                    </div>
                    <div class="accordion-content" id="prioritiesContent" style="padding: 15px; border: 2px solid #f59e0b; border-radius: 10px; background: #fffbeb;">
                        <div id="week1Priorities" class="priority-week-group">
                            <h4 style="color: #f59e0b; margin-bottom: 10px;">📅 Hafta 1</h4>
                            <div id="week1PriorityList"></div>
                            <button type="button" class="add-day-btn" style="background: #f59e0b; margin-top: 10px;" onclick="addPriority(1)">+ Öncelik Ekle</button>
                        </div>

                        <div id="week2Priorities" class="priority-week-group" style="margin-top: 20px;">
                            <h4 style="color: #f59e0b; margin-bottom: 10px;">📅 Hafta 2</h4>
                            <div id="week2PriorityList"></div>
                            <button type="button" class="add-day-btn" style="background: #f59e0b; margin-top: 10px;" onclick="addPriority(2)">+ Öncelik Ekle</button>
                        </div>

                        <div id="week3Priorities" class="priority-week-group" style="margin-top: 20px;">
                            <h4 style="color: #f59e0b; margin-bottom: 10px;">📅 Hafta 3</h4>
                            <div id="week3PriorityList"></div>
                            <button type="button" class="add-day-btn" style="background: #f59e0b; margin-top: 10px;" onclick="addPriority(3)">+ Öncelik Ekle</button>
                        </div>

                        <div id="week4Priorities" class="priority-week-group" style="margin-top: 20px;">
                            <h4 style="color: #f59e0b; margin-bottom: 10px;">📅 Hafta 4</h4>
                            <div id="week4PriorityList"></div>
                            <button type="button" class="add-day-btn" style="background: #f59e0b; margin-top: 10px;" onclick="addPriority(4)">+ Öncelik Ekle</button>
                        </div>
                    </div>
                </div>

                <!-- Manuel Ders Atamaları -->
                <div class="form-group">
                    <div class="accordion-header" onclick="toggleFormAccordion('manual')" style="margin-bottom: 10px; cursor: pointer;">
                        <span>
                            📌 Manuel Ders Atamaları (İsteğe Bağlı)
                            <span class="badge badge-info" id="manualCountBadge" style="display: none; margin-left: 8px; background: rgba(255,255,255,0.3); color: white; font-size: 0.85em;">0</span>
                        </span>
                        <span class="accordion-arrow" id="manualArrow">▼</span>
                    </div>
                    <div class="accordion-content" id="manualContent" style="padding: 15px; border: 2px solid #3b82f6; border-radius: 10px; background: #eff6ff;">
                        <div id="manualLessonGroups"></div>
                        <button type="button" class="add-day-btn" style="background: #3b82f6; margin-top: 10px;" onclick="addManualLesson()">+ Manuel Ders Ekle</button>
                    </div>
                </div>

                <!-- Öğretmen Engellemeleri -->
                <div class="form-group">
                    <div class="accordion-header" onclick="toggleFormAccordion('studentTeacherBlocks')" style="margin-bottom: 10px; cursor: pointer;">
                        <span>
                            🚫 Öğretmen Engellemeleri (İsteğe Bağlı)
                            <span class="badge badge-info" id="teacherBlockCountBadge" style="display: none; margin-left: 8px; background: rgba(255,255,255,0.3); color: white; font-size: 0.85em;">0</span>
                        </span>
                        <span class="accordion-arrow" id="studentTeacherBlocksArrow">▼</span>
                    </div>
                    <div class="accordion-content" id="studentTeacherBlocksContent" style="padding: 15px; border: 2px solid #dc2626; border-radius: 10px; background: #fef2f2;">
                        <p style="font-size: 0.9em; color: #991b1b; margin-bottom: 15px; font-weight: 500;">
                            ℹ️ Bu öğrencinin belirli öğretmenlerden ders almasını engelleyebilirsiniz
                        </p>
                        <div id="studentTeacherBlockGroups"></div>
                        <button type="button" class="add-day-btn" style="background: #dc2626; margin-top: 10px;" onclick="addStudentTeacherBlock()">+ Engelleme Ekle</button>
                    </div>
                </div>

                <button type="submit" class="submit-btn">Kaydet</button>
            </form>
        </div>
    </div>

    <!-- 💾 PROGRAM KAYDETME MODALI -->
    <!-- 📅 YENİ PROGRAM OLUŞTUR MODALI -->
    <div class="modal" id="generateScheduleModal">
        <div class="modal-content" style="max-width: 500px;">
            <span class="close-btn" onclick="closeGenerateScheduleModal()">&times;</span>
            <h2 style="color: #667eea; margin-bottom: 20px;">📅 Yeni Program Oluştur</h2>
            <div class="form-group">
                <label style="font-weight: 600; font-size: 1.1em;">📅 Program Başlangıç Tarihi (Pazartesi):</label>
                <input type="date" id="generateStartDate" style="width: 100%; padding: 14px; border: 2px solid #667eea; border-radius: 10px; font-size: 1.1em; margin-top: 10px;">
                <div style="margin-top: 12px; padding: 12px; background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); border-radius: 8px; border-left: 4px solid #3b82f6;">
                    <small style="display: block; color: #1e40af; font-size: 0.95em; line-height: 1.6;">
                        <i class="fas fa-info-circle"></i> <strong>Önemli:</strong> Lütfen bir <strong>Pazartesi</strong> günü seçin.<br>
                        Program bu tarihten başlayarak <strong>4 hafta</strong> boyunca devam edecektir.
                    </small>
                </div>
            </div>
            <button onclick="generateScheduleWithDate()" class="submit-btn" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); margin-top: 20px; font-size: 1.1em; padding: 14px;">
                <i class="fas fa-calendar-alt"></i> Program Oluştur
            </button>
        </div>
    </div>

    <!-- 📚 SINIF DERSİ ATA MODALI -->
    <div class="modal" id="classLessonModal">
        <div class="modal-content" style="max-width: 600px;">
            <span class="close-btn" onclick="closeClassLessonModal()">&times;</span>
            <h2 style="color: #10b981; margin-bottom: 20px;">📚 Sınıf Dersi Ata</h2>

            <div class="form-group">
                <label>1️⃣ Sınıf Seçin:</label>
                <select id="classLessonClass" onchange="updateClassStudentCount()" style="width: 100%; padding: 12px; border: 2px solid #10b981; border-radius: 10px; font-size: 1em;">
                    <option value="">Sınıf seçin...</option>
                </select>
                <small id="classStudentCount" style="display: block; margin-top: 8px; color: #6b7280; font-size: 0.9em;"></small>
            </div>

            <div class="form-group">
                <label>2️⃣ Öğretmen Seçin:</label>
                <select id="classLessonTeacher" onchange="updateTeacherSchedule()" style="width: 100%; padding: 12px; border: 2px solid #10b981; border-radius: 10px; font-size: 1em;">
                    <option value="">Öğretmen seçin...</option>
                </select>
                <small id="teacherBranchInfo" style="display: block; margin-top: 8px; color: #6b7280; font-size: 0.9em;"></small>
            </div>

            <div class="form-group" id="daySelectionGroup" style="display: none;">
                <label>3️⃣ Gün Seçin:</label>
                <div id="availableDays" style="display: flex; gap: 10px; flex-wrap: wrap; margin-top: 10px;">
                    <!-- Günler buraya gelecek -->
                </div>
            </div>

            <div class="form-group" id="timeSelectionGroup" style="display: none;">
                <label>4️⃣ Saat Seçin:</label>
                <div id="availableTimes" style="margin-top: 10px;">
                    <!-- Saatler buraya gelecek -->
                </div>
            </div>

            <div class="form-group" id="weekSelectionGroup" style="display: none;">
                <label>5️⃣ Hangi Haftalarda?</label>
                <div style="margin-top: 10px;">
                    <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; margin-bottom: 10px;">
                        <input type="checkbox" id="classLessonAllWeeks" onclick="toggleWeekSelection()" style="width: 20px; height: 20px; cursor: pointer;">
                        <span style="font-weight: 600;">Her hafta (1, 2, 3, 4)</span>
                    </label>
                    <div id="individualWeeks" style="display: flex; gap: 10px; margin-top: 10px;">
                        <label style="display: flex; align-items: center; gap: 6px; cursor: pointer;">
                            <input type="checkbox" class="week-checkbox" value="1" style="width: 18px; height: 18px; cursor: pointer;">
                            <span>Hafta 1</span>
                        </label>
                        <label style="display: flex; align-items: center; gap: 6px; cursor: pointer;">
                            <input type="checkbox" class="week-checkbox" value="2" style="width: 18px; height: 18px; cursor: pointer;">
                            <span>Hafta 2</span>
                        </label>
                        <label style="display: flex; align-items: center; gap: 6px; cursor: pointer;">
                            <input type="checkbox" class="week-checkbox" value="3" style="width: 18px; height: 18px; cursor: pointer;">
                            <span>Hafta 3</span>
                        </label>
                        <label style="display: flex; align-items: center; gap: 6px; cursor: pointer;">
                            <input type="checkbox" class="week-checkbox" value="4" style="width: 18px; height: 18px; cursor: pointer;">
                            <span>Hafta 4</span>
                        </label>
                    </div>
                </div>
            </div>

            <button onclick="saveClassLesson()" class="submit-btn" id="saveClassLessonBtn" style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); margin-top: 20px; display: none;">
                <i class="fas fa-save"></i> Sınıf Dersini Kaydet
            </button>
        </div>
    </div>

    <!-- 💾 PROGRAMI KAYDET MODALI -->
    <div class="modal" id="saveScheduleModal">
        <div class="modal-content" style="max-width: 500px;">
            <span class="close-btn" onclick="closeSaveScheduleModal()">&times;</span>
            <h2 style="color: #10b981; margin-bottom: 20px;">💾 Programı Kaydet</h2>
            <div class="form-group">
                <label>Program Adı:</label>
                <input type="text" id="scheduleName" placeholder="Örn: Ocak 2025 Programı" style="width: 100%; padding: 12px; border: 2px solid #10b981; border-radius: 10px; font-size: 1em;">
            </div>
            <button onclick="saveCurrentSchedule()" class="submit-btn" style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); margin-top: 20px;">
                💾 Kaydet
            </button>
        </div>
    </div>

    <!-- 📚 GEÇMİŞ PROGRAMLAR MODALI -->
    <div class="modal" id="savedSchedulesModal">
        <div class="modal-content" style="max-width: 900px; max-height: 85vh;">
            <span class="close-btn" onclick="closeSavedSchedulesModal()">&times;</span>
            <h2 style="color: #f59e0b; margin-bottom: 20px;">📚 Geçmiş Programlar</h2>
            <div id="savedSchedulesList" style="max-height: 60vh; overflow-y: auto;"></div>
        </div>
    </div>

    <!-- 🆕 ÇAKIŞMA DASHBOARD MODALI -->
        <div class="modal" id="conflictDashboardModal">
            <div class="modal-content" style="max-width: 1200px; max-height: 90vh;">
                <span class="close-btn" onclick="closeConflictDashboard()">&times;</span>

                <h2 style="color: #ef4444; margin-bottom: 20px; display: flex; align-items: center; gap: 10px;">
                    <i class="fas fa-exclamation-triangle"></i> İhlal Kontrol Paneli
                </h2>

                <!-- YÜKLEME GÖSTERGESI -->
                <div id="conflictLoading" style="text-align: center; padding: 40px; display: none;">
                    <div style="font-size: 3em; color: #667eea; margin-bottom: 20px;">
                        <i class="fas fa-spinner fa-spin"></i>
                    </div>
                    <p style="font-size: 1.2em; color: #6b7280;">Çakışmalar kontrol ediliyor...</p>
                </div>

                <!-- İSTATİSTİK KARTLARI -->
                <div id="conflictStats" style="display: none; margin-bottom: 30px;">
                    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px;">
                        <div class="stat-card" style="border-left-color: #ef4444;">
                            <div class="stat-icon" style="color: #ef4444;">🔴</div>
                            <div class="stat-value" id="criticalCount">0</div>
                            <div class="stat-label">Kritik</div>
                        </div>
                        <div class="stat-card" style="border-left-color: #f59e0b;">
                            <div class="stat-icon" style="color: #f59e0b;">⚠️</div>
                            <div class="stat-value" id="highCount">0</div>
                            <div class="stat-label">Yüksek</div>
                        </div>
                        <div class="stat-card" style="border-left-color: #3b82f6;">
                            <div class="stat-icon" style="color: #3b82f6;">ℹ️</div>
                            <div class="stat-value" id="mediumCount">0</div>
                            <div class="stat-label">Orta</div>
                        </div>
                        <div class="stat-card" style="border-left-color: #10b981;">
                            <div class="stat-icon" style="color: #10b981;">📊</div>
                            <div class="stat-value" id="totalConflictCount">0</div>
                            <div class="stat-label">Toplam</div>
                        </div>
                    </div>

                    <!-- 🆕 ONAYLANMIŞ SINIF GRUP DERSLERİ -->
                    <div id="groupLessonsSection" style="display: none; background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); border-left: 4px solid #10b981; border-radius: 12px; padding: 20px; margin-bottom: 20px;">
                        <h3 style="margin: 0 0 15px 0; color: #059669; display: flex; align-items: center; gap: 10px;">
                            <i class="fas fa-check-circle"></i>
                            Onaylanmış Sınıf Grup Dersleri
                            <span id="groupLessonsCount" style="background: #10b981; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.9em; font-weight: bold;">0</span>
                        </h3>
                        <p style="margin: 0 0 15px 0; color: #047857; font-size: 0.95em;">
                            <i class="fas fa-info-circle"></i> Bu dersler manuel olarak grup dersi olarak onaylanmıştır.
                        </p>
                        <div id="groupLessonsList" style="display: flex; flex-direction: column; gap: 10px;">
                            <!-- JavaScript ile doldurulacak -->
                        </div>
                    </div>

                    <!-- 🆕 ONAYLANMIŞ İHLALLİ SINIF DERSLERİ -->
                    <div id="approvedViolationsSection" style="display: none; background: linear-gradient(135deg, #faf5ff 0%, #f3e8ff 100%); border-left: 4px solid #8b5cf6; border-radius: 12px; padding: 20px; margin-bottom: 20px;">
                        <h3 style="margin: 0 0 15px 0; color: #7c3aed; display: flex; align-items: center; gap: 10px;">
                            <i class="fas fa-check-circle"></i>
                            Onaylanmış İhlalli Sınıf Dersleri
                            <span id="approvedViolationsCount" style="background: #8b5cf6; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.9em; font-weight: bold;">0</span>
                        </h3>
                        <p style="margin: 0 0 15px 0; color: #6b21a8; font-size: 0.95em;">
                            <i class="fas fa-info-circle"></i> Bu dersler ihlaller olmasına rağmen manuel olarak onaylanmıştır.
                        </p>
                        <div id="approvedViolationsList" style="display: flex; flex-direction: column; gap: 10px;">
                            <!-- JavaScript ile doldurulacak -->
                        </div>
                    </div>

                    <!-- CİDDİYET GÖSTERGESİ -->
                    <div id="severityIndicator" style="padding: 20px; border-radius: 12px; margin-bottom: 20px; text-align: center; font-weight: bold; font-size: 1.2em;">
                        <!-- JavaScript ile doldurulacak -->
                    </div>

                    <!-- HAFTALIK DAĞILIM -->
                    <div style="background: white; border-radius: 12px; padding: 20px; margin-bottom: 20px;">
                        <h3 style="margin-bottom: 15px; color: #1f2937;">📅 Haftalık Dağılım</h3>
                        <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px;">
                            <div style="text-align: center; padding: 15px; background: #f3f4f6; border-radius: 8px;">
                                <div style="font-size: 0.9em; color: #6b7280; margin-bottom: 5px;">Hafta 1</div>
                                <div style="font-size: 1.8em; font-weight: bold; color: #667eea;" id="week1Conflicts">0</div>
                            </div>
                            <div style="text-align: center; padding: 15px; background: #f3f4f6; border-radius: 8px;">
                                <div style="font-size: 0.9em; color: #6b7280; margin-bottom: 5px;">Hafta 2</div>
                                <div style="font-size: 1.8em; font-weight: bold; color: #667eea;" id="week2Conflicts">0</div>
                            </div>
                            <div style="text-align: center; padding: 15px; background: #f3f4f6; border-radius: 8px;">
                                <div style="font-size: 0.9em; color: #6b7280; margin-bottom: 5px;">Hafta 3</div>
                                <div style="font-size: 1.8em; font-weight: bold; color: #667eea;" id="week3Conflicts">0</div>
                            </div>
                            <div style="text-align: center; padding: 15px; background: #f3f4f6; border-radius: 8px;">
                                <div style="font-size: 0.9em; color: #6b7280; margin-bottom: 5px;">Hafta 4</div>
                                <div style="font-size: 1.8em; font-weight: bold; color: #667eea;" id="week4Conflicts">0</div>
                            </div>
                        </div>
                    </div>

                    <!-- AKSIYONLAR -->
                    <div style="display: flex; gap: 15px; margin-bottom: 20px;">
                        <button onclick="autoFixConflicts()" style="flex: 1; background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; border: none; padding: 15px 25px; border-radius: 10px; font-weight: bold; cursor: pointer; font-size: 1em;">
                            <i class="fas fa-magic"></i> Otomatik Düzelt
                        </button>
                        <button onclick="exportConflictReport()" style="flex: 1; background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%); color: white; border: none; padding: 15px 25px; border-radius: 10px; font-weight: bold; cursor: pointer; font-size: 1em;">
                            <i class="fas fa-file-excel"></i> Excel İndir
                        </button>
                        <button onclick="refreshConflictCheck()" style="flex: 1; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border: none; padding: 15px 25px; border-radius: 10px; font-weight: bold; cursor: pointer; font-size: 1em;">
                            <i class="fas fa-sync-alt"></i> Yenile
                        </button>
                    </div>
                </div>

                <!-- ÇAKIŞMA LİSTESİ -->
                <div id="conflictList" style="max-height: 50vh; overflow-y: auto; display: none;"></div>

                <!-- BOŞ DURUM -->
                <div id="noConflicts" style="text-align: center; padding: 60px 20px; display: none;">
                    <div style="font-size: 5em; margin-bottom: 20px;">✅</div>
                    <h3 style="color: #10b981; font-size: 1.8em; margin-bottom: 10px;">Harika! Çakışma Bulunamadı</h3>
                    <p style="color: #6b7280; font-size: 1.1em;">Programınız çakışmasız ve düzenli.</p>
                </div>
            </div>
        </div>

    <!-- ═══════════════════════════════════════════════════════════════
         ✅ BAŞARI POPUP MODAL
         ═══════════════════════════════════════════════════════════════ -->
    <div class="swap-modal" id="successModal">
        <div class="swap-modal-content" style="max-width: 500px;">
            <div class="swap-modal-header">
                <div class="swap-modal-icon" style="font-size: 4em;">✅</div>
                <div class="swap-modal-title" style="color: #10b981;">İşlem Başarılı!</div>
            </div>

            <div class="swap-modal-body" style="border-left-color: #10b981; background: #f0fdf4; max-height: 400px; overflow-y: auto;">
                <p id="successModalMessage" style="font-size: 1.05em; line-height: 1.6; color: #1f2937; text-align: center;">
                    Başarı mesajı buraya gelecek
                </p>
            </div>

            <div class="swap-modal-buttons">
                <button class="swap-btn swap-btn-confirm" onclick="closeSuccessModal()" style="width: 100%;">
                    ✓ Tamam
                </button>
            </div>
        </div>
    </div>

    <!-- ═══════════════════════════════════════════════════════════════
         🚨 HATA POPUP MODAL
         ═══════════════════════════════════════════════════════════════ -->
    <div class="swap-modal" id="errorModal">
        <div class="swap-modal-content" style="max-width: 500px;">
            <div class="swap-modal-header">
                <div class="swap-modal-icon" style="font-size: 4em;">🚫</div>
                <div class="swap-modal-title" style="color: #dc2626;">İşlem Yapılamıyor!</div>
            </div>

            <div class="swap-modal-body" style="border-left-color: #dc2626; max-height: 400px; overflow-y: auto;">
                <p id="errorModalMessage" style="font-size: 1.05em; line-height: 1.6; color: #1f2937; text-align: left; white-space: pre-line;">
                    Hata mesajı buraya gelecek
                </p>
            </div>

            <div class="swap-modal-buttons">
                <button class="swap-btn" onclick="closeErrorModal()" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); width: 100%;">
                    ✓ Anladım
                </button>
            </div>
        </div>
    </div>

    <!-- ═══════════════════════════════════════════════════════════════
         ⚠️ UYARI MODAL (KURALLARA AYKIRI SWAP - DEVAM ET / İPTAL ET)
         ═══════════════════════════════════════════════════════════════ -->
    <div class="swap-modal" id="warningModal">
        <div class="swap-modal-content" style="max-width: 550px;">
            <div class="swap-modal-header">
                <div class="swap-modal-icon" style="font-size: 4em;">⚠️</div>
                <div class="swap-modal-title" style="color: #f59e0b;">Kurallara Aykırı Swap!</div>
            </div>

            <div class="swap-modal-body" style="border-left-color: #f59e0b; max-height: 450px; overflow-y: auto;">
                <p id="warningModalMessage" style="font-size: 1.05em; line-height: 1.6; color: #1f2937; text-align: left; white-space: pre-line;">
                    Uyarı mesajı buraya gelecek
                </p>
            </div>

            <div class="swap-modal-buttons" style="display: flex; gap: 10px;">
                <button class="swap-btn" onclick="cancelWarning()" style="background: linear-gradient(135deg, #6b7280 0%, #4b5563 100%); flex: 1;">
                    ✕ İptal Et
                </button>
                <button class="swap-btn" onclick="confirmWarning()" style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); flex: 1;">
                    ⚠️ Devam Et
                </button>
            </div>
        </div>
    </div>

    <!-- ═══════════════════════════════════════════════════════════════
         🎨 YER DEĞİŞTİRME ONAY POPUP'I
         ═══════════════════════════════════════════════════════════════ -->
    <div class="swap-modal" id="swapConfirmModal">
        <div class="swap-modal-content" style="max-width: 500px;">
            <div class="swap-modal-header">
                <div class="swap-modal-icon">⚠️</div>
                <div class="swap-modal-title">Bu Slot Dolu!</div>
            </div>

            <div class="swap-modal-body" style="max-height: 400px; overflow-y: auto;">
                <div class="swap-info">
                    <span id="swapStudent1"></span>
                </div>
                <div class="swap-arrow">⇅</div>
                <div class="swap-info">
                    <span id="swapStudent2"></span>
                </div>
            </div>

            <p style="text-align: center; color: #6b7280; margin-bottom: 20px; font-size: 1em;">
                Bu iki öğrenci yer değiştirecek. Onaylıyor musunuz?
            </p>

            <div class="swap-modal-buttons">
                <button class="swap-btn swap-btn-confirm" onclick="confirmSwap()">
                    ✓ Yer Değiştir
                </button>
                <button class="swap-btn swap-btn-cancel" onclick="cancelSwap()">
                    ✗ İptal
                </button>
            </div>
        </div>
    </div>

    <!-- 🆕 GRUP DERSİ ONAY MODALI -->
    <div class="modal" id="groupLessonConfirmModal" style="display: none;">
        <div class="modal-content" style="max-width: 500px; max-height: 90vh; overflow-y: auto; width: 90%; min-width: 300px; margin: auto; position: relative; transform: none !important;">
            <h2 style="color: #f59e0b; margin-bottom: 20px;">
                <i class="fas fa-exclamation-triangle"></i> Öğretmen Çakışması!
            </h2>

            <div id="groupLessonMessage" style="background: #fff7ed; border-left: 4px solid #f59e0b; padding: 15px; border-radius: 8px; margin-bottom: 20px; word-wrap: break-word; max-height: 200px; overflow-y: auto;">
                <!-- JavaScript ile doldurulacak -->
            </div>

            <div style="background: #f0fdf4; border-left: 4px solid #10b981; padding: 15px; border-radius: 8px; margin-bottom: 20px; word-wrap: break-word;">
                <h3 style="color: #059669; margin: 0 0 10px 0; font-size: 1.1em;">
                    <i class="fas fa-link"></i> Grup Dersi Olarak Kaydet
                </h3>
                <p style="margin: 0; color: #047857;">
                    Bu öğretmen aynı gün ve saatte <strong id="groupClassList"></strong> sınıflarına birlikte ders verecek.
                </p>
            </div>

            <div style="display: flex; gap: 15px; justify-content: center; flex-wrap: wrap;">
                <button onclick="cancelGroupLesson()" style="flex: 1; min-width: 120px; padding: 12px 24px; background: #6b7280; color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em; transition: all 0.3s;">
                    <i class="fas fa-times"></i> İptal Et
                </button>
                <button onclick="confirmGroupLesson()" style="flex: 1; min-width: 120px; padding: 12px 24px; background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em; transition: all 0.3s;">
                    <i class="fas fa-check"></i> Grup Dersi Olarak Kaydet
                </button>
            </div>
        </div>
    </div>

    <!-- 🆕 ÖĞRENCİ UYARILARI MODALI -->
    <div class="modal" id="studentWarningsModal" style="display: none;">
        <div class="modal-content" style="max-width: 600px; max-height: 90vh; overflow-y: auto; width: 90%; min-width: 300px; margin: auto; position: relative; transform: none !important;">
            <h2 style="color: #f59e0b; margin-bottom: 20px;">
                <i class="fas fa-exclamation-triangle"></i> Öğrenci Uyarıları!
            </h2>

            <div style="background: #fff7ed; border-left: 4px solid #f59e0b; padding: 15px; border-radius: 8px; margin-bottom: 20px; word-wrap: break-word;">
                <p style="margin: 0 0 10px 0; color: #92400e; font-weight: 600;">
                    Bu atama aşağıdaki öğrenci kısıtlamaları ve engellemelerle çakışıyor:
                </p>
                <div id="warningsList" style="max-height: 300px; overflow-y: auto; margin-top: 10px;">
                    <!-- JavaScript ile doldurulacak -->
                </div>
            </div>

            <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                <p style="margin: 0; color: #92400e; font-size: 0.95em;">
                    <i class="fas fa-info-circle"></i> <strong>Not:</strong> Bu uyarıları göz ardı edip atamayı yaparsanız, çakışma kontrolünde "Onaylanmış Çakışma" olarak görünecektir.
                </p>
            </div>

            <div style="display: flex; gap: 15px; justify-content: center; flex-wrap: wrap;">
                <button onclick="cancelWarnings()" style="flex: 1; min-width: 120px; padding: 12px 24px; background: #6b7280; color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em; transition: all 0.3s;">
                    <i class="fas fa-times"></i> İptal Et
                </button>
                <button onclick="confirmWithWarnings()" style="flex: 1; min-width: 150px; padding: 12px 24px; background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em; transition: all 0.3s;">
                    <i class="fas fa-check-circle"></i> Yine de Kaydet
                </button>
            </div>
        </div>
    </div>

    <!-- 🆕 ÖĞRETMEN ÇAKIŞMASI ONAY MODALI -->
    <div class="modal" id="teacherConflictModal" style="display: none;">
        <div class="modal-content" style="max-width: 500px; max-height: 90vh; overflow-y: auto; width: 90%; min-width: 300px; margin: auto; position: relative; transform: none !important;">
            <h2 style="color: #f59e0b; margin-bottom: 20px;">
                <i class="fas fa-exclamation-triangle"></i> Öğretmen Çakışması!
            </h2>

            <div id="teacherConflictMessage" style="background: #fff7ed; border-left: 4px solid #f59e0b; padding: 15px; border-radius: 8px; margin-bottom: 20px; word-wrap: break-word; max-height: 300px; overflow-y: auto; white-space: pre-line; line-height: 1.6;">
                <!-- JavaScript ile doldurulacak -->
            </div>

            <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                <p style="margin: 0; color: #92400e; font-size: 0.95em;">
                    <i class="fas fa-info-circle"></i> <strong>Not:</strong> Sınıf dersi eklemek doğal olarak öğretmen çakışması oluşturur. Bu normal bir durumdur.
                </p>
            </div>

            <div style="display: flex; gap: 15px; justify-content: center; flex-wrap: wrap;">
                <button onclick="cancelTeacherConflict()" style="flex: 1; min-width: 120px; padding: 12px 24px; background: #6b7280; color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em; transition: all 0.3s;">
                    <i class="fas fa-times"></i> İptal Et
                </button>
                <button onclick="confirmTeacherConflict()" style="flex: 1; min-width: 150px; padding: 12px 24px; background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em; transition: all 0.3s;">
                    <i class="fas fa-check-circle"></i> Devam Et
                </button>
            </div>
        </div>
    </div>

    <!-- 🆕 YENİ İHLAL KONTROL PANELİ V2 -->
    <div class="modal" id="conflictPanelV2">
        <div class="modal-content" style="max-width: 1000px; max-height: 90vh; overflow-y: auto;">
            <span class="close-btn" onclick="closeConflictPanelV2()">&times;</span>

            <h2 style="color: #ef4444; margin-bottom: 20px; display: flex; align-items: center; gap: 10px; border-bottom: 3px solid #ef4444; padding-bottom: 15px;">
                <i class="fas fa-exclamation-triangle"></i> İhlal Kontrol Paneli
            </h2>

            <!-- YÜKLEME -->
            <div id="panelV2Loading" style="text-align: center; padding: 60px; display: none;">
                <div style="font-size: 4em; color: #667eea; margin-bottom: 20px;">
                    <i class="fas fa-spinner fa-spin"></i>
                </div>
                <p style="font-size: 1.3em; color: #6b7280; font-weight: 500;">Kontrol ediliyor...</p>
            </div>

            <!-- İÇERİK -->
            <div id="panelV2Content" style="display: none;">

                <!-- ÖZET KARTLARI -->
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 15px; margin-bottom: 30px;">
                    <!-- Kritik -->
                    <div style="background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%); border-left: 4px solid #ef4444; padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgba(239, 68, 68, 0.1);">
                        <div style="font-size: 2.5em; margin-bottom: 5px;">🔴</div>
                        <div id="v2CriticalCount" style="font-size: 2em; font-weight: 700; color: #dc2626;">0</div>
                        <div style="color: #991b1b; font-weight: 600; font-size: 0.9em;">Kritik</div>
                    </div>

                    <!-- Yüksek -->
                    <div style="background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%); border-left: 4px solid #f59e0b; padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgba(245, 158, 11, 0.1);">
                        <div style="font-size: 2.5em; margin-bottom: 5px;">⚠️</div>
                        <div id="v2HighCount" style="font-size: 2em; font-weight: 700; color: #d97706;">0</div>
                        <div style="color: #92400e; font-weight: 600; font-size: 0.9em;">Yüksek</div>
                    </div>

                    <!-- Orta -->
                    <div style="background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%); border-left: 4px solid #3b82f6; padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgba(59, 130, 246, 0.1);">
                        <div style="font-size: 2.5em; margin-bottom: 5px;">ℹ️</div>
                        <div id="v2MediumCount" style="font-size: 2em; font-weight: 700; color: #2563eb;">0</div>
                        <div style="color: #1e40af; font-weight: 600; font-size: 0.9em;">Orta</div>
                    </div>

                    <!-- Toplam -->
                    <div style="background: linear-gradient(135deg, #f3e8ff 0%, #e9d5ff 100%); border-left: 4px solid #a855f7; padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgba(168, 85, 247, 0.1);">
                        <div style="font-size: 2.5em; margin-bottom: 5px;">📊</div>
                        <div id="v2TotalCount" style="font-size: 2em; font-weight: 700; color: #9333ea;">0</div>
                        <div style="color: #6b21a8; font-weight: 600; font-size: 0.9em;">Toplam</div>
                    </div>
                </div>

                <!-- KARTLAR -->
                <div id="panelV2Cards" style="display: flex; flex-direction: column; gap: 20px;">
                    <!-- Kartlar buraya eklenecek -->
                </div>

                <!-- BOŞ DURUM -->
                <div id="panelV2Empty" style="display: none; text-align: center; padding: 60px 20px; background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); border-radius: 16px; border: 2px dashed #86efac;">
                    <div style="font-size: 5em; margin-bottom: 20px;">✅</div>
                    <h3 style="color: #15803d; font-size: 1.8em; margin-bottom: 10px;">Tebrikler!</h3>
                    <p style="color: #166534; font-size: 1.2em;">Hiçbir ihlal bulunamadı.</p>
                </div>
            </div>
        </div>
    </div>

    <script>


        // ═══════════════════════════════════════════════════════════════
        // 🔢 TABLO SIRALAMA SİSTEMİ
        // ═══════════════════════════════════════════════════════════════

        // Sıralama durumlarını tut
        let teacherSortState = { column: 'name', direction: 'asc' };
        let studentSortState = { column: 'name', direction: 'asc' };

        /**
         * Öğretmenleri belirli bir kolona göre sırala
         */
        function sortTeachers(column) {
            // Aynı kolona tıklanırsa direction değiştir (asc ↔ desc)
            if (teacherSortState.column === column) {
                teacherSortState.direction = teacherSortState.direction === 'asc' ? 'desc' : 'asc';
            } else {
                // Yeni kolona geçiş - varsayılan asc
                teacherSortState.column = column;
                teacherSortState.direction = 'asc';
            }

            // Listeyi yeniden yükle
            loadTeachers();
        }

        /**
         * Öğrencileri belirli bir kolona göre sırala
         */
        function sortStudents(column) {
            // Aynı kolona tıklanırsa direction değiştir
            if (studentSortState.column === column) {
                studentSortState.direction = studentSortState.direction === 'asc' ? 'desc' : 'asc';
            } else {
                studentSortState.column = column;
                studentSortState.direction = 'asc';
            }

            loadStudents();
        }

        /**
         * Sıralama ok simgesini göster
         */
        function getSortIcon(currentColumn, targetColumn, direction) {
            if (currentColumn === targetColumn) {
                return direction === 'asc' ? ' ▲' : ' ▼';
            }
            return '';
        }

        // ═══════════════════════════════════════════════════════════════
        // 🔢 ACCORDION SAYAÇ FONKSİYONLARI
        // ═══════════════════════════════════════════════════════════════

        /**
         * Kısıtlama sayısını güncelle
         */
        function updateRestrictionCount() {
            const container = document.getElementById('restrictionGroups');
            if (!container) return;

            const count = container.querySelectorAll('.restriction-group').length;
            const badge = document.getElementById('restrictionCountBadge');

            if (badge) {
                badge.textContent = count;
                if (count > 0) {
                    badge.style.display = 'inline-block';
                } else {
                    badge.style.display = 'none';
                }
            }
        }

        /**
         * Öncelik sayısını güncelle (Tüm haftalar toplamı)
         */
        function updatePriorityCount() {
            let totalCount = 0;

            for (let week = 1; week <= 4; week++) {
                const container = document.getElementById(`week${week}PriorityList`);
                if (container) {
                    const weekCount = container.querySelectorAll('.restriction-group').length;
                    totalCount += weekCount;
                }
            }

            const badge = document.getElementById('priorityCountBadge');

            if (badge) {
                badge.textContent = totalCount;
                if (totalCount > 0) {
                    badge.style.display = 'inline-block';
                } else {
                    badge.style.display = 'none';
                }
            }
        }

        /**
         * Manuel ders sayısını güncelle
         */
        function updateManualLessonCount() {
            const container = document.getElementById('manualLessonGroups');
            if (!container) return;

            const count = container.querySelectorAll('[id^="manual_"]').length;
            const badge = document.getElementById('manualCountBadge');

            if (badge) {
                badge.textContent = count;
                if (count > 0) {
                    badge.style.display = 'inline-block';
                } else {
                    badge.style.display = 'none';
                }
            }
        }

        /**
         * Öğretmen engelleme sayısını güncelle
         */
        function updateStudentTeacherBlockCount() {
            const container = document.getElementById('studentTeacherBlockGroups');
            if (!container) return;

            const count = container.querySelectorAll('[id^="studentTeacherBlock"]').length;
            const badge = document.getElementById('teacherBlockCountBadge');

            if (badge) {
                badge.textContent = count;
                if (count > 0) {
                    badge.style.display = 'inline-block';
                } else {
                    badge.style.display = 'none';
                }
            }
        }

        /**
         * Tüm sayaçları güncelle (Tek seferde hepsini çağırmak için)
         */
        function updateAllStudentCounts() {
            updateRestrictionCount();
            updatePriorityCount();
            updateManualLessonCount();
            updateStudentTeacherBlockCount();
        }

        /**
         * 👨‍🏫 ÖĞRETMEN MODAL SAYAÇLARI
         */

        /**
         * Gün sayısını güncelle
         */
        function updateDayGroupCount() {
            const container = document.getElementById('dayGroups');
            if (!container) return;

            const count = container.querySelectorAll('.day-group').length;
            const badge = document.getElementById('dayCountBadge');

            if (badge) {
                badge.textContent = count;
                if (count > 0) {
                    badge.style.display = 'inline-block';
                } else {
                    badge.style.display = 'none';
                }
            }
        }

        /**
         * Öğretmen bloklama sayısını güncelle
         */
        function updateTeacherBlockCount() {
            const container = document.getElementById('teacherBlockGroups');
            if (!container) return;

            const count = container.querySelectorAll('[id^="teacherBlock"]').length;
            const badge = document.getElementById('blockCountBadge');

            if (badge) {
                badge.textContent = count;
                if (count > 0) {
                    badge.style.display = 'inline-block';
                } else {
                    badge.style.display = 'none';
                }
            }
        }

        /**
         * Tüm öğretmen sayaçlarını güncelle
         */
        function updateAllTeacherCounts() {
            updateDayGroupCount();
            updateTeacherBlockCount();
        }

        // ═══════════════════════════════════════════════════════════════
        // 🔢 SAYAÇ FONKSİYONLARI SONU
        // ═══════════════════════════════════════════════════════════════

        let dayGroupCounter = 0;
        let restrictionCounter = 0;
        let priorityCounters = {1: 0, 2: 0, 3: 0, 4: 0};
        let manualLessonCounter = 0;
        let teacherBlockCounter = 0;
        let studentTeacherBlockCounter = 0;

        function openTeacherModal(teacherId = null) {
            document.getElementById('teacherModal').style.display = 'block';
            if (teacherId) {
                document.getElementById('teacherModalTitle').textContent = 'Öğretmen Düzenle';
                loadTeacherData(teacherId);
            } else {
                document.getElementById('teacherModalTitle').textContent = 'Yeni Öğretmen Ekle';
                document.getElementById('teacherId').value = '';
                document.getElementById('dayGroups').innerHTML = '';
                addDayGroup();

                // 🔢 SAYAÇLARI SIFIRLA - YENİ EKLENEN
                setTimeout(() => {
                    updateAllTeacherCounts();
                }, 100);
            }
        }

        async function loadTeacherData(teacherId) {
            const response = await fetch('/get_teachers');
            const data = await response.json();
            const teacher = data.teachers.find(t => t.id == teacherId);

            if (teacher) {
                document.getElementById('teacherId').value = teacher.id;
                document.getElementById('teacherName').value = teacher.name;
                document.getElementById('teacherSurname').value = teacher.surname;
                document.getElementById('teacherBranch').value = teacher.branch;

                document.getElementById('dayGroups').innerHTML = '';

                // ✅ GÜNLERE GÖRE SIRALA
                const dayOrder = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar'];
                const sortedSchedule = [...teacher.schedule].sort((a, b) =>
                    dayOrder.indexOf(a.day) - dayOrder.indexOf(b.day)
                );

                sortedSchedule.forEach(daySchedule => {
                    const groupId = dayGroupCounter++;
                    const dayGroup = document.createElement('div');
                    dayGroup.className = 'day-group';
                    dayGroup.id = 'dayGroup' + groupId;

                    dayGroup.innerHTML = `
                        <div class="day-group-header">
                            <div class="form-group" style="flex: 1; margin: 0; margin-right: 10px;">
                                <select class="day-select" onchange="updateDefaultDuration(this)" required>
                                    <option value="">Gün Seçiniz...</option>
                                    ${['Pazartesi','Salı','Çarşamba','Perşembe','Cuma','Cumartesi','Pazar'].map(d =>
                                        `<option value="${d}" ${daySchedule.day === d ? 'selected' : ''}>${d}</option>`
                                    ).join('')}
                                </select>
                            </div>
                            <button type="button" class="remove-day-btn" onclick="removeDayGroup(${groupId})">Günü Sil</button>
                        </div>
                        <div class="lessons-container" id="lessonsContainer${groupId}"></div>
                        <button type="button" class="add-day-btn" style="background: #2196F3; margin-top: 10px;" onclick="addLesson(${groupId})">+ Ders Ekle</button>
                    `;

                    document.getElementById('dayGroups').appendChild(dayGroup);

                    // ✅ SAATE GÖRE SIRALA
                    const sortedLessons = [...daySchedule.lessons].sort((a, b) =>
                        a.start_time.localeCompare(b.start_time)
                    );

                    sortedLessons.forEach((lesson, idx) => {
                        const container = document.getElementById('lessonsContainer' + groupId);
                        const lessonNumber = idx + 1;
                        const duration = lesson.duration || 35;

                        const lessonSlot = document.createElement('div');
                        lessonSlot.className = 'lesson-slot';
                        lessonSlot.id = `lesson${groupId}_${lessonNumber}`;

                        lessonSlot.innerHTML = `
                            <input type="time" class="lesson-start" data-lesson="${lessonNumber}" value="${lesson.start_time}" onchange="calcEnd(this)" required style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                            <select class="duration-select" data-lesson="${lessonNumber}" onchange="calcEnd(this.parentElement.querySelector('.lesson-start'))" style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                                ${[10,15,20,25,30,35,40,45,50,55].map(d =>
                                    `<option value="${d}" ${d == duration ? 'selected' : ''}>${d} dk</option>`
                                ).join('')}
                            </select>
                            <input type="time" class="lesson-end" data-lesson="${lessonNumber}" value="${lesson.end_time}" readonly required style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                            <button type="button" class="remove-lesson-btn" onclick="removeLesson(${groupId}, ${lessonNumber})" style="background: #f44336; color: white; border: none; padding: 5px 10px; border-radius: 5px; cursor: pointer; font-size: 0.9em;">Sil</button>
                        `;

                        container.appendChild(lessonSlot);
                    });
                });

                // 🆕 BLOKLAMALARI YÜKLE
                document.getElementById('teacherBlockGroups').innerHTML = '';
                teacherBlockCounter = 0;

                if (teacher.blocked_slots && teacher.blocked_slots.length > 0) {
                    teacher.blocked_slots.forEach(block => {
                        const bId = teacherBlockCounter++;
                        const bGroup = document.createElement('div');
                        bGroup.className = 'restriction-group';
                        bGroup.id = 'teacherBlock' + bId;
                        bGroup.style.background = '#fee2e2';
                        bGroup.style.border = '2px solid #ef4444';

                        // TYPE BELİRLE
                        const isCustom = block.type === 'custom';
                        const typeOpts = isCustom ?
                            '<option value="weekly">Her hafta</option><option value="custom" selected>Özel hafta seçimi</option>' :
                            '<option value="weekly" selected>Her hafta</option><option value="custom">Özel hafta seçimi</option>';

                        // HAFTA CHECKBOX'LARI
                        let weekCheckboxHTML = '';
                        if (isCustom && block.weeks) {
                            weekCheckboxHTML = `
                                <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px;">
                                    ${[1,2,3,4].map(w => `
                                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                                            <input type="checkbox" class="block-week-checkbox" data-bid="${bId}" value="${w}" ${block.weeks.includes(w) ? 'checked' : ''} style="width: auto;">
                                            Hafta ${w}
                                        </label>
                                    `).join('')}
                                </div>
                            `;
                        }

                        // DERS CHECKBOX'LARI
                        const blockedSlots = block.blocked_slots || [];
                        let lessonCheckboxHTML = '';
                        if (blockedSlots.length > 0) {
                            lessonCheckboxHTML = `
                                <div style="margin-bottom: 10px;"><strong style="display: block; margin-bottom: 8px;">Bloklanacak Dersler:</strong></div>
                                <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px;">
                                    ${blockedSlots.map(slot => `
                                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                                            <input type="checkbox" class="block-lesson-checkbox" data-bid="${bId}" value="${slot}" checked style="width: auto;">
                                            ${slot}
                                        </label>
                                    `).join('')}
                                </div>
                            `;
                        }

                        bGroup.innerHTML = `
                            <div class="restriction-header">
                                <span class="restriction-title" style="color: #ef4444;">🚫 Bloklama ${bId + 1}</span>
                                <button type="button" class="remove-restriction-btn" onclick="removeTeacherBlock(${bId})">Sil</button>
                            </div>

                            <!-- Hafta Seçimi -->
                            <div style="margin-bottom: 15px;">
                                <strong style="display: block; margin-bottom: 8px;">Hafta Seçimi:</strong>
                                <select class="block-type" data-bid="${bId}" onchange="toggleBlockWeekSelect(${bId})" style="width: 100%; padding: 10px; border: 2px solid #ddd; border-radius: 6px;">
                                    ${typeOpts}
                                </select>
                                <div id="blockWeekContainer${bId}" style="margin-top: 10px;">${weekCheckboxHTML}</div>
                            </div>

                            <!-- Gün Seçimi -->
                            <div style="margin-bottom: 15px;">
                                <strong style="display: block; margin-bottom: 8px;">Gün Seçimi:</strong>
                                <select class="block-day" data-bid="${bId}" onchange="updateBlockLessonList(${bId})" style="width: 100%; padding: 10px; border: 2px solid #ddd; border-radius: 6px;">
                                    <option value="">Gün Seçiniz...</option>
                                    ${['Pazartesi','Salı','Çarşamba','Perşembe','Cuma','Cumartesi','Pazar'].map(d =>
                                        `<option value="${d}" ${block.day === d ? 'selected' : ''}>${d}</option>`
                                    ).join('')}
                                </select>
                            </div>

                            <!-- Ders Listesi -->
                            <div id="blockLessonList${bId}" style="margin-top: 10px;">${lessonCheckboxHTML}</div>
                        `;

                        document.getElementById('teacherBlockGroups').appendChild(bGroup);
                    });
                }
            }

            // 🔢 TÜM SAYAÇLARI GÜNCELLE (Düzenleme modunda mevcut verileri göster) - YENİ EKLENEN
            setTimeout(() => {
                updateAllTeacherCounts();
            }, 200);
        }


        function closeTeacherModal() {
            document.getElementById('teacherModal').style.display = 'none';
            document.getElementById('teacherForm').reset();
            document.getElementById('dayGroups').innerHTML = '';
            dayGroupCounter = 0;

            // 🔢 SAYAÇLARI SIFIRLA - YENİ EKLENEN
            setTimeout(() => {
                updateAllTeacherCounts();
            }, 100);
        }

        function openStudentModal(studentId = null) {
            document.getElementById('studentModal').style.display = 'block';
            if (studentId) {
                document.getElementById('studentModalTitle').textContent = 'Öğrenci Düzenle';
                loadStudentData(studentId);
            } else {
                document.getElementById('studentModalTitle').textContent = 'Yeni Öğrenci Ekle';
                document.getElementById('studentId').value = '';
                document.getElementById('restrictionGroups').innerHTML = '';
                // Öncelik ve manuel ders listelerini temizle
                for (let week = 1; week <= 4; week++) {
                    const container = document.getElementById(`week${week}PriorityList`);
                    if (container) container.innerHTML = '';
                    priorityCounters[week] = 0;
                }
                document.getElementById('manualLessonGroups').innerHTML = '';
                manualLessonCounter = 0;
                document.getElementById('studentTeacherBlockGroups').innerHTML = '';
                studentTeacherBlockCounter = 0;
                restrictionCounter = 0;

                // 🔢 TÜM SAYAÇLARI SIFIRLA
                setTimeout(() => {
                    updateAllStudentCounts();
                }, 100);
            }
        }

        async function loadStudentData(studentId) {
            const response = await fetch('/get_students');
            const data = await response.json();
            const student = data.students.find(s => s.id == studentId);

            if (student) {
                document.getElementById('studentId').value = student.id;
                document.getElementById('studentName').value = student.name;
                document.getElementById('studentSurname').value = student.surname;
                document.getElementById('studentClass').value = student.class;

                // KISITLAMALARI YÜKLE
                document.getElementById('restrictionGroups').innerHTML = '';
                restrictionCounter = 0;

                if (student.restrictions && student.restrictions.length > 0) {
                    student.restrictions.forEach(restriction => {
                        const rId = restrictionCounter++;
                        const rGroup = document.createElement('div');
                        rGroup.className = 'restriction-group';
                        rGroup.id = 'restriction' + rId;

                        // ✅ TYPE BELİRLE
                        const isCustom = restriction.type === 'custom' || (restriction.weeks && restriction.weeks.length > 0);
                        const typeOpts = isCustom ?
                            '<option value="weekly">Her hafta</option><option value="custom" selected>Özel hafta seçimi</option>' :
                            '<option value="weekly" selected>Her hafta</option><option value="custom">Özel hafta seçimi</option>';

                        // ✅ HAFTA CHECKBOX'LARI
                        let weekCheckboxHTML = '';
                        if (isCustom) {
                            const weeks = restriction.weeks || [restriction.week];
                            weekCheckboxHTML = `
                                <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; margin-top: 10px;">
                                    ${[1,2,3,4].map(w => `
                                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                                            <input type="checkbox" class="restriction-week-checkbox" data-rid="${rId}" value="${w}" ${weeks.includes(w) ? 'checked' : ''} style="width: auto;">
                                            Hafta ${w}
                                        </label>
                                    `).join('')}
                                </div>
                            `;
                        }

                        // ✅ GÜN CHECKBOX'LARI
                        const days = restriction.days || [restriction.day];
                        const dayCheckboxHTML = `
                            <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px;">
                                ${['Pazartesi','Salı','Çarşamba','Perşembe','Cuma','Cumartesi','Pazar'].map(d => `
                                    <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                                        <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="${d}" ${days.includes(d) ? 'checked' : ''} style="width: auto;">
                                        ${d}
                                    </label>
                                `).join('')}
                            </div>
                        `;

                        rGroup.innerHTML = `
                            <div class="restriction-header">
                                <span class="restriction-title">Kısıtlama ${rId + 1}</span>
                                <button type="button" class="remove-restriction-btn" onclick="removeRestriction(${rId})">Sil</button>
                            </div>

                            <!-- Hafta Seçimi -->
                            <div style="margin-bottom: 15px;">
                                <strong style="display: block; margin-bottom: 8px;">Hafta Seçimi:</strong>
                                <select class="restriction-type" data-rid="${rId}" onchange="toggleWeekSelect(${rId})" style="width: 100%; padding: 10px; border: 2px solid #ddd; border-radius: 6px;">
                                    ${typeOpts}
                                </select>
                                <div id="weekNumContainer${rId}">${weekCheckboxHTML}</div>
                            </div>

                            <!-- Gün Seçimi -->
                            <div style="margin-bottom: 15px;">
                                <strong style="display: block; margin-bottom: 8px;">Günler:</strong>
                                ${dayCheckboxHTML}
                            </div>

                            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: 10px;">
                                <input type="time" class="restriction-start" data-rid="${rId}" value="${restriction.start_time || ''}" style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                                <input type="time" class="restriction-end" data-rid="${rId}" value="${restriction.end_time || ''}" style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                            </div>
                        `;

                        document.getElementById('restrictionGroups').appendChild(rGroup);
                    });
                }

                // ÖNCELİKLERİ YÜKLE
                for (let week = 1; week <= 4; week++) {
                    const container = document.getElementById(`week${week}PriorityList`);
                    if (container) {
                        container.innerHTML = '';
                        priorityCounters[week] = 0;
                    }
                }

                if (student.priorities) {
                    for (let week = 1; week <= 4; week++) {
                        const weekKey = `week${week}`;
                        const weekPriorities = student.priorities[weekKey] || [];

                        weekPriorities.forEach(branch => {
                            const pId = priorityCounters[week]++;
                            const container = document.getElementById(`week${week}PriorityList`);

                            const priorityDiv = document.createElement('div');
                            priorityDiv.className = 'restriction-group';
                            priorityDiv.id = `priority_${week}_${pId}`;
                            priorityDiv.style.background = '#fef3c7';
                            priorityDiv.style.border = '2px solid #f59e0b';

                            priorityDiv.innerHTML = `
                                <div class="restriction-header">
                                    <span class="restriction-title" style="color: #f59e0b;">${pId + 1}. Öncelik</span>
                                    <button type="button" class="remove-restriction-btn" onclick="removePriority(${week}, ${pId})">Sil</button>
                                </div>
                                <select class="priority-branch" data-week="${week}" data-pid="${pId}" style="width: 100%; padding: 10px; border: 2px solid #f59e0b; border-radius: 8px;">
                                    <option value="">Branş Seçiniz...</option>
                                    <option value="Matematik" ${branch === 'Matematik' ? 'selected' : ''}>Matematik</option>
                                    <option value="Matematik-1" ${branch === 'Matematik-1' ? 'selected' : ''}>Matematik-1</option>
                                    <option value="Matematik-2" ${branch === 'Matematik-2' ? 'selected' : ''}>Matematik-2</option>
                                    <option value="Geometri" ${branch === 'Geometri' ? 'selected' : ''}>Geometri</option>
                                    <option value="Türkçe" ${branch === 'Türkçe' ? 'selected' : ''}>Türkçe</option>
                                    <option value="Edebiyat" ${branch === 'Edebiyat' ? 'selected' : ''}>Edebiyat</option>
                                    <option value="İngilizce" ${branch === 'İngilizce' ? 'selected' : ''}>İngilizce</option>
                                    <option value="Fen Bilgisi" ${branch === 'Fen Bilgisi' ? 'selected' : ''}>Fen Bilgisi</option>
                                    <option value="Fizik" ${branch === 'Fizik' ? 'selected' : ''}>Fizik</option>
                                    <option value="Kimya" ${branch === 'Kimya' ? 'selected' : ''}>Kimya</option>
                                    <option value="Biyoloji" ${branch === 'Biyoloji' ? 'selected' : ''}>Biyoloji</option>
                                    <option value="Sosyal Bilgiler" ${branch === 'Sosyal Bilgiler' ? 'selected' : ''}>Sosyal Bilgiler</option>
                                    <option value="Tarih" ${branch === 'Tarih' ? 'selected' : ''}>Tarih</option>
                                    <option value="Coğrafya" ${branch === 'Coğrafya' ? 'selected' : ''}>Coğrafya</option>
                                    <option value="Felsefe" ${branch === 'Felsefe' ? 'selected' : ''}>Felsefe</option>
                                    <option value="Din Kültürü" ${branch === 'Din Kültürü' ? 'selected' : ''}>Din Kültürü</option>
                                </select>
                            `;

                            container.appendChild(priorityDiv);
                        });
                    }
                }

                // MANUEL DERSLERİ YÜKLE
                document.getElementById('manualLessonGroups').innerHTML = '';
                manualLessonCounter = 0;

                if (student.manual_lessons && student.manual_lessons.length > 0) {
                    for (const manualLesson of student.manual_lessons) {
                        const mId = manualLessonCounter++;
                        const container = document.getElementById('manualLessonGroups');

                        const manualDiv = document.createElement('div');
                        manualDiv.className = 'restriction-group';
                        manualDiv.id = `manual_${mId}`;
                        manualDiv.style.background = '#dbeafe';
                        manualDiv.style.border = '2px solid #3b82f6';

                        manualDiv.innerHTML = `
                            <div class="restriction-header">
                                <span class="restriction-title" style="color: #3b82f6;">Manuel Ders ${mId + 1}</span>
                                <button type="button" class="remove-restriction-btn" onclick="removeManualLesson(${mId})">Sil</button>
                            </div>
                            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 10px;">
                                <select class="manual-week" data-mid="${mId}" style="padding: 8px; border: 2px solid #3b82f6; border-radius: 6px;">
                                    <option value="">Hafta Seçiniz...</option>
                                    <option value="1" ${manualLesson.week == 1 ? 'selected' : ''}>Hafta 1</option>
                                    <option value="2" ${manualLesson.week == 2 ? 'selected' : ''}>Hafta 2</option>
                                    <option value="3" ${manualLesson.week == 3 ? 'selected' : ''}>Hafta 3</option>
                                    <option value="4" ${manualLesson.week == 4 ? 'selected' : ''}>Hafta 4</option>
                                </select>
                                <select class="manual-day" data-mid="${mId}" onchange="updateManualTeachers(${mId})" style="padding: 8px; border: 2px solid #3b82f6; border-radius: 6px;">
                                    <option value="">Gün Seçiniz...</option>
                                    <option value="Pazartesi" ${manualLesson.day === 'Pazartesi' ? 'selected' : ''}>Pazartesi</option>
                                    <option value="Salı" ${manualLesson.day === 'Salı' ? 'selected' : ''}>Salı</option>
                                    <option value="Çarşamba" ${manualLesson.day === 'Çarşamba' ? 'selected' : ''}>Çarşamba</option>
                                    <option value="Perşembe" ${manualLesson.day === 'Perşembe' ? 'selected' : ''}>Perşembe</option>
                                    <option value="Cuma" ${manualLesson.day === 'Cuma' ? 'selected' : ''}>Cuma</option>
                                    <option value="Cumartesi" ${manualLesson.day === 'Cumartesi' ? 'selected' : ''}>Cumartesi</option>
                                    <option value="Pazar" ${manualLesson.day === 'Pazar' ? 'selected' : ''}>Pazar</option>
                                </select>
                            </div>
                            <select class="manual-teacher" data-mid="${mId}" id="manualTeacher_${mId}" onchange="updateManualTimes(${mId})" style="width: 100%; padding: 8px; border: 2px solid #3b82f6; border-radius: 6px; margin-bottom: 10px;">
                                <option value="">Yükleniyor...</option>
                            </select>
                            <select class="manual-time" data-mid="${mId}" id="manualTime_${mId}" style="width: 100%; padding: 8px; border: 2px solid #3b82f6; border-radius: 6px;">
                                <option value="">Önce öğretmen seçiniz...</option>
                            </select>
                        `;

                        container.appendChild(manualDiv);

                        // Öğretmenleri yükle
                        await updateManualTeachersForEdit(mId, manualLesson);
                    }
                }

                // ÖĞRETMEN ENGELLEMELERİNİ YÜKLE
                document.getElementById('studentTeacherBlockGroups').innerHTML = '';
                studentTeacherBlockCounter = 0;

                if (student.teacher_blocks && student.teacher_blocks.length > 0) {
                    for (const block of student.teacher_blocks) {
                        const tbId = studentTeacherBlockCounter++;
                        const tbGroup = document.createElement('div');
                        tbGroup.className = 'restriction-group';
                        tbGroup.id = 'studentTeacherBlock' + tbId;
                        tbGroup.style.background = '#fee2e2';
                        tbGroup.style.border = '2px solid #dc2626';

                        // TYPE BELİRLE
                        const isCustom = block.type === 'custom';
                        const typeOpts = isCustom ?
                            '<option value="weekly">Her hafta</option><option value="custom" selected>Özel hafta seçimi</option>' :
                            '<option value="weekly" selected>Her hafta</option><option value="custom">Özel hafta seçimi</option>';

                        // HAFTA CHECKBOX'LARI
                        let weekCheckboxHTML = '';
                        if (isCustom && block.weeks) {
                            weekCheckboxHTML = `
                                <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px;">
                                    ${[1,2,3,4].map(w => `
                                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                                            <input type="checkbox" class="teacher-block-week-checkbox" data-tbid="${tbId}" value="${w}" ${block.weeks.includes(w) ? 'checked' : ''} style="width: auto;">
                                            Hafta ${w}
                                        </label>
                                    `).join('')}
                                </div>
                            `;
                        }

                        // DERS CHECKBOX'LARI
                        const blockedSlots = block.blocked_slots || [];
                        let lessonCheckboxHTML = '';
                        if (blockedSlots.length > 0) {
                            lessonCheckboxHTML = '<div style="margin-bottom: 10px;"><strong style="display: block; margin-bottom: 8px;">Engellenecek Dersler:</strong></div>';
                            lessonCheckboxHTML += '<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px;">';

                            blockedSlots.forEach(slot => {
                                const parts = slot.split('_');
                                const day = parts[0];
                                const time = parts[1];
                                const displayText = block.day === 'all' ? `${day} ${time}` : time;

                                lessonCheckboxHTML += `
                                    <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                                        <input type="checkbox" class="teacher-block-lesson-checkbox" data-tbid="${tbId}" value="${slot}" checked style="width: auto;">
                                        ${displayText}
                                    </label>
                                `;
                            });

                            lessonCheckboxHTML += '</div>';
                        }

                        tbGroup.innerHTML = `
                            <div class="restriction-header">
                                <span class="restriction-title" style="color: #dc2626;">🚫 Engelleme ${tbId + 1}</span>
                                <button type="button" class="remove-restriction-btn" onclick="removeStudentTeacherBlock(${tbId})">Sil</button>
                            </div>

                            <!-- Öğretmen Seçimi -->
                            <div style="margin-bottom: 15px;">
                                <strong style="display: block; margin-bottom: 8px;">Öğretmen Seçimi:</strong>
                                <select class="teacher-block-select" data-tbid="${tbId}" onchange="updateTeacherBlockSlots(${tbId})" style="width: 100%; padding: 10px; border: 2px solid #dc2626; border-radius: 6px;">
                                    <option value="">Yükleniyor...</option>
                                </select>
                            </div>

                            <!-- Hafta Seçimi -->
                            <div style="margin-bottom: 15px;">
                                <strong style="display: block; margin-bottom: 8px;">Hafta Seçimi:</strong>
                                <select class="teacher-block-type" data-tbid="${tbId}" onchange="toggleTeacherBlockWeekSelect(${tbId})" style="width: 100%; padding: 10px; border: 2px solid #dc2626; border-radius: 6px;">
                                    ${typeOpts}
                                </select>
                                <div id="teacherBlockWeekContainer${tbId}" style="margin-top: 10px;">${weekCheckboxHTML}</div>
                            </div>

                            <!-- Gün Seçimi -->
                            <div style="margin-bottom: 15px;">
                                <strong style="display: block; margin-bottom: 8px;">Gün Seçimi:</strong>
                                <select class="teacher-block-day" data-tbid="${tbId}" onchange="updateTeacherBlockLessonList(${tbId})" style="width: 100%; padding: 10px; border: 2px solid #dc2626; border-radius: 6px;">
                                    <option value="all" ${block.day === 'all' ? 'selected' : ''}>Tüm günler</option>
                                    ${['Pazartesi','Salı','Çarşamba','Perşembe','Cuma','Cumartesi','Pazar'].map(d =>
                                        `<option value="${d}" ${block.day === d ? 'selected' : ''}>${d}</option>`
                                    ).join('')}
                                </select>
                            </div>

                            <!-- Ders Listesi -->
                            <div id="teacherBlockLessonList${tbId}">${lessonCheckboxHTML}</div>
                        `;

                        document.getElementById('studentTeacherBlockGroups').appendChild(tbGroup);

                        // Öğretmenleri yükle ve seçili olanı işaretle
                        (async () => {
                            await loadTeachersForBlock(tbId);
                            const teacherSelect = document.querySelector(`.teacher-block-select[data-tbid="${tbId}"]`);
                            if (teacherSelect && block.teacher_id) {
                                teacherSelect.value = block.teacher_id;
                            }
                        })();
                    }
                }

            }

            // 🔢 TÜM SAYAÇLARI GÜNCELLE (Düzenleme modunda mevcut verileri göster)
            setTimeout(() => {
                updateAllStudentCounts();
            }, 200);
        }




        // Manuel ders düzenleme için öğretmen ve saat yükleme
        async function updateManualTeachersForEdit(mId, manualLesson) {
            const teacherSelect = document.getElementById(`manualTeacher_${mId}`);
            const timeSelect = document.getElementById(`manualTime_${mId}`);

            if (!teacherSelect || !manualLesson.day) return;

            const response = await fetch('/get_teachers');
            const data = await response.json();

            teacherSelect.innerHTML = '<option value="">Öğretmen Seçiniz...</option>';

            // ✅ ALFABETIK SIRALA
            const sortedTeachers = data.teachers
                .filter(teacher => teacher.schedule.some(s => s.day === manualLesson.day))
                .sort((a, b) => {
                    const nameA = `${a.name} ${a.surname}`.toLocaleLowerCase('tr');
                    const nameB = `${b.name} ${b.surname}`.toLocaleLowerCase('tr');
                    return nameA.localeCompare(nameB, 'tr');
                });

            sortedTeachers.forEach(teacher => {
                const option = document.createElement('option');
                option.value = teacher.id;
                option.setAttribute('data-schedule', JSON.stringify(teacher.schedule));
                option.textContent = `${teacher.name} ${teacher.surname} (${teacher.branch})`;

                if (teacher.id == manualLesson.teacher_id) {
                    option.selected = true;
                }

                teacherSelect.appendChild(option);
            });

            if (manualLesson.teacher_id) {
                const selectedOption = teacherSelect.options[teacherSelect.selectedIndex];
                if (selectedOption && selectedOption.value) {
                    const schedule = JSON.parse(selectedOption.getAttribute('data-schedule'));
                    const daySchedule = schedule.find(s => s.day === manualLesson.day);

                    timeSelect.innerHTML = '<option value="">Ders Saati Seçiniz...</option>';

                    if (daySchedule && daySchedule.lessons) {
                        daySchedule.lessons.forEach(lesson => {
                            const option = document.createElement('option');
                            const timeValue = `${lesson.start_time}-${lesson.end_time}`;
                            option.value = timeValue;
                            option.textContent = `${lesson.start_time}-${lesson.end_time}`;

                            if (timeValue === manualLesson.time) {
                                option.selected = true;
                            }

                            timeSelect.appendChild(option);
                        });
                    }
                }
            }
        }



        function closeStudentModal() {
            document.getElementById('studentModal').style.display = 'none';
            document.getElementById('studentForm').reset();
            document.getElementById('restrictionGroups').innerHTML = '';
            for (let week = 1; week <= 4; week++) {
                const container = document.getElementById(`week${week}PriorityList`);
                if (container) container.innerHTML = '';
                priorityCounters[week] = 0;
            }
            document.getElementById('manualLessonGroups').innerHTML = '';
            manualLessonCounter = 0;
            document.getElementById('studentTeacherBlockGroups').innerHTML = '';
            studentTeacherBlockCounter = 0;
            restrictionCounter = 0;
        }

        function addTeacherBlock() {
            const bId = teacherBlockCounter++;
            const bGroup = document.createElement('div');
            bGroup.className = 'restriction-group';
            bGroup.id = 'teacherBlock' + bId;
            bGroup.style.background = '#fee2e2';
            bGroup.style.border = '2px solid #ef4444';

            bGroup.innerHTML = `
                <div class="restriction-header">
                    <span class="restriction-title" style="color: #ef4444;">🚫 Bloklama ${bId + 1}</span>
                    <button type="button" class="remove-restriction-btn" onclick="removeTeacherBlock(${bId})">Sil</button>
                </div>

                <!-- Hafta Seçimi -->
                <div style="margin-bottom: 15px;">
                    <strong style="display: block; margin-bottom: 8px;">Hafta Seçimi:</strong>
                    <select class="block-type" data-bid="${bId}" onchange="toggleBlockWeekSelect(${bId})" style="width: 100%; padding: 10px; border: 2px solid #ddd; border-radius: 6px;">
                        <option value="weekly">Her hafta</option>
                        <option value="custom">Özel hafta seçimi</option>
                    </select>
                    <div id="blockWeekContainer${bId}" style="margin-top: 10px;"></div>
                </div>

                <!-- Gün Seçimi -->
                <div style="margin-bottom: 15px;">
                    <strong style="display: block; margin-bottom: 8px;">Gün Seçimi:</strong>
                    <select class="block-day" data-bid="${bId}" onchange="updateBlockLessonList(${bId})" style="width: 100%; padding: 10px; border: 2px solid #ddd; border-radius: 6px;">
                        <option value="">Gün Seçiniz...</option>
                        <option value="Pazartesi">Pazartesi</option>
                        <option value="Salı">Salı</option>
                        <option value="Çarşamba">Çarşamba</option>
                        <option value="Perşembe">Perşembe</option>
                        <option value="Cuma">Cuma</option>
                        <option value="Cumartesi">Cumartesi</option>
                        <option value="Pazar">Pazar</option>
                    </select>
                </div>

                <!-- Ders Listesi (Gün seçildikten sonra doldurulacak) -->
                <div id="blockLessonList${bId}" style="margin-top: 10px;"></div>
            `;

            document.getElementById('teacherBlockGroups').appendChild(bGroup);

            // 🆕 OTOMATIK SCROLL
            setTimeout(() => {
                const blockContainer = document.getElementById('teacherBlockGroups');
                if (blockContainer) {
                    blockContainer.scrollTop = blockContainer.scrollHeight;
                }
            }, 100);

            // 🔢 SAYACI GÜNCELLE - YENİ EKLENEN
            updateTeacherBlockCount();
        }

        function removeTeacherBlock(bId) {
            const elem = document.getElementById('teacherBlock' + bId);
            if (elem) elem.remove();

            // 🔢 SAYACI GÜNCELLE - YENİ EKLENEN
            updateTeacherBlockCount();
        }

        function toggleBlockWeekSelect(bId) {
            const typeSelect = document.querySelector(`.block-type[data-bid="${bId}"]`);
            const container = document.getElementById('blockWeekContainer' + bId);

            if (typeSelect.value === 'custom') {
                container.innerHTML = `
                    <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px;">
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="block-week-checkbox" data-bid="${bId}" value="1" style="width: auto;">
                            Hafta 1
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="block-week-checkbox" data-bid="${bId}" value="2" style="width: auto;">
                            Hafta 2
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="block-week-checkbox" data-bid="${bId}" value="3" style="width: auto;">
                            Hafta 3
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="block-week-checkbox" data-bid="${bId}" value="4" style="width: auto;">
                            Hafta 4
                        </label>
                    </div>
                `;
            } else {
                container.innerHTML = '';
            }
        }

        function updateBlockLessonList(bId) {
            const daySelect = document.querySelector(`.block-day[data-bid="${bId}"]`);
            const container = document.getElementById('blockLessonList' + bId);
            const selectedDay = daySelect.value;

            if (!selectedDay) {
                container.innerHTML = '';
                return;
            }

            // Öğretmenin bu gündeki derslerini bul
            const teacherId = document.getElementById('teacherId').value;

            // Eğer yeni öğretmen ekleme modundaysa, mevcut form verilerinden al
            const dayGroups = document.querySelectorAll('.day-group');
            let lessons = [];

            dayGroups.forEach(group => {
                const day = group.querySelector('.day-select').value;
                if (day === selectedDay) {
                    const lessonSlots = group.querySelectorAll('.lesson-slot');
                    lessonSlots.forEach(slot => {
                        const start = slot.querySelector('.lesson-start')?.value;
                        const end = slot.querySelector('.lesson-end')?.value;
                        if (start && end) {
                            lessons.push(`${start}-${end}`);
                        }
                    });
                }
            });

            if (lessons.length === 0) {
                container.innerHTML = '<p style="color: #991b1b; font-size: 0.9em; padding: 10px; background: white; border-radius: 5px;">Bu gün için henüz ders tanımlanmamış.</p>';
                return;
            }

            // Checkbox listesi oluştur
            let html = '<div style="margin-bottom: 10px;"><strong style="display: block; margin-bottom: 8px;">Bloklanacak Dersler:</strong></div>';
            html += '<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px;">';

            lessons.forEach(lesson => {
                html += `
                    <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                        <input type="checkbox" class="block-lesson-checkbox" data-bid="${bId}" value="${lesson}" style="width: auto;">
                        ${lesson}
                    </label>
                `;
            });

            html += '</div>';
            container.innerHTML = html;
        }

        function toggleFormAccordion(section) {
            const content = document.getElementById(section + 'Content');
            const arrow = document.getElementById(section + 'Arrow');
            if (content && arrow) {
                content.classList.toggle('open');
                arrow.classList.toggle('open');
            }
        }

        function addPriority(week) {
            const pId = priorityCounters[week]++;
            const container = document.getElementById(`week${week}PriorityList`);

            const priorityDiv = document.createElement('div');
            priorityDiv.className = 'restriction-group';
            priorityDiv.id = `priority_${week}_${pId}`;
            priorityDiv.style.background = '#fef3c7';
            priorityDiv.style.border = '2px solid #f59e0b';

            priorityDiv.innerHTML = `
                <div class="restriction-header">
                    <span class="restriction-title" style="color: #f59e0b;">${pId + 1}. Öncelik</span>
                    <button type="button" class="remove-restriction-btn" onclick="removePriority(${week}, ${pId})">Sil</button>
                </div>
                <select class="priority-branch" data-week="${week}" data-pid="${pId}" style="width: 100%; padding: 10px; border: 2px solid #f59e0b; border-radius: 8px;">
                    <option value="">Branş Seçiniz...</option>
                    <option value="Matematik">Matematik</option>
                    <option value="Matematik-1">Matematik-1</option>
                    <option value="Matematik-2">Matematik-2</option>
                    <option value="Geometri">Geometri</option>
                    <option value="Türkçe">Türkçe</option>
                    <option value="Edebiyat">Edebiyat</option>
                    <option value="İngilizce">İngilizce</option>
                    <option value="Fen Bilgisi">Fen Bilgisi</option>
                    <option value="Fizik">Fizik</option>
                    <option value="Kimya">Kimya</option>
                    <option value="Biyoloji">Biyoloji</option>
                    <option value="Sosyal Bilgiler">Sosyal Bilgiler</option>
                    <option value="Tarih">Tarih</option>
                    <option value="Coğrafya">Coğrafya</option>
                    <option value="Felsefe">Felsefe</option>
                    <option value="Din Kültürü">Din Kültürü</option>
                </select>
            `;

            container.appendChild(priorityDiv);

            updatePriorityNumbers(week);

            // 🔢 SAYACI GÜNCELLE
            updatePriorityCount();

            // 🆕 OTOMATIK SCROLL
            setTimeout(() => {
                if (container) {
                    container.scrollTop = container.scrollHeight;
                }
            }, 100);
        }

        function removePriority(week, pId) {
            const elem = document.getElementById(`priority_${week}_${pId}`);
            if (elem) {
                elem.remove();
                updatePriorityNumbers(week);

                // 🔢 SAYACI GÜNCELLE
                updatePriorityCount();
            }
        }

        function updatePriorityNumbers(week) {
            const container = document.getElementById(`week${week}PriorityList`);
            const priorities = container.querySelectorAll('.restriction-group');
            priorities.forEach((priority, index) => {
                const title = priority.querySelector('.restriction-title');
                if (title) {
                    title.textContent = `${index + 1}. Öncelik`;
                }
            });
        }

        function addManualLesson() {
            const mId = manualLessonCounter++;
            const container = document.getElementById('manualLessonGroups');

            const manualDiv = document.createElement('div');
            manualDiv.className = 'restriction-group';
            manualDiv.id = `manual_${mId}`;
            manualDiv.style.background = '#dbeafe';
            manualDiv.style.border = '2px solid #3b82f6';

            manualDiv.innerHTML = `
                <div class="restriction-header">
                    <span class="restriction-title" style="color: #3b82f6;">Manuel Ders ${mId + 1}</span>
                    <button type="button" class="remove-restriction-btn" onclick="removeManualLesson(${mId})">Sil</button>
                </div>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 10px;">
                    <select class="manual-week" data-mid="${mId}" style="padding: 8px; border: 2px solid #3b82f6; border-radius: 6px;">
                        <option value="">Hafta Seçiniz...</option>
                        <option value="1">Hafta 1</option>
                        <option value="2">Hafta 2</option>
                        <option value="3">Hafta 3</option>
                        <option value="4">Hafta 4</option>
                    </select>
                    <select class="manual-day" data-mid="${mId}" onchange="updateManualTeachers(${mId})" style="padding: 8px; border: 2px solid #3b82f6; border-radius: 6px;">
                        <option value="">Gün Seçiniz...</option>
                        <option value="Pazartesi">Pazartesi</option>
                        <option value="Salı">Salı</option>
                        <option value="Çarşamba">Çarşamba</option>
                        <option value="Perşembe">Perşembe</option>
                        <option value="Cuma">Cuma</option>
                        <option value="Cumartesi">Cumartesi</option>
                        <option value="Pazar">Pazar</option>
                    </select>
                </div>
                <select class="manual-teacher" data-mid="${mId}" id="manualTeacher_${mId}" onchange="updateManualTimes(${mId})" style="width: 100%; padding: 8px; border: 2px solid #3b82f6; border-radius: 6px; margin-bottom: 10px;">
                    <option value="">Önce gün seçiniz...</option>
                </select>
                <select class="manual-time" data-mid="${mId}" id="manualTime_${mId}" style="width: 100%; padding: 8px; border: 2px solid #3b82f6; border-radius: 6px;">
                    <option value="">Önce öğretmen seçiniz...</option>
                </select>
            `;

            container.appendChild(manualDiv);

            // 🔢 SAYACI GÜNCELLE
            updateManualLessonCount();

            // 🆕 OTOMATIK SCROLL
            setTimeout(() => {
                const manualContainer = document.getElementById('manualLessonGroups');
                if (manualContainer) {
                    manualContainer.scrollTop = manualContainer.scrollHeight;
                }
            }, 100);
        }

        async function updateManualTeachers(mId) {
            const daySelect = document.querySelector(`.manual-day[data-mid="${mId}"]`);
            const teacherSelect = document.getElementById(`manualTeacher_${mId}`);
            const timeSelect = document.getElementById(`manualTime_${mId}`);

            const selectedDay = daySelect.value;

            if (!selectedDay) {
                teacherSelect.innerHTML = '<option value="">Önce gün seçiniz...</option>';
                timeSelect.innerHTML = '<option value="">Önce öğretmen seçiniz...</option>';
                return;
            }

            const response = await fetch('/get_teachers');
            const data = await response.json();

            teacherSelect.innerHTML = '<option value="">Öğretmen Seçiniz...</option>';

            // ✅ ALFABETIK SIRALA
            const sortedTeachers = data.teachers
                .filter(teacher => teacher.schedule.some(s => s.day === selectedDay))
                .sort((a, b) => {
                    const nameA = `${a.name} ${a.surname}`.toLocaleLowerCase('tr');
                    const nameB = `${b.name} ${b.surname}`.toLocaleLowerCase('tr');
                    return nameA.localeCompare(nameB, 'tr');
                });

            sortedTeachers.forEach(teacher => {
                const option = document.createElement('option');
                option.value = teacher.id;
                option.setAttribute('data-schedule', JSON.stringify(teacher.schedule));
                option.textContent = `${teacher.name} ${teacher.surname} (${teacher.branch})`;
                teacherSelect.appendChild(option);
            });

            timeSelect.innerHTML = '<option value="">Önce öğretmen seçiniz...</option>';
        }

        function updateManualTimes(mId) {
            const daySelect = document.querySelector(`.manual-day[data-mid="${mId}"]`);
            const teacherSelect = document.getElementById(`manualTeacher_${mId}`);
            const timeSelect = document.getElementById(`manualTime_${mId}`);

            const selectedDay = daySelect.value;
            const selectedOption = teacherSelect.options[teacherSelect.selectedIndex];

            if (!selectedOption.value) {
                timeSelect.innerHTML = '<option value="">Önce öğretmen seçiniz...</option>';
                return;
            }

            const schedule = JSON.parse(selectedOption.getAttribute('data-schedule'));
            const daySchedule = schedule.find(s => s.day === selectedDay);

            timeSelect.innerHTML = '<option value="">Ders Saati Seçiniz...</option>';

            if (daySchedule && daySchedule.lessons) {
                daySchedule.lessons.forEach(lesson => {
                    const option = document.createElement('option');
                    option.value = `${lesson.start_time}-${lesson.end_time}`;
                    option.textContent = `${lesson.start_time}-${lesson.end_time}`;
                    timeSelect.appendChild(option);
                });
            }
        }

        function removeManualLesson(mId) {
            const elem = document.getElementById(`manual_${mId}`);
            if (elem) {
                elem.remove();

                // 🔢 SAYACI GÜNCELLE
                updateManualLessonCount();
            }
        }

        function addStudentTeacherBlock() {
            const tbId = studentTeacherBlockCounter++;
            const tbGroup = document.createElement('div');
            tbGroup.className = 'restriction-group';
            tbGroup.id = 'studentTeacherBlock' + tbId;
            tbGroup.style.background = '#fee2e2';
            tbGroup.style.border = '2px solid #dc2626';

            tbGroup.innerHTML = `
                <div class="restriction-header">
                    <span class="restriction-title" style="color: #dc2626;">🚫 Engelleme ${tbId + 1}</span>
                    <button type="button" class="remove-restriction-btn" onclick="removeStudentTeacherBlock(${tbId})">Sil</button>
                </div>

                <!-- Öğretmen Seçimi -->
                <div style="margin-bottom: 15px;">
                    <strong style="display: block; margin-bottom: 8px;">Öğretmen Seçimi:</strong>
                    <select class="teacher-block-select" data-tbid="${tbId}" onchange="updateTeacherBlockSlots(${tbId})" style="width: 100%; padding: 10px; border: 2px solid #dc2626; border-radius: 6px;">
                        <option value="">Öğretmen Seçiniz...</option>
                    </select>
                </div>

                <!-- Hafta Seçimi -->
                <div style="margin-bottom: 15px;">
                    <strong style="display: block; margin-bottom: 8px;">Hafta Seçimi:</strong>
                    <select class="teacher-block-type" data-tbid="${tbId}" onchange="toggleTeacherBlockWeekSelect(${tbId})" style="width: 100%; padding: 10px; border: 2px solid #dc2626; border-radius: 6px;">
                        <option value="weekly">Her hafta</option>
                        <option value="custom">Özel hafta seçimi</option>
                    </select>
                    <div id="teacherBlockWeekContainer${tbId}" style="margin-top: 10px;"></div>
                </div>

                <!-- Gün Seçimi -->
                <div style="margin-bottom: 15px;">
                    <strong style="display: block; margin-bottom: 8px;">Gün Seçimi:</strong>
                    <select class="teacher-block-day" data-tbid="${tbId}" onchange="updateTeacherBlockLessonList(${tbId})" style="width: 100%; padding: 10px; border: 2px solid #dc2626; border-radius: 6px;">
                        <option value="all">Tüm günler</option>
                        <option value="Pazartesi">Pazartesi</option>
                        <option value="Salı">Salı</option>
                        <option value="Çarşamba">Çarşamba</option>
                        <option value="Perşembe">Perşembe</option>
                        <option value="Cuma">Cuma</option>
                        <option value="Cumartesi">Cumartesi</option>
                        <option value="Pazar">Pazar</option>
                    </select>
                </div>

                <!-- Ders Saatleri -->
                <div id="teacherBlockLessonList${tbId}" style="margin-top: 10px;"></div>
            `;

            document.getElementById('studentTeacherBlockGroups').appendChild(tbGroup);

            // Öğretmenleri yükle
            loadTeachersForBlock(tbId);

            // 🔢 SAYACI GÜNCELLE
            updateStudentTeacherBlockCount();

            // 🆕 OTOMATIK SCROLL
            setTimeout(() => {
                const teacherBlockContainer = document.getElementById('studentTeacherBlockGroups');
                if (teacherBlockContainer) {
                    teacherBlockContainer.scrollTop = teacherBlockContainer.scrollHeight;
                }
            }, 100);
        }

        function removeStudentTeacherBlock(tbId) {
            const elem = document.getElementById('studentTeacherBlock' + tbId);
            if (elem) {
                elem.remove();

                // 🔢 SAYACI GÜNCELLE
                updateStudentTeacherBlockCount();
            }
        }

        function toggleTeacherBlockWeekSelect(tbId) {
            const typeSelect = document.querySelector(`.teacher-block-type[data-tbid="${tbId}"]`);
            const container = document.getElementById('teacherBlockWeekContainer' + tbId);

            if (typeSelect.value === 'custom') {
                container.innerHTML = `
                    <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px;">
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="teacher-block-week-checkbox" data-tbid="${tbId}" value="1" style="width: auto;">
                            Hafta 1
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="teacher-block-week-checkbox" data-tbid="${tbId}" value="2" style="width: auto;">
                            Hafta 2
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="teacher-block-week-checkbox" data-tbid="${tbId}" value="3" style="width: auto;">
                            Hafta 3
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="teacher-block-week-checkbox" data-tbid="${tbId}" value="4" style="width: auto;">
                            Hafta 4
                        </label>
                    </div>
                `;
            } else {
                container.innerHTML = '';
            }
        }

        async function loadTeachersForBlock(tbId) {
            const response = await fetch('/get_teachers');
            const data = await response.json();
            const teacherSelect = document.querySelector(`.teacher-block-select[data-tbid="${tbId}"]`);

            let optionsHTML = '<option value="">Öğretmen Seçiniz...</option>';

            // ✅ ALFABETIK SIRALA
            const sortedTeachers = data.teachers.sort((a, b) => {
                const nameA = `${a.name} ${a.surname}`.toLocaleLowerCase('tr');
                const nameB = `${b.name} ${b.surname}`.toLocaleLowerCase('tr');
                return nameA.localeCompare(nameB, 'tr');
            });

            sortedTeachers.forEach(teacher => {
                optionsHTML += `<option value="${teacher.id}" data-schedule='${JSON.stringify(teacher.schedule)}'>${teacher.name} ${teacher.surname} (${teacher.branch})</option>`;
            });

            teacherSelect.innerHTML = optionsHTML;
        }

        function updateTeacherBlockSlots(tbId) {
            const teacherSelect = document.querySelector(`.teacher-block-select[data-tbid="${tbId}"]`);
            const daySelect = document.querySelector(`.teacher-block-day[data-tbid="${tbId}"]`);

            if (teacherSelect.value && daySelect.value) {
                updateTeacherBlockLessonList(tbId);
            }
        }

        function updateTeacherBlockLessonList(tbId) {
            const teacherSelect = document.querySelector(`.teacher-block-select[data-tbid="${tbId}"]`);
            const daySelect = document.querySelector(`.teacher-block-day[data-tbid="${tbId}"]`);
            const container = document.getElementById('teacherBlockLessonList' + tbId);

            if (!teacherSelect.value) {
                container.innerHTML = '<p style="color: #991b1b; font-size: 0.9em;">Önce öğretmen seçiniz</p>';
                return;
            }

            const selectedOption = teacherSelect.options[teacherSelect.selectedIndex];
            const schedule = JSON.parse(selectedOption.getAttribute('data-schedule'));
            const selectedDay = daySelect.value;

            let lessons = [];

            if (selectedDay === 'all') {
                // Tüm günlerdeki dersler
                schedule.forEach(daySchedule => {
                    daySchedule.lessons.forEach(lesson => {
                        lessons.push({
                            day: daySchedule.day,
                            time: `${lesson.start_time}-${lesson.end_time}`
                        });
                    });
                });
            } else {
                // Belirli gündeki dersler
                const daySchedule = schedule.find(s => s.day === selectedDay);
                if (daySchedule) {
                    daySchedule.lessons.forEach(lesson => {
                        lessons.push({
                            day: selectedDay,
                            time: `${lesson.start_time}-${lesson.end_time}`
                        });
                    });
                }
            }

            if (lessons.length === 0) {
                container.innerHTML = '<p style="color: #991b1b; font-size: 0.9em; padding: 10px; background: white; border-radius: 5px;">Bu seçimde ders bulunamadı.</p>';
                return;
            }

            let html = '<div style="margin-bottom: 10px;"><strong style="display: block; margin-bottom: 8px;">Engellenecek Dersler:</strong></div>';
            html += '<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px;">';

            lessons.forEach(lesson => {
                const lessonKey = selectedDay === 'all' ? `${lesson.day} ${lesson.time}` : lesson.time;
                html += `
                    <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                        <input type="checkbox" class="teacher-block-lesson-checkbox" data-tbid="${tbId}" value="${lesson.day}_${lesson.time}" checked style="width: auto;">
                        ${lessonKey}
                    </label>
                `;
            });

            html += '</div>';
            container.innerHTML = html;
        }

        function addDayGroup() {
            const groupId = dayGroupCounter++;
            const dayGroup = document.createElement('div');
            dayGroup.className = 'day-group';
            dayGroup.id = 'dayGroup' + groupId;

            dayGroup.innerHTML = `
                <div class="day-group-header">
                    <div class="form-group" style="flex: 1; margin: 0; margin-right: 10px;">
                        <select class="day-select" onchange="updateDefaultDuration(this)" required>
                            <option value="">Gün Seçiniz...</option>
                            <option value="Pazartesi">Pazartesi</option>
                            <option value="Salı">Salı</option>
                            <option value="Çarşamba">Çarşamba</option>
                            <option value="Perşembe">Perşembe</option>
                            <option value="Cuma">Cuma</option>
                            <option value="Cumartesi">Cumartesi</option>
                            <option value="Pazar">Pazar</option>
                        </select>
                    </div>
                    <button type="button" class="remove-day-btn" onclick="removeDayGroup(${groupId})">Günü Sil</button>
                </div>
                <div class="lessons-container" id="lessonsContainer${groupId}"></div>
                <button type="button" class="add-day-btn" style="background: #2196F3; margin-top: 10px;" onclick="addLesson(${groupId})">+ Ders Ekle</button>
            `;

            document.getElementById('dayGroups').appendChild(dayGroup);
            addLesson(groupId);

            // 🆕 OTOMATIK SCROLL
            setTimeout(() => {
                const container = document.getElementById('dayGroups');
                if (container) {
                    container.scrollTop = container.scrollHeight;
                }
            }, 100);

            // 🔢 SAYACI GÜNCELLE - YENİ EKLENEN
            updateDayGroupCount();
        }

        function removeDayGroup(groupId) {
            document.getElementById('dayGroup' + groupId)?.remove();

            // 🔢 SAYACI GÜNCELLE - YENİ EKLENEN
            updateDayGroupCount();
        }

        function addLesson(groupId) {
            const container = document.getElementById('lessonsContainer' + groupId);
            const currentLessons = container.querySelectorAll('.lesson-slot').length;
            const lessonNumber = currentLessons + 1;

            const lessonSlot = document.createElement('div');
            lessonSlot.className = 'lesson-slot';
            lessonSlot.id = `lesson${groupId}_${lessonNumber}`;

            const dayGroup = document.getElementById('dayGroup' + groupId);
            const selectedDay = dayGroup.querySelector('.day-select').value;
            let defaultDuration = 35;
            if (selectedDay === 'Cumartesi' || selectedDay === 'Pazar') {
                defaultDuration = 40;
            }

            lessonSlot.innerHTML = `
                <input type="time" class="lesson-start" data-lesson="${lessonNumber}" onchange="calcEnd(this)" required style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                <select class="duration-select" data-lesson="${lessonNumber}" onchange="calcEnd(this.parentElement.querySelector('.lesson-start'))" style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                    ${[10,15,20,25,30,35,40,45,50,55].map(d =>
                        `<option value="${d}" ${d == defaultDuration ? 'selected' : ''}>${d} dk</option>`
                    ).join('')}
                </select>
                <input type="time" class="lesson-end" data-lesson="${lessonNumber}" readonly required style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                <button type="button" class="remove-lesson-btn" onclick="removeLesson(${groupId}, ${lessonNumber})" style="background: #f44336; color: white; border: none; padding: 5px 10px; border-radius: 5px; cursor: pointer; font-size: 0.9em;">Sil</button>
            `;

            container.appendChild(lessonSlot);

            // 🆕 OTOMATIK SCROLL
            setTimeout(() => {
                const dayGroupsContainer = document.getElementById('dayGroups');
                if (dayGroupsContainer) {
                    dayGroupsContainer.scrollTop = dayGroupsContainer.scrollHeight;
                }
            }, 100);
        }

        function removeLesson(groupId, lessonNumber) {
            const lessonSlot = document.getElementById(`lesson${groupId}_${lessonNumber}`);
            if (lessonSlot) {
                lessonSlot.remove();
                const container = document.getElementById('lessonsContainer' + groupId);
                const lessons = container.querySelectorAll('.lesson-slot');
                lessons.forEach((lesson, index) => {
                    const newNumber = index + 1;
                    lesson.id = `lesson${groupId}_${newNumber}`;
                    lesson.querySelector('.lesson-start').setAttribute('data-lesson', newNumber);
                    lesson.querySelector('.duration-select').setAttribute('data-lesson', newNumber);
                    lesson.querySelector('.lesson-end').setAttribute('data-lesson', newNumber);
                    const removeBtn = lesson.querySelector('.remove-lesson-btn');
                    removeBtn.setAttribute('onclick', `removeLesson(${groupId}, ${newNumber})`);
                });
            }
        }

        function updateDefaultDuration(daySelect) {
            const selectedDay = daySelect.value;
            const dayGroup = daySelect.closest('.day-group');

            let defaultDuration = 35;
            if (selectedDay === 'Cumartesi' || selectedDay === 'Pazar') {
                defaultDuration = 40;
            }

            const durationSelects = dayGroup.querySelectorAll('.duration-select');
            durationSelects.forEach(select => {
                select.value = defaultDuration;
                const lessonNum = select.getAttribute('data-lesson');
                const startInput = dayGroup.querySelector(`.lesson-start[data-lesson="${lessonNum}"]`);
                if (startInput && startInput.value) {
                    calcEnd(startInput);
                }
            });
        }

        function addRestriction() {
            const rId = restrictionCounter++;
            const rGroup = document.createElement('div');
            rGroup.className = 'restriction-group';
            rGroup.id = 'restriction' + rId;

            rGroup.innerHTML = `
                <div class="restriction-header">
                    <span class="restriction-title">Kısıtlama ${rId + 1}</span>
                    <button type="button" class="remove-restriction-btn" onclick="removeRestriction(${rId})">Sil</button>
                </div>
                <!-- Hafta Seçimi -->
                <div style="margin-bottom: 15px;">
                    <strong style="display: block; margin-bottom: 8px;">Hafta Seçimi:</strong>
                    <select class="restriction-type" data-rid="${rId}" onchange="toggleWeekSelect(${rId})" style="width: 100%; padding: 10px; border: 2px solid #ddd; border-radius: 6px;">
                        <option value="weekly">Her hafta</option>
                        <option value="custom">Özel hafta seçimi</option>
                    </select>
                    <div id="weekNumContainer${rId}" style="margin-top: 10px;"></div>
                </div>

                <!-- Gün Seçimi (Çoklu Checkbox) -->
                <div style="margin-bottom: 15px;">
                    <strong style="display: block; margin-bottom: 8px;">Günler:</strong>
                    <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px;">
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="Pazartesi" style="width: auto;">
                            Pazartesi
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="Salı" style="width: auto;">
                            Salı
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="Çarşamba" style="width: auto;">
                            Çarşamba
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="Perşembe" style="width: auto;">
                            Perşembe
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="Cuma" style="width: auto;">
                            Cuma
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="Cumartesi" style="width: auto;">
                            Cumartesi
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-day-checkbox" data-rid="${rId}" value="Pazar" style="width: auto;">
                            Pazar
                        </label>
                    </div>
                </div>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: 10px;">
                    <input type="time" class="restriction-start" data-rid="${rId}" placeholder="Başlangıç" style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                    <input type="time" class="restriction-end" data-rid="${rId}" placeholder="Bitiş" style="padding: 8px; border: 2px solid #ddd; border-radius: 6px;">
                </div>
            `;

            document.getElementById('restrictionGroups').appendChild(rGroup);

            // 🔢 SAYACI GÜNCELLE
            updateRestrictionCount();

            // 🆕 OTOMATIK SCROLL
            setTimeout(() => {
                const container = document.getElementById('restrictionGroups');
                if (container) {
                    container.scrollTop = container.scrollHeight;
                }
            }, 100);
        }

        function removeRestriction(rId) {
            document.getElementById('restriction' + rId)?.remove();

            // 🔢 SAYACI GÜNCELLE
            updateRestrictionCount();
        }

        function toggleWeekSelect(rId) {
            const typeSelect = document.querySelector(`.restriction-type[data-rid="${rId}"]`);
            const container = document.getElementById('weekNumContainer' + rId);

            if (typeSelect.value === 'custom') {
                container.innerHTML = `
                    <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px;">
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-week-checkbox" data-rid="${rId}" value="1" style="width: auto;">
                            Hafta 1
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-week-checkbox" data-rid="${rId}" value="2" style="width: auto;">
                            Hafta 2
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-week-checkbox" data-rid="${rId}" value="3" style="width: auto;">
                            Hafta 3
                        </label>
                        <label style="display: flex; align-items: center; gap: 5px; padding: 8px; background: white; border-radius: 5px; cursor: pointer;">
                            <input type="checkbox" class="restriction-week-checkbox" data-rid="${rId}" value="4" style="width: auto;">
                            Hafta 4
                        </label>
                    </div>
                `;
            } else {
                container.innerHTML = '';
            }
        }

        function calcEnd(startInput) {
            const lessonNum = startInput.getAttribute('data-lesson');
            const group = startInput.closest('.day-group');
            const durationSelect = group.querySelector(`.duration-select[data-lesson="${lessonNum}"]`);
            const endInput = group.querySelector(`.lesson-end[data-lesson="${lessonNum}"]`);

            if (startInput.value && durationSelect) {
                const [h, m] = startInput.value.split(':').map(Number);
                const duration = parseInt(durationSelect.value);
                const totalMins = h * 60 + m + duration;
                const endH = Math.floor(totalMins / 60);
                const endM = totalMins % 60;
                endInput.value = `${String(endH).padStart(2,'0')}:${String(endM).padStart(2,'0')}`;
            }
        }

        async function saveTeacher(event) {
            event.preventDefault();
            const teacherId = document.getElementById('teacherId').value;
            const name = document.getElementById('teacherName').value;
            const surname = document.getElementById('teacherSurname').value;
            const branch = document.getElementById('teacherBranch').value;

            const schedule = [];
            document.querySelectorAll('.day-group').forEach(group => {
                const day = group.querySelector('.day-select').value;
                if (!day) return;

                const lessons = [];
                const lessonSlots = group.querySelectorAll('.lesson-slot');

                lessonSlots.forEach((slot, index) => {
                    const start = slot.querySelector('.lesson-start')?.value;
                    const end = slot.querySelector('.lesson-end')?.value;
                    const duration = slot.querySelector('.duration-select')?.value;

                    if (start && end) {
                        lessons.push({
                            start_time: start,
                            end_time: end,
                            duration: parseInt(duration || 35)
                        });
                    }
                });

                if (lessons.length > 0) {
                    // ✅ KAYDETMEDEN ÖNCE SAATE GÖRE SIRALA
                    lessons.sort((a, b) => a.start_time.localeCompare(b.start_time));
                    schedule.push({ day, lessons });
                }
            });

            if (schedule.length === 0) {
                showError('Lütfen en az bir gün ve ders saati ekleyin!');
                return;
            }

            // 🆕 BLOKLAMALARI TOPLA
            const blocked_slots = [];
            document.querySelectorAll('[id^="teacherBlock"]').forEach(bGroup => {
                const bId = bGroup.id.replace('teacherBlock', '');

                const typeSelect = bGroup.querySelector(`.block-type[data-bid="${bId}"]`);
                const daySelect = bGroup.querySelector(`.block-day[data-bid="${bId}"]`);

                if (!typeSelect || !daySelect || !daySelect.value) return;

                const type = typeSelect.value;
                const day = daySelect.value;

                // Hafta bilgisi topla
                let weeks = [];
                if (type === 'custom') {
                    const weekCheckboxes = bGroup.querySelectorAll(`.block-week-checkbox[data-bid="${bId}"]:checked`);
                    weeks = Array.from(weekCheckboxes).map(cb => parseInt(cb.value));

                    if (weeks.length === 0) {
                        alert('Bloklama için en az bir hafta seçmelisiniz!');
                        return;
                    }
                }

                // Bloklanacak dersleri topla
                const lessonCheckboxes = bGroup.querySelectorAll(`.block-lesson-checkbox[data-bid="${bId}"]:checked`);
                const blocked_lessons = Array.from(lessonCheckboxes).map(cb => cb.value);

                if (blocked_lessons.length === 0) return;

                blocked_slots.push({
                    type: type,
                    weeks: weeks,
                    day: day,
                    blocked_slots: blocked_lessons
                });
            });

            const url = teacherId ? '/update_teacher' : '/add_teacher';
            const payload = teacherId ?
                { id: parseInt(teacherId), name, surname, branch, schedule, blocked_slots } :
                { name, surname, branch, schedule, blocked_slots };

            const response = await fetch(url, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(payload)
            });

            const result = await response.json();
            showSuccess(result.message);
            closeTeacherModal();
            loadTeachers();
        }

        async function deleteTeacher(teacherId) {
            if (!confirm('Bu öğretmeni silmek istediğinizden emin misiniz?')) return;
            const response = await fetch('/delete_teacher', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ id: teacherId })
            });
            const result = await response.json();
            showSuccess(result.message);
            loadTeachers();
        }

        async function saveStudent(event) {
            event.preventDefault();
            const studentId = document.getElementById('studentId').value;
            const name = document.getElementById('studentName').value;
            const surname = document.getElementById('studentSurname').value;
            const className = document.getElementById('studentClass').value;

            // KISITLAMALARI TOPLA
            const restrictions = [];
            document.querySelectorAll('.restriction-group').forEach(rGroup => {
                const rId = rGroup.querySelector('.restriction-type')?.getAttribute('data-rid');
                if (!rId) return;

                const type = rGroup.querySelector(`.restriction-type[data-rid="${rId}"]`)?.value;

                // ✅ ÇOKLU GÜN TOPLAMA
                const dayCheckboxes = rGroup.querySelectorAll(`.restriction-day-checkbox[data-rid="${rId}"]:checked`);
                const days = Array.from(dayCheckboxes).map(cb => cb.value);

                if (days.length === 0) {
                    alert(`Kısıtlama ${parseInt(rId) + 1} için en az bir gün seçmelisiniz!`);
                    return;
                }

                // ✅ ÇOKLU HAFTA TOPLAMA
                let weeks = [];
                if (type === 'custom') {
                    const weekCheckboxes = rGroup.querySelectorAll(`.restriction-week-checkbox[data-rid="${rId}"]:checked`);
                    weeks = Array.from(weekCheckboxes).map(cb => parseInt(cb.value));

                    if (weeks.length === 0) {
                        alert(`Kısıtlama ${parseInt(rId) + 1} için en az bir hafta seçmelisiniz!`);
                        return;
                    }
                }

                const startTime = rGroup.querySelector(`.restriction-start[data-rid="${rId}"]`)?.value;
                const endTime = rGroup.querySelector(`.restriction-end[data-rid="${rId}"]`)?.value;

                if (startTime && endTime && days.length > 0) {
                    const restriction = {
                        type,
                        days,              // ✅ Artık array
                        weeks,             // ✅ Artık array
                        start_time: startTime,
                        end_time: endTime
                    };
                    restrictions.push(restriction);
                }
            });

            // ÖNCELİKLERİ TOPLA
            const priorities = {
                week1: [],
                week2: [],
                week3: [],
                week4: []
            };

            for (let week = 1; week <= 4; week++) {
                const container = document.getElementById(`week${week}PriorityList`);
                if (container) {
                    const priorityElements = container.querySelectorAll('.priority-branch');
                    priorityElements.forEach(select => {
                        const branch = select.value;
                        if (branch) {
                            priorities[`week${week}`].push(branch);
                        }
                    });
                }
            }

            // MANUEL DERSLERİ TOPLA
            const manualLessons = [];
            document.querySelectorAll('[id^="manual_"]').forEach(manualDiv => {
                const mId = manualDiv.id.replace('manual_', '');

                const weekSelect = manualDiv.querySelector(`.manual-week[data-mid="${mId}"]`);
                const daySelect = manualDiv.querySelector(`.manual-day[data-mid="${mId}"]`);
                const teacherSelect = document.getElementById(`manualTeacher_${mId}`);
                const timeSelect = document.getElementById(`manualTime_${mId}`);

                if (weekSelect && daySelect && teacherSelect && timeSelect) {
                    const week = weekSelect.value;
                    const day = daySelect.value;
                    const teacherId = teacherSelect.value;
                    const time = timeSelect.value;

                    if (week && day && teacherId && time) {
                        const teacherOption = teacherSelect.options[teacherSelect.selectedIndex];
                        const teacherText = teacherOption.textContent;

                        manualLessons.push({
                            week: parseInt(week),
                            day: day,
                            teacher_id: parseInt(teacherId),
                            teacher_name: teacherText,
                            time: time
                        });
                    }
                }
            });

            // ÖĞRETMEN ENGELLEMELERİNİ TOPLA
            const teacherBlocks = [];
            document.querySelectorAll('[id^="studentTeacherBlock"]').forEach(tbGroup => {
                const tbId = tbGroup.id.replace('studentTeacherBlock', '');

                const teacherSelect = tbGroup.querySelector(`.teacher-block-select[data-tbid="${tbId}"]`);
                const typeSelect = tbGroup.querySelector(`.teacher-block-type[data-tbid="${tbId}"]`);
                const daySelect = tbGroup.querySelector(`.teacher-block-day[data-tbid="${tbId}"]`);

                if (!teacherSelect || !teacherSelect.value) return;

                const teacherId = teacherSelect.value;
                const type = typeSelect.value;
                const day = daySelect.value;

                // Hafta bilgisi topla
                let weeks = [];
                if (type === 'custom') {
                    const weekCheckboxes = tbGroup.querySelectorAll(`.teacher-block-week-checkbox[data-tbid="${tbId}"]:checked`);
                    weeks = Array.from(weekCheckboxes).map(cb => parseInt(cb.value));

                    if (weeks.length === 0) {
                        alert('Öğretmen engelleme için en az bir hafta seçmelisiniz!');
                        return;
                    }
                }

                // Engellenecek dersleri topla
                const lessonCheckboxes = tbGroup.querySelectorAll(`.teacher-block-lesson-checkbox[data-tbid="${tbId}"]:checked`);
                const blockedSlots = Array.from(lessonCheckboxes).map(cb => cb.value);

                if (blockedSlots.length === 0) return;

                teacherBlocks.push({
                    teacher_id: parseInt(teacherId),
                    type: type,
                    weeks: weeks,
                    day: day,
                    blocked_slots: blockedSlots
                });
            });

            // KAYDET
            const url = studentId ? '/update_student' : '/add_student';
            const payload = studentId ?
                {
                    id: parseInt(studentId),
                    name,
                    surname,
                    class: className,
                    restrictions,
                    priorities,
                    manual_lessons: manualLessons,
                    teacher_blocks: teacherBlocks
                } :
                {
                    name,
                    surname,
                    class: className,
                    restrictions,
                    priorities,
                    manual_lessons: manualLessons,
                    teacher_blocks: teacherBlocks
                };

            const response = await fetch(url, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(payload)
            });

            const result = await response.json();
            showSuccess(result.message);
            closeStudentModal();
            loadStudents();
        }

        async function deleteStudent(studentId) {
            if (!confirm('Bu öğrenciyi silmek istediğinizden emin misiniz?')) return;
            const response = await fetch('/delete_student', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ id: studentId })
            });
            const result = await response.json();
            showSuccess(result.message);
            loadStudents();
        }

        async function loadTeachers() {
            const response = await fetch('/get_teachers');
            const data = await response.json();
            document.getElementById('teacherCount').textContent = data.teachers.length;

            // ✅ SIRALAMA - Mevcut duruma göre
            data.teachers.sort((a, b) => {
                let compareResult = 0;

                if (teacherSortState.column === 'name') {
                    // Ad-Soyad'a göre sırala
                    const nameA = `${a.name} ${a.surname}`.toLocaleLowerCase('tr');
                    const nameB = `${b.name} ${b.surname}`.toLocaleLowerCase('tr');
                    compareResult = nameA.localeCompare(nameB, 'tr');
                } else if (teacherSortState.column === 'branch') {
                    // Branş'a göre sırala (eşitse ad-soyada bak)
                    compareResult = a.branch.localeCompare(b.branch, 'tr');
                    if (compareResult === 0) {
                        const nameA = `${a.name} ${a.surname}`.toLocaleLowerCase('tr');
                        const nameB = `${b.name} ${b.surname}`.toLocaleLowerCase('tr');
                        compareResult = nameA.localeCompare(nameB, 'tr');
                    }
                }

                // Direction'a göre ters çevir
                return teacherSortState.direction === 'asc' ? compareResult : -compareResult;
            });

            const tbody = document.getElementById('teacherTableBody');

            if (data.teachers.length === 0) {
                tbody.innerHTML = '<tr><td colspan="5" style="text-align: center; padding: 20px; color: #999;">Henüz öğretmen eklenmedi.</td></tr>';
                return;
            }

            // ✅ TABLO BAŞLIKLARINI GÜNCELLe (Sıralanabilir yap)
            const theadRow = tbody.closest('table').querySelector('thead tr');
            if (theadRow && theadRow.children.length >= 2) {
                // Ad Soyad başlığı
                const nameHeader = theadRow.children[0];
                nameHeader.style.cursor = 'pointer';
                nameHeader.style.userSelect = 'none';
                nameHeader.onclick = () => sortTeachers('name');
                nameHeader.innerHTML = `Ad Soyad${getSortIcon(teacherSortState.column, 'name', teacherSortState.direction)}`;

                // Branş başlığı
                const branchHeader = theadRow.children[1];
                branchHeader.style.cursor = 'pointer';
                branchHeader.style.userSelect = 'none';
                branchHeader.onclick = () => sortTeachers('branch');
                branchHeader.innerHTML = `Branş${getSortIcon(teacherSortState.column, 'branch', teacherSortState.direction)}`;
            }

            tbody.innerHTML = data.teachers.map(t => {
                // 🆕 BLOKLAMA KONTROLÜ
                const hasBlocks = t.blocked_slots && t.blocked_slots.length > 0;

                return `
                    <tr>
                        <td>
                            <strong>${t.name} ${t.surname}</strong>
                            ${globalScheduleData ? `
                                <br>
                                <button class="view-btn" onclick="showTeacherScheduleDetail('${t.name} ${t.surname}', '${t.branch}')" style="margin-top: 8px; padding: 6px 12px; font-size: 0.85em;">
                                    📋 Tüm Dersler
                                </button>
                            ` : ''}
                        </td>
                        <td>${t.branch}</td>
                        <td>
                            <button class="view-btn" onclick='showTeacherDetail(${JSON.stringify(t).replace(/'/g, "&apos;")})'>
                                Göster
                            </button>
                        </td>
                        <td style="padding: 12px 15px;">
                             ${hasBlocks ?
                                `<button class="view-btn" style="background: #ef4444;" onclick='showTeacherBlocks(${JSON.stringify(t).replace(/'/g, "&apos;")})'>Göster</button>`
                                : '<span style="color: #999;">Yok</span>'}
                        </td>
                        <td>
                            <div class="action-buttons">
                                <button class="edit-btn-small" onclick="openTeacherModal(${t.id})">Düzenle</button>
                                <button class="delete-btn-small" onclick="deleteTeacher(${t.id})">Sil</button>
                            </div>
                        </td>
                    </tr>
                `;
            }).join('');
        }

        // 🔍 ÖĞRETMEN ARAMA/FİLTRELEME
        function filterTeachers() {
            const searchInput = document.getElementById('teacherSearchInput').value.toLocaleLowerCase('tr').trim();
            const tbody = document.getElementById('teacherTableBody');
            const rows = tbody.getElementsByTagName('tr');
            let visibleCount = 0;
            let totalCount = rows.length;

            for (let i = 0; i < rows.length; i++) {
                const row = rows[i];
                const cells = row.getElementsByTagName('td');

                // Boş satır kontrolü
                if (cells.length < 2) continue;

                // ✅ DÜZELTME: Sadece <strong> tag'inin içindeki ismi al + Türkçe locale
                const strongElement = cells[0].querySelector('strong');
                const nameText = strongElement
                    ? strongElement.textContent.toLocaleLowerCase('tr')
                    : cells[0].textContent.toLocaleLowerCase('tr');

                const branchText = cells[1].textContent.toLocaleLowerCase('tr');
                const fullText = nameText + ' ' + branchText;

                // Arama metnini kontrol et
                if (fullText.includes(searchInput)) {
                    row.style.display = '';
                    visibleCount++;
                } else {
                    row.style.display = 'none';
                }
            }

            // Sonuç sayacını güncelle
            const searchCount = document.getElementById('teacherSearchCount');
            if (searchInput === '') {
                searchCount.textContent = '';
            } else {
                searchCount.textContent = `${visibleCount} / ${totalCount} öğretmen gösteriliyor`;
                if (visibleCount === 0) {
                    searchCount.innerHTML = '<span style="color: #ef4444;">❌ Eşleşen öğretmen bulunamadı</span>';
                }
            }
        }

        async function loadStudents() {
            const response = await fetch('/get_students');
            const data = await response.json();
            document.getElementById('studentCount').textContent = data.students.length;

            // ✅ SIRALAMA - Mevcut duruma göre
            data.students.sort((a, b) => {
                let compareResult = 0;

                if (studentSortState.column === 'name') {
                    // Ad-Soyad'a göre sırala
                    const nameA = `${a.name} ${a.surname}`.toLocaleLowerCase('tr');
                    const nameB = `${b.name} ${b.surname}`.toLocaleLowerCase('tr');
                    compareResult = nameA.localeCompare(nameB, 'tr');
                } else if (studentSortState.column === 'class') {
                    // Sınıf'a göre sırala (eşitse ad-soyada bak)
                    compareResult = a.class.localeCompare(b.class, 'tr');
                    if (compareResult === 0) {
                        const nameA = `${a.name} ${a.surname}`.toLocaleLowerCase('tr');
                        const nameB = `${b.name} ${b.surname}`.toLocaleLowerCase('tr');
                        compareResult = nameA.localeCompare(nameB, 'tr');
                    }
                }

                return studentSortState.direction === 'asc' ? compareResult : -compareResult;
            });

            const tbody = document.getElementById('studentTableBody');

            if (data.students.length === 0) {
                tbody.innerHTML = '<tr><td colspan="7" style="text-align: center; padding: 20px; color: #999;">Henüz öğrenci eklenmedi.</td></tr>';
                return;
            }

            // ✅ TABLO BAŞLIKLARINI GÜNCELLe
            const theadRow = tbody.closest('table').querySelector('thead tr');
            if (theadRow && theadRow.children.length >= 2) {
                // Ad Soyad başlığı
                const nameHeader = theadRow.children[0];
                nameHeader.style.cursor = 'pointer';
                nameHeader.style.userSelect = 'none';
                nameHeader.onclick = () => sortStudents('name');
                nameHeader.innerHTML = `Ad Soyad${getSortIcon(studentSortState.column, 'name', studentSortState.direction)}`;

                // Sınıf başlığı
                const classHeader = theadRow.children[1];
                classHeader.style.cursor = 'pointer';
                classHeader.style.userSelect = 'none';
                classHeader.onclick = () => sortStudents('class');
                classHeader.innerHTML = `Sınıf${getSortIcon(studentSortState.column, 'class', studentSortState.direction)}`;
            }

            tbody.innerHTML = data.students.map(s => {
                const hasRestrictions = s.restrictions && s.restrictions.length > 0;
                const hasPriorities = s.priorities && Object.values(s.priorities).some(arr => arr.length > 0);
                const hasManualLessons = s.manual_lessons && s.manual_lessons.length > 0;
                const hasTeacherBlocks = s.teacher_blocks && s.teacher_blocks.length > 0;

                return `
                    <tr>
                        <td>
                            <strong>${s.name} ${s.surname}</strong>
                            ${globalScheduleData ? `
                                <br>
                                <button class="view-btn" onclick="showStudentScheduleDetail('${s.name} ${s.surname}', '${s.class}')" style="margin-top: 8px; padding: 6px 12px; font-size: 0.85em;">
                                    📋 Tüm Dersler
                                </button>
                            ` : ''}
                        </td>
                        <td>${s.class}</td>
                        <td>
                            ${hasRestrictions ?
                                `<button class="view-btn" onclick='showStudentRestrictions(${JSON.stringify(s).replace(/'/g, "&apos;")})'>Göster</button>`
                                : '<span style="color: #999;">Yok</span>'}
                        </td>
                        <td>
                            ${hasPriorities ?
                                `<button class="view-btn" onclick='showStudentPriorities(${JSON.stringify(s).replace(/'/g, "&apos;")})'>Göster</button>`
                                : '<span style="color: #999;">Yok</span>'}
                        </td>
                        <td>
                            ${hasManualLessons ?
                                `<button class="view-btn" onclick='showStudentManualLessons(${JSON.stringify(s).replace(/'/g, "&apos;")})'>Göster</button>`
                                : '<span style="color: #999;">Yok</span>'}
                        </td>
                        <td>
                            ${hasTeacherBlocks ?
                                `<button class="view-btn" onclick='showStudentTeacherBlocks(${JSON.stringify(s).replace(/'/g, "&apos;")})'>Göster</button>`
                                : '<span style="color: #999;">Yok</span>'}
                        </td>
                        <td>
                            <div class="action-buttons">
                                <button class="edit-btn-small" onclick="openStudentModal(${s.id})">Düzenle</button>
                                <button class="delete-btn-small" onclick="deleteStudent(${s.id})">Sil</button>
                            </div>
                        </td>
                    </tr>
                `;
            }).join('');
        }

        // 📅 GÜN SEÇİCİ STATE
        let selectedDay = null; // null = bugün, yoksa seçilen gün adı
        let savedScheduleData = null; // Kaydedilmiş program verisi

        // 📅 SEÇİLİ GÜNÜN DERSLERİNİ YÜKLE
        async function loadTodayLessons(selectedDateStr = null) {
            try {
                // ============== GLOBAL SCHEDULE VAR MI KONTROL ==============
                if (!globalScheduleData || !globalScheduleData.start_date) {
                    document.getElementById('daySelector').style.display = 'none';
                    document.getElementById('todayLessonsCount').textContent = '0';
                    document.getElementById('todayLessonsWidget').innerHTML = `
                        <div style="text-align: center; padding: 40px; color: #6b7280;">
                            <i class="fas fa-calendar-day" style="font-size: 3em; margin-bottom: 15px; opacity: 0.3;"></i>
                            <p>Program bulunamadı.</p>
                            <p style="font-size: 0.9em; opacity: 0.7;">Lütfen önce bir program oluşturun.</p>
                        </div>
                    `;
                    return;
                }

                // ============== 28 GÜNLÜK TARİH LİSTESİ OLUŞTUR ==============
                const startDate = new Date(globalScheduleData.start_date + 'T00:00:00');
                const allDates = [];
                const dayNamesLong = ['Pazar', 'Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi'];
                const monthNames = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık'];

                for (let i = 0; i < 28; i++) {
                    const currentDate = new Date(startDate);
                    currentDate.setDate(startDate.getDate() + i);

                    const dayName = dayNamesLong[currentDate.getDay()];
                    const day = currentDate.getDate();
                    const month = monthNames[currentDate.getMonth()];
                    const year = currentDate.getFullYear();
                    const weekNum = Math.floor(i / 7) + 1;

                    // ✅ LOCAL TARİH FORMATLA (timezone sorunu yok)
                    const dateStr = `${year}-${String(currentDate.getMonth() + 1).padStart(2, '0')}-${String(day).padStart(2, '0')}`;

                    allDates.push({
                        dateObj: currentDate,
                        dateStr: dateStr,
                        displayStr: `${day} ${month} ${year} ${dayName}`,
                        dayName: dayName,
                        weekNum: weekNum
                    });
                }

                // ============== DROPDOWN'U DOLDUR ==============
                const dropdown = document.getElementById('dayDropdown');
                dropdown.innerHTML = '';

                allDates.forEach(dateInfo => {
                    const option = document.createElement('option');
                    option.value = dateInfo.dateStr;
                    option.textContent = dateInfo.displayStr;
                    dropdown.appendChild(option);
                });

                // ============== BUGÜNÜ VEYA SEÇİLİ TARİHİ BUL ==============
                let selectedDate;
                if (selectedDateStr) {
                    selectedDate = allDates.find(d => d.dateStr === selectedDateStr);
                } else {
                    // Bugüne en yakın tarihi bul
                    const today = new Date();
                    selectedDate = allDates.find(d => d.dateObj.toDateString() === today.toDateString());
                    if (!selectedDate) {
                        selectedDate = allDates[0]; // İlk tarihi varsayılan yap
                    }
                }

                dropdown.value = selectedDate.dateStr;

                // Gün seçiciyi göster
                document.getElementById('daySelector').style.display = 'flex';

                // ============== SEÇİLİ TARİHİN DERSLERİNİ AL ==============
                let dayLessons = [];
                if (globalScheduleData.weeks && globalScheduleData.weeks[selectedDate.weekNum - 1]) {
                    const week = globalScheduleData.weeks[selectedDate.weekNum - 1];
                    week.forEach(lesson => {
                        if (lesson.day === selectedDate.dayName) {
                            dayLessons.push({
                                ...lesson,
                                weekNum: selectedDate.weekNum
                            });
                        }
                    });
                }

                // Badge'i güncelle
                document.getElementById('todayLessonsCount').textContent = dayLessons.length;

                if (dayLessons.length === 0) {
                    document.getElementById('todayLessonsWidget').innerHTML = `
                        <div style="text-align: center; padding: 40px; color: #6b7280;">
                            <i class="fas fa-calendar-day" style="font-size: 3em; margin-bottom: 15px; color: #10b981; opacity: 0.3;"></i>
                            <p style="font-size: 1.2em; font-weight: 600;">Bu gün ders yok! 🎉</p>
                            <p style="font-size: 0.9em; opacity: 0.7;">${selectedDate.displayStr} için ders bulunmuyor.</p>
                        </div>
                    `;
                    return;
                }

                // ============== BRANŞLARA GÖRE GRUPLA ==============
                const lessonsByBranch = {};
                dayLessons.forEach(lesson => {
                    const branch = lesson.branch;
                    if (!lessonsByBranch[branch]) {
                        lessonsByBranch[branch] = [];
                    }
                    lessonsByBranch[branch].push(lesson);
                });

                // Branşları alfabetik sırala
                const sortedBranches = Object.keys(lessonsByBranch).sort((a, b) =>
                    a.toLocaleLowerCase('tr').localeCompare(b.toLocaleLowerCase('tr'), 'tr')
                );

                // Her branş içindeki dersleri saate göre sırala
                sortedBranches.forEach(branch => {
                    lessonsByBranch[branch].sort((a, b) => a.time.localeCompare(b.time));
                });

                // ============== HTML OLUŞTUR ==============
                let html = `
                    <div style="background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); padding: 20px; border-radius: 12px; margin-bottom: 20px; border-left: 4px solid #10b981;">
                        <h3 style="margin: 0 0 8px 0; color: #065f46; display: flex; align-items: center; gap: 10px;">
                            <i class="fas fa-calendar-check"></i> ${selectedDate.displayStr}
                        </h3>
                        <p style="margin: 0; color: #047857; font-size: 0.95em;">
                            <strong>${dayLessons.length} ders</strong> planlanmış • Hafta ${selectedDate.weekNum}
                        </p>
                    </div>

                    <div style="display: flex; flex-direction: column; gap: 20px;">
                `;

                // Her branş için grup oluştur
                sortedBranches.forEach(branch => {
                    const branchLessons = lessonsByBranch[branch];

                    html += `
                        <div style="background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); border-radius: 15px; padding: 20px; border-left: 5px solid #3b82f6;">
                            <h4 style="margin: 0 0 15px 0; color: #1e40af; font-size: 1.3em; display: flex; align-items: center; gap: 10px;">
                                <i class="fas fa-book-open"></i> ${branch}
                                <span style="background: #3b82f6; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7em; font-weight: 600;">${branchLessons.length} ders</span>
                            </h4>
                            <div style="display: flex; flex-direction: column; gap: 12px;">
                    `;

                    // 🆕 AYNI SLOT'TAKİ DERSLERİ GRUPLA (GÜNLÜK DERSLER)
                    const groupedLessons = [];
                    const processedSlots = new Set();

                    branchLessons.forEach(lesson => {
                        const slotKey = `${lesson.time}_${lesson.teacher_name}`;

                        if (processedSlots.has(slotKey)) {
                            return; // Zaten işlendi
                        }

                        // Aynı slot'taki tüm dersleri bul
                        const sameslotLessons = branchLessons.filter(l =>
                            l.time === lesson.time && l.teacher_name === lesson.teacher_name
                        );

                        processedSlots.add(slotKey);

                        if (sameslotLessons.length === 1) {
                            // Tek ders
                            groupedLessons.push(lesson);
                        } else {
                            // 🆕 GRUP DERSİ - TÜM SINIFLARI TOPLA
                            const uniqueClasses = [...new Set(sameslotLessons.map(l => l.student_class).filter(c => c))];
                            let displayName;
                            if (uniqueClasses.length > 0) {
                                const classesStr = uniqueClasses.sort().join(', ');
                                displayName = `${classesStr} (${sameslotLessons.length} öğr)`;
                            } else {
                                displayName = `${sameslotLessons.length} öğrenci`;
                            }
                            groupedLessons.push({
                                ...lesson,
                                student_name: displayName,
                                is_grouped: true
                            });
                        }
                    });

                    groupedLessons.forEach(lesson => {
                        html += `
                            <div style="background: white; border-radius: 10px; padding: 15px; box-shadow: 0 2px 8px rgba(0,0,0,0.05); transition: all 0.3s;"
                                 onmouseover="this.style.transform='translateX(5px)'; this.style.boxShadow='0 4px 15px rgba(0,0,0,0.1)'"
                                 onmouseout="this.style.transform='translateX(0)'; this.style.boxShadow='0 2px 8px rgba(0,0,0,0.05)'">
                                <div style="display: grid; grid-template-columns: 140px 1fr 1fr; gap: 15px; align-items: center;">
                                    <div style="background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%); color: white; padding: 8px 4px; border-radius: 8px; text-align: center; font-weight: 700; font-size: 0.95em; white-space: nowrap;">
                                        ⏰ ${lesson.time}
                                    </div>
                                    <div style="padding: 8px;">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">${lesson.is_grouped ? '🎓 Sınıf' : '👨‍🎓 Öğrenci'}</div>
                                        <div style="font-weight: 600; color: #1f2937; font-size: 1em;">${lesson.student_name}</div>
                                    </div>
                                    <div style="padding: 8px;">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">👨‍🏫 Öğretmen</div>
                                        <div style="font-weight: 600; color: #1f2937; font-size: 1em;">${lesson.teacher_name}</div>
                                    </div>
                                </div>
                            </div>
                        `;
                    });

                    html += `
                            </div>
                        </div>
                    `;
                });

                html += `</div>`;

                document.getElementById('todayLessonsWidget').innerHTML = html;

            } catch (error) {
                console.error('Günlük dersler yüklenirken hata:', error);
                document.getElementById('todayLessonsCount').textContent = '0';
                document.getElementById('todayLessonsWidget').innerHTML = `
                    <div style="text-align: center; padding: 40px; color: #ef4444;">
                        <i class="fas fa-exclamation-triangle" style="font-size: 3em; margin-bottom: 15px;"></i>
                        <p>Dersler yüklenirken bir hata oluştu.</p>
                    </div>
                `;
            }
        }

        // 📅 GÜN DEĞİŞTİRME FONKSİYONLARI
        // 📅 GÜN DEĞİŞTİRME FONKSİYONLARI (TARİH BAZLI)

        function selectDayByDate() {
            const selectedDateStr = document.getElementById('dayDropdown').value;
            loadTodayLessons(selectedDateStr);
        }

        function changeDayView(direction) {
            const dropdown = document.getElementById('dayDropdown');
            const options = Array.from(dropdown.options);
            const currentIndex = dropdown.selectedIndex;

            let newIndex = currentIndex + direction;
            if (newIndex < 0) newIndex = options.length - 1;
            if (newIndex >= options.length) newIndex = 0;

            dropdown.selectedIndex = newIndex;
            loadTodayLessons(options[newIndex].value);
        }

        function goToToday() {
            const today = new Date();

            // ✅ LOCAL TARİH (timezone sorunu çözüldü)
            const year = today.getFullYear();
            const month = String(today.getMonth() + 1).padStart(2, '0');
            const day = String(today.getDate()).padStart(2, '0');
            const todayStr = `${year}-${month}-${day}`;

            // ⚠️ KRİTİK KONTROL: Dropdown var mı ve dolu mu?
            const dropdown = document.getElementById('dayDropdown');
            if (!dropdown || dropdown.options.length === 0) {
                console.error('Dropdown boş! Program henüz yüklenmemiş.');
                showError('Lütfen önce bir program oluşturun.');
                return;
            }

            // Dropdown'da bugünü bul ve seç
            const options = Array.from(dropdown.options);
            const todayOption = options.find(opt => opt.value === todayStr);

            if (todayOption) {
                // ✅ Bugün program içinde
                dropdown.value = todayStr;
                loadTodayLessons(todayStr);
            } else {
                // ❌ Bugün program dışında - MODAL GÖSTER

                // ✅ Dropdown boş mu kontrol et
                if (options.length === 0) {
                    showError('Program tarihleri yüklenemedi. Lütfen sayfayı yenileyin.');
                    return;
                }

                // ✅ Bugünün timestamp'ini al
                const todayTime = today.getTime();

                const firstOption = options[0];
                const lastOption = options[options.length - 1];

                const firstDate = new Date(firstOption.value);
                const lastDate = new Date(lastOption.value);

                const firstDay = firstDate.getDate();
                const firstMonth = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık'][firstDate.getMonth()];
                const firstYear = firstDate.getFullYear();
                const firstDayName = ['Pazar', 'Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi'][firstDate.getDay()];

                const lastDay = lastDate.getDate();
                const lastMonth = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık'][lastDate.getMonth()];
                const lastYear = lastDate.getFullYear();
                const lastDayName = ['Pazar', 'Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi'][lastDate.getDay()];

                let message = '';
                let targetDate = '';

                if (firstDate.getTime() > todayTime) {
                    // Program henüz başlamadı
                    message = `⚠️ Bugün Program Dışında<br><br>Program henüz başlamadı.<br><br><strong>İlk ders: ${firstDay} ${firstMonth} ${firstYear} ${firstDayName}</strong>`;
                    targetDate = firstOption.value;
                } else {
                    // Program bitti
                    message = `⚠️ Bugün Program Dışında<br><br>Program sona erdi.<br><br><strong>Son ders: ${lastDay} ${lastMonth} ${lastYear} ${lastDayName}</strong>`;
                    targetDate = lastOption.value;
                }

                showTodayWarningModal(message, targetDate);
            }
        }

        // 🔍 ÖĞRENCİ ARAMA/FİLTRELEME (ANA SAYFA - TABLO)
        function filterStudentsTable() {
            const searchInput = document.getElementById('mainStudentSearchInput').value.toLocaleLowerCase('tr').trim();
            const tbody = document.getElementById('studentTableBody');
            const rows = tbody.getElementsByTagName('tr');
            let visibleCount = 0;
            let totalCount = rows.length;

            for (let i = 0; i < rows.length; i++) {
                const row = rows[i];
                const cells = row.getElementsByTagName('td');

                // Boş satır kontrolü
                if (cells.length < 2) continue;

                // ✅ DÜZELTME: Sadece <strong> tag'inin içindeki ismi al + Türkçe locale
                const strongElement = cells[0].querySelector('strong');
                const nameText = strongElement
                    ? strongElement.textContent.toLocaleLowerCase('tr')
                    : cells[0].textContent.toLocaleLowerCase('tr');

                const classText = cells[1].textContent.toLocaleLowerCase('tr');
                const fullText = nameText + ' ' + classText;

                // Arama metnini kontrol et
                if (fullText.includes(searchInput)) {
                    row.style.display = '';
                    visibleCount++;
                } else {
                    row.style.display = 'none';
                }
            }

            // Sonuç sayacını güncelle
            const searchCount = document.getElementById('studentSearchCount');
            if (searchInput === '') {
                searchCount.textContent = '';
            } else {
                searchCount.textContent = `${visibleCount} / ${totalCount} öğrenci gösteriliyor`;
                if (visibleCount === 0) {
                    searchCount.innerHTML = '<span style="color: #ef4444;">❌ Eşleşen öğrenci bulunamadı</span>';
                }
            }
        }

        // ============== YENİ PROGRAM OLUŞTUR MODAL ==============

        let programStartDate = null; // Global değişken: Program başlangıç tarihi

        function openGenerateScheduleModal() {
            // Varsayılan başlangıç tarihi: Bu haftanın Pazartesi'si
            const today = new Date();
            const dayOfWeek = today.getDay(); // 0=Pazar, 1=Pazartesi, ...

            // Bu haftanın Pazartesi'sine kaç gün var/geçti?
            const daysUntilMonday = dayOfWeek === 0 ? -6 : dayOfWeek === 1 ? 0 : 1 - dayOfWeek;

            const thisMonday = new Date(today);
            thisMonday.setDate(today.getDate() + daysUntilMonday);

            // YYYY-MM-DD formatında
            const year = thisMonday.getFullYear();
            const month = String(thisMonday.getMonth() + 1).padStart(2, '0');
            const day = String(thisMonday.getDate()).padStart(2, '0');
            document.getElementById('generateStartDate').value = `${year}-${month}-${day}`;

            document.getElementById('generateScheduleModal').style.display = 'block';
        }

        function closeGenerateScheduleModal() {
            document.getElementById('generateScheduleModal').style.display = 'none';
        }

        async function generateScheduleWithDate() {
            const startDate = document.getElementById('generateStartDate').value;

            if (!startDate) {
                showError('Lütfen program başlangıç tarihini seçin!');
                return;
            }

            // Tarihin Pazartesi olup olmadığını kontrol et
            const selectedDate = new Date(startDate + 'T00:00:00');
            const dayOfWeek = selectedDate.getDay();
            if (dayOfWeek !== 1) {
                showError('⚠️ Başlangıç tarihi Pazartesi olmalıdır! Lütfen bir Pazartesi günü seçin.');
                return;
            }

            // Global değişkene kaydet
            programStartDate = startDate;

            // Modal'ı kapat
            closeGenerateScheduleModal();

            // Program oluştur (eski fonksiyon)
            await generateSchedule();
        }

        async function generateSchedule() {
            const response = await fetch('/generate_schedule');
            const result = await response.json();

            if (result.error) {
                showError(result.error);
                return;
            }

            showSuccess('4 haftalık program başarıyla oluşturuldu!');

            // 🧹 YENİ PROGRAM OLUŞTURULDU - ESKİ AYKIRI SWAP VERİLERİNİ TEMİZLE
            try {
                sessionStorage.removeItem('aykiriSwapViolations');
                console.log('✅ Eski aykırı swap verileri temizlendi');
            } catch (e) {
                console.error('⚠️ SessionStorage temizleme hatası:', e);
            }

            // 📊 İSTATİSTİK KARTLARINI GÖSTER VE GÜNCELLE
            updateStatisticsCards(result.schedule);

            displayModernSchedule(result.schedule);

            // Haftalık tabloyu göster
            globalScheduleData = result.schedule;

            // 📅 BAŞLANGIÇ TARİHİNİ EKLE
            if (programStartDate) {
                globalScheduleData.start_date = programStartDate;
            }

            document.getElementById('weeklyScheduleSection').style.display = 'block';
            currentWeekView = 1;

            // ✅ BİRAZ BEKLE VE SONRA RENDER ET
            setTimeout(() => {
                renderWeeklyTable(1);
            }, 100);

            // 🆕 OTOMATİK ÇAKIŞMA KONTROLÜ
            setTimeout(() => {
                checkConflictsInBackground();
            }, 500);
            // ✅ Öğrenci ve öğretmen listelerini güncelle
            await loadStudents();
            await loadTeachers();

            // ✅ Günlük Dersler widget'ını yükle
            await loadTodayLessons();
        }

        function displayModernSchedule(schedule) {
            const resultsDiv = document.getElementById('resultsSection');
            let html = '';

            // ============== ÖĞRENCİ BAZLI ÖĞRETMEN DAĞILIMI ==============
            html += `
                <h2 style="text-align: center; font-size: 2em; color: white; background: linear-gradient(135deg, #7c3aed 0%, #6d28d9 100%); padding: 20px; border-radius: 15px; margin: 50px 0 30px 0; box-shadow: 0 4px 15px rgba(0,0,0,0.2);">Öğrenci Bazlı Öğretmen Dağılımı</h2>

                <div style="background: white; border-radius: 15px; padding: 25px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                    <div style="display: flex; gap: 15px; align-items: flex-start; flex-wrap: wrap;">
                        <input type="text" id="studentSearchInput" placeholder="🔍👨‍🎓 Öğrenci ara..." style="flex: 1; min-width: 250px; padding: 12px 20px; border: 2px solid #7c3aed; border-radius: 10px; font-size: 1em;" oninput="filterStudents()">

                        <div style="position: relative;">
                            <button id="studentDropdownBtn" onclick="toggleStudentDropdown()" style="padding: 12px 20px; border: 2px solid #7c3aed; border-radius: 10px; font-size: 1em; background: white; cursor: pointer; min-width: 200px; display: flex; justify-content: space-between; align-items: center;">
                                <span id="selectedCountText">📋👨‍🎓 Öğrenci Seç</span>
                                <span>▼</span>
                            </button>
                            <div id="studentDropdownMenu" style="display: none; position: absolute; top: 100%; left: 0; background: white; border: 2px solid #7c3aed; border-radius: 10px; margin-top: 5px; padding: 15px; min-width: 300px; max-height: 400px; overflow-y: auto; z-index: 1000; box-shadow: 0 4px 15px rgba(0,0,0,0.2);">
                                <div style="margin-bottom: 10px; padding-bottom: 10px; border-bottom: 2px solid #e5e7eb;">
                                    <label style="display: flex; align-items: center; gap: 8px; font-weight: bold; cursor: pointer; padding: 8px;">
                                        <input type="checkbox" id="selectAllCheckbox" onchange="selectAllStudents()" style="width: 18px; height: 18px; cursor: pointer;">
                                        <span>Tümünü Seç</span>
                                    </label>
                                </div>
                                <div id="studentCheckboxList"></div>
                                <button onclick="applyStudentFilter()"
                                    onmouseover="this.style.transform='translateY(-3px)'; this.style.boxShadow='0 8px 25px rgba(124,58,237,0.6)'"
                                    onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 4px 15px rgba(0,0,0,0.1)'"
                                    style="width: 100%; margin-top: 10px; padding: 10px; background: linear-gradient(135deg, #7c3aed 0%, #6d28d9 100%); color: white; border: none; border-radius: 8px; font-weight: bold; cursor: pointer; transition: all 0.3s; position: relative; overflow: hidden;">Uygula</button>
                            </div>
                        </div>

                        <button onclick="resetStudentFilter()"
                            onmouseover="this.style.transform='translateY(-3px)'; this.style.boxShadow='0 8px 25px rgba(124,58,237,0.6)'"
                            onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 4px 15px rgba(0,0,0,0.1)'"
                            style="padding: 12px 25px; background: linear-gradient(135deg, #7c3aed 0%, #6d28d9 100%); color: white; border: none; border-radius: 10px; font-weight: bold; cursor: pointer; font-size: 1em; transition: all 0.3s; position: relative; overflow: hidden;">Sıfırla</button>
                    </div>
                </div>

                <div id="studentCardsContainer">
            `;

            const teacherDist = getTeacherDistribution(schedule);

            teacherDist.forEach((item, idx) => {
                html += `
                    <div class="student-card" data-student-name="${item.student_name.toLocaleLowerCase('tr')}" data-student-index="${idx}" style="background: white; border-radius: 15px; padding: 25px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); display: none;">
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; padding-bottom: 15px; border-bottom: 3px solid #7c3aed;">
                            <div>
                                <h3 style="margin: 0; font-size: 1.5em; color: #1f2937;">${item.student_name}</h3>
                                <p style="margin: 5px 0 0 0; color: #6b7280; font-size: 1.1em;">Sınıf: ${item.student_class}</p>
                            </div>
                            <div style="background: linear-gradient(135deg, #7c3aed 0%, #6d28d9 100%); color: white; padding: 15px 25px; border-radius: 15px; text-align: center;">
                                <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">Toplam Ders</div>
                                <div style="font-size: 2em; font-weight: bold;">${item.total}</div>
                            </div>
                        </div>
                        <div style="display: grid; gap: 15px;">
                `;

                item.teachers.forEach(teacher => {
                    html += `
                        <div style="background: linear-gradient(135deg, #f9fafb 0%, #f3f4f6 100%); border-radius: 12px; padding: 20px; border-left: 5px solid #7c3aed;">
                            <div style="display: flex; justify-content: space-between; align-items: start; margin-bottom: 12px;">
                                <div>
                                    <div style="font-size: 1.2em; font-weight: bold; color: #1f2937; margin-bottom: 5px;">
                                        ${teacher.teacher_name}
                                    </div>
                                    <div style="display: inline-block; background: #7c3aed; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.9em;">
                                        ${teacher.branch}
                                    </div>
                                </div>
                            </div>
                            <div style="color: #4b5563; font-size: 0.95em; line-height: 1.8;">
                                ${teacher.schedule}
                            </div>
                        </div>
                    `;
                });

                html += `</div></div>`;
            });

            html += `</div>`;

// ============== BRANŞ DAĞILIMI - FİLTRELEME İLE ==============
            html += `
                <h2 style="text-align: center; font-size: 2em; color: white; background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%); padding: 20px; border-radius: 15px; margin: 50px 0 30px 0; box-shadow: 0 4px 15px rgba(0,0,0,0.2);">Öğrenci Bazlı Branş Dağılımı</h2>

                <div style="background: white; border-radius: 15px; padding: 25px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                    <div style="display: flex; gap: 15px; align-items: flex-start; flex-wrap: wrap;">
                        <input type="text" id="branchSearchInput" placeholder="🔍👨‍🎓 Öğrenci ara..." style="flex: 1; min-width: 250px; padding: 12px 20px; border: 2px solid #dc2626; border-radius: 10px; font-size: 1em;" oninput="filterBranchStudents()">

                        <div style="position: relative;">
                            <button id="branchDropdownBtn" onclick="toggleBranchDropdown()" style="padding: 12px 20px; border: 2px solid #dc2626; border-radius: 10px; font-size: 1em; background: white; cursor: pointer; min-width: 200px; display: flex; justify-content: space-between; align-items: center;">
                                <span id="branchSelectedCountText">📋👨‍🎓 Öğrenci Seç</span>
                                <span>▼</span>
                            </button>
                            <div id="branchDropdownMenu" style="display: none; position: absolute; top: 100%; left: 0; background: white; border: 2px solid #dc2626; border-radius: 10px; margin-top: 5px; padding: 15px; min-width: 300px; max-height: 400px; overflow-y: auto; z-index: 1000; box-shadow: 0 4px 15px rgba(0,0,0,0.2);">
                                <div style="margin-bottom: 10px; padding-bottom: 10px; border-bottom: 2px solid #e5e7eb;">
                                    <label style="display: flex; align-items: center; gap: 8px; font-weight: bold; cursor: pointer; padding: 8px;">
                                        <input type="checkbox" id="branchSelectAllCheckbox" onchange="selectAllBranchStudents()" style="width: 18px; height: 18px; cursor: pointer;">
                                        <span>Tümünü Seç</span>
                                    </label>
                                </div>
                                <div id="branchCheckboxList"></div>
                                <button onclick="applyBranchFilter()"
                                    onmouseover="this.style.transform='translateY(-3px)'; this.style.boxShadow='0 8px 25px rgba(220,38,38,0.6)'"
                                    onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 4px 15px rgba(0,0,0,0.1)'"
                                    style="width: 100%; margin-top: 10px; padding: 10px; background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%); color: white; border: none; border-radius: 8px; font-weight: bold; cursor: pointer; transition: all 0.3s; position: relative; overflow: hidden;">Uygula</button>
                            </div>
                        </div>

                        <button onclick="resetBranchFilter()"
                            onmouseover="this.style.transform='translateY(-3px)'; this.style.boxShadow='0 8px 25px rgba(220,38,38,0.6)'"
                            onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 4px 15px rgba(0,0,0,0.1)'"
                            style="padding: 12px 25px; background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%); color: white; border: none; border-radius: 10px; font-weight: bold; cursor: pointer; font-size: 1em; transition: all 0.3s; position: relative; overflow: hidden;">Sıfırla</button>
                    </div>
                </div>

                <div id="branchCardsContainer">
            `;

            const allBranches = getAllBranches(schedule);
            const branchDist = getBranchDistribution(schedule, allBranches);

            const branchColors = {
                'Matematik': '#3b82f6',
                'Fizik': '#8b5cf6',
                'Kimya': '#ec4899',
                'Biyoloji': '#10b981',
                'Geometri': '#f59e0b',
                'Türkçe': '#ef4444',
                'Edebiyat': '#06b6d4',
                'İngilizce': '#6366f1',
                'Fen Bilgisi': '#14b8a6',
                'Sosyal Bilgiler': '#f97316',
                'Tarih': '#84cc16',
                'Coğrafya': '#22c55e',
                'Felsefe': '#a855f7',
                'Din Kültürü': '#0ea5e9'
            };

            branchDist.forEach((student, idx) => {
                let totalLessons = 0;
                allBranches.forEach(branch => {
                    const branchData = student.branches[branch];
                    if (branchData) {
                        totalLessons += branchData.total;
                    }
                });

                html += `
                    <div class="branch-card" data-branch-student-name="${student.student_name.toLocaleLowerCase('tr')}" data-branch-student-index="${idx}" style="background: white; border-radius: 15px; padding: 25px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); display: none;">
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; padding-bottom: 15px; border-bottom: 3px solid #dc2626;">
                            <div>
                                <h3 style="margin: 0; font-size: 1.5em; color: #1f2937;">${student.student_name}</h3>
                                <p style="margin: 5px 0 0 0; color: #6b7280; font-size: 1.1em;">Sınıf: ${student.student_class}</p>
                            </div>
                            <div style="background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%); color: white; padding: 15px 25px; border-radius: 15px; text-align: center;">
                                <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">Toplam Ders</div>
                                <div style="font-size: 2em; font-weight: bold;">${totalLessons}</div>
                            </div>
                        </div>
                        <div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 15px;">
                `;

                allBranches.forEach(branch => {
                    const branchData = student.branches[branch];
                    if (branchData && branchData.total > 0) {
                        const color = branchColors[branch] || '#6b7280';
                        html += `
                            <div style="background: linear-gradient(135deg, ${color}15 0%, ${color}08 100%); border-left: 4px solid ${color}; border-radius: 10px; padding: 15px;">
                                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
                                    <span style="font-weight: bold; color: ${color}; font-size: 1.1em;">${branch}</span>
                                    <span style="background: ${color}; color: white; padding: 4px 12px; border-radius: 20px; font-weight: bold; font-size: 0.9em;">${branchData.total}</span>
                                </div>
                                <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; margin-top: 10px;">
                                    <div style="text-align: center; background: white; padding: 8px; border-radius: 6px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">H1</div>
                                        <div style="font-weight: bold; color: ${color}; font-size: 1.2em;">${branchData.week1}</div>
                                    </div>
                                    <div style="text-align: center; background: white; padding: 8px; border-radius: 6px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">H2</div>
                                        <div style="font-weight: bold; color: ${color}; font-size: 1.2em;">${branchData.week2}</div>
                                    </div>
                                    <div style="text-align: center; background: white; padding: 8px; border-radius: 6px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">H3</div>
                                        <div style="font-weight: bold; color: ${color}; font-size: 1.2em;">${branchData.week3}</div>
                                    </div>
                                    <div style="text-align: center; background: white; padding: 8px; border-radius: 6px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">H4</div>
                                        <div style="font-weight: bold; color: ${color}; font-size: 1.2em;">${branchData.week4}</div>
                                    </div>
                                </div>
                            </div>
                        `;
                    }
                });

                html += `
                        </div>
                    </div>
                `;
            });

            html += `</div>`;

            // ============== 4 HAFTALIK DERS PROGRAMI ==============
            html += `<h2 style="text-align: center; font-size: 2em; color: white; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 15px; margin: 50px 0 30px 0; box-shadow: 0 4px 15px rgba(0,0,0,0.2);">4 HAFTALIK DERS PROGRAMI</h2>`;

            const organizedByDayWeek = {};
            schedule.weeks.forEach((week, weekIdx) => {
                week.forEach(lesson => {
                    const day = lesson.day;
                    const teacherKey = `${lesson.teacher_name}_${lesson.branch}`;

                    if (!organizedByDayWeek[day]) {
                        organizedByDayWeek[day] = {};
                    }
                    if (!organizedByDayWeek[day][teacherKey]) {
                        organizedByDayWeek[day][teacherKey] = {
                            teacher_name: lesson.teacher_name,
                            branch: lesson.branch,
                            weeks: [[], [], [], []]
                        };
                    }

                    organizedByDayWeek[day][teacherKey].weeks[weekIdx].push({
                        time: lesson.time,
                        student_name: lesson.student_name,
                        student_class: lesson.student_class
                    });
                });
            });

            // ✅ HER GÜN İÇİNDEKİ ÖĞRETMENLERI ALFABETİK SIRALA
            const dayOrder4 = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar'];

            dayOrder4.forEach(day => {
                if (organizedByDayWeek[day]) {
                    // Öğretmenleri array'e çevir
                    const teachersArray = Object.values(organizedByDayWeek[day]);

                    // Alfabetik sırala (önce branş, sonra isim)
                    teachersArray.sort((a, b) => {
                        // Önce branşa göre sırala
                        if (a.branch !== b.branch) {
                            return a.branch.localeCompare(b.branch, 'tr');
                        }
                        // Aynı branşta isime göre sırala
                        return a.teacher_name.localeCompare(b.teacher_name, 'tr');
                    });

                    // Sıralı object'e geri dönüştür
                    const sortedTeachers = {};
                    teachersArray.forEach(teacher => {
                        const key = `${teacher.teacher_name}_${teacher.branch}`;
                        sortedTeachers[key] = teacher;
                    });

                    // Eski nesneyi değiştir
                    organizedByDayWeek[day] = sortedTeachers;
                }
            });



            html += `<div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 40px;">`;

            for (let weekIdx = 0; weekIdx < 4; weekIdx++) {
                html += `
                    <div style="background: #667eea; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.15);">
                        <div style="background: #667eea; color: white; padding: 15px; text-align: center; font-weight: bold; font-size: 1.2em;">
                            HAFTA ${weekIdx + 1}
                        </div>
                        <div style="background: white; padding: 15px;">
                `;

                dayOrder4.forEach(day => {
                    if (organizedByDayWeek[day]) {
                        html += `
                            <div style="background: #667eea; color: white; padding: 10px; border-radius: 8px; font-weight: bold; margin-bottom: 15px; text-align: center;">
                                ${day}
                            </div>
                        `;

                        Object.values(organizedByDayWeek[day]).forEach(teacher => {
                            const weekLessons = teacher.weeks[weekIdx] || [];

                            // 🆕 O günün TÜM SAATLERİNİ BUL
                            const allTimesForDay = new Set();
                            teacher.weeks.forEach(w => {
                                w.forEach(l => allTimesForDay.add(l.time));
                            });
                            const sortedTimes = Array.from(allTimesForDay).sort();

                            if (weekLessons.length > 0 || sortedTimes.length > 0) {
                                html += `
                                    <div style="background: #f8f9fa; border-radius: 8px; padding: 12px; margin-bottom: 15px;">
                                        <div style="font-weight: bold; color: #333; margin-bottom: 8px; font-size: 0.9em;">
                                            ${teacher.branch} - ${teacher.teacher_name}
                                        </div>
                                `;

                                // 🆕 TÜM SAATLERİ KONTROL ET
                                sortedTimes.forEach(time => {
                                    const lessonsAtTime = weekLessons.filter(l => l.time === time);

                                    if (lessonsAtTime.length === 0) {
                                        // 🆕 BOŞ SLOT - SİLİK GÖSTER
                                        html += `
                                            <div style="display: flex; justify-content: space-between; align-items: center; padding: 6px 8px; margin-bottom: 4px; font-size: 0.8em; opacity: 0.4;">
                                                <span style="color: #666; white-space: nowrap; flex-shrink: 0;">${time}</span>
                                                <div style="display: flex; gap: 6px; align-items: center; min-width: 0; margin-left: 8px;">
                                                    <span style="font-style: italic; color: #999; text-align: right;">Uygun eşleşme yok</span>
                                                </div>
                                            </div>
                                        `;
                                    } else if (lessonsAtTime.length === 1) {
                                        // TEK ÖĞRENCİ - NORMAL GÖSTER
                                        const lesson = lessonsAtTime[0];
                                        html += `
                                            <div style="display: flex; justify-content: space-between; align-items: center; padding: 6px 8px; margin-bottom: 4px; font-size: 0.8em;">
                                                <span style="color: #666; white-space: nowrap; flex-shrink: 0;">${lesson.time}</span>
                                                <div style="display: flex; gap: 6px; align-items: center; min-width: 0; margin-left: 8px;">
                                                    <span style="font-weight: 600; color: #333; text-align: right; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">${lesson.student_name}</span>
                                                    <span style="color: #999; flex-shrink: 0;">${lesson.student_class}</span>
                                                </div>
                                            </div>
                                        `;
                                    } else {
                                        // 🆕 GRUP DERSİ - TÜM SINIFLARI TOPLA
                                        const uniqueClasses = [...new Set(lessonsAtTime.map(l => l.student_class).filter(c => c))];
                                        let displayText;
                                        if (uniqueClasses.length > 0) {
                                            const classesStr = uniqueClasses.sort().join(', ');
                                            displayText = `${classesStr} (${lessonsAtTime.length} öğr)`;
                                        } else {
                                            displayText = `${lessonsAtTime.length} öğrenci`;
                                        }
                                        html += `
                                            <div style="display: flex; justify-content: space-between; align-items: center; padding: 6px 8px; margin-bottom: 4px; font-size: 0.8em;">
                                                <span style="color: #666; white-space: nowrap; flex-shrink: 0;">${time}</span>
                                                <div style="display: flex; gap: 6px; align-items: center; min-width: 0; margin-left: 8px;">
                                                    <span style="font-weight: 600; color: #333; text-align: right; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">${displayText}</span>
                                                </div>
                                            </div>
                                        `;
                                    }
                                });

                                html += `</div>`;
                            }
                        });
                    }
                });

                html += `</div></div>`;
            }

            html += `</div>`;

            // ============== 4 HAFTALIK GENEL ÖZET ==============
            html += `
                <h2 style="text-align: center; font-size: 2em; color: white; background: linear-gradient(135deg, #16a34a 0%, #15803d 100%); padding: 20px; border-radius: 15px; margin: 50px 0 30px 0; box-shadow: 0 4px 15px rgba(0,0,0,0.2);">4 HAFTALIK GENEL ÖZET</h2>
                <div style="background: white; border-radius: 15px; padding: 30px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); overflow-x: auto;">
                    <table style="width: 100%; border-collapse: collapse; border-radius: 10px; overflow: hidden;">
                        <thead>
                            <tr style="background: linear-gradient(135deg, #16a34a 0%, #15803d 100%);">
                                <th style="padding: 18px 20px; text-align: left; color: white; font-weight: bold; font-size: 1.1em; border: 1px solid #15803d;">Öğrenci</th>
                                <th style="padding: 18px 20px; text-align: center; color: white; font-weight: bold; font-size: 1.1em; border: 1px solid #15803d;">Sınıf</th>
                                <th style="padding: 18px 20px; text-align: center; color: white; font-weight: bold; font-size: 1.1em; border: 1px solid #15803d;">H1</th>
                                <th style="padding: 18px 20px; text-align: center; color: white; font-weight: bold; font-size: 1.1em; border: 1px solid #15803d;">H2</th>
                                <th style="padding: 18px 20px; text-align: center; color: white; font-weight: bold; font-size: 1.1em; border: 1px solid #15803d;">H3</th>
                                <th style="padding: 18px 20px; text-align: center; color: white; font-weight: bold; font-size: 1.1em; border: 1px solid #15803d;">H4</th>
                                <th style="padding: 18px 20px; text-align: center; color: white; font-weight: bold; font-size: 1.1em; border: 1px solid #15803d; background: #15803d;">Toplam</th>
                            </tr>
                        </thead>
                        <tbody>
            `;

            const studentWeeklyCounts = {};
            schedule.weeks.forEach((week, weekIdx) => {
                week.forEach(lesson => {
                    const studentKey = `${lesson.student_name}_${lesson.student_class}`;
                    if (!studentWeeklyCounts[studentKey]) {
                        studentWeeklyCounts[studentKey] = {
                            name: lesson.student_name,
                            class: lesson.student_class,
                            weeks: [0, 0, 0, 0],
                            total: 0
                        };
                    }
                    studentWeeklyCounts[studentKey].weeks[weekIdx]++;
                    studentWeeklyCounts[studentKey].total++;
                });
            });

            const sortedStudents = Object.values(studentWeeklyCounts).sort((a, b) => {
                if (a.class !== b.class) return a.class.localeCompare(b.class);
                return a.name.localeCompare(b.name);
            });

            sortedStudents.forEach((student, idx) => {
                const rowColor = idx % 2 === 0 ? '#ffffff' : '#f9fafb';
                html += `
                    <tr style="background: ${rowColor}; transition: all 0.3s;" onmouseover="this.style.background='#f0fdf4'" onmouseout="this.style.background='${rowColor}'">
                        <td style="padding: 15px 20px; border: 1px solid #e5e7eb; font-weight: 600; color: #1f2937;">${student.name}</td>
                        <td style="padding: 15px 20px; text-align: center; border: 1px solid #e5e7eb; color: #6b7280;">${student.class}</td>
                        <td style="padding: 15px 20px; text-align: center; border: 1px solid #e5e7eb;">
                            <span style="background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%); padding: 8px 16px; border-radius: 20px; font-weight: bold; color: #1e40af; display: inline-block; min-width: 40px;">${student.weeks[0]}</span>
                        </td>
                        <td style="padding: 15px 20px; text-align: center; border: 1px solid #e5e7eb;">
                            <span style="background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%); padding: 8px 16px; border-radius: 20px; font-weight: bold; color: #1e40af; display: inline-block; min-width: 40px;">${student.weeks[1]}</span>
                        </td>
                        <td style="padding: 15px 20px; text-align: center; border: 1px solid #e5e7eb;">
                            <span style="background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%); padding: 8px 16px; border-radius: 20px; font-weight: bold; color: #1e40af; display: inline-block; min-width: 40px;">${student.weeks[2]}</span>
                        </td>
                        <td style="padding: 15px 20px; text-align: center; border: 1px solid #e5e7eb;">
                            <span style="background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%); padding: 8px 16px; border-radius: 20px; font-weight: bold; color: #1e40af; display: inline-block; min-width: 40px;">${student.weeks[3]}</span>
                        </td>
                        <td style="padding: 15px 20px; text-align: center; border: 1px solid #e5e7eb; background: #dcfce7;">
                            <span style="font-weight: bold; font-size: 1.1em; color: #15803d;">${student.total}</span>
                        </td>
                    </tr>
                `;
            });

            html += `
                        </tbody>
                    </table>
                </div>
            `;





            resultsDiv.innerHTML = html;

// ============== BRANŞ FİLTRE FONKSİYONLARI ==============
            const branchCheckboxList = document.getElementById('branchCheckboxList');
            if (branchCheckboxList) {
                const sortedBranchDist = branchDist.map((item, originalIdx) => ({
                    ...item,
                    originalIndex: originalIdx
                })).sort((a, b) =>
                    a.student_name.localeCompare(b.student_name, 'tr')
                );

                let branchCheckboxHTML = '';
                sortedBranchDist.forEach((item) => {
                    branchCheckboxHTML += `
                        <label style="display: flex; align-items: center; gap: 8px; padding: 8px; cursor: pointer; border-radius: 6px; transition: all 0.2s;" onmouseover="this.style.background='#f3f4f6'" onmouseout="this.style.background='white'">
                            <input type="checkbox" class="branch-checkbox" value="${item.originalIndex}" onchange="updateBranchSelectedCount()" style="width: 18px; height: 18px; cursor: pointer;">
                            <span>${item.student_name} (${item.student_class})</span>
                        </label>
                    `;
                });
                branchCheckboxList.innerHTML = branchCheckboxHTML;
            }

            const branchCards = document.querySelectorAll('.branch-card');
            branchCards.forEach(card => card.style.display = 'none');

            window.toggleBranchDropdown = function() {
                const menu = document.getElementById('branchDropdownMenu');
                menu.style.display = menu.style.display === 'none' ? 'block' : 'none';
            };

            window.selectAllBranchStudents = function() {
                const selectAll = document.getElementById('branchSelectAllCheckbox');
                const checkboxes = document.querySelectorAll('.branch-checkbox');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
                updateBranchSelectedCount();
            };

            window.updateBranchSelectedCount = function() {
                const checkboxes = document.querySelectorAll('.branch-checkbox:checked');
                const count = checkboxes.length;
                const text = document.getElementById('branchSelectedCountText');

                if (count === 0) {
                    text.textContent = '📋 Öğrenci Seç';
                } else if (count === 1) {
                    text.textContent = '1 Öğrenci Seçildi';
                } else {
                    text.textContent = `${count} Öğrenci Seçildi`;
                }

                const allCheckboxes = document.querySelectorAll('.branch-checkbox');
                const selectAll = document.getElementById('branchSelectAllCheckbox');
                selectAll.checked = count === allCheckboxes.length;
            };

            window.applyBranchFilter = function() {
                const selectedCheckboxes = document.querySelectorAll('.branch-checkbox:checked');
                const cards = document.querySelectorAll('.branch-card');

                if (selectedCheckboxes.length === 0) {
                    cards.forEach(card => card.style.display = 'none');
                } else {
                    const selectedIndices = Array.from(selectedCheckboxes).map(cb => cb.value);

                    // ✅ Seçilen kartları alfabetik sırala
                    const selectedCards = [];
                    cards.forEach(card => {
                        const cardIndex = card.getAttribute('data-branch-student-index');
                        if (selectedIndices.includes(cardIndex)) {
                            selectedCards.push({
                                element: card,
                                name: card.getAttribute('data-branch-student-name')
                            });
                            card.style.display = 'none'; // Önce gizle
                        } else {
                            card.style.display = 'none';
                        }
                    });

                    // Alfabetik sırala
                    selectedCards.sort((a, b) => a.name.localeCompare(b.name, 'tr'));

                    // Sırayla göster
                    const container = document.getElementById('branchCardsContainer');
                    selectedCards.forEach(item => {
                        container.appendChild(item.element); // Sırayla ekle
                        item.element.style.display = 'block';
                    });
                }

                document.getElementById('branchSearchInput').value = '';
                toggleBranchDropdown();
           };

            window.filterBranchStudents = function() {
                const searchTerm = document.getElementById('branchSearchInput').value.toLocaleLowerCase('tr').trim();
                const cards = document.querySelectorAll('.branch-card');

                if (searchTerm === '') {
                    cards.forEach(card => card.style.display = 'none');
                } else {
                    cards.forEach(card => {
                        const studentName = card.getAttribute('data-branch-student-name');
                        if (studentName.includes(searchTerm)) {
                            card.style.display = 'block';
                        } else {
                            card.style.display = 'none';
                        }
                    });
                }

                const checkboxes = document.querySelectorAll('.branch-checkbox');
                checkboxes.forEach(cb => cb.checked = false);
                updateBranchSelectedCount();
            };

            window.resetBranchFilter = function() {
                document.getElementById('branchSearchInput').value = '';
                const checkboxes = document.querySelectorAll('.branch-checkbox');
                checkboxes.forEach(cb => cb.checked = false);
                updateBranchSelectedCount();
                const cards = document.querySelectorAll('.branch-card');
                cards.forEach(card => card.style.display = 'none');
                document.getElementById('branchDropdownMenu').style.display = 'none';
            };

            document.addEventListener('click', function(event) {
                const menu = document.getElementById('studentDropdownMenu');
                const btn = document.getElementById('studentDropdownBtn');
                if (menu && btn && !menu.contains(event.target) && !btn.contains(event.target)) {
                    menu.style.display = 'none';
                }

                const branchMenu = document.getElementById('branchDropdownMenu');
                const branchBtn = document.getElementById('branchDropdownBtn');
                if (branchMenu && branchBtn && !branchMenu.contains(event.target) && !branchBtn.contains(event.target)) {
                    branchMenu.style.display = 'none';
                }
            });

            // JavaScript fonksiyonlarını ekle
            const checkboxList = document.getElementById('studentCheckboxList');
            if (checkboxList) {
                // ✅ Alfabetik sırala AMA orijinal index'i sakla
                const sortedTeacherDist = teacherDist.map((item, originalIdx) => ({
                    ...item,
                    originalIndex: originalIdx  // 🔑 Orijinal index'i sakla
                })).sort((a, b) =>
                    a.student_name.localeCompare(b.student_name, 'tr')
                );

                let checkboxHTML = '';
                sortedTeacherDist.forEach((item) => {
                    checkboxHTML += `
                        <label style="display: flex; align-items: center; gap: 8px; padding: 8px; cursor: pointer; border-radius: 6px; transition: all 0.2s;" onmouseover="this.style.background='#f3f4f6'" onmouseout="this.style.background='white'">
                            <input type="checkbox" class="student-checkbox" value="${item.originalIndex}" onchange="updateSelectedCount()" style="width: 18px; height: 18px; cursor: pointer;">
                            <span>${item.student_name} (${item.student_class})</span>
                        </label>
                    `;
                });
                checkboxList.innerHTML = checkboxHTML;
            }

            // İLK AÇILIŞTA TÜM KARTLARI GİZLE
            const cards = document.querySelectorAll('.student-card');
            cards.forEach(card => card.style.display = 'none');

            window.toggleStudentDropdown = function() {
                const menu = document.getElementById('studentDropdownMenu');
                menu.style.display = menu.style.display === 'none' ? 'block' : 'none';
            };

            window.selectAllStudents = function() {
                const selectAll = document.getElementById('selectAllCheckbox');
                const checkboxes = document.querySelectorAll('.student-checkbox');
                checkboxes.forEach(cb => cb.checked = selectAll.checked);
                updateSelectedCount();
            };

            window.updateSelectedCount = function() {
                const checkboxes = document.querySelectorAll('.student-checkbox:checked');
                const count = checkboxes.length;
                const text = document.getElementById('selectedCountText');

                if (count === 0) {
                    text.textContent = '📋 Öğrenci Seç';
                } else if (count === 1) {
                    text.textContent = '1 Öğrenci Seçildi';
                } else {
                    text.textContent = `${count} Öğrenci Seçildi`;
                }

                const allCheckboxes = document.querySelectorAll('.student-checkbox');
                const selectAll = document.getElementById('selectAllCheckbox');
                selectAll.checked = count === allCheckboxes.length;
            };

            window.applyStudentFilter = function() {
                const selectedCheckboxes = document.querySelectorAll('.student-checkbox:checked');
                const cards = document.querySelectorAll('.student-card');

                if (selectedCheckboxes.length === 0) {
                    cards.forEach(card => card.style.display = 'none');
                } else {
                    const selectedIndices = Array.from(selectedCheckboxes).map(cb => cb.value);

                    // ✅ Seçilen kartları alfabetik sırala
                    const selectedCards = [];
                    cards.forEach(card => {
                        const cardIndex = card.getAttribute('data-student-index');
                        if (selectedIndices.includes(cardIndex)) {
                            selectedCards.push({
                                element: card,
                                name: card.getAttribute('data-student-name')
                            });
                            card.style.display = 'none'; // Önce gizle
                        } else {
                            card.style.display = 'none';
                        }
                    });

                    // Alfabetik sırala
                    selectedCards.sort((a, b) => a.name.localeCompare(b.name, 'tr'));

                    // Sırayla göster
                    const container = document.getElementById('studentCardsContainer');
                    selectedCards.forEach(item => {
                        container.appendChild(item.element); // Sırayla ekle
                        item.element.style.display = 'block';
                    });
                }

                document.getElementById('studentSearchInput').value = '';
                toggleStudentDropdown();
            };



            window.filterStudents = function() {
                const searchTerm = document.getElementById('studentSearchInput').value.toLocaleLowerCase('tr').trim();
                const cards = document.querySelectorAll('.student-card');

                if (searchTerm === '') {
                    cards.forEach(card => card.style.display = 'none');
                } else {
                    cards.forEach(card => {
                        const studentName = card.getAttribute('data-student-name');
                        if (studentName.includes(searchTerm)) {
                            card.style.display = 'block';
                        } else {
                            card.style.display = 'none';
                        }
                    });
                }

                const checkboxes = document.querySelectorAll('.student-checkbox');
                checkboxes.forEach(cb => cb.checked = false);
                updateSelectedCount();
            };

            window.resetStudentFilter = function() {
                document.getElementById('studentSearchInput').value = '';
                const checkboxes = document.querySelectorAll('.student-checkbox');
                checkboxes.forEach(cb => cb.checked = false);
                updateSelectedCount();
                const cards = document.querySelectorAll('.student-card');
                cards.forEach(card => card.style.display = 'none');
                document.getElementById('studentDropdownMenu').style.display = 'none';
            };

            document.addEventListener('click', function(event) {
                const menu = document.getElementById('studentDropdownMenu');
                const btn = document.getElementById('studentDropdownBtn');
                if (menu && btn && !menu.contains(event.target) && !btn.contains(event.target)) {
                    menu.style.display = 'none';
                }
            });
        }

        function getTeacherDistribution(schedule) {
            const studentMap = {};

            schedule.weeks.forEach((week, weekIdx) => {
                week.forEach(lesson => {
                    const studentKey = `${lesson.student_name}_${lesson.student_class}`;
                    if (!studentMap[studentKey]) {
                        studentMap[studentKey] = {
                            student_name: lesson.student_name,
                            student_class: lesson.student_class,
                            teachers: {}
                        };
                    }

                    const teacherKey = `${lesson.teacher_name}_${lesson.branch}`;
                    if (!studentMap[studentKey].teachers[teacherKey]) {
                        studentMap[studentKey].teachers[teacherKey] = {
                            teacher_name: lesson.teacher_name,
                            branch: lesson.branch,
                            lessons: []
                        };
                    }

                    studentMap[studentKey].teachers[teacherKey].lessons.push({
                        week: weekIdx + 1,
                        day: lesson.day,
                        time: lesson.time
                    });
                });
            });

            const result = [];
            Object.values(studentMap).forEach(student => {
                const teachers = Object.values(student.teachers).map(teacher => {
                    const schedule = teacher.lessons.map(l =>
                        `H${l.week} ${l.day} (${l.time})`
                    ).join(', ');

                    return {
                        teacher_name: teacher.teacher_name,
                        branch: teacher.branch,
                        schedule: schedule
                    };
                });

                // ✅ ÖĞRETMENLERI SIRALA: ÖNCE BRANŞ, SONRA AD
                teachers.sort((a, b) => {
                    // Önce branşa göre sırala
                    if (a.branch !== b.branch) {
                        return a.branch.localeCompare(b.branch, 'tr');
                    }
                    // Aynı branşta isime göre sırala
                    return a.teacher_name.localeCompare(b.teacher_name, 'tr');
                });

                const total = teachers.reduce((sum, t) => sum + t.schedule.split(',').length, 0);

                result.push({
                    student_name: student.student_name,
                    student_class: student.student_class,
                    teachers: teachers,
                    total: total
                });
            });

            return result;
        }
        function getAllBranches(schedule) {
            const branches = new Set();
            schedule.weeks.forEach(week => {
                week.forEach(lesson => {
                    branches.add(lesson.branch);
                });
            });
            // ✅ TÜRKÇE ALFABEYE GÖRE SIRALA
            return Array.from(branches).sort((a, b) => a.localeCompare(b, 'tr'));
        }

        function getBranchDistribution(schedule, allBranches) {
            const studentMap = {};

            schedule.weeks.forEach((week, weekIdx) => {
                week.forEach(lesson => {
                    const studentKey = `${lesson.student_name}_${lesson.student_class}`;
                    if (!studentMap[studentKey]) {
                        studentMap[studentKey] = {
                            student_name: lesson.student_name,
                            student_class: lesson.student_class,
                            branches: {}
                        };
                    }

                    if (!studentMap[studentKey].branches[lesson.branch]) {
                        studentMap[studentKey].branches[lesson.branch] = {
                            week1: 0, week2: 0, week3: 0, week4: 0, total: 0
                        };
                    }

                    studentMap[studentKey].branches[lesson.branch][`week${weekIdx + 1}`]++;
                    studentMap[studentKey].branches[lesson.branch].total++;
                });
            });

            return Object.values(studentMap);
        }
        function exportToExcel() { window.location.href = '/export_excel'; }
        function exportToHTML() { window.location.href = '/export_html'; }

        function showSuccess(msg) {
            const div = document.getElementById('successMessage');
            div.textContent = msg;
            div.style.display = 'block';
            setTimeout(() => div.style.display = 'none', 3000);
        }

        function showError(msg) {
            const div = document.getElementById('errorMessage');
            div.textContent = msg;
            div.style.display = 'block';
            setTimeout(() => div.style.display = 'none', 5000);
        }

        function toggleAccordion(section) {
            const content = document.getElementById(section + 'Content');
            const arrow = document.getElementById(section + 'Arrow');

            content.classList.toggle('open');
            arrow.classList.toggle('open');
        }

        function showTeacherDetail(teacher) {
            let scheduleHTML = '<h3 style="color: #667eea; margin-bottom: 15px;">Ders Programı</h3>';

            teacher.schedule.forEach(day => {
                // ✅ DERSLERİ SAATE GÖRE SIRALA
                const sortedLessons = [...day.lessons].sort((a, b) =>
                    a.start_time.localeCompare(b.start_time)
                );

                scheduleHTML += `
                    <div class="detail-section">
                        <h4>${day.day}</h4>
                        ${sortedLessons.map((lesson, idx) => `
                            <div class="detail-item">
                                <strong>${lesson.start_time} - ${lesson.end_time}</strong>
                                <span style="color: #667eea;">(${lesson.duration} dk)</span>
                            </div>
                        `).join('')}
                    </div>
                `;
            });

            document.getElementById('detailContent').innerHTML = scheduleHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        function showStudentRestrictions(student) {
            let restrictionHTML = '<h3 style="color: #ff9800; margin-bottom: 15px;">🚫 Kısıtlamalar</h3>';

            if (!student.restrictions || student.restrictions.length === 0) {
                restrictionHTML += '<p style="color: #999;">Kısıtlama bulunmuyor.</p>';
            } else {
                student.restrictions.forEach((r, idx) => {
                    // ✅ ÇOKLU HAFTA GÖSTER
                    let weekText = 'Her hafta';
                    if (r.type === 'custom') {
                        const weeks = r.weeks || [];
                        weekText = weeks.length > 0 ? `Hafta: ${weeks.join(', ')}` : 'Hafta seçilmemiş';
                    }

                    // ✅ ÇOKLU GÜN GÖSTER
                    const days = r.days || [r.day];
                    const dayText = days.join(', ');

                    restrictionHTML += `
                        <div class="detail-section" style="border-left-color: #ff9800;">
                            <h4 style="color: #ff9800;">Kısıtlama ${idx + 1}</h4>
                            <div class="detail-item">
                                <strong>Zaman:</strong> ${weekText}
                            </div>
                            <div class="detail-item">
                                <strong>Günler:</strong> ${dayText}
                            </div>
                            <div class="detail-item">
                                <strong>Saat:</strong> ${r.start_time} - ${r.end_time}
                            </div>
                        </div>
                    `;
                });
            }

            document.getElementById('detailContent').innerHTML = restrictionHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        function showStudentPriorities(student) {
            let priorityHTML = '<h3 style="color: #f59e0b; margin-bottom: 15px;">⭐ Ders Öncelikleri</h3>';

            if (!student.priorities || Object.values(student.priorities).every(arr => arr.length === 0)) {
                priorityHTML += '<p style="color: #999;">Öncelik bulunmuyor.</p>';
            } else {
                for (let week = 1; week <= 4; week++) {
                    const weekKey = `week${week}`;
                    const weekPriorities = student.priorities[weekKey] || [];

                    if (weekPriorities.length > 0) {
                        priorityHTML += `
                            <div class="detail-section" style="border-left-color: #f59e0b;">
                                <h4 style="color: #f59e0b;">Hafta ${week}</h4>
                                ${weekPriorities.map((branch, idx) => `
                                    <div class="detail-item">
                                        <strong>${idx + 1}. Öncelik:</strong> ${branch}
                                    </div>
                                `).join('')}
                            </div>
                        `;
                    }
                }
            }

            document.getElementById('detailContent').innerHTML = priorityHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        function showStudentManualLessons(student) {
            let manualHTML = '<h3 style="color: #3b82f6; margin-bottom: 15px;">📌 Manuel Ders Atamaları</h3>';

            if (!student.manual_lessons || student.manual_lessons.length === 0) {
                manualHTML += '<p style="color: #999;">Manuel ders bulunmuyor.</p>';
            } else {
                student.manual_lessons.forEach((manual, idx) => {
                    manualHTML += `
                        <div class="detail-section" style="border-left-color: #3b82f6;">
                            <h4 style="color: #3b82f6;">Manuel Ders ${idx + 1}</h4>
                            <div class="detail-item">
                                <strong>Hafta:</strong> ${manual.week}
                            </div>
                            <div class="detail-item">
                                <strong>Gün:</strong> ${manual.day}
                            </div>
                            <div class="detail-item">
                                <strong>Öğretmen:</strong> ${manual.teacher_name}
                            </div>
                            <div class="detail-item">
                                <strong>Saat:</strong> ${manual.time}
                            </div>
                        </div>
                    `;
                });
            }

            document.getElementById('detailContent').innerHTML = manualHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        async function showStudentTeacherBlocks(student) {
            let blockHTML = '<h3 style="color: #dc2626; margin-bottom: 15px;">🚫 Öğretmen Engellemeleri</h3>';

            if (!student.teacher_blocks || student.teacher_blocks.length === 0) {
                blockHTML += '<p style="color: #999;">Öğretmen engeli bulunmuyor.</p>';
            } else {
                // Öğretmenleri çek
                const response = await fetch('/get_teachers');
                const data = await response.json();
                const teachers = data.teachers;

                student.teacher_blocks.forEach((block, idx) => {
                    const teacher = teachers.find(t => t.id == block.teacher_id);
                    const teacherName = teacher ? `${teacher.name} ${teacher.surname} (${teacher.branch})` : 'Bilinmeyen Öğretmen';

                    // Hafta bilgisi
                    let weekText = 'Her hafta';
                    if (block.type === 'custom' && block.weeks && block.weeks.length > 0) {
                        weekText = `Hafta: ${block.weeks.join(', ')}`;
                    }

                    // Gün bilgisi
                    const dayText = block.day === 'all' ? 'Tüm günler' : block.day;

                    // Engellenen dersler
                    const blockedLessons = block.blocked_slots || [];
                    let lessonsText = '';
                    if (blockedLessons.length > 0) {
                        lessonsText = blockedLessons.map(slot => {
                            const parts = slot.split('_');
                            const day = parts[0];
                            const time = parts[1];
                            return block.day === 'all' ? `${day} ${time}` : time;
                        }).join(', ');
                    }

                    blockHTML += `
                        <div class="detail-section" style="border-left-color: #dc2626;">
                            <h4 style="color: #dc2626;">🚫 Engelleme ${idx + 1}</h4>
                            <div class="detail-item">
                                <strong>Öğretmen:</strong> ${teacherName}
                            </div>
                            <div class="detail-item">
                                <strong>Zaman:</strong> ${weekText}
                            </div>
                            <div class="detail-item">
                                <strong>Gün:</strong> ${dayText}
                            </div>
                            <div class="detail-item">
                                <strong>Engellenen Saatler:</strong> ${lessonsText}
                            </div>
                        </div>
                    `;
                });
            }

            document.getElementById('detailContent').innerHTML = blockHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        function showTeacherBlocks(teacher) {
            let blockHTML = '<h3 style="color: #ef4444; margin-bottom: 15px;">🚫 Slot Bloklamaları</h3>';

            if (!teacher.blocked_slots || teacher.blocked_slots.length === 0) {
                blockHTML += '<p style="color: #999;">Bloklama bulunmuyor.</p>';
            } else {
                teacher.blocked_slots.forEach((block, idx) => {
                    // HAFTA BİLGİSİ
                    let weekText = 'Her hafta';
                    if (block.type === 'custom' && block.weeks && block.weeks.length > 0) {
                        weekText = `Hafta: ${block.weeks.join(', ')}`;
                    }

                    // BLOKLANMIŞ DERSLER
                    const blockedLessons = block.blocked_slots || [];
                    const lessonsText = blockedLessons.join(', ');

                    blockHTML += `
                        <div class="detail-section" style="border-left-color: #ef4444;">
                            <h4 style="color: #ef4444;">🚫 Bloklama ${idx + 1}</h4>
                            <div class="detail-item">
                                <strong>Zaman:</strong> ${weekText}
                            </div>
                            <div class="detail-item">
                                <strong>Gün:</strong> ${block.day}
                            </div>
                            <div class="detail-item">
                                <strong>Bloklanmış Saatler:</strong> ${lessonsText}
                            </div>
                        </div>
                    `;
                });
            }

            document.getElementById('detailContent').innerHTML = blockHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        // ═══════════════════════════════════════════════════════════════
        // 🆕 ÖĞRENCİ DETAY SAYFASI - TÜM DERSLERİ GÖSTER
        // ═══════════════════════════════════════════════════════════════

        function showStudentScheduleDetail(studentName, studentClass) {
            if (!globalScheduleData || !globalScheduleData.weeks) {
                showError('Program bulunamadı!');
                return;
            }

            let detailHTML = `
                <h3 style="color: #667eea; margin-bottom: 20px; display: flex; align-items: center; gap: 10px;">
                    <i class="fas fa-calendar-check"></i> ${studentName} - Tüm Dersler
                </h3>
                <div style="background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%); padding: 15px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #3b82f6;">
                    <strong>Sınıf:</strong> ${studentClass}
                </div>
            `;

            let totalLessons = 0;

            // Her hafta için
            for (let weekNum = 1; weekNum <= 4; weekNum++) {
                const weekData = globalScheduleData.weeks[weekNum - 1];
                const weekLessons = weekData.filter(lesson => lesson.student_name === studentName);

                if (weekLessons.length === 0) continue;

                totalLessons += weekLessons.length;

                // Günlere göre grupla
                const lessonsByDay = {};
                const dayOrder = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar'];

                weekLessons.forEach(lesson => {
                    if (!lessonsByDay[lesson.day]) {
                        lessonsByDay[lesson.day] = [];
                    }
                    lessonsByDay[lesson.day].push(lesson);
                });

                // Her gün için saate göre sırala
                Object.keys(lessonsByDay).forEach(day => {
                    lessonsByDay[day].sort((a, b) => a.time.localeCompare(b.time));
                });

                detailHTML += `
                    <div class="detail-section" style="border-left-color: #667eea;">
                        <h4 style="color: #667eea; display: flex; align-items: center; gap: 8px;">
                            <i class="fas fa-calendar-week"></i> Hafta ${weekNum}
                            <span style="background: #667eea; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.85em;">
                                ${weekLessons.length} Ders
                            </span>
                        </h4>
                `;

                // Günlere göre göster
                dayOrder.forEach(day => {
                    if (!lessonsByDay[day]) return;

                    detailHTML += `
                        <div style="margin-top: 15px; margin-bottom: 15px; padding-left: 15px; border-left: 3px solid #e5e7eb;">
                            <div style="font-weight: 600; color: #1f2937; margin-bottom: 10px; font-size: 1.05em;">
                                📅 ${day}
                            </div>
                    `;

                    lessonsByDay[day].forEach(lesson => {
                        detailHTML += `
                            <div class="detail-item" style="margin-bottom: 8px; display: flex; align-items: center; gap: 15px; padding: 12px; background: linear-gradient(135deg, #f9fafb 0%, #ffffff 100%);">
                                <div style="min-width: 100px; font-weight: 600; color: #667eea;">
                                    <i class="fas fa-clock"></i> ${lesson.time}
                                </div>
                                <div style="flex: 1;">
                                    <div style="font-weight: 600; color: #1f2937; margin-bottom: 4px;">
                                        📚 ${lesson.branch}
                                    </div>
                                    <div style="color: #6b7280; font-size: 0.9em;">
                                        👨‍🏫 ${lesson.teacher_name}
                                    </div>
                                </div>
                            </div>
                        `;
                    });

                    detailHTML += `</div>`;
                });

                detailHTML += `</div>`;
            }

            // Toplam özet
            detailHTML += `
                <div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 20px; border-radius: 12px; margin-top: 20px; text-align: center;">
                    <div style="font-size: 2.5em; font-weight: bold; margin-bottom: 5px;">
                        ${totalLessons}
                    </div>
                    <div style="font-size: 1.1em; opacity: 0.9;">
                        Toplam Ders (4 Hafta)
                    </div>
                </div>
            `;

            document.getElementById('detailContent').innerHTML = detailHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        // ═══════════════════════════════════════════════════════════════
        // 🆕 ÖĞRETMEN DETAY SAYFASI - TÜM ÖĞRENCİLERİ GÖSTER
        // ═══════════════════════════════════════════════════════════════

        function showTeacherScheduleDetail(teacherName, branch) {
            if (!globalScheduleData || !globalScheduleData.weeks) {
                showError('Program bulunamadı!');
                return;
            }

            let detailHTML = `
                <h3 style="color: #667eea; margin-bottom: 20px; display: flex; align-items: center; gap: 10px;">
                    <i class="fas fa-chalkboard-teacher"></i> ${teacherName} - Tüm Dersler
                </h3>
                <div style="background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%); padding: 15px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #3b82f6;">
                    <strong>Branş:</strong> ${branch}
                </div>
            `;

            let totalLessons = 0;
            let uniqueStudents = new Set();

            // Her hafta için
            for (let weekNum = 1; weekNum <= 4; weekNum++) {
                const weekData = globalScheduleData.weeks[weekNum - 1];
                const weekLessons = weekData.filter(lesson => lesson.teacher_name === teacherName);

                if (weekLessons.length === 0) continue;

                totalLessons += weekLessons.length;
                weekLessons.forEach(lesson => uniqueStudents.add(lesson.student_name));

                // Günlere göre grupla
                const lessonsByDay = {};
                const dayOrder = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar'];

                weekLessons.forEach(lesson => {
                    if (!lessonsByDay[lesson.day]) {
                        lessonsByDay[lesson.day] = [];
                    }
                    lessonsByDay[lesson.day].push(lesson);
                });

                // Her gün için saate göre sırala
                Object.keys(lessonsByDay).forEach(day => {
                    lessonsByDay[day].sort((a, b) => a.time.localeCompare(b.time));
                });

                detailHTML += `
                    <div class="detail-section" style="border-left-color: #667eea;">
                        <h4 style="color: #667eea; display: flex; align-items: center; gap: 8px;">
                            <i class="fas fa-calendar-week"></i> Hafta ${weekNum}
                            <span style="background: #667eea; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.85em;">
                                ${weekLessons.length} Ders
                            </span>
                        </h4>
                `;

                // Günlere göre göster
                dayOrder.forEach(day => {
                    if (!lessonsByDay[day]) return;

                    detailHTML += `
                        <div style="margin-top: 15px; margin-bottom: 15px; padding-left: 15px; border-left: 3px solid #e5e7eb;">
                            <div style="font-weight: 600; color: #1f2937; margin-bottom: 10px; font-size: 1.05em;">
                                📅 ${day}
                            </div>
                    `;

                    lessonsByDay[day].forEach(lesson => {
                        detailHTML += `
                            <div class="detail-item" style="margin-bottom: 8px; display: flex; align-items: center; gap: 15px; padding: 12px; background: linear-gradient(135deg, #f9fafb 0%, #ffffff 100%);">
                                <div style="min-width: 100px; font-weight: 600; color: #667eea;">
                                    <i class="fas fa-clock"></i> ${lesson.time}
                                </div>
                                <div style="flex: 1;">
                                    <div style="font-weight: 600; color: #1f2937; margin-bottom: 4px;">
                                        👨‍🎓 ${lesson.student_name}
                                    </div>
                                    <div style="color: #6b7280; font-size: 0.9em;">
                                        🏫 ${lesson.student_class}
                                    </div>
                                </div>
                            </div>
                        `;
                    });

                    detailHTML += `</div>`;
                });

                detailHTML += `</div>`;
            }

            // Toplam özet
            detailHTML += `
                <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin-top: 20px;">
                    <div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 20px; border-radius: 12px; text-align: center;">
                        <div style="font-size: 2.5em; font-weight: bold; margin-bottom: 5px;">
                            ${totalLessons}
                        </div>
                        <div style="font-size: 1em; opacity: 0.9;">
                            Toplam Ders
                        </div>
                    </div>
                    <div style="background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%); color: white; padding: 20px; border-radius: 12px; text-align: center;">
                        <div style="font-size: 2.5em; font-weight: bold; margin-bottom: 5px;">
                            ${uniqueStudents.size}
                        </div>
                        <div style="font-size: 1em; opacity: 0.9;">
                            Farklı Öğrenci
                        </div>
                    </div>
                </div>
            `;

            document.getElementById('detailContent').innerHTML = detailHTML;
            document.getElementById('detailModal').style.display = 'flex';
        }

        function closeDetailModal() {
            document.getElementById('detailModal').style.display = 'none';
        }

        // ⚠️ BUGÜN BUTONU UYARI MODALI FONKSİYONLARI
        function showTodayWarningModal(message, targetDate) {
            const modal = document.getElementById('todayWarningModal');
            const content = document.getElementById('todayWarningContent');

            content.innerHTML = `
                <div style="margin-bottom: 20px;">
                    <i class="fas fa-exclamation-triangle" style="font-size: 4em; color: #f59e0b; margin-bottom: 15px;"></i>
                    <h3 style="color: #1f2937; margin: 15px 0;">${message}</h3>
                </div>
                <div style="display: flex; gap: 10px; justify-content: center;">
                    <button onclick="closeTodayWarningModal()"
                        style="background: #e5e7eb; color: #6b7280; border: none; padding: 12px 24px; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em;">
                        İptal
                    </button>
                    <button onclick="goToTargetDate('${targetDate}')"
                        style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; border: none; padding: 12px 24px; border-radius: 8px; cursor: pointer; font-weight: 600; font-size: 1em;">
                        Tamam
                    </button>
                </div>
            `;

            modal.style.display = 'block';
        }

        function closeTodayWarningModal() {
            document.getElementById('todayWarningModal').style.display = 'none';
        }

        function goToTargetDate(dateStr) {
            closeTodayWarningModal();
            const dropdown = document.getElementById('dayDropdown');
            dropdown.value = dateStr;
            loadTodayLessons(dateStr);
        }


        // ============== HAFTALIK PROGRAM GÖRÜNTÜLEYICI ==============
        let currentWeekView = 1;
        let globalScheduleData = null;

        // Takvim için global değişkenler
        let currentView = 'table'; // 'table' veya 'calendar'
        let currentCalendarDate = new Date(); // Şu anki görüntülenen ay

        function renderWeeklyTable(weekNum) {
            const container = document.getElementById('weeklyScheduleTable');

            if (!globalScheduleData) {
                container.innerHTML = '<p style="text-align: center; padding: 20px; color: #999;">Önce program oluşturun.</p>';
                return;
            }

            // ============== TARİH HESAPLAMA FONKSİYONU ==============
            function calculateDateForDay(dayName, weekNumber) {
                if (!globalScheduleData.start_date) {
                    return ''; // Tarih yoksa boş döndür
                }

                const dayMap = {
                    'Pazartesi': 0,
                    'Salı': 1,
                    'Çarşamba': 2,
                    'Perşembe': 3,
                    'Cuma': 4,
                    'Cumartesi': 5,
                    'Pazar': 6
                };

                const startDate = new Date(globalScheduleData.start_date + 'T00:00:00');
                const dayOffset = dayMap[dayName];
                const weekOffset = (weekNumber - 1) * 7;
                const totalOffset = weekOffset + dayOffset;

                const targetDate = new Date(startDate);
                targetDate.setDate(startDate.getDate() + totalOffset);

                // Formatla: "2 Ara 24"
                const day = targetDate.getDate();
                const monthNames = ['Oca', 'Şub', 'Mar', 'Nis', 'May', 'Haz', 'Tem', 'Ağu', 'Eyl', 'Eki', 'Kas', 'Ara'];
                const month = monthNames[targetDate.getMonth()];
                const year = String(targetDate.getFullYear()).slice(-2);

                return `${day} ${month} ${year}`;
            }

            // Öğretmenleri çek
            fetch('/get_teachers')
                .then(response => response.json())
                .then(data => {
                    const teachers = data.teachers;

                    // Alfabetik sırala
                    teachers.sort((a, b) => {
                        if (a.branch !== b.branch) {
                            return a.branch.localeCompare(b.branch, 'tr');
                        }
                        return a.name.localeCompare(b.name, 'tr');
                    });

                    // Tüm slotları topla
                    const allSlots = [];
                    teachers.forEach(teacher => {
                        teacher.schedule.forEach(daySchedule => {
                            daySchedule.lessons.forEach(lesson => {
                                const slotKey = `${daySchedule.day}_${lesson.start_time}_${lesson.end_time}`;
                                const existingSlot = allSlots.find(s => s.key === slotKey);
                                if (!existingSlot) {
                                    allSlots.push({
                                        day: daySchedule.day,
                                        start_time: lesson.start_time,
                                        end_time: lesson.end_time,
                                        key: slotKey
                                    });
                                }
                            });
                        });
                    });

                    // Günlere göre sırala
                    const dayOrder = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar'];
                    allSlots.sort((a, b) => {
                        const dayDiff = dayOrder.indexOf(a.day) - dayOrder.indexOf(b.day);
                        if (dayDiff !== 0) return dayDiff;
                        return a.start_time.localeCompare(b.start_time);
                    });

                    // Hafta datasını al
                    const weekData = globalScheduleData.weeks[weekNum - 1];

                    // Tablo oluştur (Hafta başlığı caption ile birlikte)
                    let html = `
                        <table id="weeklyPrintTable" style="width: 100%; border-collapse: collapse; box-shadow: 0 4px 15px rgba(0,0,0,0.1); overflow: hidden; table-layout: fixed; border-radius: 15px; transform: scale(0.888); transform-origin: top left;">
                            <caption style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px 20px; text-align: center; font-weight: bold; font-size: 1.5em; caption-side: top; border-radius: 10px 10px 0 0; margin: 0;">
                                <span style="display: inline-block; background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 10px 25px; border-radius: 50px; font-weight: bold; font-size: 1em; animation: pulseScale 2s ease-in-out infinite; box-shadow: 0 4px 15px rgba(16, 185, 129, 0.4);">
                                    HAFTA ${weekNum}
                                </span>
                            </caption>
                            <colgroup>
                                <col style="width: 120px;">
                    `;

                    // Öğretmen kolonları için dinamik genişlik
                    const teacherColWidth = '120px';
                    teachers.forEach(() => {
                        html += `<col style="width: ${teacherColWidth};">`;
                    });

                    html += `
                            </colgroup>
                            <thead>
                                <tr style="background: linear-gradient(135deg, #4472C4 0%, #5B9BD5 100%);">
                                    <th style="color: white; padding: 12px 15px; text-align: left; font-weight: bold; font-size: 0.9em; border-right: 1px solid rgba(255,255,255,0.2);">
                                        GÜN / SAAT
                                    </th>
                    `;

                    // Öğretmen başlıkları
                    teachers.forEach((teacher, idx) => {
                        const isLast = idx === teachers.length - 1;
                        html += `
                            <th style="color: white; padding: 12px 4px; text-align: center; font-weight: bold; font-size: 0.85em; text-transform: uppercase; ${!isLast ? 'border-right: 1px solid rgba(255,255,255,0.2);' : ''} white-space: normal; word-wrap: break-word; line-height: 1.3;">
                                ${teacher.branch}<br>
                                <span style="font-size: 0.8em; opacity: 0.9;">(${teacher.name} ${teacher.surname})</span>
                            </th>
                        `;
                    });

                    html += `
                                </tr>
                            </thead>
                            <tbody>
                    `;

                    // Satırları oluştur
                    let currentDay = null;
                    let rowIndex = 0;  // 🎬 Animasyon için satır sayacı
                    allSlots.forEach(slot => {
                        // Gün başlığı
                        if (slot.day !== currentDay) {
                            const dateStr = calculateDateForDay(slot.day, weekNum);
                            const fullDayTitle = dateStr ? `${slot.day} - ${dateStr}` : slot.day;

                            html += `
                                <tr>
                                    <td colspan="${teachers.length + 1}" style="background: #9575CD !important; color: #FFFFFF !important; font-weight: bold !important; font-size: 1em !important; padding: 10px 15px !important; text-align: center !important; border: 1px solid #7E57C2 !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important;">
                                        ${fullDayTitle}
                                    </td>
                                </tr>
                            `;
                            currentDay = slot.day;
                        }

                        // Saat satırı
                        rowIndex++;  // 🎬 Her satır için sayacı artır
                        const animDelay = (rowIndex * 0.03).toFixed(2);  // 🎬 Her satır için 0.03s gecikme
                        html += `
                            <tr style="background: #E3F2FD; animation: slideInRow 0.5s ease backwards; animation-delay: ${animDelay}s;">
                                <td style="font-weight: 600; color: #1565C0; padding: 10px 15px; border: 1px solid #e5e7eb; font-size: 0.9em;">
                                    ${slot.start_time}-${slot.end_time}
                                </td>
                        `;

                        // Her öğretmen için öğrenci
                        teachers.forEach(teacher => {
                            const teacherFullName = `${teacher.name} ${teacher.surname}`;
                            let studentName = '';

                            // 🆕 TÜM eşleşen dersleri bul (tek değil!)
                            const matchingLessons = weekData.filter(l =>
                                l.teacher_name === teacherFullName &&
                                l.day === slot.day &&
                                l.time === `${slot.start_time}-${slot.end_time}`
                            );

                            // 🆕 Gruplama mantığı
                            if (matchingLessons.length === 0) {
                                studentName = '';
                            } else if (matchingLessons.length === 1) {
                                studentName = matchingLessons[0].student_name;
                            } else {
                                // 🆕 GRUP DERSİ - TÜM SINIFLARI TOPLA
                                const uniqueClasses = [...new Set(matchingLessons.map(l => l.student_class).filter(c => c))];
                                if (uniqueClasses.length > 0) {
                                    const classesStr = uniqueClasses.sort().join(', ');
                                    studentName = `${classesStr} (${matchingLessons.length} öğr)`;
                                } else {
                                    studentName = `${matchingLessons[0].student_name} +${matchingLessons.length - 1}`;
                                }
                            }

                            html += `
                                <td style="padding: 10px 8px; text-align: center; border: 1px solid #e5e7eb; font-size: 0.85em; font-weight: 600; color: #1f2937; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">
                                    ${studentName}
                                </td>
                            `;
                        });

                        html += `</tr>`;
                    });

                    html += `
                            </tbody>
                        </table>
                    `;

                    container.innerHTML = html;

                    // ✅ BUTONLARI GÜNCELLE
                    if (typeof updateWeekButtons === 'function') {
                        updateWeekButtons();
                    }
                    // 🎯 DRAG & DROP'U AKTİF ET VE BORDER'LARI RESTORE ET
                    setTimeout(() => {
                        enableDragAndDrop();
                        restoreAykiriSwapBorders(weekNum);  // ✅ Aykırı swap border'larını geri yükle
                    }, 100);
                });
        }

        // ═══════════════════════════════════════════════════════════════
        // 🎨 AYKIRI SWAP BORDER RESTORE SİSTEMİ
        // ═══════════════════════════════════════════════════════════════

        /**
         * 🎨 AYKIRI SWAP BORDER'LARINI RESTORE ET
         * Hafta değiştiğinde sessionStorage'daki ihlallere göre border'ları geri ekler
         * @param {number} weekNum - Gösterilecek hafta numarası (1-4)
         */
        function restoreAykiriSwapBorders(weekNum) {
            try {
                const stored = sessionStorage.getItem('aykiriSwapViolations');
                if (!stored) {
                    console.log('✅ Restore edilecek aykırı swap yok');
                    return;
                }

                const violations = JSON.parse(stored);

                // Bu haftadaki ihlalleri filtrele
                const weekViolations = violations.filter(v => v.week === weekNum);

                if (weekViolations.length === 0) {
                    console.log(`✅ Hafta ${weekNum} için aykırı swap yok`);
                    return;
                }

                console.log(`🔄 ${weekViolations.length} aykırı swap border'ı restore ediliyor...`);

                let restoredCount = 0;

                // Her ihlal için border'ı geri ekle
                weekViolations.forEach(violation => {
                    // ⚠️ teacherName filtresini KALDIRDIK - çakışmalar farklı öğretmen sütunlarında olabilir
                    const cells = findCellsByDayAndTime(violation.day, violation.time);

                    cells.forEach(cell => {
                        const cellText = cell.textContent.trim().toLocaleUpperCase('tr');
                        let isConflicting = false;

                        // 1️⃣ Bireysel ders kontrolü: Öğrenci adını içeriyor mu?
                        isConflicting = violation.conflictingStudents.some(studentName => {
                            const studentUpper = studentName.toLocaleUpperCase('tr');
                            return cellText.includes(studentUpper);
                        });

                        // 2️⃣ Sınıf/Grup dersi kontrolü: Sınıf adını içeriyor mu?
                        if (!isConflicting && violation.studentClass) {
                            // Hücre formatı: "11A, 11B (7 öğr)" veya "11A (5 öğr)"
                            // Sınıfı kontrol et
                            isConflicting = cellText.includes(violation.studentClass.toLocaleUpperCase('tr'));
                        }

                        if (isConflicting) {
                            // Border'ı geri ekle
                            cell.style.border = `4px solid ${violation.borderColor}`;
                            cell.style.boxSizing = 'border-box';
                            restoredCount++;
                        }
                    });
                });

                console.log(`✅ ${restoredCount} border başarıyla restore edildi`);

            } catch (e) {
                console.error('❌ Border restore hatası:', e);
            }
        }

        /**
         * 🔍 GÜN VE SAATE GÖRE HÜCRELERİ BUL
         * Tabloda belirli gün ve saatteki tüm öğrenci hücrelerini döndürür
         * @param {string} day - Gün adı (örn: "Pazartesi")
         * @param {string} time - Saat aralığı (örn: "10:00-11:00")
         * @returns {Array} Bulunan hücre elemanları
         */
        function findCellsByDayAndTime(day, time) {
            const table = document.getElementById('weeklyPrintTable');
            if (!table) return [];

            const cells = [];
            const rows = table.querySelectorAll('tbody tr');
            let currentDay = '';

            for (const row of rows) {
                // Gün başlığı mı?
                const dayCell = row.querySelector('td[colspan]');
                if (dayCell) {
                    const fullDayText = dayCell.textContent.trim();
                    currentDay = fullDayText;
                    continue;
                }

                // Saat satırı mı?
                const timeCell = row.querySelector('td:first-child');
                if (timeCell) {
                    const rowTime = timeCell.textContent.trim();

                    // Gün adını temizle (tarih bilgisi varsa ayır)
                    const cleanCurrentDay = extractDayName(currentDay).toLocaleUpperCase('tr');
                    const cleanTargetDay = extractDayName(day).toLocaleUpperCase('tr');

                    // Gün ve saat eşleşiyor mu?
                    if (cleanCurrentDay === cleanTargetDay && rowTime === time) {
                        // Bu satırdaki tüm öğrenci hücrelerini al (ilk sütun hariç)
                        const studentCells = row.querySelectorAll('td:not(:first-child)');
                        cells.push(...studentCells);
                        break; // Bu gün/saati bulduk, döngüden çık
                    }
                }
            }

            return cells;
        }

        // 📊 GRAFİK DEĞİŞKENLERİ
        let branchChartInstance = null;
        let weeklyChartInstance = null;

        // 📊 İSTATİSTİK KARTLARINI GÜNCELLE
        function updateStatisticsCards(schedule) {
            // Stat kartlarını göster
            document.getElementById('statsCardsSection').style.display = 'block';

            // Toplam ders sayısı
            let totalLessons = 0;
            schedule.weeks.forEach(week => {
                totalLessons += week.length;
            });
            document.getElementById('totalLessons').textContent = totalLessons;

            // Öğretmen ve öğrenci sayıları
            fetch('/get_teachers').then(r => r.json()).then(data => {
                document.getElementById('totalTeachers').textContent = data.teachers.length;
            });

            fetch('/get_students').then(r => r.json()).then(data => {
                document.getElementById('totalStudents').textContent = data.students.length;
            });

            // Grafikleri oluştur
            createBranchChart(schedule);
            createWeeklyChart(schedule);
        }

        // 📊 BRANŞ DAĞILIM GRAFİĞİ
        function createBranchChart(schedule) {
            const ctx = document.getElementById('branchChart');

            // Eski grafiği temizle
            if (branchChartInstance) {
                branchChartInstance.destroy();
            }

            // Branş sayılarını topla
            const branchCounts = {};
            schedule.weeks.forEach(week => {
                week.forEach(lesson => {
                    branchCounts[lesson.branch] = (branchCounts[lesson.branch] || 0) + 1;
                });
            });

            // Renk paleti
            const colors = [
                '#667eea', '#764ba2', '#f093fb', '#4facfe',
                '#43e97b', '#fa709a', '#fee140', '#30cfd0',
                '#a8edea', '#fed6e3', '#c471f5', '#12c2e9'
            ];

            branchChartInstance = new Chart(ctx, {
                type: 'pie',
                data: {
                    labels: Object.keys(branchCounts),
                    datasets: [{
                        data: Object.values(branchCounts),
                        backgroundColor: colors,
                        borderWidth: 2,
                        borderColor: '#fff'
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: true,
                    plugins: {
                        legend: {
                            position: 'bottom',
                            labels: {
                                padding: 15,
                                font: {
                                    size: 12,
                                    family: "'Inter', sans-serif"
                                }
                            }
                        },
                        tooltip: {
                            callbacks: {
                                label: function(context) {
                                    const label = context.label || '';
                                    const value = context.parsed || 0;
                                    const total = context.dataset.data.reduce((a, b) => a + b, 0);
                                    const percentage = ((value / total) * 100).toFixed(1);
                                    return `${label}: ${value} ders (${percentage}%)`;
                                }
                            }
                        }
                    }
                }
            });
        }

        // 📊 HAFTALIK DERS DAĞILIM GRAFİĞİ
        function createWeeklyChart(schedule) {
            const ctx = document.getElementById('weeklyChart');

            // Eski grafiği temizle
            if (weeklyChartInstance) {
                weeklyChartInstance.destroy();
            }

            // Haftalık ders sayıları
            const weeklyData = schedule.weeks.map((week, index) => ({
                week: `Hafta ${index + 1}`,
                count: week.length
            }));

            weeklyChartInstance = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: weeklyData.map(d => d.week),
                    datasets: [{
                        label: 'Ders Sayısı',
                        data: weeklyData.map(d => d.count),
                        backgroundColor: [
                            'rgba(102, 126, 234, 0.8)',
                            'rgba(118, 75, 162, 0.8)',
                            'rgba(240, 147, 251, 0.8)',
                            'rgba(79, 172, 254, 0.8)'
                        ],
                        borderColor: [
                            'rgb(102, 126, 234)',
                            'rgb(118, 75, 162)',
                            'rgb(240, 147, 251)',
                            'rgb(79, 172, 254)'
                        ],
                        borderWidth: 2,
                        borderRadius: 8
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: true,
                    scales: {
                        y: {
                            beginAtZero: true,
                            ticks: {
                                stepSize: 1,
                                font: {
                                    family: "'Inter', sans-serif"
                                }
                            },
                            grid: {
                                color: 'rgba(0, 0, 0, 0.05)'
                            }
                        },
                        x: {
                            ticks: {
                                font: {
                                    family: "'Inter', sans-serif"
                                }
                            },
                            grid: {
                                display: false
                            }
                        }
                    },
                    plugins: {
                        legend: {
                            display: false
                        },
                        tooltip: {
                            callbacks: {
                                label: function(context) {
                                    return `${context.parsed.y} ders`;
                                }
                            }
                        }
                    }
                }
            });
        }

        // 🌙 DARK MODE FONKSİYONLARI
        function toggleDarkMode() {
            const body = document.body;
            const icon = document.getElementById('darkModeIcon');
            const text = document.getElementById('darkModeText');

            body.classList.toggle('dark-mode');

            if (body.classList.contains('dark-mode')) {
                icon.className = 'fas fa-sun';
                text.textContent = 'Açık Tema';
                localStorage.setItem('darkMode', 'enabled');
            } else {
                icon.className = 'fas fa-moon';
                text.textContent = 'Koyu Tema';
                localStorage.setItem('darkMode', 'disabled');
            }
        }

        // 🌙 SAYFA YÜKLENİNCE DARK MODE DURUMUNU KONTROL ET
        function checkDarkMode() {
            const darkMode = localStorage.getItem('darkMode');
            const body = document.body;
            const icon = document.getElementById('darkModeIcon');
            const text = document.getElementById('darkModeText');

            if (darkMode === 'enabled') {
                body.classList.add('dark-mode');
                if (icon) icon.className = 'fas fa-sun';
                if (text) text.textContent = 'Açık Tema';
            }
        }

        // Sayfa yüklenince kontrol et
        document.addEventListener('DOMContentLoaded', function() {
            checkDarkMode();
            loadTodayLessons(); // Bugünün derslerini yükle
        });

        // CAKISMA KONTROL - CALISIR VERSIYON
        function openConflictDashboard() {
            if (!globalScheduleData) {
                alert('Once program olusturun!');
                return;
            }
            document.getElementById('conflictDashboardModal').style.display = 'block';
            checkConflictsNow();
        }

        function closeConflictDashboard() {
            document.getElementById('conflictDashboardModal').style.display = 'none';
        }

        // 🆕 BADGE TÜRKÇELEŞTIRME
        function translateBadge(type) {
            const translations = {
                'restriction': 'KISITLAMA',
                'teacher_block': 'ÖĞRETMEN ENGELİ',
                'student': 'ÖĞRENCİ ÇAKIŞMASI',
                'teacher': 'ÖĞRETMEN ÇAKIŞMASI'
            };
            return translations[type] || type.toLocaleUpperCase('tr');
        }

        function checkConflictsNow() {
            const loadingEl = document.getElementById('conflictLoading');
            const statsEl = document.getElementById('conflictStats');
            const listEl = document.getElementById('conflictList');
            const noConflictsEl = document.getElementById('noConflicts');

            if (loadingEl) loadingEl.style.display = 'block';
            if (statsEl) statsEl.style.display = 'none';
            if (listEl) listEl.style.display = 'none';
            if (noConflictsEl) noConflictsEl.style.display = 'none';

            fetch('/check_conflicts', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' }
            })
            .then(response => response.json())
            .then(data => {
                if (loadingEl) loadingEl.style.display = 'none';

                // 🆕 GRUP DERSLERİNİ KONTROL ET
                const hasGroupLessons = data.group_lessons && data.group_lessons.length > 0;
                const hasConflicts = data.conflicts && data.conflicts.length > 0;

                if (!hasConflicts && !hasGroupLessons) {
                    if (noConflictsEl) noConflictsEl.style.display = 'block';
                    return;
                }

                if (!hasConflicts && hasGroupLessons) {
                    // Sadece grup dersleri var, çakışma yok
                    if (noConflictsEl) noConflictsEl.style.display = 'block';
                }

                // KARTLARI GÜNCELLE
                if (statsEl) statsEl.style.display = 'block';

                document.getElementById('criticalCount').textContent = data.summary.critical || 0;
                document.getElementById('highCount').textContent = data.summary.high || 0;
                document.getElementById('mediumCount').textContent = data.summary.medium || 0;
                document.getElementById('totalConflictCount').textContent = data.summary.total || 0;

                // 🆕 BADGE'İ GÜNCELLE
                const totalConflicts = data.summary.total || 0;
                updateConflictBadge(totalConflicts);

                // HAFTALIK DAĞILIM
                document.getElementById('week1Conflicts').textContent = data.summary.by_week.week1 || 0;
                document.getElementById('week2Conflicts').textContent = data.summary.by_week.week2 || 0;
                document.getElementById('week3Conflicts').textContent = data.summary.by_week.week3 || 0;
                document.getElementById('week4Conflicts').textContent = data.summary.by_week.week4 || 0;

                // CİDDİYET GÖSTERGESİ
                const indicator = document.getElementById('severityIndicator');
                if (indicator) {
                    if (data.severity === 'critical') {
                        indicator.style.background = 'linear-gradient(135deg, #dc2626 0%, #991b1b 100%)';
                        indicator.style.color = 'white';
                        indicator.textContent = '🔴 KRİTİK SEVİYE - Acil Müdahale Gerekli!';
                    } else if (data.severity === 'high') {
                        indicator.style.background = 'linear-gradient(135deg, #f59e0b 0%, #d97706 100%)';
                        indicator.style.color = 'white';
                        indicator.textContent = '⚠️ YÜKSEK SEVİYE - Düzeltme Önerilir';
                    } else if (data.severity === 'medium') {
                        indicator.style.background = 'linear-gradient(135deg, #3b82f6 0%, #2563eb 100%)';
                        indicator.style.color = 'white';
                        indicator.textContent = 'ℹ️ ORTA SEVİYE - Kontrol Edilmeli';
                    } else {
                        indicator.style.background = 'linear-gradient(135deg, #10b981 0%, #059669 100%)';
                        indicator.style.color = 'white';
                        indicator.textContent = '✅ DÜŞÜK SEVİYE - Sorun Yok';
                    }
                }

                // 🆕 GRUP DERSLERİNİ GÖSTER
                const groupLessonsSection = document.getElementById('groupLessonsSection');
                const groupLessonsList = document.getElementById('groupLessonsList');
                const groupLessonsCount = document.getElementById('groupLessonsCount');

                if (data.group_lessons && data.group_lessons.length > 0) {
                    if (groupLessonsSection) groupLessonsSection.style.display = 'block';
                    if (groupLessonsCount) groupLessonsCount.textContent = data.group_lessons.length;

                    if (groupLessonsList) {
                        let groupHtml = '';

                        data.group_lessons.forEach((groupLesson, index) => {
                            const classesText = groupLesson.classes.join(', ');

                            // Bu grup dersine ait ihlalleri bul
                            const key = `${groupLesson.teacher}_${groupLesson.day}_${groupLesson.time}_${groupLesson.week}`;
                            const groupViolations = data.grouped_violations && data.grouped_violations[key]
                                ? data.grouped_violations[key].violations
                                : [];

                            const violationCount = groupViolations.length;
                            const accordionId = `groupViolations_${index}`;

                            groupHtml += `
                                <div style="background: white; border-left: 3px solid #10b981; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                                    <div style="display: flex; justify-content: space-between; align-items: start; margin-bottom: 10px;">
                                        <div style="flex: 1;">
                                            <div style="font-weight: bold; color: #059669; font-size: 1.05em; margin-bottom: 5px;">
                                                <i class="fas fa-chalkboard-teacher"></i> ${groupLesson.teacher}
                                            </div>
                                            <div style="color: #047857; font-size: 0.95em; margin-bottom: 5px;">
                                                <i class="fas fa-book"></i> ${groupLesson.branch || 'Ders'}
                                            </div>
                                            <div style="color: #6b7280; font-size: 0.9em;">
                                                <i class="fas fa-calendar"></i> ${groupLesson.day} ${groupLesson.time} - Hafta ${groupLesson.week}
                                            </div>
                                        </div>
                                        <div style="background: #10b981; color: white; padding: 6px 12px; border-radius: 20px; font-size: 0.85em; font-weight: bold;">
                                            ${groupLesson.classes.length} Sınıf
                                        </div>
                                    </div>

                                    <div style="background: #f0fdf4; padding: 10px; border-radius: 6px; margin-bottom: ${violationCount > 0 ? '10px' : '0'};">
                                        <div style="color: #047857; font-size: 0.9em; font-weight: 600; margin-bottom: 5px;">
                                            <i class="fas fa-users"></i> Katılımcı Sınıflar:
                                        </div>
                                        <div style="color: #059669; font-size: 0.95em;">
                                            ${classesText}
                                        </div>
                                    </div>
            `;

                            // İhlaller varsa göster
                            if (violationCount > 0) {
                                groupHtml += `
                                    <div style="border-top: 1px solid #d1fae5; padding-top: 10px;">
                                        <div style="cursor: pointer; display: flex; justify-content: space-between; align-items: center; padding: 8px; background: #fef3c7; border-radius: 6px;"
                                             onclick="document.getElementById('${accordionId}').style.display = document.getElementById('${accordionId}').style.display === 'none' ? 'block' : 'none'; this.querySelector('.accordion-icon').textContent = document.getElementById('${accordionId}').style.display === 'none' ? '▼' : '▲';">
                                            <span style="color: #92400e; font-weight: 600; font-size: 0.9em;">
                                                <i class="fas fa-exclamation-triangle"></i> İhlaller (${violationCount})
                                            </span>
                                            <span class="accordion-icon" style="color: #92400e; font-weight: bold;">▼</span>
                                        </div>

                                        <div id="${accordionId}" style="display: none; margin-top: 8px; background: #fef3c7; border-radius: 6px; padding: 12px;">
                `;

                                // İhlalleri listele
                                groupViolations.forEach((violation, vIdx) => {
                                    const badgeText = translateBadge(violation.type);
                                    groupHtml += `
                                        <div style="display: flex; align-items: start; gap: 10px; margin-bottom: ${vIdx < violationCount - 1 ? '10px' : '0'}; padding-bottom: ${vIdx < violationCount - 1 ? '10px' : '0'}; border-bottom: ${vIdx < violationCount - 1 ? '1px solid #fde68a' : 'none'};">
                                            <span class="conflict-type-badge ${violation.type}" style="flex-shrink: 0;">${badgeText}</span>
                                            <div style="color: #78350f; line-height: 1.6; font-size: 0.9em;">
                                                ${violation.message}
                                            </div>
                                        </div>
                                    `;
                                });

                                groupHtml += `
                                        </div>
                                    </div>
                                `;
                            }

                            groupHtml += `
                                </div>
                            `;
                        });

                        groupLessonsList.innerHTML = groupHtml;
                    }
                } else {
                    if (groupLessonsSection) groupLessonsSection.style.display = 'none';
                }

                // 🆕 ONAYLANMIŞ İHLALLİ SINIF DERSLERİNİ GÖSTER
                const approvedSection = document.getElementById('approvedViolationsSection');
                const approvedList = document.getElementById('approvedViolationsList');
                const approvedCount = document.getElementById('approvedViolationsCount');

                if (data.approved_violations && data.approved_violations.length > 0) {
                    if (approvedSection) approvedSection.style.display = 'block';
                    if (approvedCount) approvedCount.textContent = data.approved_violations.length;

                    if (approvedList) {
                        let approvedHtml = '';

                        data.approved_violations.forEach((approved, index) => {
                            const violationCount = approved.violations ? approved.violations.length : 0;
                            const accordionId = `approvedViolations_${index}`;

                            approvedHtml += `
                                <div style="background: white; border-left: 3px solid #8b5cf6; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                                    <div style="display: flex; justify-content: space-between; align-items: start; margin-bottom: 10px;">
                                        <div style="flex: 1;">
                                            <div style="font-weight: bold; color: #7c3aed; font-size: 1.05em; margin-bottom: 5px;">
                                                <i class="fas fa-chalkboard-teacher"></i> ${approved.teacher}
                                            </div>
                                            <div style="color: #6b21a8; font-size: 0.95em; margin-bottom: 5px;">
                                                <i class="fas fa-book"></i> ${approved.branch || 'Ders'}
                                            </div>
                                            <div style="color: #6b7280; font-size: 0.9em;">
                                                <i class="fas fa-calendar"></i> ${approved.day} ${approved.time} - Hafta ${approved.week}
                                            </div>
                                        </div>
                                        <div style="background: #8b5cf6; color: white; padding: 6px 12px; border-radius: 20px; font-size: 0.85em; font-weight: bold;">
                                            ${violationCount} İhlal
                                        </div>
                                    </div>

                                    <div style="background: #faf5ff; padding: 10px; border-radius: 6px; margin-bottom: ${violationCount > 0 ? '10px' : '0'};">
                                        <div style="color: #7c3aed; font-size: 0.9em; font-weight: 600; margin-bottom: 5px;">
                                            <i class="fas fa-school"></i> Sınıf:
                                        </div>
                                        <div style="color: #8b5cf6; font-size: 0.95em;">
                                            ${approved.class_name}
                                        </div>
                                    </div>
            `;

                            // İhlaller varsa göster
                            if (violationCount > 0) {
                                approvedHtml += `
                                    <div style="border-top: 1px solid #e9d5ff; padding-top: 10px;">
                                        <div style="cursor: pointer; display: flex; justify-content: space-between; align-items: center; padding: 8px; background: #fef3c7; border-radius: 6px;"
                                             onclick="document.getElementById('${accordionId}').style.display = document.getElementById('${accordionId}').style.display === 'none' ? 'block' : 'none'; this.querySelector('.accordion-icon').textContent = document.getElementById('${accordionId}').style.display === 'none' ? '▼' : '▲';">
                                            <span style="color: #92400e; font-weight: 600; font-size: 0.9em;">
                                                <i class="fas fa-exclamation-triangle"></i> İhlaller (${violationCount})
                                            </span>
                                            <span class="accordion-icon" style="color: #92400e; font-weight: bold;">▼</span>
                                        </div>

                                        <div id="${accordionId}" style="display: none; margin-top: 8px; background: #fef3c7; border-radius: 6px; padding: 12px;">
                `;

                                // İhlalleri listele
                                approved.violations.forEach((violation, vIdx) => {
                                    const badgeText = translateBadge(violation.type);
                                    approvedHtml += `
                                        <div style="display: flex; align-items: start; gap: 10px; margin-bottom: ${vIdx < violationCount - 1 ? '10px' : '0'}; padding-bottom: ${vIdx < violationCount - 1 ? '10px' : '0'}; border-bottom: ${vIdx < violationCount - 1 ? '1px solid #fde68a' : 'none'};">
                                            <span class="conflict-type-badge ${violation.type}" style="flex-shrink: 0;">${badgeText}</span>
                                            <div style="color: #78350f; line-height: 1.6; font-size: 0.9em;">
                                                ${violation.message}
                                            </div>
                                        </div>
                                    `;
                                });

                                approvedHtml += `
                                        </div>
                                    </div>
                                `;
                            }

                            approvedHtml += `
                                </div>
                            `;
                        });

                        approvedList.innerHTML = approvedHtml;
                    }
                } else {
                    if (approvedSection) approvedSection.style.display = 'none';
                }

                // ═══════════════════════════════════════════════════════════
                // ⚠️ AYKIRI SWAP İHLALLERİNİ GÖSTER (YUMUŞAK KIRMIZI KARTLAR)
                // ═══════════════════════════════════════════════════════════
                displayAykiriSwapViolations();
            })
            .catch(error => {
                if (loadingEl) loadingEl.style.display = 'none';
                alert('Hata: ' + error);
                console.error('Çakışma kontrolü hatası:', error);
            });
        }

        // 🆕 BADGE GÜNCELLEME
        function updateConflictBadge(value) {
            const badge = document.getElementById('conflictBadge');
            if (badge) {
                if (value && value !== '0' && value !== 0) {
                    badge.textContent = value;
                    badge.style.display = 'inline-block';
                } else {
                    badge.style.display = 'none';
                }
            }
        }

        // 🆕 ARKA PLANDA KONTROL (YENİ SİSTEM)
        function checkConflictsInBackground() {
            fetch('/check_conflicts_v2', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' }
            })
            .then(response => response.json())
            .then(data => {
                // API'den gelen veriler
                let totalCards = data.summary?.total || 0;
                let totalIssues = data.summary?.total_issues || 0;

                // 🆕 AYKIRI SWAP KARTLARINI EKLE
                try {
                    const stored = sessionStorage.getItem('aykiriSwapViolations');
                    if (stored) {
                        const violations = JSON.parse(stored);
                        // Her aykırı swap = 1 kart
                        totalCards += violations.length;
                        // Her aykırı swap'taki çakışan öğrenci = ihlal
                        violations.forEach(violation => {
                            totalIssues += (violation.conflictingStudents?.length || 0);
                        });
                    }
                } catch (e) {
                    console.error('Aykırı swap sayacı hatası:', e);
                }

                if (totalCards > 0) {
                    // 🆕 KARMA FORMAT: "KART_SAYISI / İHLAL_SAYISI"
                    const badgeText = `${totalCards} / ${totalIssues}`;
                    updateConflictBadge(badgeText);
                } else {
                    updateConflictBadge('0');
                }
            })
            .catch(error => {
                console.error('Çakışma kontrolü hatası:', error);
            });
        }

        // 🔧 OTOMATİK DÜZELT
        window.autoFixConflicts = function() {
            if (!confirm('⚠️ Çakışan dersler otomatik olarak düzeltilecek. Onaylıyor musunuz?')) {
                return;
            }

            const btn = event.target.closest('button');
            const originalText = btn.innerHTML;
            btn.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Düzeltiliyor...';
            btn.disabled = true;

            fetch('/auto_fix_conflicts', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' }
            })
            .then(response => response.json())
            .then(data => {
                btn.innerHTML = originalText;
                btn.disabled = false;

                if (data.fixed > 0) {
                    showSuccess('✅ ' + data.fixed + ' çakışma düzeltildi!');
                    setTimeout(() => {
                        location.reload();
                    }, 1500);
                } else {
                    showError('❌ Düzeltilecek çakışma bulunamadı.');
                }

                if (data.remaining > 0) {
                    updateConflictBadge(data.remaining);
                } else {
                    updateConflictBadge(0);
                }
            })
            .catch(error => {
                btn.innerHTML = originalText;
                btn.disabled = false;
                showError('Düzeltme hatası: ' + error);
                console.error('Otomatik düzelt hatası:', error);
            });
        };

        // 📊 EXCEL RAPOR
        window.exportConflictReport = function() {
            showSuccess('Excel raporu hazırlanıyor...');
            window.location.href = '/export_conflict_report';
        };

        // 🔄 YENİLE
        window.refreshConflictCheck = function() {
            closeConflictDashboard();
            setTimeout(() => {
                openConflictDashboard();
            }, 300);
        };

        // ═══════════════════════════════════════════════════════════════
        // 🔍 ÖĞRENCİ ARAMA FONKSİYONLARI
        // ═══════════════════════════════════════════════════════════════

        function searchStudentInTable() {
            const searchText = document.getElementById('studentSearchBox').value.trim().toLocaleLowerCase('tr-TR');
            const resultInfo = document.getElementById('searchResultInfo');

            // Tüm hücreleri normal haline döndür
            const allCells = document.querySelectorAll('#weeklyPrintTable td');
            allCells.forEach(cell => {
                cell.classList.remove('student-highlight');
            });

            // Arama kutusu boşsa temizle
            if (searchText === '') {
                resultInfo.innerHTML = '';
                return;
            }

            // Arama yap
            let foundCount = 0;
            const table = document.getElementById('weeklyPrintTable');

            allCells.forEach(cell => {
                const cellText = cell.textContent.trim();
                const cellTextLower = cellText.toLocaleLowerCase('tr-TR');

                // Boş hücre veya saat hücresi değilse devam et
                if (!cellText || cellText.includes(':') || cellText.includes('-')) {
                    return;
                }

                let isMatch = false;

                // DIREKT ESLESME (bireysel ders)
                if (cellTextLower.includes(searchText)) {
                    isMatch = true;
                }

                // SINIF DERSI KONTROLU
                // Pattern: "11A (3 öğr)" veya "(3 öğrenci)" içeren hücreler
                if (!isMatch && (cellText.includes('(') && /\d+\s*(öğrenci|öğr)/.test(cellText))) {
                    console.log('🔍 Sınıf dersi bulundu:', cellText);

                    // Bu hucrede sinif dersi var, o slottaki ogrencileri kontrol et
                    let cellDay = '';
                    let cellTime = '';

                    // Hucrenin gun/saat bilgisini al
                    let row = cell.parentElement;
                    while (row) {
                        const dayCell = row.querySelector('.day-header, td[colspan]');
                        if (dayCell && dayCell.textContent.trim()) {
                            cellDay = dayCell.textContent.trim();
                            console.log('  📅 Gün:', cellDay);
                            break;
                        }
                        row = row.previousElementSibling;
                    }

                    const timeCell = cell.parentElement.querySelector('td:first-child');
                    if (timeCell) {
                        cellTime = timeCell.textContent.trim();
                        console.log('  🕐 Saat:', cellTime);
                    }

                    const cellIndex = Array.from(cell.parentElement.children).indexOf(cell);

                    // Bu slottaki ogrencileri bul
                    const cleanCellDay = extractDayName(cellDay);
                    const weekData = globalScheduleData ? globalScheduleData.weeks[currentWeekView - 1] : null;

                    console.log('  🗓️ Temiz gün:', cleanCellDay);
                    console.log('  📊 Hafta data var mı:', !!weekData);

                    if (weekData && cleanCellDay && cellTime) {
                        const headerRow = table.querySelector('thead tr');
                        const teacherHeader = headerRow ? headerRow.children[cellIndex] : null;
                        const teacherText = teacherHeader ? teacherHeader.textContent : '';

                        // Ogretmen adini parantezden ayikla
                        const teacherMatch = teacherText.match(/\(([^)]+)\)/);
                        const headerTeacherName = teacherMatch ? teacherMatch[1].trim() : '';

                        console.log('  👨‍🏫 Öğretmen:', headerTeacherName);

                        // Bu slottaki dersleri kontrol et
                        let foundInSlot = false;
                        for (const lesson of weekData) {
                            const lessonDay = lesson.day ? lesson.day.trim() : '';
                            const lessonTeacher = lesson.teacher_name ? lesson.teacher_name.trim() : '';

                            // Gün eşleşmesi (büyük/küçük harf duyarsız, Türkçe karakterler)
                            const dayMatch = lessonDay.toLocaleLowerCase('tr') === cleanCellDay.toLocaleLowerCase('tr');
                            // Saat eşleşmesi
                            const timeMatch = lesson.time === cellTime;
                            // Öğretmen eşleşmesi (büyük/küçük harf duyarsız, Türkçe karakterler)
                            const teacherMatch = lessonTeacher.toLocaleLowerCase('tr') === headerTeacherName.toLocaleLowerCase('tr');

                            if (dayMatch && timeMatch && teacherMatch) {
                                // Ogrenci adini kucuk harfe cevir ve karsilastir
                                const studentNameLower = lesson.student_name.toLocaleLowerCase('tr-TR');
                                if (studentNameLower.includes(searchText)) {
                                    console.log('  ✅ ÖĞRENCİ BULUNDU:', lesson.student_name);
                                    isMatch = true;
                                    foundInSlot = true;
                                    break;
                                }
                            }
                        }

                        if (!foundInSlot) {
                            console.log('  ❌ Bu slotta aranan öğrenci bulunamadı');
                        }
                    } else {
                        console.log('  ⚠️ Eksik bilgi - hafta:', !!weekData, 'gün:', !!cleanCellDay, 'saat:', !!cellTime);
                    }
                }

                if (isMatch) {
                    cell.classList.add('student-highlight');
                    foundCount++;
                }
            });

            // Sonuç bilgisi göster
            if (foundCount > 0) {
                resultInfo.innerHTML = `✅ <strong>${foundCount}</strong> ders bulundu ve vurgulandı.`;
                resultInfo.style.color = '#059669';
            } else {
                resultInfo.innerHTML = `❌ <strong>"${document.getElementById('studentSearchBox').value}"</strong> için sonuç bulunamadı.`;
                resultInfo.style.color = '#dc2626';
            }
        }

        function clearStudentSearch() {
            // Arama kutusunu temizle
            document.getElementById('studentSearchBox').value = '';

            // Vurguları kaldır
            const allCells = document.querySelectorAll('#weeklyPrintTable td');
            allCells.forEach(cell => {
                cell.classList.remove('student-highlight');
            });

            // Sonuç bilgisini temizle
            document.getElementById('searchResultInfo').innerHTML = '';
        }

        // 👨‍🏫 ÖĞRETMEN ARAMA FONKSİYONLARI
        function searchTeacherInTable() {
            const searchText = document.getElementById('teacherSearchBox').value.trim().toLocaleLowerCase('tr-TR');
            const resultInfo = document.getElementById('teacherSearchResultInfo');

            // Önceki vurgulamaları temizle
            const allCells = document.querySelectorAll('#weeklyPrintTable th');
            allCells.forEach(cell => {
                cell.classList.remove('teacher-highlight');
            });

            // Arama kutusu boşsa temizle
            if (searchText === '') {
                resultInfo.innerHTML = '';
                return;
            }

            // Arama yap (öğretmen başlıklarında)
            let foundCount = 0;
            const teacherHeaders = document.querySelectorAll('#weeklyPrintTable thead th');

            teacherHeaders.forEach(cell => {
                const cellText = cell.textContent.trim().toLocaleLowerCase('tr-TR');

                // "GÜN/SAAT" hücresini atla
                if (cellText.includes('gün') || cellText.includes('saat')) {
                    return;
                }

                if (cellText && cellText.includes(searchText)) {
                    cell.classList.add('teacher-highlight');
                    foundCount++;
                }
            });

            // Sonuç bilgisi göster
            if (foundCount > 0) {
                resultInfo.innerHTML = `✅ <strong>${foundCount}</strong> öğretmen bulundu ve vurgulandı.`;
                resultInfo.style.color = '#059669';
            } else {
                resultInfo.innerHTML = `❌ <strong>"${document.getElementById('teacherSearchBox').value}"</strong> için sonuç bulunamadı.`;
                resultInfo.style.color = '#dc2626';
            }
        }

        function clearTeacherSearch() {
            // Arama kutusunu temizle
            document.getElementById('teacherSearchBox').value = '';

            // Vurguları kaldır
            const allCells = document.querySelectorAll('#weeklyPrintTable th');
            allCells.forEach(cell => {
                cell.classList.remove('teacher-highlight');
            });

            // Sonuç bilgisini temizle
            document.getElementById('teacherSearchResultInfo').innerHTML = '';
        }

        // ═══════════════════════════════════════════════════════════════
        // 🎯 SÜRÜKLE-BIRAK SİSTEMİ - DRAG & DROP
        // ═══════════════════════════════════════════════════════════════

        let draggedCell = null;
        let draggedData = null;
        let swapPendingData = null;

        function enableDragAndDrop() {
            const table = document.getElementById('weeklyPrintTable');
            if (!table) return;

            // Tüm öğrenci hücrelerini sürüklenebilir yap
            const cells = table.querySelectorAll('tbody td:not(:first-child)');

            cells.forEach(cell => {
                // Zaten event listener varsa ekleme
                if (cell.getAttribute('data-drag-enabled')) return;
                cell.setAttribute('data-drag-enabled', 'true');

                const studentName = cell.textContent.trim();

                // Boş hücre
                if (!studentName) {
                    cell.classList.add('empty-slot');
                    cell.setAttribute('draggable', false);
                } else {
                    cell.classList.add('draggable-cell');
                    cell.setAttribute('draggable', true);
                }

                // DRAG START
                cell.addEventListener('dragstart', function(e) {
                    if (!this.textContent.trim()) return;

                    draggedCell = this;
                    this.classList.add('dragging');

                    // Veri topla
                    const row = this.parentElement;
                    const cellIndex = Array.from(row.children).indexOf(this);

                    // Gün ve saat bilgisini bul
                    let currentRow = this.parentElement;
                    let day = '';
                    let time = '';

                    // Gün başlığını bul (yukarı doğru tara)
                    while (currentRow) {
                        const dayCell = currentRow.querySelector('.day-header, td[colspan]');
                        if (dayCell && dayCell.textContent.trim()) {
                            day = dayCell.textContent.trim().toLocaleUpperCase('tr');
                            break;
                        }
                        currentRow = currentRow.previousElementSibling;
                    }

                    // Saat bilgisini al
                    const timeCell = this.parentElement.querySelector('td:first-child');
                    if (timeCell) {
                        time = timeCell.textContent.trim();
                    }

                    const cellText = this.textContent.trim();

                    // ✅ ÖĞRETMEN BİLGİSİNİ HER ZAMAN AL (sınıf ve bireysel dersler için)
                    const headerRow = this.closest('table').querySelector('thead tr');
                    const teacherHeader = headerRow.children[cellIndex];
                    const teacherText = teacherHeader ? teacherHeader.textContent : '';

                    // Ogretmen adini parantezden ayikla: "Matematik (DILEK COKUN)" -> "DILEK COKUN"
                    const teacherMatch = teacherText.match(/\(([^)]+)\)/);
                    const teacherFullName = teacherMatch ? teacherMatch[1].trim() : '';

                    // SINIF DERSI MI KONTROL ET - parantez icinde sayi varsa sinif dersidir
                    const isClassLesson = /\(\d+\s/.test(cellText);
                    let studentNames = [];

                    if (isClassLesson) {
                        // Sinif dersi - Bu slottaki TUM ogrencileri bul
                        const cleanDay = extractDayName(day);
                        const weekData = globalScheduleData.weeks[currentWeekView - 1];
                        const teacherUpper = teacherFullName.toLocaleUpperCase('tr');

                        // Bu slottaki tum ogrencileri topla
                        weekData.forEach(lesson => {
                            // GUN KARSILASTIRMASINI BUYUK HARFLE YAP
                            const lessonDayUpper = lesson.day ? lesson.day.toLocaleUpperCase('tr') : '';
                            const cleanDayUpper = cleanDay.toLocaleUpperCase('tr');
                            const lessonTeacherUpper = lesson.teacher_name ? lesson.teacher_name.toLocaleUpperCase('tr') : '';

                            if (lessonDayUpper === cleanDayUpper &&
                                lesson.time === time &&
                                lessonTeacherUpper === teacherUpper) {
                                studentNames.push(lesson.student_name);
                            }
                        });
                    } else {
                        // Bireysel ders
                        studentNames = [cellText];
                    }

                    draggedData = {
                        studentName: cellText,
                        studentNames: studentNames,
                        isClassLesson: isClassLesson,
                        teacherName: teacherFullName,  // ✅ Öğretmen bilgisi eklendi
                        day: day,
                        time: time,
                        cellIndex: cellIndex,
                        week: currentWeekView
                    };

                    e.dataTransfer.effectAllowed = 'move';
                    e.dataTransfer.setData('text/html', this.innerHTML);
                });

                // DRAG END
                cell.addEventListener('dragend', function() {
                    this.classList.remove('dragging');

                    // Tüm drag-over sınıflarını temizle
                    document.querySelectorAll('.drag-over').forEach(el => {
                        el.classList.remove('drag-over');
                    });
                });

                // DRAG OVER
                cell.addEventListener('dragover', function(e) {
                    if (draggedCell === this) return;

                    e.preventDefault();
                    e.dataTransfer.dropEffect = 'move';

                    this.classList.add('drag-over');
                });

                // DRAG LEAVE
                cell.addEventListener('dragleave', function() {
                    this.classList.remove('drag-over');
                });

                // DROP
                cell.addEventListener('drop', function(e) {
                    e.preventDefault();
                    this.classList.remove('drag-over');

                    if (draggedCell === this) return;

                    const targetStudentName = this.textContent.trim();

                    // Hedef slot dolu mu?
                    if (targetStudentName) {
                        // ONAY POPUP AÇ
                        showSwapConfirmation(this, targetStudentName);
                    } else {
                        // BOŞ SLOT - DİREKT TAŞI
                        performMove(this);
                    }
                });
            });
        }

        function showSwapConfirmation(targetCell, targetStudentName) {
            // Hedef hücrenin bilgilerini topla
            const row = targetCell.parentElement;
            let currentRow = targetCell.parentElement;
            let targetDay = '';
            let targetTime = '';

            // Gün başlığını bul
            while (currentRow) {
                const dayCell = currentRow.querySelector('.day-header, td[colspan]');
                if (dayCell && dayCell.textContent.trim()) {
                    targetDay = dayCell.textContent.trim().toLocaleUpperCase('tr');
                    break;
                }
                currentRow = currentRow.previousElementSibling;
            }

            // Saat bilgisini al
            const timeCell = targetCell.parentElement.querySelector('td:first-child');
            if (timeCell) {
                targetTime = timeCell.textContent.trim();
            }

            // ✅ HEDEF ÖĞRETMENİ BUL (cellIndex'ten)
            const targetCellIndex = Array.from(targetCell.parentElement.children).indexOf(targetCell);
            const headerRow = targetCell.closest('table').querySelector('thead tr');
            const targetTeacherHeader = headerRow.children[targetCellIndex];
            const targetTeacherText = targetTeacherHeader ? targetTeacherHeader.textContent : '';
            const targetTeacherMatch = targetTeacherText.match(/\(([^)]+)\)/);
            const targetTeacherName = targetTeacherMatch ? targetTeacherMatch[1].trim() : '';

            // Popup'ı doldur
            document.getElementById('swapStudent1').textContent =
                `${draggedData.studentName} (${draggedData.day} ${draggedData.time})`;
            document.getElementById('swapStudent2').textContent =
                `${targetStudentName} (${targetDay} ${targetTime})`;

            // Hedef bilgilerini sakla
            swapPendingData = {
                targetCell: targetCell,
                targetStudentName: targetStudentName,
                targetDay: targetDay,
                targetTime: targetTime,
                targetTeacherName: targetTeacherName  // ✅ Hedef öğretmen bilgisi eklendi
            };

            // Popup'ı göster
            document.getElementById('swapConfirmModal').style.display = 'flex';
        }

        // ═══════════════════════════════════════════════════════════
        // 🎨 RANDOM RENK ÜRETİCİ (SARI HARİÇ - Öğrenci arama ile karışmasın)
        // ═══════════════════════════════════════════════════════════
        function generateRandomBorderColor() {
            const colors = [
                '#ef4444', // Kırmızı
                '#f97316', // Turuncu
                '#10b981', // Yeşil
                '#3b82f6', // Mavi
                '#8b5cf6', // Mor
                '#ec4899', // Pembe
                '#06b6d4', // Cyan
                '#84cc16', // Lime
                '#f43f5e', // Rose
                '#6366f1', // Indigo
                '#14b8a6', // Teal
                '#a855f7', // Purple
                '#fb923c', // Orange-400
                '#22c55e', // Green-500
                '#0ea5e9', // Sky
                '#d946ef'  // Fuchsia
            ];

            return colors[Math.floor(Math.random() * colors.length)];
        }

        // ═══════════════════════════════════════════════════════════
        // 🔍 ÇAKIŞAN SLOTLARI BUL (Tabloda aynı gün/saat çakışan öğrenciler)
        // ═══════════════════════════════════════════════════════════
        function findConflictingSlots(targetDay, targetTime, swappedStudentNames) {
            const conflictingSlots = [];
            const table = document.getElementById('weeklyPrintTable');
            if (!table) return conflictingSlots;

            const cleanTargetDay = extractDayName(targetDay).toLocaleUpperCase('tr');
            const cells = table.querySelectorAll('tbody td:not(:first-child)');

            cells.forEach(cell => {
                const cellText = cell.textContent.trim();
                if (!cellText) return; // Boş slot

                // Bu cell'in gün ve saatini bul
                let cellDay = '';
                let cellTime = '';

                let row = cell.parentElement;
                while (row) {
                    const dayCell = row.querySelector('.day-header, td[colspan]');
                    if (dayCell && dayCell.textContent.trim()) {
                        cellDay = dayCell.textContent.trim().toLocaleUpperCase('tr');
                        break;
                    }
                    row = row.previousElementSibling;
                }

                const timeCell = cell.parentElement.querySelector('td:first-child');
                if (timeCell) {
                    cellTime = timeCell.textContent.trim();
                }

                const cleanCellDay = extractDayName(cellDay).toLocaleUpperCase('tr');

                // Aynı gün ve aynı saat mı?
                if (cleanCellDay === cleanTargetDay && cellTime === targetTime) {
                    // Bu slot'taki öğrenci swap yapılan öğrencilerden farklı mı?
                    const isSwappedStudent = swappedStudentNames && swappedStudentNames.some(name => cellText.includes(name));

                    if (!isSwappedStudent) {
                        // Çakışan slot bulundu!
                        conflictingSlots.push({
                            cell: cell,
                            studentName: cellText,
                            day: targetDay,
                            time: targetTime
                        });
                    }
                }
            });

            return conflictingSlots;
        }

        // ═══════════════════════════════════════════════════════════
        // 💾 İHLAL VERİLERİNİ KAYDET (sessionStorage)
        // ═══════════════════════════════════════════════════════════
        function saveViolationToPanel(violationData) {
            // Mevcut ihlalleri al
            let violations = [];
            try {
                const stored = sessionStorage.getItem('aykiriSwapViolations');
                if (stored) {
                    violations = JSON.parse(stored);
                }
            } catch (e) {
                console.error('İhlal verileri okunamadı:', e);
            }

            // Yeni ihlali ekle
            violations.push(violationData);

            // Kaydet
            try {
                sessionStorage.setItem('aykiriSwapViolations', JSON.stringify(violations));
                console.log('✅ İhlal kaydedildi:', violationData);
            } catch (e) {
                console.error('İhlal kaydedilemedi:', e);
            }
        }

        // ═══════════════════════════════════════════════════════════
        // 🔍 ÖĞRENCİNİN BELİRLİ SAATTEKİ ÇAKIŞMALARINI BUL
        // ═══════════════════════════════════════════════════════════
        function findStudentConflictsAtTime(studentName, targetDay, targetTime, excludeCell, studentClass = null) {
            const conflicts = [];
            const table = document.getElementById('weeklyPrintTable');
            if (!table) return conflicts;

            const cleanTargetDay = extractDayName(targetDay).toLocaleUpperCase('tr');
            const studentUpper = studentName.toLocaleUpperCase('tr');
            const cells = table.querySelectorAll('tbody td:not(:first-child)');

            cells.forEach(cell => {
                if (cell === excludeCell) return;
                const cellText = cell.textContent.trim();
                if (!cellText) return;

                const cellTextUpper = cellText.toLocaleUpperCase('tr');

                // 1️⃣ BİREYSEL DERS: Öğrenci ismini içeriyor mu?
                let isMatch = cellTextUpper.includes(studentUpper);

                // 2️⃣ SINIF DERSİ: Öğrencinin sınıfı ile eşleşiyor mu?
                if (!isMatch && studentClass) {
                    // Grup dersi: "12A, 12B (7 öğr)" formatını kontrol et
                    // Sınıf listesini çıkar: "12A, 12B (7 öğr)" → "12A, 12B"
                    const classListMatch = cellText.match(/^([^(]+)\s*\(/);
                    if (classListMatch) {
                        // "12A, 12B" → ["12A", "12B"]
                        const classList = classListMatch[1].split(',').map(c => c.trim());
                        // Öğrencinin sınıfı bu listede var mı?
                        isMatch = classList.includes(studentClass);
                    }
                }

                if (!isMatch) return;

                let cellDay = '', cellTime = '';
                let row = cell.parentElement;
                while (row) {
                    const dayCell = row.querySelector('.day-header, td[colspan]');
                    if (dayCell && dayCell.textContent.trim()) {
                        cellDay = dayCell.textContent.trim();
                        break;
                    }
                    row = row.previousElementSibling;
                }

                const timeCell = cell.parentElement.querySelector('td:first-child');
                if (timeCell) cellTime = timeCell.textContent.trim();

                const cleanCellDay = extractDayName(cellDay).toLocaleUpperCase('tr');
                if (cleanCellDay === cleanTargetDay && cellTime === targetTime) {
                    conflicts.push({ cell, studentName, day: targetDay, time: targetTime });
                }
            });

            return conflicts;
        }

        // ═══════════════════════════════════════════════════════════
        // 🎨 AYKIRI SWAP İHLALLERİNİ PANELDE GÖSTER (YUMUŞAK KIRMIZI)
        // ═══════════════════════════════════════════════════════════
        function displayAykiriSwapViolations() {
            // sessionStorage'dan aykırı swap ihlallerini al
            let violations = [];
            try {
                const stored = sessionStorage.getItem('aykiriSwapViolations');
                if (stored) {
                    violations = JSON.parse(stored);
                }
            } catch (e) {
                console.error('Aykırı swap ihlalleri okunamadı:', e);
                return;
            }

            if (violations.length === 0) return;

            // ✅ YENİ: Violation'ları (day + time) kombinasyonuna göre grupla (DEDUPLICATE)
            const groupedViolations = {};
            violations.forEach(violation => {
                const cleanDay = extractDayName(violation.day).toLocaleUpperCase('tr');
                const key = `${cleanDay}|${violation.time}`;

                if (!groupedViolations[key]) {
                    groupedViolations[key] = {
                        day: violation.day,
                        time: violation.time,
                        students: new Set(),
                        colors: new Set(),
                        conflictingStudents: new Set()
                    };
                }

                // Öğrencileri ekle
                violation.swappedStudents.forEach(s => groupedViolations[key].students.add(s));
                violation.conflictingStudents.forEach(s => groupedViolations[key].conflictingStudents.add(s));
                groupedViolations[key].colors.add(violation.borderColor);
            });

            // Grupları array'e çevir
            const groupedArray = Object.values(groupedViolations).map(group => ({
                day: group.day,
                time: group.time,
                students: Array.from(group.students),
                conflictingStudents: Array.from(group.conflictingStudents),
                colors: Array.from(group.colors)
            }));

            // Panelde gösterim için section oluştur veya mevcut olanı bul
            let aykiriSection = document.getElementById('aykiriSwapSection');
            if (!aykiriSection) {
                // Section yoksa oluştur ve conflictList'in başına ekle
                const conflictList = document.getElementById('conflictList');
                if (!conflictList) return;

                aykiriSection = document.createElement('div');
                aykiriSection.id = 'aykiriSwapSection';
                aykiriSection.style.marginBottom = '30px';
                conflictList.insertBefore(aykiriSection, conflictList.firstChild);
            }

            // Kartları oluştur
            let html = `
                <div style="background: linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%); padding: 20px; border-radius: 12px; border-left: 5px solid #ef4444; box-shadow: 0 4px 6px rgba(239, 68, 68, 0.1);">
                    <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 15px;">
                        <div style="background: #ef4444; color: white; width: 40px; height: 40px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 1.3em;">
                            ⚠️
                        </div>
                        <div>
                            <div style="color: #dc2626; font-weight: bold; font-size: 1.2em;">Aykırı Swap İhlalleri</div>
                            <div style="color: #991b1b; font-size: 0.9em;">Manuel onaylanarak yapılmış kuraldışı yer değiştirmeler</div>
                        </div>
                        <div style="margin-left: auto; background: #dc2626; color: white; padding: 8px 16px; border-radius: 20px; font-weight: bold;">
                            ${groupedArray.length}
                        </div>
                    </div>
            `;

            groupedArray.forEach((group, index) => {
                // Birden fazla öğrenci varsa sayısını göster
                const studentCount = group.students.length;
                const studentsText = studentCount > 1
                    ? `${group.students.join(', ')} (${studentCount} öğrenci)`
                    : group.students[0];

                const conflictingText = group.conflictingStudents.length > 0
                    ? group.conflictingStudents.join(', ')
                    : 'Yok';

                // Birden fazla renk varsa hepsini göster
                const colorBoxes = group.colors.map(color =>
                    `<div style="width: 24px; height: 24px; border-radius: 4px; border: 3px solid ${color};" title="Tabloda bu renkle işaretli"></div>`
                ).join('');

                html += `
                    <div style="background: white; padding: 15px; border-radius: 8px; margin-bottom: ${index < groupedArray.length - 1 ? '12px' : '0'}; border: 2px solid #fecaca;">
                        <div style="display: flex; justify-content: space-between; align-items: start; margin-bottom: 10px;">
                            <div style="flex: 1;">
                                <div style="color: #991b1b; font-weight: 600; font-size: 1.05em; margin-bottom: 8px;">
                                    <i class="fas fa-exchange-alt"></i> ${studentsText}
                                </div>
                                <div style="color: #6b7280; font-size: 0.9em; display: flex; gap: 15px; flex-wrap: wrap;">
                                    <span><i class="fas fa-calendar"></i> ${group.day}</span>
                                    <span><i class="fas fa-clock"></i> ${group.time}</span>
                                </div>
                            </div>
                            <div style="display: flex; gap: 4px;">
                                ${colorBoxes}
                            </div>
                        </div>

                        ${group.conflictingStudents.length > 0 ? `
                        <div style="background: #fff5f5; padding: 10px; border-radius: 6px; border-left: 3px solid #ef4444;">
                            <div style="color: #dc2626; font-weight: 600; font-size: 0.9em; margin-bottom: 5px;">
                                <i class="fas fa-user-times"></i> Çakışan Öğrenciler:
                            </div>
                            <div style="color: #991b1b; font-size: 0.9em;">
                                ${conflictingText}
                            </div>
                            <div style="color: #b91c1c; font-size: 0.85em; margin-top: 5px; font-style: italic;">
                                Bu öğrencilerin aynı saatte başka dersleri var
                            </div>
                        </div>
                        ` : ''}
                    </div>
                `;
            });

            html += `</div>`;
            aykiriSection.innerHTML = html;

            // Badge'i güncelle (toplam ihlal sayısına aykırı swap'leri ekle)
            const badge = document.getElementById('conflictBadge');
            if (badge) {
                const currentCount = parseInt(badge.textContent) || 0;
                const newCount = currentCount + groupedArray.length;
                updateConflictBadge(newCount);
            }
        }

        async function confirmSwap() {
            console.log('🔵 confirmSwap ÇAĞRILDI');
            if (!swapPendingData || !draggedCell) return;

            const targetCell = swapPendingData.targetCell;
            const draggedContent = draggedCell.textContent.trim();
            const targetContent = targetCell.textContent.trim();
            console.log('📌 draggedContent:', draggedContent);
            console.log('📌 targetContent:', targetContent);

            // HEDEF HUCRE SINIF DERSI MI KONTROL ET
            const targetIsClassLesson = /\(\d+\s/.test(targetContent);

            let draggedStudentNames = draggedData.studentNames || [draggedContent];
            let targetStudentNames = [];

            // HEDEF SINIF DERSIYSE, O SLOTTAKI OGRENCILERI BUL
            if (targetIsClassLesson) {
                const cleanTargetDay = extractDayName(swapPendingData.targetDay);
                const weekData = globalScheduleData.weeks[currentWeekView - 1];

                // Hedef hucrenin ogretmenini bul
                const targetCellIndex = Array.from(targetCell.parentElement.children).indexOf(targetCell);
                const headerRow = targetCell.closest('table').querySelector('thead tr');
                const targetTeacherHeader = headerRow ? headerRow.children[targetCellIndex] : null;
                const targetTeacherText = targetTeacherHeader ? targetTeacherHeader.textContent : '';

                // Ogretmen adini parantezden ayikla: "Matematik (DILEK COKUN)" -> "DILEK COKUN"
                const teacherMatch = targetTeacherText.match(/\(([^)]+)\)/);
                const targetTeacherName = teacherMatch ? teacherMatch[1].trim().toLocaleUpperCase('tr') : '';

                // Hedef slottaki tum ogrencileri topla
                weekData.forEach(lesson => {
                    const lessonDayUpper = lesson.day ? lesson.day.toLocaleUpperCase('tr') : '';
                    const cleanTargetDayUpper = cleanTargetDay.toLocaleUpperCase('tr');
                    const lessonTeacherUpper = lesson.teacher_name ? lesson.teacher_name.toLocaleUpperCase('tr') : '';

                    if (lessonDayUpper === cleanTargetDayUpper &&
                        lesson.time === swapPendingData.targetTime &&
                        lessonTeacherUpper === targetTeacherName) {
                        targetStudentNames.push(lesson.student_name);
                    }
                });
            } else {
                // Hedef bireysel ders
                targetStudentNames = [targetContent];
            }

            const errors = [];

            // ═══════════════════════════════════════════════════════════
            // 🆕 1️⃣ ÖNCE DOM-BASED ÇAKIŞMA KONTROLÜ (globalScheduleData DEĞİL!)
            // ═══════════════════════════════════════════════════════════
            console.log('🔍 DOM-based çakışma kontrolü başlıyor...');

            // Öğrencilerin sınıf bilgilerini al
            const studentClassMap = {};
            if (globalScheduleData && globalScheduleData.weeks) {
                const weekData = globalScheduleData.weeks[currentWeekView - 1];
                if (weekData) {
                    weekData.forEach(lesson => {
                        if (lesson.student_name && lesson.student_class) {
                            studentClassMap[lesson.student_name.toLocaleUpperCase('tr')] = lesson.student_class;
                        }
                    });
                }
            }

            // KAYNAK öğrencilerin yeni yerinde (hedef slot) çakışma var mı?
            for (const studentName of draggedStudentNames) {
                const studentClass = studentClassMap[studentName.toLocaleUpperCase('tr')];
                console.log(`🔎 DOM kontrol - KAYNAK: ${studentName}, sınıf: ${studentClass}`);

                const conflicts = findStudentConflictsAtTime(
                    studentName,
                    swapPendingData.targetDay,
                    swapPendingData.targetTime,
                    targetCell,  // Bu slotu hariç tut
                    studentClass
                );

                console.log(`🔎 ${studentName} için DOM'da çakışma: ${conflicts.length}`);

                if (conflicts.length > 0) {
                    errors.push(`${studentName}: Çakışma - ${studentName} zaten ${swapPendingData.targetDay} ${swapPendingData.targetTime}'de başka bir derste!`);
                }
            }

            // HEDEF öğrencilerin yeni yerinde (kaynak slot) çakışma var mı?
            for (const studentName of targetStudentNames) {
                const studentClass = studentClassMap[studentName.toLocaleUpperCase('tr')];
                console.log(`🔎 DOM kontrol - HEDEF: ${studentName}, sınıf: ${studentClass}`);

                const conflicts = findStudentConflictsAtTime(
                    studentName,
                    draggedData.day,
                    draggedData.time,
                    draggedCell,  // Bu slotu hariç tut
                    studentClass
                );

                console.log(`🔎 ${studentName} için DOM'da çakışma: ${conflicts.length}`);

                if (conflicts.length > 0) {
                    errors.push(`${studentName}: Çakışma - ${studentName} zaten ${draggedData.day} ${draggedData.time}'de başka bir derste!`);
                }
            }

            console.log('🔍 DOM-based çakışma kontrolü bitti. Bulunan hatalar:', errors.length);

            // ═══════════════════════════════════════════════════════════
            // 2️⃣ SONRA validateMove İLE KISITLAMA KONTROLÜ
            // ═══════════════════════════════════════════════════════════

            // KAYNAK TARAFTAKI TUM OGRENCILER ICIN KONTROL
            for (const studentName of draggedStudentNames) {
                const validation = await validateMove(
                    studentName,
                    draggedData.day,
                    draggedData.time,
                    swapPendingData.targetDay,
                    swapPendingData.targetTime,
                    draggedCell,
                    targetCell
                );

                console.log(`🔍 validateMove (KAYNAK ${studentName}):`, validation);

                if (!validation.valid) {
                    errors.push(`${studentName}: ${validation.message}`);
                }
            }

            // HEDEF TARAFTAKI TUM OGRENCILER ICIN KONTROL
            for (const studentName of targetStudentNames) {
                const validation = await validateMove(
                    studentName,
                    swapPendingData.targetDay,
                    swapPendingData.targetTime,
                    draggedData.day,
                    draggedData.time,
                    targetCell,
                    draggedCell
                );

                console.log(`🔍 validateMove (HEDEF ${studentName}):`, validation);

                if (!validation.valid) {
                    errors.push(`${studentName}: ${validation.message}`);
                }
            }

            console.log('⚠️ TOPLAM HATA SAYISI:', errors.length);
            console.log('⚠️ HATALAR:', errors);

            // HATA VARSA UYARI MODAL GÖSTER (DEVAM ET / İPTAL ET)
            if (errors.length > 0) {
                console.log('🚨 AYKIRI SWAP TESPİT EDİLDİ - Modal gösteriliyor');
                let errorMsg = '⚠️ KURALLARA AYKIRI SWAP!\\n\\nAşağıdaki sorunlar tespit edildi:\\n\\n';
                errors.forEach(err => {
                    errorMsg += err + '\\n\\n';
                });
                errorMsg += '⚠️ DİKKAT: Devam ederseniz bu swap kurallara aykırı olarak işaretlenecek!\\n\\nDevam etmek istiyor musunuz?';

                // ✅ ÖNEMLİ: Verileri closure ile yakala (swapPendingData sonra null olacak)
                const savedDraggedCell = draggedCell;
                const savedTargetCell = targetCell;
                const savedDraggedContent = draggedContent;
                const savedTargetContent = targetContent;
                const savedTargetStudentNames = targetStudentNames;
                const savedTargetIsClassLesson = targetIsClassLesson;
                const savedDraggedData = {...draggedData};
                savedDraggedData.studentNames = draggedStudentNames; // ← EKLE!
                const savedSwapPendingData = {...swapPendingData};

                // ✅ İLK MODALI KAPAT
                document.getElementById('swapConfirmModal').style.display = 'none';

                // Uyarı modalını göster
                showWarningModal(errorMsg, () => {
                    console.log('✅ DEVAM ET butonuna basıldı - performSwapWithWarning çağrılıyor');
                    // DEVAM ET - Aykırı swap yap ve renklendir
                    performSwapWithWarning(savedDraggedCell, savedTargetCell, savedDraggedContent, savedTargetContent, savedTargetStudentNames, savedTargetIsClassLesson, savedDraggedData, savedSwapPendingData);
                }, () => {
                    console.log('❌ İPTAL ET butonuna basıldı');
                    // İPTAL ET - Tüm değişkenleri temizle
                    draggedCell = null;
                    draggedData = null;
                    swapPendingData = null;
                });
                return;
            }

            console.log('✅ HATA YOK - Normal swap yapılıyor');
            // HATA YOKSA NORMAL SWAP YAP
            performNormalSwap(draggedCell, targetCell, draggedContent, targetContent, targetStudentNames, targetIsClassLesson);
        }

        // ═══════════════════════════════════════════════════════════
        // 🔄 NORMAL SWAP (KURALLARA UYGUN - RENKLENDİRME YOK)
        // ═══════════════════════════════════════════════════════════
        // ═══════════════════════════════════════════════════════════
        // 🧹 SWAP YAPAN ÖĞRENCİLERİN AYKIRI SWAP KAYITLARINI TEMİZLE
        // ═══════════════════════════════════════════════════════════
        function clearAykiriSwapForStudents(slotInfoList) {
            // ✅ YENİ PARAMETRE: Slot bilgileri listesi [{studentName, day, time, teacherName}, ...]
            // Geriye dönük uyumluluk: Eğer eski formatta (sadece string array) gelirse çevir
            if (!slotInfoList || slotInfoList.length === 0) return;

            // Eski format kontrolü (string array mı?)
            const isOldFormat = typeof slotInfoList[0] === 'string';
            let slotsToRemove = [];

            if (isOldFormat) {
                // Eski format: sadece öğrenci isimleri
                // Bu durumda tüm violation'ları temizle (eski davranış)
                slotsToRemove = slotInfoList.map(name => ({
                    studentName: name,
                    day: null,
                    time: null,
                    teacherName: null,
                    removeAll: true // Tüm kayıtları sil flag'i
                }));
            } else {
                // Yeni format: slot bilgileri
                slotsToRemove = slotInfoList;
            }

            // SessionStorage'dan aykırı swap kayıtlarını oku
            let violations = [];
            try {
                const stored = sessionStorage.getItem('aykiriSwapViolations');
                if (stored) {
                    violations = JSON.parse(stored);
                }
            } catch (e) {
                return;
            }

            if (violations.length === 0) return;

            // ✅ YENİ: Slot bazlı violation'ları filtrele
            const remainingViolations = [];

            violations.forEach(violation => {
                let shouldRemove = false;

                for (const slot of slotsToRemove) {
                    const studentUpper = slot.studentName.toLocaleUpperCase('tr');

                    // Eski format: Tüm öğrenciye ait violation'ları sil
                    if (slot.removeAll) {
                        const hasStudent = violation.swappedStudents.some(s =>
                            s.toLocaleUpperCase('tr') === studentUpper
                        );
                        if (hasStudent) {
                            shouldRemove = true;
                            break;
                        }
                    }
                    // Yeni format: slotIdentifier ile karşılaştır
                    else if (violation.slotIdentifier) {
                        const id = violation.slotIdentifier;
                        const cleanSlotDay = extractDayName(slot.day).toLocaleUpperCase('tr');
                        const cleanViolationDay = extractDayName(id.day).toLocaleUpperCase('tr');

                        if (id.studentName.toLocaleUpperCase('tr') === studentUpper &&
                            cleanViolationDay === cleanSlotDay &&
                            id.time === slot.time &&
                            id.teacherName.toLocaleUpperCase('tr') === slot.teacherName.toLocaleUpperCase('tr')) {
                            shouldRemove = true;
                            break;
                        }
                    }
                    // Eski violation kaydı (slotIdentifier yok): Öğrenci+gün+saat eşleşmesine bak
                    else {
                        const hasStudent = violation.swappedStudents.some(s =>
                            s.toLocaleUpperCase('tr') === studentUpper
                        );
                        const cleanSlotDay = extractDayName(slot.day).toLocaleUpperCase('tr');
                        const cleanViolationDay = extractDayName(violation.day).toLocaleUpperCase('tr');

                        if (hasStudent &&
                            cleanViolationDay === cleanSlotDay &&
                            violation.time === slot.time) {
                            shouldRemove = true;
                            break;
                        }
                    }
                }

                if (!shouldRemove) {
                    remainingViolations.push(violation);
                }
            });

            // 🔄 TÜM BORDER'LARI TEMİZLE
            const table = document.getElementById('weeklyPrintTable');
            if (table) {
                const cells = table.querySelectorAll('tbody td');
                cells.forEach(cell => {
                    if (cell.style.border) {
                        cell.style.border = '';
                        cell.style.boxSizing = '';
                    }
                });
            }

            // 🎨 KALAN VIOLATION'LAR İÇİN BORDER'LARI YENİDEN UYGULA
            if (remainingViolations.length > 0 && table) {
                remainingViolations.forEach(violation => {
                    const studentName = violation.swappedStudents[0];
                    const day = violation.day;
                    const time = violation.time;
                    const color = violation.borderColor;

                    // ✅ YENİ: globalScheduleData'dan öğrencinin bu gün/saatteki GERÇEK slot'larını bul
                    const studentActualSlots = [];
                    if (globalScheduleData && globalScheduleData.weeks) {
                        const weekData = globalScheduleData.weeks[currentWeekView - 1];
                        if (weekData) {
                            const cleanTargetDay = extractDayName(day).toLocaleUpperCase('tr');

                            weekData.forEach(lesson => {
                                const lessonDayClean = extractDayName(lesson.day).toLocaleUpperCase('tr');

                                if (lesson.student_name &&
                                    lesson.student_name.toLocaleUpperCase('tr') === studentName.toLocaleUpperCase('tr') &&
                                    lessonDayClean === cleanTargetDay &&
                                    lesson.time === time) {
                                    // Bu öğrenci bu gün/saatte bu öğretmenle ders yapıyor
                                    studentActualSlots.push({
                                        teacherName: lesson.teacher_name,
                                        studentClass: lesson.student_class
                                    });
                                }
                            });
                        }
                    }

                    // ✅ Sadece öğrencinin GERÇEKTEN olduğu slot'ları renklendir
                    const cells = table.querySelectorAll('tbody td:not(:first-child)');
                    cells.forEach(cell => {
                        const cellText = cell.textContent.trim();
                        if (!cellText) return;

                        // Gün ve saat bilgisini bul
                        let cellDay = '', cellTime = '';
                        let row = cell.parentElement;
                        while (row) {
                            const dayCell = row.querySelector('.day-header, td[colspan]');
                            if (dayCell && dayCell.textContent.trim()) {
                                cellDay = dayCell.textContent.trim();
                                break;
                            }
                            row = row.previousElementSibling;
                        }

                        const timeCell = cell.parentElement.querySelector('td:first-child');
                        if (timeCell) cellTime = timeCell.textContent.trim();

                        const cleanCellDay = extractDayName(cellDay).toLocaleUpperCase('tr');
                        const cleanTargetDay = extractDayName(day).toLocaleUpperCase('tr');

                        // Gün ve saat eşleşiyor mu?
                        if (cleanCellDay === cleanTargetDay && cellTime === time) {
                            // ✅ Bu hücrenin öğretmenini bul
                            const cellIndex = Array.from(cell.parentElement.children).indexOf(cell);
                            const headerRow = table.querySelector('thead tr');
                            const teacherHeader = headerRow ? headerRow.children[cellIndex] : null;
                            const teacherText = teacherHeader ? teacherHeader.textContent : '';
                            const teacherMatch = teacherText.match(/\(([^)]+)\)/);
                            const cellTeacherName = teacherMatch ? teacherMatch[1].trim() : '';

                            // ✅ KONTROL: Bu öğrenci bu öğretmenle gerçekten ders yapıyor mu?
                            const isStudentActuallyHere = studentActualSlots.some(slot =>
                                slot.teacherName.toLocaleUpperCase('tr') === cellTeacherName.toLocaleUpperCase('tr')
                            );

                            if (isStudentActuallyHere) {
                                // ✅ EVET, bu öğrenci gerçekten bu slot'ta → Renklendir
                                cell.style.border = `4px solid ${color}`;
                                cell.style.boxSizing = 'border-box';
                            }
                        }
                    });
                });
            }

            // Kalan violation'ları sessionStorage'a geri yaz
            try {
                sessionStorage.setItem('aykiriSwapViolations', JSON.stringify(remainingViolations));
                // İhlal panelini güncelle
                checkConflictsInBackground();
            } catch (e) {
                console.error('SessionStorage güncellenemedi:', e);
            }
        }

        // ═══════════════════════════════════════════════════════════
        // ✅ NORMAL SWAP (KURALLARA UYGUN)
        // ═══════════════════════════════════════════════════════════
        function performNormalSwap(draggedCell, targetCell, draggedContent, targetContent, targetStudentNames, targetIsClassLesson) {
            // swapPendingData'ya hedef ogrenci listesini ekle
            swapPendingData.targetStudentNames = targetStudentNames;
            swapPendingData.targetIsClassLesson = targetIsClassLesson;

            // ✅ 1️⃣ İLK ÖNCE: Hedef öğrencilerin öğretmenlerini al (globalScheduleData henüz ESKİ!)
            const slotsToRemove = [];

            // Kaynak slot bilgileri (draggedData)
            const draggedStudentNames = draggedData.studentNames || [draggedData.studentName];
            draggedStudentNames.forEach(studentName => {
                slotsToRemove.push({
                    studentName: studentName,
                    day: draggedData.day,
                    time: draggedData.time,
                    teacherName: draggedData.teacherName
                });
            });

            // Hedef slot bilgileri (swapPendingData)
            // ✅ ÖNEMLİ: Hedef öğrencilerin öğretmenlerini ŞİMDİ al!
            // updateGlobalScheduleDataAfterSwap() çağrılınca değişecek!
            if (targetStudentNames && targetStudentNames.length > 0) {
                targetStudentNames.forEach(studentName => {
                    // globalScheduleData'dan bu öğrencinin bu gün/saatteki öğretmenini bul
                    let studentTeacherName = swapPendingData.targetTeacherName || '';

                    if (globalScheduleData && globalScheduleData.weeks) {
                        const weekData = globalScheduleData.weeks[draggedData.week - 1];
                        if (weekData) {
                            const cleanTargetDay = extractDayName(swapPendingData.targetDay).toLocaleUpperCase('tr');

                            for (const lesson of weekData) {
                                const lessonDayClean = extractDayName(lesson.day).toLocaleUpperCase('tr');

                                if (lesson.student_name === studentName &&
                                    lessonDayClean === cleanTargetDay &&
                                    lesson.time === swapPendingData.targetTime) {
                                    // Bu öğrencinin bu slottaki öğretmenini bulduk
                                    studentTeacherName = lesson.teacher_name;
                                    break;
                                }
                            }
                        }
                    }

                    slotsToRemove.push({
                        studentName: studentName,
                        day: swapPendingData.targetDay,
                        time: swapPendingData.targetTime,
                        teacherName: studentTeacherName
                    });
                });
            }

            // ✅ 2️⃣ SONRA: globalScheduleData'yı güncelle
            updateGlobalScheduleDataAfterSwap(draggedData, swapPendingData);

            // ✅ 3️⃣ EN SONRA: Aykırı swap kayıtlarını temizle (doğru öğretmen bilgileriyle)
            // Border yeniden uygularken globalScheduleData güncel olacak
            clearAykiriSwapForStudents(slotsToRemove);

            // YER DEĞİŞTİR
            draggedCell.textContent = targetContent;
            targetCell.textContent = draggedContent;

            // Sınıfları güncelle
            if (targetContent.trim()) {
                draggedCell.classList.add('draggable-cell');
                draggedCell.classList.remove('empty-slot');
                draggedCell.setAttribute('draggable', true);
            } else {
                draggedCell.classList.remove('draggable-cell');
                draggedCell.classList.add('empty-slot');
                draggedCell.setAttribute('draggable', false);
            }

            if (draggedContent.trim()) {
                targetCell.classList.add('draggable-cell');
                targetCell.classList.remove('empty-slot');
                targetCell.setAttribute('draggable', true);
            } else {
                targetCell.classList.remove('draggable-cell');
                targetCell.classList.add('empty-slot');
                targetCell.setAttribute('draggable', false);
            }

            // Backend'e değişikliği gönder
            saveSwapToBackend(draggedData, swapPendingData);

            // Popup'ı kapat
            cancelSwap();

            // Başarı mesajı
            showSuccessModal('Dersler başarıyla yer değiştirdi!');
        }

        // ═══════════════════════════════════════════════════════════
        // ⚠️ AYKIRI SWAP (KURALLARA AYKIRI - 4PX RENKLİ BORDER)
        // ═══════════════════════════════════════════════════════════
        function performSwapWithWarning(draggedCell, targetCell, draggedContent, targetContent, targetStudentNames, targetIsClassLesson, savedDraggedData, savedSwapPendingData) {
            console.log('🟡 performSwapWithWarning ÇAĞRILDI');
            console.log('📌 draggedContent:', draggedContent);
            console.log('📌 targetContent:', targetContent);
            console.log('📌 targetStudentNames:', targetStudentNames);

            // Kaydedilen swap verilerini güncelle
            savedSwapPendingData.targetStudentNames = targetStudentNames;
            savedSwapPendingData.targetIsClassLesson = targetIsClassLesson;

            // ═══════════════════════════════════════════════════════════
            // 📋 ÖĞRENCİLERİN SINIF BİLGİLERİNİ AL (SWAP YAPMADAN ÖNCE!)
            // ═══════════════════════════════════════════════════════════
            const studentClassMap = {}; // {studentName: className}

            // globalScheduleData'dan tüm öğrencilerin sınıf bilgisini al
            if (globalScheduleData && globalScheduleData.weeks) {
                const weekData = globalScheduleData.weeks[currentWeekView - 1];
                if (weekData) {
                    weekData.forEach(lesson => {
                        if (lesson.student_name && lesson.student_class) {
                            studentClassMap[lesson.student_name.toLocaleUpperCase('tr')] = lesson.student_class;
                        }
                    });
                }
            }

            console.log('📚 studentClassMap:', studentClassMap);

            // ═══════════════════════════════════════════════════════════
            // 🔍 ÖNCE ÇAKIŞMALARI TESPİT ET (SWAP YAPMADAN ÖNCE!)
            // ═══════════════════════════════════════════════════════════
            const conflictGroups = []; // [{studentName, conflicts: [], color}]
            const conflictingStudentNames = []; // Sadece çakışan öğrenci isimleri

            // 1️⃣ HEDEF ÖĞRENCİLERİN YENİ YERİNDEKİ ÇAKIŞMALARI
            if (targetStudentNames && targetStudentNames.length > 0) {
                targetStudentNames.forEach(name => {
                    const studentClass = studentClassMap[name.toLocaleUpperCase('tr')];
                    console.log(`🔎 HEDEF öğrenci: ${name}, sınıf: ${studentClass}`);
                    const conflicts = findStudentConflictsAtTime(
                        name,
                        savedDraggedData.day,           // ← YENİ YERİ (kaynak slot)
                        savedDraggedData.time,          // ← YENİ YERİ
                        draggedCell,                    // ← Bu slotu hariç tut
                        studentClass                    // ← Sınıf bilgisi
                    );

                    console.log(`🔎 ${name} için çakışma sayısı:`, conflicts.length);

                    if (conflicts.length > 0) {
                        const color = generateRandomBorderColor();
                        conflictGroups.push({
                            studentName: name,
                            swappedCell: draggedCell,  // Bu öğrencinin yeni yeri
                            conflicts: conflicts,
                            color: color
                        });
                        conflictingStudentNames.push(name);
                    }
                });
            }

            // 2️⃣ KAYNAK ÖĞRENCİLERİN YENİ YERİNDEKİ ÇAKIŞMALARI
            const draggedStudentNames = savedDraggedData.studentNames || [savedDraggedData.studentName];
            draggedStudentNames.forEach(name => {
                const studentClass = studentClassMap[name.toLocaleUpperCase('tr')];
                console.log(`🔎 KAYNAK öğrenci: ${name}, sınıf: ${studentClass}`);
                const conflicts = findStudentConflictsAtTime(
                    name,
                    savedSwapPendingData.targetDay,    // ← YENİ YERİ (hedef slot)
                    savedSwapPendingData.targetTime,   // ← YENİ YERİ
                    targetCell,                        // ← Bu slotu hariç tut
                    studentClass                       // ← Sınıf bilgisi
                );

                console.log(`🔎 ${name} için çakışma sayısı:`, conflicts.length);

                if (conflicts.length > 0) {
                    const color = generateRandomBorderColor();
                    conflictGroups.push({
                        studentName: name,
                        swappedCell: targetCell,  // Bu öğrencinin yeni yeri
                        conflicts: conflicts,
                        color: color
                    });
                    conflictingStudentNames.push(name);
                }
            });

            console.log('📊 TOPLAM conflictGroups sayısı:', conflictGroups.length);

            // ═══════════════════════════════════════════════════════════
            // 🔄 GLOBAL SCHEDULE DATA GÜNCELLE VE ESKİ VIOLATION'LARI TEMİZLE
            // ═══════════════════════════════════════════════════════════
            // ✅ 1️⃣ İLK ÖNCE: ESKİ VIOLATION SİLMEK İÇİN SLOT BİLGİLERİNİ TOPLA
            // (globalScheduleData güncellenmeden ÖNCE öğretmen bilgisini al!)
            const slotsToRemove = [];

            // Kaynak slot bilgileri (draggedData)
            draggedStudentNames.forEach(studentName => {
                slotsToRemove.push({
                    studentName: studentName,
                    day: savedDraggedData.day,
                    time: savedDraggedData.time,
                    teacherName: savedDraggedData.teacherName
                });
            });

            // Hedef slot bilgileri (swapPendingData) - ÖĞRETMENİ globalScheduleData'dan al!
            if (targetStudentNames && targetStudentNames.length > 0) {
                const weekData = globalScheduleData?.weeks?.[currentWeekView - 1];
                const cleanTargetDay = savedSwapPendingData.targetDay.toLocaleUpperCase('tr');

                targetStudentNames.forEach(studentName => {
                    let studentTeacherName = savedSwapPendingData.targetTeacherName || '';

                    // globalScheduleData'da bu öğrencinin gerçek öğretmenini bul
                    if (weekData) {
                        for (const lesson of weekData) {
                            const lessonDayClean = lesson.day.toLocaleUpperCase('tr');
                            if (lesson.student_name === studentName &&
                                lessonDayClean === cleanTargetDay &&
                                lesson.time === savedSwapPendingData.targetTime) {
                                studentTeacherName = lesson.teacher_name;
                                console.log(`✅ HEDEF öğrenci ${studentName} için öğretmen bulundu: ${studentTeacherName}`);
                                break;
                            }
                        }
                    }

                    slotsToRemove.push({
                        studentName: studentName,
                        day: savedSwapPendingData.targetDay,
                        time: savedSwapPendingData.targetTime,
                        teacherName: studentTeacherName
                    });
                });
            }

            // ✅ 2️⃣ SONRA: globalScheduleData'yı güncelle
            updateGlobalScheduleDataAfterSwap(savedDraggedData, savedSwapPendingData);

            // ✅ 3️⃣ SON OLARAK: ESKİ VIOLATION KAYITLARINI TEMİZLE
            clearAykiriSwapForStudents(slotsToRemove);

            // ═══════════════════════════════════════════════════════════
            // 🔄 ŞİMDİ YER DEĞİŞTİR (globalScheduleData güncel, eski violation'lar temizlendi)
            // ═══════════════════════════════════════════════════════════
            draggedCell.textContent = targetContent;
            targetCell.textContent = draggedContent;

            // Sınıfları güncelle
            if (targetContent.trim()) {
                draggedCell.classList.add('draggable-cell');
                draggedCell.classList.remove('empty-slot');
                draggedCell.setAttribute('draggable', true);
            } else {
                draggedCell.classList.remove('draggable-cell');
                draggedCell.classList.add('empty-slot');
                draggedCell.setAttribute('draggable', false);
            }

            if (draggedContent.trim()) {
                targetCell.classList.add('draggable-cell');
                targetCell.classList.remove('empty-slot');
                targetCell.setAttribute('draggable', true);
            } else {
                targetCell.classList.remove('draggable-cell');
                targetCell.classList.add('empty-slot');
                targetCell.setAttribute('draggable', false);
            }

            // ═══════════════════════════════════════════════════════════
            // 🎨 ÇAKIŞAN SLOTLARI RENKLENDİR (Her grup farklı renk)
            // ═══════════════════════════════════════════════════════════
            let totalMarkedSlots = 0;

            conflictGroups.forEach(group => {
                // Öğrencinin yeni yerini işaretle (4px border)
                group.swappedCell.style.border = `4px solid ${group.color}`;
                group.swappedCell.style.boxSizing = 'border-box';

                // Çakıştığı slotları işaretle (4px border - aynı renk!)
                group.conflicts.forEach(conflict => {
                    conflict.cell.style.border = `4px solid ${group.color}`;
                    conflict.cell.style.boxSizing = 'border-box';
                });

                totalMarkedSlots += 1 + group.conflicts.length; // Öğrencinin slotu + çakışan slotlar
            });

            // ═══════════════════════════════════════════════════════════
            // 💾 İHLAL VERİLERİNİ KAYDET (sessionStorage)
            // ═══════════════════════════════════════════════════════════
            // Her çakışma grubu için ayrı ihlal kaydı
            conflictGroups.forEach(group => {
                // Öğrencinin sınıf bilgisini al
                const studentClass = studentClassMap[group.studentName.toLocaleUpperCase('tr')];

                // Cell'in hangi öğretmene ait olduğunu bul
                const cellIndex = Array.from(group.swappedCell.parentElement.children).indexOf(group.swappedCell);
                const table = document.getElementById('weeklyPrintTable');
                const headerRow = table ? table.querySelector('thead tr') : null;
                const teacherHeader = headerRow ? headerRow.children[cellIndex] : null;
                const teacherText = teacherHeader ? teacherHeader.textContent : '';

                // Öğretmen adını parse et: "Matematik (DILEK COKUN)" -> "DILEK COKUN"
                const teacherMatch = teacherText.match(/\(([^)]+)\)/);
                const teacherName = teacherMatch ? teacherMatch[1].trim() : '';

                // ✅ Slot'u benzersiz tanımlamak için identifier oluştur
                const slotDay = group.swappedCell === draggedCell ? savedDraggedData.day : savedSwapPendingData.targetDay;
                const slotTime = group.swappedCell === draggedCell ? savedDraggedData.time : savedSwapPendingData.targetTime;

                saveViolationToPanel({
                    type: 'aykiri_swap',
                    week: currentWeekView,
                    swappedStudents: [group.studentName],
                    conflictingStudents: [group.studentName],
                    studentClass: studentClass,  // ✅ Sınıf bilgisi eklendi
                    teacherName: teacherName,    // ✅ Öğretmen bilgisi eklendi
                    day: slotDay,
                    time: slotTime,
                    borderColor: group.color,
                    // ✅ YENİ: Slot identifier - Her slot'u benzersiz tanımlar
                    slotIdentifier: {
                        day: slotDay,
                        time: slotTime,
                        studentName: group.studentName,
                        teacherName: teacherName
                    },
                    timestamp: new Date().toISOString()
                });
            });

            // 🔄 SAYACI HEMEN GÜNCELLE
            checkConflictsInBackground();

            // Backend'e değişikliği gönder (kaydedilmiş verileri kullan)
            saveSwapToBackend(savedDraggedData, savedSwapPendingData);

            // Global değişkenleri temizle
            draggedCell = null;
            draggedData = null;
            swapPendingData = null;

            // Uyarı mesajı
            const totalStudents = draggedStudentNames.length + (targetStudentNames ? targetStudentNames.length : 0);
            showSuccessModal(`⚠️ Aykırı swap yapıldı!\n\n` +
                `• ${totalStudents} öğrenci yer değiştirdi\n` +
                `• ${conflictingStudentNames.length} öğrenci çakışma yaşadı\n` +
                `• ${totalMarkedSlots} slot renkli border ile işaretlendi`);
        }

        function cancelSwap() {
            document.getElementById('swapConfirmModal').style.display = 'none';
            swapPendingData = null;
            draggedCell = null;
            draggedData = null;
        }

        function showError(message) {
            document.getElementById('errorModalMessage').textContent = message;
            document.getElementById('errorModal').style.display = 'flex';
        }

        function closeErrorModal() {
            document.getElementById('errorModal').style.display = 'none';
        }

        function showSuccessModal(message) {
            document.getElementById('successModalMessage').textContent = message;
            document.getElementById('successModal').style.display = 'flex';
        }

        function closeSuccessModal() {
            document.getElementById('successModal').style.display = 'none';
        }

        // ═══════════════════════════════════════════════════════════
        // ⚠️ WARNING MODAL FONKSİYONLARI
        // ═══════════════════════════════════════════════════════════
        let warningModalCallbacks = {
            onConfirm: null,
            onCancel: null
        };

        function showWarningModal(message, onConfirm, onCancel) {
            // Callback'leri kaydet
            warningModalCallbacks.onConfirm = onConfirm;
            warningModalCallbacks.onCancel = onCancel;

            // Mesajı göster
            document.getElementById('warningModalMessage').textContent = message;
            document.getElementById('warningModal').style.display = 'flex';
        }

        function confirmWarning() {
            // Devam Et butonuna basıldı
            document.getElementById('warningModal').style.display = 'none';
            if (warningModalCallbacks.onConfirm) {
                warningModalCallbacks.onConfirm();
            }
        }

        function cancelWarning() {
            // İptal Et butonuna basıldı
            document.getElementById('warningModal').style.display = 'none';
            if (warningModalCallbacks.onCancel) {
                warningModalCallbacks.onCancel();
            }
        }

        async function performMove(targetCell) {
            if (!draggedCell) return;

            // Hedef hücrenin bilgilerini topla
            const row = targetCell.parentElement;
            let currentRow = targetCell.parentElement;
            let targetDay = '';
            let targetTime = '';

            while (currentRow) {
                const dayCell = currentRow.querySelector('.day-header, td[colspan]');
                if (dayCell && dayCell.textContent.trim()) {
                    targetDay = dayCell.textContent.trim().toLocaleUpperCase('tr');
                    break;
                }
                currentRow = currentRow.previousElementSibling;
            }

            const timeCell = targetCell.parentElement.querySelector('td:first-child');
            if (timeCell) {
                targetTime = timeCell.textContent.trim();
            }

            // SINIF DERSI KONTROLU - HER OGRENCI ICIN DOGRULA
            if (draggedData.isClassLesson) {
                const errors = [];

                for (const studentName of draggedData.studentNames) {
                    const validationResult = await validateMove(
                        studentName,
                        draggedData.day,
                        draggedData.time,
                        targetDay,
                        targetTime,
                        draggedCell,
                        targetCell
                    );

                    if (!validationResult.valid) {
                        errors.push(`${studentName}: ${validationResult.message}`);
                    }
                }

                if (errors.length > 0) {
                    showError(`SINIF DERSI TASINAMAZ!\\n\\nAsagidaki ogrenciler icin sorun var:\\n\\n${errors.join('\\n\\n')}`);
                    draggedCell = null;
                    draggedData = null;
                    return;
                }
            } else {
                // TEK OGRENCI KONTROLU
                const validationResult = await validateMove(
                    draggedData.studentName,
                    draggedData.day,
                    draggedData.time,
                    targetDay,
                    targetTime,
                    draggedCell,
                    targetCell
                );

                if (!validationResult.valid) {
                    showError(validationResult.message);
                    draggedCell = null;
                    draggedData = null;
                    return;
                }
            }

            // TAŞI
            targetCell.textContent = draggedCell.textContent;
            draggedCell.textContent = '';

            // Sınıfları güncelle
            targetCell.classList.add('draggable-cell');
            targetCell.classList.remove('empty-slot');
            targetCell.setAttribute('draggable', true);

            draggedCell.classList.remove('draggable-cell');
            draggedCell.classList.add('empty-slot');
            draggedCell.setAttribute('draggable', false);

            // Backend'e kaydet
            const targetData = {
                targetDay: targetDay,
                targetTime: targetTime
            };
            await saveSwapToBackend(draggedData, targetData);

            // Başarı mesajı
            showSuccessModal('✓ Ders başarıyla taşındı!');

            draggedCell = null;
            draggedData = null;
        }

        async function validateMove(studentName, fromDay, fromTime, toDay, toTime, skipCell = null, targetCellToCheck = null) {
            try {
                // ✅ HATA TOPLAMA SİSTEMİ - Birden fazla engeli tek modalda göster
                const errors = [];

                // ✅ GÜN ADLARINI TARİHTEN AYIR
                // "SALI - 28 EKİM 25" → "SALI"
                const cleanFromDay = extractDayName(fromDay);
                const cleanToDay = extractDayName(toDay);

                const [studentsRes, teachersRes] = await Promise.all([
                    fetch('/get_students'),
                    fetch('/get_teachers')
                ]);

                const studentsData = await studentsRes.json();
                const teachersData = await teachersRes.json();

                const student = studentsData.students.find(s =>
                    `${s.name} ${s.surname}` === studentName
                );

                if (!student) {
                    return {valid: false, message: 'Öğrenci bulunamadı!'};
                }

                const [toStartTime, toEndTime] = toTime.split('-');

                // ═══════════════════════════════════════════════════════════
                // 1️⃣ ÖĞRENCİ KISITLAMALARINI KONTROL ET
                // ═══════════════════════════════════════════════════════════
                if (student.restrictions && student.restrictions.length > 0) {
                    for (const restriction of student.restrictions) {
                        const days = restriction.days || (restriction.day ? [restriction.day] : []);
                        const normalizedDays = days.map(d => d.toLocaleUpperCase('tr'));
                        const normalizedToDay = cleanToDay.toLocaleUpperCase('tr');

                        if (!normalizedDays.includes(normalizedToDay)) continue;

                        if (restriction.type === 'custom') {
                            const weeks = restriction.weeks || [];
                            if (weeks.length > 0 && !weeks.includes(currentWeekView)) {
                                continue;
                            }
                        }

                        const restStart = restriction.start_time || '';
                        const restEnd = restriction.end_time || '';

                        if (restStart && restEnd) {
                            if (checkTimeOverlap(toStartTime, toEndTime, restStart, restEnd)) {
                                // ✅ HEMEN RETURN ETME, HATAYA EKLE
                                errors.push({
                                    type: 'restriction',
                                    icon: '🚫',
                                    title: 'Öğrenci Kısıtlaması',
                                    message: `${cleanToDay} ${toTime} kısıtlı saat!\nÖğrenci bu saatte ders alamaz (${restStart}-${restEnd})`
                                });
                                break; // Aynı hatayı tekrar eklememek için çık
                            }
                        }
                    }
                }

                // CAKISMA KONTROLU
                const table = document.getElementById('weeklyPrintTable');
                if (table) {
                    const cells = table.querySelectorAll('tbody td:not(:first-child)');

                    for (const cell of cells) {
                        const cellText = cell.textContent.trim();

                        // DIREKT ESLESME (bireysel ders)
                        let isMatch = cellText === studentName;

                        // SINIF DERSI ESLEMESI - hucrede sinif dersi varsa, o siniftaki ogrencileri kontrol et
                        if (!isMatch && /\(\d+\s/.test(cellText)) {
                            // Bu hucrede sinif dersi var, o slottaki ogrencileri bul
                            let cellDay = '';
                            let cellTime = '';
                            let cellIndex = -1;

                            // Hucrenin gun/saat/ogretmen bilgisini al
                            let row = cell.parentElement;
                            while (row) {
                                const dayCell = row.querySelector('.day-header, td[colspan]');
                                if (dayCell && dayCell.textContent.trim()) {
                                    cellDay = dayCell.textContent.trim().toLocaleUpperCase('tr');
                                    break;
                                }
                                row = row.previousElementSibling;
                            }

                            const timeCell = cell.parentElement.querySelector('td:first-child');
                            if (timeCell) {
                                cellTime = timeCell.textContent.trim();
                            }

                            cellIndex = Array.from(cell.parentElement.children).indexOf(cell);

                            // Bu slottaki ogrencileri bul
                            const cleanCellDay = extractDayName(cellDay);
                            const weekData = globalScheduleData.weeks[currentWeekView - 1];

                            if (weekData && cleanCellDay && cellTime) {
                                const headerRow = table.querySelector('thead tr');
                                const teacherHeader = headerRow ? headerRow.children[cellIndex] : null;
                                const teacherText = teacherHeader ? teacherHeader.textContent : '';

                                // Ogretmen adini parantezden ayikla
                                const teacherMatch = teacherText.match(/\(([^)]+)\)/);
                                const headerTeacherName = teacherMatch ? teacherMatch[1].trim().toLocaleUpperCase('tr') : '';

                                // Bu slottaki dersleri kontrol et
                                for (const lesson of weekData) {
                                    // GUN KARSILASTIRMASINI BUYUK HARFLE YAP (Turkce karakterler icin toLocaleUpperCase)
                                    const lessonDayUpper = lesson.day ? lesson.day.toLocaleUpperCase('tr') : '';
                                    const cleanCellDayUpper = cleanCellDay.toLocaleUpperCase('tr');
                                    const lessonTeacherUpper = lesson.teacher_name ? lesson.teacher_name.toLocaleUpperCase('tr') : '';

                                    const dayMatch = lessonDayUpper === cleanCellDayUpper;
                                    const timeMatch = lesson.time === cellTime;
                                    const teacherMatch = lessonTeacherUpper === headerTeacherName;

                                    if (dayMatch && timeMatch && teacherMatch) {
                                        if (lesson.student_name === studentName) {
                                            isMatch = true;
                                            break;
                                        }
                                    }
                                }
                            }
                        }

                        if (!isMatch) continue;

                        if (skipCell && cell === skipCell) {
                            continue;
                        }

                        if (!skipCell && cell === draggedCell) {
                            continue;
                        }

                        let cellDay = '';
                        let cellTime = '';

                        let row = cell.parentElement;
                        while (row) {
                            const dayCell = row.querySelector('.day-header, td[colspan]');
                            if (dayCell && dayCell.textContent.trim()) {
                                cellDay = dayCell.textContent.trim().toLocaleUpperCase('tr');
                                break;
                            }
                            row = row.previousElementSibling;
                        }

                        const timeCell = cell.parentElement.querySelector('td:first-child');
                        if (timeCell) {
                            cellTime = timeCell.textContent.trim();
                        }

                        // GUN ADLARINI TARIHTEN AYIR
                        const cleanCellDay = extractDayName(cellDay);

                        if (cleanCellDay === cleanToDay && cellTime === toTime) {
                            // HEMEN RETURN ETME, HATAYA EKLE
                            errors.push({
                                type: 'conflict',
                                icon: '',
                                title: 'Cakisma',
                                message: `${studentName} zaten ${cleanToDay} ${toTime}'de baska bir ders aliyor!`
                            });
                            break; // Ayni hatayi tekrar eklememek icin cik
                        }

                    }
                }

                // ═══════════════════════════════════════════════════════════
                // 3️⃣ ÖĞRETMEN UYGUNLUĞU + ÖĞRETMEN ENGELLEME KONTROLÜ
                // ═══════════════════════════════════════════════════════════
                if (targetCellToCheck) {
                    const cellIndex = Array.from(targetCellToCheck.parentElement.children).indexOf(targetCellToCheck);
                    const table = document.getElementById('weeklyPrintTable');
                    const headerRow = table.querySelector('thead tr');
                    const teacherHeader = headerRow.children[cellIndex];

                    if (teacherHeader) {
                        const teacherText = teacherHeader.textContent.trim();
                        const teacherMatch = teacherText.match(/\(([^)]+)\)/);
                        if (teacherMatch) {
                            const teacherFullName = teacherMatch[1].trim().toLocaleUpperCase('tr');

                            // 3A. ÖĞRETMEN DERSİ VAR MI?
                            const hasLesson = await checkTeacherAvailability(teacherFullName, cleanToDay, toTime, teachersData.teachers);

                            if (!hasLesson) {
                                // ✅ HEMEN RETURN ETME, HATAYA EKLE
                                errors.push({
                                    type: 'teacher_availability',
                                    icon: '👨‍🏫',
                                    title: 'Öğretmen Uygunluğu',
                                    message: `Bu öğretmenin ${cleanToDay} ${toTime} saatinde dersi yok!`
                                });
                            }

                            // 3B. ÖĞRETMEN ENGELLEME KONTROLÜ
                            if (student.teacher_blocks && student.teacher_blocks.length > 0) {
                                // Öğretmeni bul (case-insensitive)
                                const teacher = teachersData.teachers.find(t =>
                                    `${t.name} ${t.surname}`.toLocaleUpperCase('tr') === teacherFullName
                                );

                                if (teacher) {
                                    // Her engellemeyi kontrol et
                                    for (const block of student.teacher_blocks) {
                                        // 1. Öğretmen ID kontrolü
                                        if (block.teacher_id !== teacher.id) {
                                            continue;
                                        }

                                        // 2. Hafta kontrolü
                                        if (block.type === 'custom') {
                                            const weeks = block.weeks || [];
                                            if (weeks.length > 0 && !weeks.includes(currentWeekView)) {
                                                continue; // Bu hafta için engelleme yok
                                            }
                                        }
                                        // type === 'weekly' ise her hafta engelli demektir

                                        // 3. Gün kontrolü (✅ cleanToDay kullan)
                                        const blockDay = block.day || 'all';
                                        if (blockDay !== 'all' && blockDay.toLocaleUpperCase('tr') !== cleanToDay) {
                                            continue; // Bu gün için engelleme yok
                                        }

                                        // 4. Slot kontrolü (✅ cleanToDay kullan)
                                        const targetSlot = `${cleanToDay}_${toTime}`;
                                        const blockedSlots = block.blocked_slots || [];

                                        // Slot listesinde bu slot var mı?
                                        const isBlocked = blockedSlots.some(slot => {
                                            // Slot formatı: "Gün_Saat" veya sadece "Saat"
                                            if (slot === targetSlot) return true;
                                            if (slot === toTime) return true;

                                            // Eğer day='all' ise sadece saat kısmını kontrol et
                                            if (blockDay === 'all') {
                                                const slotParts = slot.split('_');
                                                if (slotParts.length > 1) {
                                                    return slotParts[1] === toTime;
                                                }
                                            }

                                            return false;
                                        });

                                        if (isBlocked) {
                                            // ✅ HEMEN RETURN ETME, HATAYA EKLE
                                            let weekInfo = block.type === 'weekly' ? 'Her hafta' : `Hafta ${currentWeekView}`;
                                            let dayInfo = blockDay === 'all' ? 'Tüm günlerde' : cleanToDay;

                                            errors.push({
                                                type: 'teacher_block',
                                                icon: '🛑',
                                                title: 'Öğretmen Engelleme',
                                                message: `${studentName} bu öğretmenden ders alamaz!\n\nÖğretmen: ${teacher.name} ${teacher.surname} (${teacher.branch})\nKapsam: ${weekInfo}, ${dayInfo}, ${toTime}`
                                            });
                                            break; // Aynı hatayı tekrar eklememek için çık
                                        }
                                    }
                                }
                            }
                        }
                    }
                }

                // ═══════════════════════════════════════════════════════════
                // 🚨 HATA VARSA TÜM HATALARI BİRLİKTE GÖSTER
                // ═══════════════════════════════════════════════════════════
                if (errors.length > 0) {
                    let message = '';

                    if (errors.length === 1) {
                        // TEK HATA - BASİT GÖSTER
                        const error = errors[0];
                        message = `${error.icon} ${error.title}\n\n${error.message}`;
                    } else {
                        // ÇOKLU HATA - NUMARALI LİSTE
                        message = `⚠️ ${errors.length} SORUN TESPİT EDİLDİ:\n\n`;

                        errors.forEach((error, index) => {
                            message += `${index + 1}. ${error.icon} ${error.title}\n`;
                            message += `   ${error.message}\n\n`;
                        });

                        message += 'Lütfen başka bir slot seçin.';
                    }

                    return {
                        valid: false,
                        message: message,
                        errorCount: errors.length
                    };
                }

                // ✅ TÜM KONTROLLER BAŞARILI
                return {valid: true, message: 'OK'};

            } catch (error) {
                console.error('Doğrulama hatası:', error);
                return {
                    valid: false,
                    message: 'Doğrulama hatası! Lütfen tekrar deneyin.'
                };
            }
        }

        function findTargetCell(day, time) {
            /**
             * Belirli gün ve saatteki hücreyi bulur
             */
            const table = document.getElementById('weeklyPrintTable');
            if (!table) return null;

            const rows = table.querySelectorAll('tbody tr');
            let currentDay = '';

            for (const row of rows) {
                // Gün başlığı mı?
                const dayCell = row.querySelector('.day-header, td[colspan]');
                if (dayCell && dayCell.textContent.trim()) {
                    currentDay = dayCell.textContent.trim().toLocaleUpperCase('tr');
                    continue;
                }

                // Saat satırı mı?
                const timeCell = row.querySelector('td:first-child');
                if (timeCell && currentDay === day.toLocaleUpperCase('tr')) {
                    const rowTime = timeCell.textContent.trim();
                    if (rowTime === time) {
                        // Bu satırı bulduk, şimdi hücreleri döndür
                        return row; // Satırın kendisini döndür
                    }
                }
            }

            return null;
        }

        async function checkTeacherAvailability(teacherFullName, day, time, teachers) {
            /**
             * Öğretmenin bu gün/saatte dersi var mı kontrol eder
             */

            // ✅ GÜN ADINI TARİHTEN AYIR
            const cleanDay = extractDayName(day);

            for (const teacher of teachers) {
                const tName = `${teacher.name} ${teacher.surname}`.toLocaleUpperCase('tr');

                if (tName === teacherFullName) {
                    // Bu öğretmenin programını kontrol et
                    for (const daySchedule of teacher.schedule) {
                        if (daySchedule.day.toLocaleUpperCase('tr') === cleanDay) {
                            // Bu günde ders veriyor mu?
                            const [targetStart, targetEnd] = time.split('-');

                            for (const lesson of daySchedule.lessons) {
                                // Saat uyuyor mu?
                                if (lesson.start_time === targetStart && lesson.end_time === targetEnd) {
                                    return true; // Öğretmenin bu saatte dersi var
                                }
                            }
                        }
                    }
                }
            }

            return false; // Öğretmenin bu saatte dersi yok
        }

        function extractDayName(dayString) {
            /**
             * Tarih içeren gün bilgisinden sadece gün adını çıkarır
             * "Salı - 28 Eki 25" → "Salı"
             * "Pazartesi" → "Pazartesi"
             */
            if (!dayString) return '';
            // 🔧 Türkçe karakterler için trim() ve normalize et
            return dayString.split('-')[0].trim();
        }

        function checkTimeOverlap(start1, end1, start2, end2) {
            /**
             * İki zaman aralığının çakışıp çakışmadığını kontrol eder
             */
            function timeToMinutes(time) {
                const [h, m] = time.split(':').map(Number);
                return h * 60 + m;
            }

            const s1 = timeToMinutes(start1);
            const e1 = timeToMinutes(end1);
            const s2 = timeToMinutes(start2);
            const e2 = timeToMinutes(end2);

            return !(e1 <= s2 || s1 >= e2);
        }

        // ═══════════════════════════════════════════════════════════
        // 🔄 GLOBALSCHEDULEDATA GÜNCELLEME FONKSİYONU
        // ═══════════════════════════════════════════════════════════
        function updateGlobalScheduleDataAfterSwap(sourceData, targetData) {
            // ✅ SWAP YAPILDIKTAN SONRA globalScheduleData'yı güncelle
            // Bu fonksiyon hem performNormalSwap hem saveSwapToBackend tarafından kullanılır

            if (globalScheduleData && globalScheduleData.weeks && globalScheduleData.weeks[sourceData.week - 1]) {
                const weekData = globalScheduleData.weeks[sourceData.week - 1];

                // GÜN ADLARINI TEMİZLE VE BÜYÜK HARFE ÇEVİR (Türkçe karakterler için)
                const cleanSourceDay = extractDayName(sourceData.day).toLocaleUpperCase('tr');
                const cleanTargetDay = extractDayName(targetData.targetDay).toLocaleUpperCase('tr');

                // ✅ ÖĞRETMEN BİLGİLERİ
                const sourceTeacher = sourceData.teacherName || '';
                const targetTeacher = targetData.targetTeacherName || '';

                // KAYNAK VE HEDEF DERSLERİ BUL
                const sourceLessons = [];
                const targetLessons = [];

                for (let i = 0; i < weekData.length; i++) {
                    const lesson = weekData[i];
                    const lessonDayClean = extractDayName(lesson.day).toLocaleUpperCase('tr');
                    const lessonTeacher = lesson.teacher_name || '';

                    // ✅ Kaynak dersleri bul - ÖĞRETMEN FİLTRESİ İLE
                    if (lessonDayClean === cleanSourceDay &&
                        lesson.time === sourceData.time &&
                        lessonTeacher === sourceTeacher) {
                        const sourceStudentNames = sourceData.studentNames || [sourceData.studentName];
                        if (sourceStudentNames.includes(lesson.student_name)) {
                            sourceLessons.push({index: i, lesson: lesson});
                        }
                    }

                    // ✅ Hedef dersleri bul - ÖĞRETMEN FİLTRESİ İLE
                    if (lessonDayClean === cleanTargetDay &&
                        lesson.time === targetData.targetTime &&
                        lessonTeacher === targetTeacher) {
                        const targetStudentNames = targetData.targetStudentNames || (targetData.targetStudentName ? [targetData.targetStudentName] : []);
                        if (targetStudentNames.length === 0 || targetStudentNames.includes(lesson.student_name)) {
                            targetLessons.push({index: i, lesson: lesson});
                        }
                    }
                }

                // SWAP YAP - Kaynak dersleri hedef slota taşı
                sourceLessons.forEach(item => {
                    // Gün formatını düzelt: "PERŞEMBE" → "Perşembe"
                    const formattedTargetDay = cleanTargetDay.charAt(0).toLocaleUpperCase('tr') +
                                              cleanTargetDay.slice(1).toLocaleLowerCase('tr');
                    weekData[item.index].day = formattedTargetDay;
                    weekData[item.index].time = targetData.targetTime;
                    weekData[item.index].teacher_name = targetTeacher;  // ✅ ÖĞRETMENİ DE DEĞİŞTİR!
                });

                // Hedef dersleri kaynak slota taşı
                targetLessons.forEach(item => {
                    // Gün formatını düzelt: "SALI" → "Salı"
                    const formattedSourceDay = cleanSourceDay.charAt(0).toLocaleUpperCase('tr') +
                                              cleanSourceDay.slice(1).toLocaleLowerCase('tr');
                    weekData[item.index].day = formattedSourceDay;
                    weekData[item.index].time = sourceData.time;
                    weekData[item.index].teacher_name = sourceTeacher;  // ✅ ÖĞRETMENİ DE DEĞİŞTİR!
                });

                console.log('✅ globalScheduleData güncellendi:', {
                    sourceCount: sourceLessons.length,
                    targetCount: targetLessons.length,
                    sourceDay: sourceData.day,
                    targetDay: targetData.targetDay
                });
            }
        }

        // ═══════════════════════════════════════════════════════════
        // 💾 BACKEND'E SWAP KAYDET
        // ═══════════════════════════════════════════════════════════
        async function saveSwapToBackend(sourceData, targetData) {
            // Bu fonksiyon backend'e değişikliği gönderir
            // Sınıf dersleri için TÜM öğrencileri gönder

            try {
                const response = await fetch('/swap_lessons', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        week: sourceData.week,
                        source: {
                            day: sourceData.day,
                            time: sourceData.time,
                            student: sourceData.studentName,
                            teacher: sourceData.teacherName,  // ✅ Öğretmen bilgisi eklendi
                            isClassLesson: sourceData.isClassLesson || false,
                            studentNames: sourceData.studentNames || [sourceData.studentName]
                        },
                        target: {
                            day: targetData.targetDay,
                            time: targetData.targetTime,
                            student: targetData.targetStudentName || null,
                            teacher: targetData.targetTeacherName || null,  // ✅ Hedef öğretmen bilgisi eklendi
                            isClassLesson: targetData.targetIsClassLesson || false,
                            studentNames: targetData.targetStudentNames || (targetData.targetStudentName ? [targetData.targetStudentName] : [])
                        }
                    })
                });

                if (!response.ok) {
                    console.warn('Backend güncellemesi başarısız, ama değişiklik tabloda görünüyor');
                } else {
                    const result = await response.json();
                    console.log('Backend swap sonucu:', result.message);
                }
            } catch (error) {
                console.warn('Backend bağlantı hatası:', error);
            }

            // ✅ BACKEND BAŞARILI OLSUN YA DA OLMASIN - GLOBALSHEDULEDATA'YI GÜNCELLE
            // Çünkü ekranda swap zaten yapıldı, validation için frontend verisini güncellememiz şart!
            updateGlobalScheduleDataAfterSwap(sourceData, targetData);
        }

        window.onload = function() {
            loadTeachers();
            loadStudents();
            loadClassLessons();
            // 🆕 Ana accordion her zaman kapalı başlar, iç accordion'lar hafızalı
        };


        // ═══════════════════════════════════════════════════
        // 💾 GEÇMİŞ PROGRAMLAR SİSTEMİ
        // ═══════════════════════════════════════════════════

        function openSaveScheduleModal() {
            if (!globalScheduleData) {
                showError('Kaydedilecek program bulunamadı! Önce program oluşturun.');
                return;
            }

            const now = new Date();
            const defaultName = `Program ${now.toLocaleDateString('tr-TR')} ${now.toLocaleTimeString('tr-TR', {hour: '2-digit', minute: '2-digit'})}`;
            document.getElementById('scheduleName').value = defaultName;

            document.getElementById('saveScheduleModal').style.display = 'block';
        }

        function closeSaveScheduleModal() {
            document.getElementById('saveScheduleModal').style.display = 'none';
        }

        async function saveCurrentSchedule() {
            const name = document.getElementById('scheduleName').value.trim();

            if (!name) {
                showError('Lütfen program adı girin!');
                return;
            }

            // Tarihi globalScheduleData'dan al
            const startDate = globalScheduleData.start_date || programStartDate;

            if (!startDate) {
                showError('Program başlangıç tarihi bulunamadı! Lütfen programı yeniden oluşturun.');
                return;
            }

            try {
                const response = await fetch('/save_current_schedule', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        name: name,
                        start_date: startDate
                    })
                });

                const result = await response.json();

                if (response.ok) {
                    showSuccess(result.message);
                    closeSaveScheduleModal();
                    savedScheduleData = null; // Cache'i temizle, yeniden yüklensin
                    loadTodayLessons(); // Bugünün derslerini güncelle
                } else {
                    showError(result.error || 'Kaydetme hatası!');
                }
            } catch (error) {
                showError('Kaydetme sırasında hata oluştu!');
                console.error(error);
            }
        }

        async function openSavedSchedulesModal() {
            document.getElementById('savedSchedulesModal').style.display = 'block';
            await loadSavedSchedules();
        }

        function closeSavedSchedulesModal() {
            document.getElementById('savedSchedulesModal').style.display = 'none';
        }

        async function loadSavedSchedules() {
            try {
                const response = await fetch('/get_saved_schedules');
                const data = await response.json();

                const container = document.getElementById('savedSchedulesList');

                if (data.schedules.length === 0) {
                    container.innerHTML = `
                        <div style="text-align: center; padding: 40px; color: #999;">
                            <div style="font-size: 4em; margin-bottom: 20px;">📭</div>
                            <p style="font-size: 1.2em;">Henüz kayıtlı program yok</p>
                        </div>
                    `;
                    return;
                }

                let html = '<div style="display: grid; gap: 15px;">';

                data.schedules.forEach(schedule => {
                    const date = new Date(schedule.created_at);
                    const formattedDate = date.toLocaleDateString('tr-TR', {
                        year: 'numeric',
                        month: 'long',
                        day: 'numeric',
                        hour: '2-digit',
                        minute: '2-digit'
                    });

                    html += `
                        <div style="background: linear-gradient(135deg, #f9fafb 0%, #f3f4f6 100%); border-radius: 12px; padding: 20px; border-left: 5px solid #f59e0b; transition: all 0.3s;" onmouseover="this.style.transform='translateX(5px)'" onmouseout="this.style.transform='translateX(0)'">
                            <div style="display: flex; justify-content: space-between; align-items: start; gap: 15px;">
                                <div style="flex: 1;">
                                    <h3 style="margin: 0 0 10px 0; color: #1f2937; font-size: 1.2em;">${schedule.name}</h3>
                                    <div style="display: flex; gap: 15px; color: #6b7280; font-size: 0.9em;">
                                        <span>📅 ${formattedDate}</span>
                                        <span>📊 ${(schedule.data_size / 1024).toFixed(1)} KB</span>
                                    </div>
                                </div>
                                <div style="display: flex; gap: 8px; flex-wrap: wrap;">
                                    <button onclick="loadSchedule(${schedule.id})" style="background: #10b981; color: white; border: none; padding: 8px 16px; border-radius: 8px; cursor: pointer; font-weight: 600; transition: all 0.2s;" onmouseover="this.style.background='#059669'" onmouseout="this.style.background='#10b981'">
                                        📂 Yükle
                                    </button>
                                    <button onclick="renameScheduleBtn(${schedule.id})" data-name="${schedule.name.replace(/"/g, '&quot;')}" style="background: #3b82f6; color: white; border: none; padding: 8px 16px; border-radius: 8px; cursor: pointer; font-weight: 600; transition: all 0.2s;" onmouseover="this.style.background='#2563eb'" onmouseout="this.style.background='#3b82f6'">
                                        ✏️ Düzenle
                                    </button>
                                    <button onclick="deleteSchedule(${schedule.id})" style="background: #ef4444; color: white; border: none; padding: 8px 16px; border-radius: 8px; cursor: pointer; font-weight: 600; transition: all 0.2s;" onmouseover="this.style.background='#dc2626'" onmouseout="this.style.background='#ef4444'">
                                        🗑️ Sil
                                    </button>
                                </div>
                            </div>
                        </div>
                    `;
                });

                html += '</div>';
                container.innerHTML = html;

            } catch (error) {
                showError('Programlar yüklenirken hata oluştu!');
                console.error(error);
            }
        }

        async function loadSchedule(scheduleId) {
            try {
                const response = await fetch(`/load_schedule/${scheduleId}`);
                const data = await response.json();

                if (response.ok) {
                    globalScheduleData = data.schedule;

                    displayModernSchedule(data.schedule);

                    document.getElementById('weeklyScheduleSection').style.display = 'block';
                    currentWeekView = 1;
                    setTimeout(() => {
                        renderWeeklyTable(1);
                    }, 100);

                    showSuccess('Program başarıyla yüklendi!');
                    await loadStudents();
                    await loadTeachers();
                    closeSavedSchedulesModal();
                    savedScheduleData = null; // Cache'i temizle, yeni programı yükle
                    loadTodayLessons(); // Bugünün derslerini güncelle

                    document.getElementById('resultsSection').scrollIntoView({ behavior: 'smooth' });
                } else {
                    showError(data.error || 'Program yüklenirken hata oluştu!');
                }
            } catch (error) {
                showError('Program yüklenirken hata oluştu!');
                console.error(error);
            }
        }

        async function deleteSchedule(scheduleId) {
            if (!confirm('Bu programı silmek istediğinizden emin misiniz?')) return;

            try {
                const response = await fetch(`/delete_schedule/${scheduleId}`, {
                    method: 'POST'
                });

                const result = await response.json();

                if (response.ok) {
                    showSuccess(result.message);
                    loadSavedSchedules();
                    savedScheduleData = null; // Cache'i temizle
                    loadTodayLessons(); // Bugünün derslerini güncelle
                } else {
                    showError(result.error || 'Silme hatası!');
                }
            } catch (error) {
                showError('Silme sırasında hata oluştu!');
                console.error(error);
            }
        }

        async function renameSchedule(scheduleId, currentName) {
            const newName = prompt('Yeni program adı:', currentName);

            if (!newName || newName.trim() === '' || newName === currentName) return;

            try {
                const response = await fetch(`/rename_schedule/${scheduleId}`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ name: newName.trim() })
                });

                const result = await response.json();

                if (response.ok) {
                    showSuccess(result.message);
                    loadSavedSchedules();
                } else {
                    showError(result.error || 'Yeniden adlandırma hatası!');
                }
            } catch (error) {
                showError('Yeniden adlandırma sırasında hata oluştu!');
                console.error(error);
            }
        }

        function renameScheduleBtn(scheduleId) {
            const button = event.target;
            const currentName = button.getAttribute('data-name').replace(/&quot;/g, '"');
            renameSchedule(scheduleId, currentName);
        }

        // ============== YAZDIRMA FONKSİYONU - TAM YENİ ==============
        window.printWeeklyTable = function() {
            // 🔥 SADECE HAFTALIK BÖLÜMÜ YAZDIRMAK İÇİN CSS EKLE
            const style = document.createElement('style');
            style.id = 'print-only-style';
            style.textContent = `
                @media print {
                    /* 🔥 HER ŞEYİ GİZLE */
                    body * {
                        visibility: hidden !important;
                    }

                    /* 🔥 SADECE HAFTALIK BÖLÜMÜ GÖSTER */
                    #weeklyScheduleSection,
                    #weeklyScheduleSection * {
                        visibility: visible !important;
                    }

                    /* 🔥 TAM SAYFA YAP */
                    #weeklyScheduleSection {
                        position: absolute !important;
                        left: 0 !important;
                        top: 0 !important;
                        width: 100% !important;
                        margin: 0 !important;
                        padding: 5mm !important;
                    }

                    /* 🔥 BUTONLARI GİZLE */
                    #weeklyScheduleSection button {
                        display: none !important;
                    }

                    /* 🔥 BAŞLIK KÜÇÜLT */
                    #weeklyScheduleSection h2 {
                        font-size: 1.2em !important;
                        margin-bottom: 10px !important;
                    }

                    /* 🔥 HAFTA NUMARASI KÜÇÜLT */
                    #weeklyScheduleSection #currentWeekNumber {
                        font-size: 0.9em !important;
                    }

                    /* 🔥 KONTROL DİVİNİ GİZLE */
                    #weeklyScheduleSection > div > div:first-child {
                        display: none !important;
                    }

                    /* 🔥 TABLO AYARLARI */
                    #weeklyScheduleSection table {
                        width: 100% !important;
                        font-size: 0.75em !important;
                    }

                    #weeklyScheduleSection table th,
                    #weeklyScheduleSection table td {
                        padding: 8px 6px !important;
                    }

                    /* 🔥 RENKLER KORUNSUN */
                    * {
                        -webkit-print-color-adjust: exact !important;
                        print-color-adjust: exact !important;
                    }

                    /* 🔥 SAYFA AYARLARI */
                    @page {
                        size: A4 landscape;
                        margin: 5mm;
                    }
                }
            `;

            document.head.appendChild(style);

            // YAZDIRMAYI BAŞLAT
            window.print();

            // STYLE'I KALDIR
            setTimeout(() => {
                const s = document.getElementById('print-only-style');
                if (s) s.remove();
            }, 1000);
        };

        // ═══════════════════════════════════════════════════════════════
        // 🎬 ANİMASYONLU HAFTA DEĞİŞTİRME FONKSİYONU - ADIM 2
        // ═══════════════════════════════════════════════════════════════
        window.changeWeek = function(direction) {
            currentWeekView += direction;
            if (currentWeekView < 1) currentWeekView = 1;
            if (currentWeekView > 4) currentWeekView = 4;

            renderWeeklyTable(currentWeekView);
            updateWeekButtons();
        };

        // Buton durumlarını güncelle
        function updateWeekButtons() {
            const prevBtn = document.getElementById('prevWeekBtn');
            const nextBtn = document.getElementById('nextWeekBtn');

            if (!prevBtn || !nextBtn) return;

            // Önceki Hafta butonu
            if (currentWeekView <= 1) {
                prevBtn.disabled = true;
                prevBtn.style.opacity = '0.4';
                prevBtn.style.cursor = 'not-allowed';
                prevBtn.style.background = 'linear-gradient(135deg, #9ca3af 0%, #6b7280 100%)';
            } else {
                prevBtn.disabled = false;
                prevBtn.style.opacity = '1';
                prevBtn.style.cursor = 'pointer';
                prevBtn.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
            }

            // Sonraki Hafta butonu
            if (currentWeekView >= 4) {
                nextBtn.disabled = true;
                nextBtn.style.opacity = '0.4';
                nextBtn.style.cursor = 'not-allowed';
                nextBtn.style.background = 'linear-gradient(135deg, #9ca3af 0%, #6b7280 100%)';
            } else {
                nextBtn.disabled = false;
                nextBtn.style.opacity = '1';
                nextBtn.style.cursor = 'pointer';
                nextBtn.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
            }
        }

        // ============== TAKVİM GÖRÜNÜMÜ FONKSİYONLARI ==============

        // Görünüm değiştirme
        function switchView(view) {
            currentView = view;

            const tableViewBtn = document.getElementById('tableViewBtn');
            const calendarViewBtn = document.getElementById('calendarViewBtn');
            const weeklyScheduleTable = document.getElementById('weeklyScheduleTable');
            const calendarView = document.getElementById('calendarView');
            const searchBoxContainer = document.getElementById('searchBoxContainer');
            const teacherSearchBoxContainer = document.getElementById('teacherSearchBoxContainer');
            const prevWeekBtn = document.getElementById('prevWeekBtn');
            const nextWeekBtn = document.getElementById('nextWeekBtn');

            if (view === 'table') {
                // Tablo görünümü
                tableViewBtn.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
                tableViewBtn.style.color = 'white';
                tableViewBtn.style.boxShadow = '0 2px 8px rgba(102,126,234,0.3)';
                calendarViewBtn.style.background = 'white';
                calendarViewBtn.style.color = '#6b7280';
                calendarViewBtn.style.boxShadow = 'none';

                weeklyScheduleTable.style.display = 'block';
                calendarView.style.display = 'none';
                searchBoxContainer.style.display = 'block';
                teacherSearchBoxContainer.style.display = 'block';
                prevWeekBtn.style.display = 'inline-block';
                nextWeekBtn.style.display = 'inline-block';

                // Buton yazılarını ve fonksiyonlarını HAFTA için ayarla
                prevWeekBtn.innerHTML = '◄ Önceki Hafta';
                nextWeekBtn.innerHTML = 'Sonraki Hafta ►';
                prevWeekBtn.onclick = () => changeWeek(-1);
                nextWeekBtn.onclick = () => changeWeek(1);

                // Hafta butonlarının durumunu güncelle (deaktif kontrol)
                updateWeekButtons();
            } else {
                // Takvim görünümü
                calendarViewBtn.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
                calendarViewBtn.style.color = 'white';
                calendarViewBtn.style.boxShadow = '0 2px 8px rgba(102,126,234,0.3)';
                tableViewBtn.style.background = 'white';
                tableViewBtn.style.color = '#6b7280';
                tableViewBtn.style.boxShadow = 'none';

                weeklyScheduleTable.style.display = 'none';
                calendarView.style.display = 'block';
                searchBoxContainer.style.display = 'none';
                teacherSearchBoxContainer.style.display = 'none';
                prevWeekBtn.style.display = 'inline-block';
                nextWeekBtn.style.display = 'inline-block';

                // Buton yazılarını ve fonksiyonlarını AY için ayarla
                prevWeekBtn.innerHTML = '◄ Önceki Ay';
                nextWeekBtn.innerHTML = 'Sonraki Ay ►';
                prevWeekBtn.onclick = () => changeMonth(-1);
                nextWeekBtn.onclick = () => changeMonth(1);

                // Takvim için butonları HER ZAMAN AKTİF YAP
                prevWeekBtn.disabled = false;
                nextWeekBtn.disabled = false;
                prevWeekBtn.style.opacity = '1';
                nextWeekBtn.style.opacity = '1';
                prevWeekBtn.style.cursor = 'pointer';
                nextWeekBtn.style.cursor = 'pointer';
                prevWeekBtn.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
                nextWeekBtn.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';

                renderCalendar();
            }
        }

        // Ay değiştirme
        function changeMonth(direction) {
            currentCalendarDate.setMonth(currentCalendarDate.getMonth() + direction);
            renderCalendar();
        }

        // Takvim oluşturma
        function renderCalendar() {
            if (!globalScheduleData) return;

            const year = currentCalendarDate.getFullYear();
            const month = currentCalendarDate.getMonth();

            // Ay başlığını güncelle
            const monthNames = ['OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN',
                               'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK'];
            document.getElementById('calendarMonthTitle').textContent = `${monthNames[month]} ${year}`;

            // Ayın ilk ve son günü
            const firstDay = new Date(year, month, 1);
            const lastDay = new Date(year, month + 1, 0);
            const daysInMonth = lastDay.getDate();

            // ✅ Pazartesi bazlı: 0=Pazartesi, 1=Salı, ... 6=Pazar
            let startDayOfWeek = firstDay.getDay() - 1; // JavaScript: 0=Pazar, 1=Pazartesi
            if (startDayOfWeek === -1) startDayOfWeek = 6; // Pazar ise sona at

            // ============== PROGRAM TARİH ARALIĞINI HESAPLA ==============
            let programStartDate = null;
            let programEndDate = null;

            if (globalScheduleData.start_date) {
                programStartDate = new Date(globalScheduleData.start_date + 'T00:00:00');
                programEndDate = new Date(programStartDate);
                programEndDate.setDate(programStartDate.getDate() + 27); // 4 hafta = 28 gün (0-27)
            }

            // Gün adları
            // ⚠️ dayNames: Takvim başlıkları için (Pazartesi'den başlar)
            // ⚠️ dayNamesLong: JavaScript getDay() ile uyumlu (PAZAR'dan başlar)
            const dayNames = ['Pzt', 'Sal', 'Çar', 'Per', 'Cum', 'Cmt', 'Paz'];
            const dayNamesLong = ['Pazar', 'Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi'];

            let html = `
                <div style="display: grid; grid-template-columns: repeat(7, 1fr); gap: 10px;">
                    <!-- Başlık satırı -->
            `;

            // Gün başlıkları
            dayNames.forEach(day => {
                html += `
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px; text-align: center; font-weight: 700; border-radius: 10px; font-size: 1.1em;">
                        ${day}
                    </div>
                `;
            });

            // Boş hücreler (ayın ilk gününden önceki)
            for (let i = 0; i < startDayOfWeek; i++) {
                html += `<div style="background: #f9fafb; border-radius: 10px; min-height: 100px;"></div>`;
            }

            // Günler
            for (let day = 1; day <= daysInMonth; day++) {
                const date = new Date(year, month, day);
                const dayOfWeek = date.getDay();
                const dayName = dayNamesLong[dayOfWeek];

                // ============== BU TARİH PROGRAM İÇİNDE Mİ? ==============
                let lessonCount = 0;
                let isInProgramRange = false;

                if (programStartDate && programEndDate) {
                    // Tarihleri karşılaştır
                    const dateOnly = new Date(year, month, day);
                    if (dateOnly >= programStartDate && dateOnly <= programEndDate) {
                        isInProgramRange = true;

                        // Bu tarihin hangi haftaya denk geldiğini hesapla
                        const diffTime = dateOnly - programStartDate;
                        const diffDays = Math.floor(diffTime / (1000 * 60 * 60 * 24));
                        const weekNum = Math.floor(diffDays / 7) + 1;

                        // Sadece o haftanın derslerini say
                        if (globalScheduleData.weeks && globalScheduleData.weeks[weekNum - 1]) {
                            const week = globalScheduleData.weeks[weekNum - 1];
                            week.forEach(lesson => {
                                if (lesson.day === dayName) {
                                    lessonCount++;
                                }
                            });
                        }
                    }
                }

                // Bugün mü?
                const today = new Date();
                const isToday = date.toDateString() === today.toDateString();

                const bgColor = isToday ? '#dcfce7' : 'white';
                const borderColor = isToday ? '#10b981' : '#e5e7eb';

                html += `
                    <div onclick="showDayDetail('${dayName}', ${day}, ${month + 1}, ${year})"
                         style="background: ${bgColor}; border: 2px solid ${borderColor}; border-radius: 10px; padding: 10px; cursor: pointer; transition: all 0.3s; min-height: 100px; display: flex; flex-direction: column;"
                         onmouseover="this.style.transform='translateY(-3px)'; this.style.boxShadow='0 6px 20px rgba(0,0,0,0.15)'"
                         onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='none'">
                        <div style="font-size: 1.3em; font-weight: 700; color: ${isToday ? '#10b981' : '#1f2937'}; margin-bottom: 8px;">
                            ${day}
                        </div>
                        ${isInProgramRange && lessonCount > 0 ? `
                            <div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 8px; border-radius: 8px; text-align: center; font-weight: 600; font-size: 0.9em; margin-top: auto;">
                                📚 ${lessonCount} Ders
                            </div>
                        ` : `
                            <div style="text-align: center; color: #9ca3af; font-size: 0.8em; margin-top: auto;">
                                Ders yok
                            </div>
                        `}
                    </div>
                `;
            }

            html += `</div>`;
            document.getElementById('calendarGrid').innerHTML = html;
        }

        // Gün detayı göster
        function showDayDetail(dayName, day, month, year) {
            if (!globalScheduleData) return;

            // ============== TARİHE GÖRE HAFTA HESAPLA ==============
            let targetWeekNum = null;

            if (globalScheduleData.start_date) {
                // Tıklanan tarih
                const clickedDate = new Date(year, month - 1, day);

                // Program başlangıç tarihi
                const startDate = new Date(globalScheduleData.start_date + 'T00:00:00');

                // Gün farkını hesapla
                const diffTime = clickedDate - startDate;
                const diffDays = Math.floor(diffTime / (1000 * 60 * 60 * 24));

                // Hangi haftaya denk geliyor? (0-6 gün = Hafta 1, 7-13 gün = Hafta 2, vs.)
                if (diffDays >= 0 && diffDays < 28) {
                    targetWeekNum = Math.floor(diffDays / 7) + 1;
                }
            }

            // Bu güne ait dersleri topla
            const dayLessons = [];
            if (globalScheduleData.weeks) {
                globalScheduleData.weeks.forEach((week, weekIndex) => {
                    const currentWeekNum = weekIndex + 1;

                    // ✅ SADECE HESAPLANAN HAFTANIN DERSLERİNİ AL
                    if (targetWeekNum === null || currentWeekNum === targetWeekNum) {
                        week.forEach(lesson => {
                            if (lesson.day === dayName) {
                                dayLessons.push({
                                    ...lesson,
                                    weekNum: currentWeekNum
                                });
                            }
                        });
                    }
                });
            }

            const monthNames = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
                               'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık'];

            let html = `
                <div style="padding: 20px;">
                    <h2 style="color: #667eea; margin-bottom: 20px; display: flex; align-items: center; gap: 10px;">
                        <i class="fas fa-calendar-day"></i> ${day} ${monthNames[month - 1]} ${year} - ${dayName}
                    </h2>
            `;

            if (dayLessons.length === 0) {
                html += `
                    <div style="text-align: center; padding: 60px 20px; color: #6b7280;">
                        <i class="fas fa-calendar-times" style="font-size: 4em; margin-bottom: 20px; opacity: 0.3;"></i>
                        <p style="font-size: 1.3em; font-weight: 600;">Bu gün için ders yok</p>
                    </div>
                `;
            } else {
                html += `
                    <div style="margin-bottom: 15px; padding: 15px; background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); border-radius: 10px; border-left: 4px solid #10b981;">
                        <strong style="color: #065f46; font-size: 1.1em;">
                            Toplam ${dayLessons.length} ders
                            ${targetWeekNum ? ` (Hafta ${targetWeekNum})` : ''}
                        </strong>
                    </div>
                `;

                // ============== BRANŞLARA GÖRE GRUPLA ==============
                const lessonsByBranch = {};
                dayLessons.forEach(lesson => {
                    const branch = lesson.branch;
                    if (!lessonsByBranch[branch]) {
                        lessonsByBranch[branch] = [];
                    }
                    lessonsByBranch[branch].push(lesson);
                });

                // Branşları alfabetik sırala
                const sortedBranches = Object.keys(lessonsByBranch).sort((a, b) =>
                    a.toLocaleLowerCase('tr').localeCompare(b.toLocaleLowerCase('tr'), 'tr')
                );

                // Her branş içindeki dersleri saate göre sırala
                sortedBranches.forEach(branch => {
                    lessonsByBranch[branch].sort((a, b) => a.time.localeCompare(b.time));
                });

                html += `<div style="display: flex; flex-direction: column; gap: 20px; max-height: 500px; overflow-y: auto; padding-right: 10px;">`;

                // Her branş için grup oluştur
                sortedBranches.forEach(branch => {
                    const branchLessons = lessonsByBranch[branch];

                    html += `
                        <div style="background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); border-radius: 15px; padding: 20px; border-left: 5px solid #3b82f6;">
                            <h4 style="margin: 0 0 15px 0; color: #1e40af; font-size: 1.3em; display: flex; align-items: center; gap: 10px;">
                                <i class="fas fa-book-open"></i> ${branch}
                                <span style="background: #3b82f6; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7em; font-weight: 600;">${branchLessons.length} ders</span>
                            </h4>
                            <div style="display: flex; flex-direction: column; gap: 12px;">
                    `;

                    // 🆕 AYNI SLOT'TAKİ DERSLERİ GRUPLA (TAKVİM DETAY)
                    const groupedLessons = [];
                    const processedSlots = new Set();

                    branchLessons.forEach(lesson => {
                        const slotKey = `${lesson.time}_${lesson.teacher_name}`;

                        if (processedSlots.has(slotKey)) {
                            return; // Zaten işlendi
                        }

                        // Aynı slot'taki tüm dersleri bul
                        const sameslotLessons = branchLessons.filter(l =>
                            l.time === lesson.time && l.teacher_name === lesson.teacher_name
                        );

                        processedSlots.add(slotKey);

                        if (sameslotLessons.length === 1) {
                            // Tek ders
                            groupedLessons.push(lesson);
                        } else {
                            // 🆕 GRUP DERSİ - TÜM SINIFLARI TOPLA
                            const uniqueClasses = [...new Set(sameslotLessons.map(l => l.student_class).filter(c => c))];
                            let displayName;
                            if (uniqueClasses.length > 0) {
                                const classesStr = uniqueClasses.sort().join(', ');
                                displayName = `${classesStr} (${sameslotLessons.length} öğr)`;
                            } else {
                                displayName = `${sameslotLessons.length} öğrenci`;
                            }
                            groupedLessons.push({
                                ...lesson,
                                student_name: displayName,
                                is_grouped: true
                            });
                        }
                    });

                    groupedLessons.forEach(lesson => {
                        html += `
                            <div style="background: white; border-radius: 10px; padding: 15px; box-shadow: 0 2px 8px rgba(0,0,0,0.05); transition: all 0.3s;"
                                 onmouseover="this.style.transform='translateX(5px)'; this.style.boxShadow='0 4px 15px rgba(0,0,0,0.1)'"
                                 onmouseout="this.style.transform='translateX(0)'; this.style.boxShadow='0 2px 8px rgba(0,0,0,0.05)'">
                                <div style="display: grid; grid-template-columns: 140px 1fr 1fr; gap: 15px; align-items: center;">
                                    <div style="background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%); color: white; padding: 8px 4px; border-radius: 8px; text-align: center; font-weight: 700; font-size: 0.95em; white-space: nowrap;">
                                        ⏰ ${lesson.time}
                                    </div>
                                    <div style="padding: 8px;">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">${lesson.is_grouped ? '🎓 Sınıf' : '👨‍🎓 Öğrenci'}</div>
                                        <div style="font-weight: 600; color: #1f2937; font-size: 1em;">${lesson.student_name}</div>
                                    </div>
                                    <div style="padding: 8px;">
                                        <div style="font-size: 0.75em; color: #6b7280; margin-bottom: 4px;">👨‍🏫 Öğretmen</div>
                                        <div style="font-weight: 600; color: #1f2937; font-size: 1em;">${lesson.teacher_name}</div>
                                    </div>
                                </div>
                            </div>
                        `;
                    });

                    html += `
                            </div>
                        </div>
                    `;
                });

                html += `</div>`;
            }

            html += `</div>`;

            document.getElementById('dayDetailContent').innerHTML = html;
            document.getElementById('dayDetailModal').style.display = 'block';
        }

        function closeDayDetailModal() {
            document.getElementById('dayDetailModal').style.display = 'none';
        }

        // ============== PDF EXPORT - TAM YENİ ==============
        window.exportWeeklyToPDF = async function() {
            const weekNum = currentWeekView;
            const sectionElement = document.getElementById('weeklyScheduleSection');
            const table = document.getElementById('weeklyPrintTable');

            if (!sectionElement || !table) {
                alert('Tablo bulunamadı!');
                return;
            }

            // 🔥 ORİJİNAL STYLE'LARI SAKLA
            const originalStyles = {
                buttons: [],
                headerDisplay: null,
                controlDisplay: null,
                sectionTransform: sectionElement.style.transform,
                sectionTransformOrigin: sectionElement.style.transformOrigin,
                sectionWidth: sectionElement.style.width,
                sectionPadding: sectionElement.style.padding,
                sectionBackground: sectionElement.style.background
            };

            // 🔥 HER ŞEYİ GİZLE
            const allButtons = sectionElement.querySelectorAll('button');
            allButtons.forEach(btn => {
                originalStyles.buttons.push(btn.style.display);
                btn.style.display = 'none';
            });

            const header = sectionElement.querySelector('h2');
            const controlDiv = sectionElement.querySelector('div[style*="display: flex"]');

            if (header) {
                originalStyles.headerDisplay = header.style.display;
                header.style.display = 'none';
            }

            if (controlDiv) {
                originalStyles.controlDisplay = controlDiv.style.display;
                controlDiv.style.display = 'none';
            }

            // 🔥 ARKA PLAN VE PADDING KALDIR
            sectionElement.style.background = 'white';
            sectionElement.style.padding = '0';

            // 🔥 TABLO GENİŞLİĞİNİ ÖLÇ
            const pageWidth = 1140;
            const pageHeight = 780;

            const tableWidth = table.offsetWidth;
            const tableHeight = table.offsetHeight + 50;

            const scaleX = pageWidth / tableWidth;
            const scaleY = pageHeight / tableHeight;

            let autoScale = Math.min(scaleX, scaleY);
            if (autoScale > 0.92) autoScale = 0.92;
            if (autoScale < 0.55) autoScale = 0.55;

            console.log('📄 PDF - Tablo:', tableWidth, 'x', tableHeight);
            console.log('🎯 PDF - Ölçek:', autoScale);

            const newWidth = 100 / autoScale;
            sectionElement.style.transform = `scale(${autoScale})`;
            sectionElement.style.transformOrigin = 'top left';
            sectionElement.style.width = `${newWidth}%`;

            try {
                await new Promise(resolve => setTimeout(resolve, 700));

                const canvas = await html2canvas(sectionElement, {
                    scale: 2.5,
                    useCORS: true,
                    logging: false,
                    backgroundColor: '#ffffff',
                    windowWidth: sectionElement.scrollWidth,
                    windowHeight: sectionElement.scrollHeight,
                    removeContainer: true
                });

                const imgData = canvas.toDataURL('image/png', 1.0);
                const pdf = new jsPDF('landscape', 'mm', 'a4');

                const pdfWidth = pdf.internal.pageSize.getWidth();
                const pdfHeight = pdf.internal.pageSize.getHeight();

                const imgWidth = canvas.width;
                const imgHeight = canvas.height;

                const margin = 5;
                const availableWidth = pdfWidth - (margin * 2);
                const availableHeight = pdfHeight - (margin * 2);

                const ratio = Math.min(
                    availableWidth / (imgWidth * 0.264583),
                    availableHeight / (imgHeight * 0.264583)
                );

                const scaledWidth = imgWidth * 0.264583 * ratio;
                const scaledHeight = imgHeight * 0.264583 * ratio;

                const imgX = (pdfWidth - scaledWidth) / 2;
                const imgY = (pdfHeight - scaledHeight) / 2;

                pdf.addImage(imgData, 'PNG', imgX, imgY, scaledWidth, scaledHeight, '', 'FAST');

                pdf.save(`Hafta_${weekNum}_Program.pdf`);

            } catch (error) {
                console.error('PDF oluşturma hatası:', error);
                alert('PDF oluşturulurken hata oluştu!');
            } finally {
                // 🔥 HER ŞEYİ GERİ AL
                allButtons.forEach((btn, index) => {
                    btn.style.display = originalStyles.buttons[index];
                });

                if (header) header.style.display = originalStyles.headerDisplay;
                if (controlDiv) controlDiv.style.display = originalStyles.controlDisplay;

                sectionElement.style.background = originalStyles.sectionBackground;
                sectionElement.style.padding = originalStyles.sectionPadding;
                sectionElement.style.transform = originalStyles.sectionTransform;
                sectionElement.style.transformOrigin = originalStyles.sectionTransformOrigin;
                sectionElement.style.width = originalStyles.sectionWidth;
            }
        };

        // ============== 4 HAFTALIK PDF EXPORT ==============
        window.exportAllWeeksToPDF = async function() {
            if (!globalScheduleData) {
                alert('Önce program oluşturun!');
                return;
            }

            const sectionElement = document.getElementById('weeklyScheduleSection');
            if (!sectionElement) {
                alert('Haftalık program bulunamadı!');
                return;
            }

            const originalWeek = currentWeekView;

            try {
                const pdf = new jsPDF('landscape', 'mm', 'a4');

                // Her hafta için
                for (let weekNum = 1; weekNum <= 4; weekNum++) {
                    // İlk sayfadan sonra yeni sayfa ekle
                    if (weekNum > 1) {
                        pdf.addPage();
                    }

                    // Bu haftayı göster
                    renderWeeklyTable(weekNum);

                    // Render için bekle
                    await new Promise(resolve => setTimeout(resolve, 300));

                    const table = document.getElementById('weeklyPrintTable');
                    if (!table) continue;

                    // 🔥 ORİJİNAL STYLE'LARI SAKLA
                    const originalStyles = {
                        buttons: [],
                        headerDisplay: null,
                        controlDisplay: null,
                        sectionTransform: sectionElement.style.transform,
                        sectionTransformOrigin: sectionElement.style.transformOrigin,
                        sectionWidth: sectionElement.style.width,
                        sectionPadding: sectionElement.style.padding,
                        sectionBackground: sectionElement.style.background
                    };

                    // 🔥 HER ŞEYİ GİZLE
                    const allButtons = sectionElement.querySelectorAll('button');
                    allButtons.forEach(btn => {
                        originalStyles.buttons.push(btn.style.display);
                        btn.style.display = 'none';
                    });

                    const header = sectionElement.querySelector('h2');
                    const controlDiv = sectionElement.querySelector('div[style*="display: flex"]');

                    if (header) {
                        originalStyles.headerDisplay = header.style.display;
                        header.style.display = 'none';
                    }

                    if (controlDiv) {
                        originalStyles.controlDisplay = controlDiv.style.display;
                        controlDiv.style.display = 'none';
                    }

                    // 🔥 ARKA PLAN VE PADDING KALDIR
                    sectionElement.style.background = 'white';
                    sectionElement.style.padding = '0';

                    // 🔥 TABLO GENİŞLİĞİNİ ÖLÇ
                    const pageWidth = 1140;
                    const pageHeight = 780;

                    const tableWidth = table.offsetWidth;
                    const tableHeight = table.offsetHeight + 50;

                    const scaleX = pageWidth / tableWidth;
                    const scaleY = pageHeight / tableHeight;

                    let autoScale = Math.min(scaleX, scaleY);
                    if (autoScale > 0.92) autoScale = 0.92;
                    if (autoScale < 0.55) autoScale = 0.55;

                    const newWidth = 100 / autoScale;
                    sectionElement.style.transform = `scale(${autoScale})`;
                    sectionElement.style.transformOrigin = 'top left';
                    sectionElement.style.width = `${newWidth}%`;

                    // Render için bekle
                    await new Promise(resolve => setTimeout(resolve, 700));

                    // Canvas oluştur
                    const canvas = await html2canvas(sectionElement, {
                        scale: 2.5,
                        useCORS: true,
                        logging: false,
                        backgroundColor: '#ffffff',
                        windowWidth: sectionElement.scrollWidth,
                        windowHeight: sectionElement.scrollHeight,
                        removeContainer: true
                    });

                    // 🔥 HER ŞEYİ GERİ AL
                    allButtons.forEach((btn, index) => {
                        btn.style.display = originalStyles.buttons[index];
                    });

                    if (header) header.style.display = originalStyles.headerDisplay;
                    if (controlDiv) controlDiv.style.display = originalStyles.controlDisplay;

                    sectionElement.style.background = originalStyles.sectionBackground;
                    sectionElement.style.padding = originalStyles.sectionPadding;
                    sectionElement.style.transform = originalStyles.sectionTransform;
                    sectionElement.style.transformOrigin = originalStyles.sectionTransformOrigin;
                    sectionElement.style.width = originalStyles.sectionWidth;

                    // PDF'e ekle
                    const imgData = canvas.toDataURL('image/png', 1.0);
                    const pdfWidth = pdf.internal.pageSize.getWidth();
                    const pdfHeight = pdf.internal.pageSize.getHeight();

                    const imgWidth = canvas.width;
                    const imgHeight = canvas.height;

                    const margin = 5;
                    const availableWidth = pdfWidth - (margin * 2);
                    const availableHeight = pdfHeight - (margin * 2);

                    const ratio = Math.min(
                        availableWidth / (imgWidth * 0.264583),
                        availableHeight / (imgHeight * 0.264583)
                    );

                    const scaledWidth = imgWidth * 0.264583 * ratio;
                    const scaledHeight = imgHeight * 0.264583 * ratio;

                    const imgX = (pdfWidth - scaledWidth) / 2;
                    const imgY = (pdfHeight - scaledHeight) / 2;

                    pdf.addImage(imgData, 'PNG', imgX, imgY, scaledWidth, scaledHeight, '', 'FAST');
                }

                // PDF'i kaydet
                pdf.save('4_Haftalik_Program.pdf');

            } catch (error) {
                console.error('PDF oluşturma hatası:', error);
                alert('PDF oluşturulurken hata oluştu: ' + error.message);
            } finally {
                // Orijinal haftaya dön
                renderWeeklyTable(originalWeek);
            }
        };

        // Hafta tablosu HTML'i oluştur
        async function generateWeekTableHTML(weekNum) {
            const response = await fetch('/get_teachers');
            const data = await response.json();
            const teachers = data.teachers;

            teachers.sort((a, b) => {
                if (a.branch !== b.branch) {
                    return a.branch.localeCompare(b.branch, 'tr');
                }
                return a.name.localeCompare(b.name, 'tr');
            });

            const allSlots = [];
            teachers.forEach(teacher => {
                teacher.schedule.forEach(daySchedule => {
                    daySchedule.lessons.forEach(lesson => {
                        const slotKey = `${daySchedule.day}_${lesson.start_time}_${lesson.end_time}`;
                        const existingSlot = allSlots.find(s => s.key === slotKey);
                        if (!existingSlot) {
                            allSlots.push({
                                day: daySchedule.day,
                                start_time: lesson.start_time,
                                end_time: lesson.end_time,
                                key: slotKey
                            });
                        }
                    });
                });
            });

            const dayOrder = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar'];
            allSlots.sort((a, b) => {
                const dayDiff = dayOrder.indexOf(a.day) - dayOrder.indexOf(b.day);
                if (dayDiff !== 0) return dayDiff;
                return a.start_time.localeCompare(b.start_time);
            });

            const weekData = globalScheduleData.weeks[weekNum - 1];

            let html = `
                <table style="width: 100%; border-collapse: collapse; font-size: 0.5em;">
                    <caption style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 8px; font-weight: bold; font-size: 1.2em;">
                        HAFTA ${weekNum}
                    </caption>
                    <thead>
                        <tr style="background: linear-gradient(135deg, #4472C4 0%, #5B9BD5 100%);">
                            <th style="color: white; padding: 5px 3px; font-size: 0.8em; border: 1px solid rgba(255,255,255,0.2);">GÜN/SAAT</th>
            `;

            teachers.forEach(teacher => {
                html += `
                    <th style="color: white; padding: 5px 2px; text-align: center; font-size: 0.7em; border: 1px solid rgba(255,255,255,0.2); line-height: 1.2;">
                        ${teacher.branch}<br><span style="font-size: 0.8em;">(${teacher.name} ${teacher.surname})</span>
                    </th>
                `;
            });

            html += `
                        </tr>
                    </thead>
                    <tbody>
            `;

            let currentDay = null;
            allSlots.forEach(slot => {
                if (slot.day !== currentDay) {
                    html += `
                        <tr>
                            <td colspan="${teachers.length + 1}" style="background: #9575CD; color: white; font-weight: bold; padding: 4px; text-align: center; font-size: 0.9em;">
                                ${slot.day}
                            </td>
                        </tr>
                    `;
                    currentDay = slot.day;
                }

                html += `
                    <tr>
                        <td style="background: #E3F2FD; color: #1565C0; font-weight: 600; padding: 4px; font-size: 0.8em; white-space: nowrap;">
                            ${slot.start_time}-${slot.end_time}
                        </td>
                `;

                teachers.forEach(teacher => {
                    const teacherFullName = `${teacher.name} ${teacher.surname}`;
                    let studentName = '';

                    const lesson = weekData.find(l =>
                        l.teacher_name === teacherFullName &&
                        l.day === slot.day &&
                        l.time === `${slot.start_time}-${slot.end_time}`
                    );

                    if (lesson) {
                        studentName = lesson.student_name;
                    }

                    html += `
                        <td style="padding: 4px 2px; text-align: center; font-size: 0.75em; font-weight: 600; border: 1px solid #e5e7eb; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">
                            ${studentName}
                        </td>
                    `;
                });

                html += `</tr>`;
            });

            html += `
                    </tbody>
                </table>
            `;

            return html;
        }

        // ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        // 📚 SINIF DERSİ YÖNETİMİ FONKSİYONLARI
        // ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

        var currentTeacherSchedule = null;
        var classLessonSelectedDay = null;
        var classLessonSelectedTime = [];  // 🆕 Çoklu saat seçimi için array


        function openClassLessonModal() {
            document.getElementById("classLessonModal").style.display = "block";
            loadClasses();
            loadTeachersForClassLesson();
        }

        function closeClassLessonModal() {
            document.getElementById("classLessonModal").style.display = "none";
            resetClassLessonForm();
        }

        function resetClassLessonForm() {
            // 🆕 DÜZENLEME MODUNU SIFIRLA
            editingClassLessonId = null;

            // 🆕 MODAL BAŞLIĞINI SIFIRLA
            document.querySelector("#classLessonModal h2").innerHTML = '📚 Sınıf Dersi Ata';
            document.getElementById("saveClassLessonBtn").innerHTML = '<i class="fas fa-save"></i> Sınıf Dersini Kaydet';

            document.getElementById("classLessonClass").value = "";
            document.getElementById("classLessonTeacher").value = "";
            document.getElementById("classStudentCount").innerHTML = "";
            document.getElementById("teacherBranchInfo").innerHTML = "";
            document.getElementById("daySelectionGroup").style.display = "none";
            document.getElementById("timeSelectionGroup").style.display = "none";
            document.getElementById("weekSelectionGroup").style.display = "none";
            document.getElementById("saveClassLessonBtn").style.display = "none";
            currentTeacherSchedule = null;
            classLessonSelectedDay = null;
            classLessonSelectedTime = [];  // Array olarak sıfırla

            // 🆕 GÜN BUTONLARINI SIFIRLA
            document.querySelectorAll("#availableDays button").forEach(btn => {
                btn.style.background = "white";
                btn.style.color = "#4b5563";
                btn.style.border = "2px solid #e5e7eb";
            });

            // 🆕 SAAT CHECKBOXLARINI SIFIRLA
            document.querySelectorAll("#availableTimes label").forEach(lbl => {
                const checkbox = lbl.querySelector("input[type='checkbox']");
                if (checkbox) {
                    checkbox.checked = false;
                }
                lbl.style.borderColor = "#e5e7eb";
                lbl.style.background = "white";
            });

            // 🆕 HAFTA CHECKBOXLARINI SIFIRLA
            document.getElementById("classLessonAllWeeks").checked = false;
            document.querySelectorAll(".week-checkbox").forEach(cb => {
                cb.checked = false;
            });
            document.getElementById("individualWeeks").style.display = "flex";

            // 🆕 AVAILABLE DAYS VE TIMES CONTAINER'LARI TEMİZLE
            document.getElementById("availableDays").innerHTML = "";
            document.getElementById("availableTimes").innerHTML = "";
        }

        async function loadClasses() {
            try {
                const response = await fetch("/get_unique_classes");
                const classes = await response.json();

                const select = document.getElementById("classLessonClass");
                select.innerHTML = "<option value=''>Sinif secin...</option>";

                classes.forEach(function(className) {
                    var option = document.createElement("option");
                    option.value = className;
                    option.textContent = className;
                    select.appendChild(option);
                });
            } catch (error) {
                showError("Siniflar yuklenirken hata!");
            }
        }

        async function updateClassStudentCount() {
            const className = document.getElementById("classLessonClass").value;
            if (!className) {
                document.getElementById("classStudentCount").innerHTML = "";
                return;
            }

            try {
                const response = await fetch("/get_students_by_class/" + className);
                const data = await response.json();

                var names = data.students.map(function(s) { return s.name; }).join(", ");
                document.getElementById("classStudentCount").innerHTML =
                    "<i class='fas fa-users'></i> " + data.count + " ogrenci: " + names;
            } catch (error) {
                showError("Ogrenci bilgileri alinirken hata!");
            }
        }

        async function loadTeachersForClassLesson() {
            try {
                const response = await fetch("/get_teachers");
                const data = await response.json();

                const select = document.getElementById("classLessonTeacher");
                select.innerHTML = "<option value=''>Ogretmen secin...</option>";

                // ✅ ALFABETİK SIRALAMA (önce branş, sonra isim)
                const sortedTeachers = data.teachers.sort((a, b) => {
                    // Önce branşa göre sırala
                    const branchCompare = a.branch.toLocaleLowerCase('tr').localeCompare(b.branch.toLocaleLowerCase('tr'), 'tr');
                    if (branchCompare !== 0) {
                        return branchCompare;
                    }
                    // Aynı branşta isime göre sırala
                    return a.name.toLocaleLowerCase('tr').localeCompare(b.name.toLocaleLowerCase('tr'), 'tr');
                });

                sortedTeachers.forEach(function(teacher) {
                    var option = document.createElement("option");
                    option.value = teacher.id;
                    option.setAttribute("data-branch", teacher.branch);
                    option.textContent = teacher.name + " " + teacher.surname + " (" + teacher.branch + ")";
                    select.appendChild(option);
                });
            } catch (error) {
                showError("Ogretmenler yuklenirken hata!");
            }
        }

        async function updateTeacherSchedule() {
            const teacherId = document.getElementById("classLessonTeacher").value;
            if (!teacherId) {
                document.getElementById("teacherBranchInfo").innerHTML = "";
                document.getElementById("daySelectionGroup").style.display = "none";
                document.getElementById("timeSelectionGroup").style.display = "none";
                document.getElementById("weekSelectionGroup").style.display = "none";
                document.getElementById("saveClassLessonBtn").style.display = "none";
                return;
            }

            const select = document.getElementById("classLessonTeacher");
            const branch = select.options[select.selectedIndex].getAttribute("data-branch");
            document.getElementById("teacherBranchInfo").innerHTML =
                "<i class='fas fa-chalkboard-teacher'></i> Brans: <strong>" + branch + "</strong>";

            try {
                const response = await fetch("/get_teachers");
                const data = await response.json();
                var teacher = data.teachers.find(function(t) { return t.id == teacherId; });

                if (teacher && teacher.schedule) {
                    currentTeacherSchedule = teacher.schedule;
                    displayAvailableDays(teacher.schedule);
                } else {
                    showError("Ogretmen programi bulunamadi!");
                }
            } catch (error) {
                showError("Ogretmen programi yuklenirken hata!");
            }
        }

        function displayAvailableDays(schedule) {
            const daysDiv = document.getElementById("availableDays");
            daysDiv.innerHTML = "";

            const allDays = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"];
            var availableDays = schedule.map(function(s) { return s.day; });

            allDays.forEach(function(day) {
                if (availableDays.includes(day)) {
                    var btn = document.createElement("button");
                    btn.textContent = day;
                    btn.style.padding = "10px 20px";
                    btn.style.border = "2px solid #10b981";
                    btn.style.background = "white";
                    btn.style.color = "#10b981";
                    btn.style.borderRadius = "8px";
                    btn.style.cursor = "pointer";
                    btn.style.fontWeight = "600";
                    btn.style.transition = "all 0.3s";
                    btn.onclick = function() { selectDay(day, btn); };
                    daysDiv.appendChild(btn);
                }
            });

            document.getElementById("daySelectionGroup").style.display = "block";
            document.getElementById("timeSelectionGroup").style.display = "none";
            document.getElementById("weekSelectionGroup").style.display = "none";
            document.getElementById("saveClassLessonBtn").style.display = "none";
        }

        function selectDay(day, button) {
            var buttons = document.querySelectorAll("#availableDays button");
            buttons.forEach(function(btn) {
                btn.style.background = "white";
                btn.style.color = "#10b981";
            });

            button.style.background = "linear-gradient(135deg, #10b981 0%, #059669 100%)";
            button.style.color = "white";

            classLessonSelectedDay = day;
            displayAvailableTimes(day);
        }

        function displayAvailableTimes(day) {
            const timesDiv = document.getElementById("availableTimes");
            timesDiv.innerHTML = "";

            var daySchedule = currentTeacherSchedule.find(function(s) { return s.day === day; });
            if (!daySchedule || !daySchedule.lessons) {
                showError("Bu gun icin ders saati bulunamadi!");
                return;
            }

            daySchedule.lessons.forEach(function(lesson) {
                var timeSlot = lesson.start_time + "-" + lesson.end_time;

                var label = document.createElement("label");
                label.style.display = "flex";
                label.style.alignItems = "center";
                label.style.gap = "10px";
                label.style.padding = "12px";
                label.style.border = "2px solid #e5e7eb";
                label.style.borderRadius = "8px";
                label.style.cursor = "pointer";
                label.style.marginBottom = "10px";
                label.style.transition = "all 0.3s";

                // 🆕 CHECKBOX - Çoklu seçim için
                var checkbox = document.createElement("input");
                checkbox.type = "checkbox";
                checkbox.value = timeSlot;
                checkbox.style.width = "20px";
                checkbox.style.height = "20px";
                checkbox.style.cursor = "pointer";
                checkbox.style.accentColor = "#10b981";  // Yeşil checkbox
                checkbox.onclick = function() { selectTime(timeSlot, label, this); };

                var span = document.createElement("span");
                span.style.fontWeight = "600";
                span.style.fontSize = "1.05em";
                span.textContent = timeSlot;

                label.appendChild(checkbox);
                label.appendChild(span);
                timesDiv.appendChild(label);
            });

            document.getElementById("timeSelectionGroup").style.display = "block";
            document.getElementById("weekSelectionGroup").style.display = "none";
            document.getElementById("saveClassLessonBtn").style.display = "none";
        }

        function selectTime(timeSlot, label, checkbox) {
            // TOGGLE - Seçili ise kaldır, değilse ekle
            const index = classLessonSelectedTime.indexOf(timeSlot);

            if (index > -1) {
                // Seçili - KALDIR
                classLessonSelectedTime.splice(index, 1);
                label.style.borderColor = "#e5e7eb";
                label.style.background = "white";
                checkbox.checked = false;  // Checkbox'ı işaretsiz yap
            } else {
                // Seçili değil - EKLE
                classLessonSelectedTime.push(timeSlot);
                label.style.borderColor = "#10b981";
                label.style.background = "#f0fdf4";
                checkbox.checked = true;  // Checkbox'ı işaretle
            }

            // En az bir saat seçilmişse göster
            if (classLessonSelectedTime.length > 0) {
                document.getElementById("weekSelectionGroup").style.display = "block";
                document.getElementById("saveClassLessonBtn").style.display = "block";
            } else {
                document.getElementById("weekSelectionGroup").style.display = "none";
                document.getElementById("saveClassLessonBtn").style.display = "none";
            }
        }

        function toggleWeekSelection() {
            const allWeeks = document.getElementById("classLessonAllWeeks").checked;
            var checkboxes = document.querySelectorAll(".week-checkbox");

            checkboxes.forEach(function(cb) {
                cb.checked = allWeeks;
                cb.disabled = allWeeks;
            });
        }

        async function saveClassLesson() {
            const className = document.getElementById("classLessonClass").value;
            const teacherId = document.getElementById("classLessonTeacher").value;
            const allWeeks = document.getElementById("classLessonAllWeeks").checked;

            // Array kontrolü
            if (!className || !teacherId || !classLessonSelectedDay || classLessonSelectedTime.length === 0) {
                showError("Lutfen tum alanlari doldurun!");
                return;
            }

            var weeks = "all";
            if (!allWeeks) {
                var checkboxes = document.querySelectorAll(".week-checkbox:checked");
                var selectedWeeks = [];
                checkboxes.forEach(function(cb) {
                    selectedWeeks.push(cb.value);
                });
                if (selectedWeeks.length === 0) {
                    showError("Lutfen en az bir hafta secin!");
                    return;
                }
                weeks = selectedWeeks.join(",");
            }

            // DÜZENLEME MODU
            if (editingClassLessonId) {
                if (classLessonSelectedTime.length !== 1) {
                    showError("Duzenleme modunda sadece bir saat secebilirsiniz!");
                    return;
                }

                var times = classLessonSelectedTime[0].split("-");
                var data = {
                    lesson_id: editingClassLessonId,
                    class_name: className,
                    teacher_id: parseInt(teacherId),
                    day: classLessonSelectedDay,
                    start_time: times[0],
                    end_time: times[1],
                    weeks: weeks
                };

                await saveOrUpdateSingleLesson(data, "/update_class_lesson", "Sinif dersi guncellendi!");
                return;
            }

            // YENİ KAYIT MODU - Çoklu saat
            var successCount = 0;
            var errorMessages = [];

            for (let i = 0; i < classLessonSelectedTime.length; i++) {
                var timeSlot = classLessonSelectedTime[i];
                var times = timeSlot.split("-");

                var data = {
                    class_name: className,
                    teacher_id: parseInt(teacherId),
                    day: classLessonSelectedDay,
                    start_time: times[0],
                    end_time: times[1],
                    weeks: weeks
                };

                try {
                    const response = await fetch("/save_class_lesson", {
                        method: "POST",
                        headers: { "Content-Type": "application/json" },
                        body: JSON.stringify(data)
                    });

                    const result = await response.json();

                    // 🆕 ÖĞRETMEN ÇAKIŞMASI UYARISI KONTROLÜ (EN ÖNCELİKLİ)
                    if (result.teacher_conflict_warning) {
                        // Döngüyü durdur ve modal göster
                        showTeacherConflictModal(result.message, data, "/save_class_lesson");
                        return; // Fonksiyondan çık
                    }

                    // 🆕 GRUP DERSİ SEÇENEĞİ KONTROLÜ
                    if (result.group_option) {
                        // Döngüyü durdur ve modal göster
                        showGroupLessonConfirm(result, data, "/save_class_lesson");
                        return; // Fonksiyondan çık
                    }

                    // 🆕 ÖĞRENCİ UYARILARI KONTROLÜ
                    if (result.warnings && result.warnings.length > 0) {
                        // Döngüyü durdur ve modal göster
                        showWarningsModal(result.warnings, data, "/save_class_lesson");
                        return; // Fonksiyondan çık
                    }

                    if (result.success) {
                        successCount++;
                    } else if (result.error) {
                        errorMessages.push(timeSlot + ": " + result.error);
                    }
                } catch (error) {
                    errorMessages.push(timeSlot + ": Kayit hatasi!");
                }
            }

            if (successCount > 0) {
                showSuccess(successCount + " ders saati basariyla kaydedildi!");
                closeClassLessonModal();
                loadClassLessons();

                // 🆕 Çakışma badge'ini güncelle
                if (globalScheduleData) {
                    setTimeout(() => {
                        checkConflictsInBackground();
                    }, 500);
                }
            }

            if (errorMessages.length > 0) {
                showError("Hatalar:\\n" + errorMessages.join("\\n"));
            }
        }

        // saveOrUpdateSingleLesson fonksiyonu
        async function saveOrUpdateSingleLesson(data, endpoint, successMessage) {
            try {
                const response = await fetch(endpoint, {
                    method: "POST",
                    headers: { "Content-Type": "application/json" },
                    body: JSON.stringify(data)
                });

                const result = await response.json();

                // 🆕 ÖĞRETMEN ÇAKIŞMASI UYARISI KONTROLÜ (EN ÖNCELİKLİ)
                if (result.teacher_conflict_warning) {
                    showTeacherConflictModal(result.message, data, endpoint);
                    return;
                }

                // 🆕 GRUP DERSİ SEÇENEĞİ
                if (result.group_option) {
                    // Özel modal göster
                    showGroupLessonConfirm(result, data, endpoint);
                    return;
                }

                // 🆕 ÖĞRENCİ UYARILARI
                if (result.warnings && result.warnings.length > 0) {
                    // Yeni modal göster
                    showWarningsModal(result.warnings, data, endpoint);
                    return;
                }

                if (result.error) {
                    showError(result.error);
                    return;
                }

                if (result.success) {
                    showSuccess(successMessage);
                    closeClassLessonModal();
                    loadClassLessons();

                    // 🆕 Çakışma badge'ini güncelle
                    if (globalScheduleData) {
                        setTimeout(() => {
                            checkConflictsInBackground();
                        }, 500);
                    }
                } else {
                    showError("Kayıt başarısız!");
                }
            } catch (error) {
                showError("Kayit sirasinda hata!");
            }
        }

        async function loadClassLessons() {
            try {
                const response = await fetch("/get_class_lessons");
                const lessons = await response.json();

                document.getElementById("classLessonCount").textContent = lessons.length;

                var listDiv = document.getElementById("classLessonsList");

                if (lessons.length === 0) {
                    var emptyDiv = document.createElement("div");
                    emptyDiv.style.textAlign = "center";
                    emptyDiv.style.padding = "40px";
                    emptyDiv.style.color = "#999";

                    var icon = document.createElement("i");
                    icon.className = "fas fa-users-class";
                    icon.style.fontSize = "3em";
                    icon.style.marginBottom = "15px";
                    icon.style.opacity = "0.3";

                    var p1 = document.createElement("p");
                    p1.style.fontSize = "1.1em";
                    p1.textContent = "Henuz sinif dersi eklenmedi.";

                    var p2 = document.createElement("p");
                    p2.style.fontSize = "0.9em";
                    p2.style.marginTop = "8px";
                    p2.innerHTML = 'Sinif dersi eklemek icin yukaridaki <strong>"Sinif Dersi Ata"</strong> butonuna tiklayin.';

                    emptyDiv.appendChild(icon);
                    emptyDiv.appendChild(p1);
                    emptyDiv.appendChild(p2);
                    listDiv.innerHTML = "";
                    listDiv.appendChild(emptyDiv);
                } else {
                    // 🆕 GRUPLAMA VE SIRALAMA

                    // 1. Sınıflara göre grupla
                    var groupedByClass = {};
                    lessons.forEach(function(lesson) {
                        if (!groupedByClass[lesson.class_name]) {
                            groupedByClass[lesson.class_name] = [];
                        }
                        groupedByClass[lesson.class_name].push(lesson);
                    });

                    // 2. Sınıf isimlerini alfabetik sırala
                    var classNames = Object.keys(groupedByClass).sort(function(a, b) {
                        return a.toLocaleLowerCase('tr').localeCompare(b.toLocaleLowerCase('tr'), 'tr');
                    });

                    // Gün sırası
                    var dayOrder = {
                        'Pazartesi': 1, 'Salı': 2, 'Çarşamba': 3, 'Perşembe': 4,
                        'Cuma': 5, 'Cumartesi': 6, 'Pazar': 7
                    };

                    // 3. Her sınıf içindeki dersleri gün ve saat sırasına göre sırala
                    classNames.forEach(function(className) {
                        groupedByClass[className].sort(function(a, b) {
                            // Önce güne göre sırala
                            var dayDiff = dayOrder[a.day] - dayOrder[b.day];
                            if (dayDiff !== 0) return dayDiff;

                            // Aynı gün ise saate göre sırala
                            return a.start_time.localeCompare(b.start_time);
                        });
                    });

                    // 4. UI'da grupları göster
                    listDiv.innerHTML = "";

                    classNames.forEach(function(className) {
                        // 🆕 UNIQUE ID
                        var safeClassName = className.replace(/[^a-zA-Z0-9]/g, '_');
                        var groupId = 'classGroup_' + safeClassName;

                        // 🆕 AKORDIYON GRUP BAŞLIĞI
                        var groupHeader = document.createElement("div");
                        groupHeader.style.background = "linear-gradient(135deg, #10b981 0%, #059669 100%)";
                        groupHeader.style.color = "white";
                        groupHeader.style.padding = "15px 20px";
                        groupHeader.style.borderRadius = "10px";
                        groupHeader.style.marginBottom = "5px";
                        groupHeader.style.marginTop = "15px";
                        groupHeader.style.fontWeight = "bold";
                        groupHeader.style.fontSize = "1.2em";
                        groupHeader.style.boxShadow = "0 2px 8px rgba(16, 185, 129, 0.3)";
                        groupHeader.style.cursor = "pointer";
                        groupHeader.style.transition = "all 0.3s";
                        groupHeader.style.display = "flex";
                        groupHeader.style.justifyContent = "space-between";
                        groupHeader.style.alignItems = "center";
                        groupHeader.style.userSelect = "none"; // 🆕 TEXT SEÇİMİ ENGELLE
                        groupHeader.style.webkitUserSelect = "none"; // Safari için
                        groupHeader.style.mozUserSelect = "none"; // Firefox için
                        groupHeader.style.msUserSelect = "none"; // IE için

                        groupHeader.innerHTML = '<span><i class="fas fa-users-class"></i> ' + className +
                            ' <span style="font-size: 0.8em; font-weight: normal; opacity: 0.9;">(' +
                            groupedByClass[className].length + ' ders)</span></span>' +
                            '<i class="fas fa-chevron-down" id="arrow_' + groupId + '" style="transition: transform 0.3s;"></i>';

                        // 🆕 HOVER EFEKT
                        groupHeader.onmouseover = function() {
                            this.style.background = "linear-gradient(135deg, #059669 0%, #047857 100%)";
                        };
                        groupHeader.onmouseout = function() {
                            this.style.background = "linear-gradient(135deg, #10b981 0%, #059669 100%)";
                        };

                        // 🆕 TIKLAMA İLE AÇILIP KAPANMA
                        groupHeader.onclick = function() {
                            toggleClassGroup(groupId);
                        };

                        listDiv.appendChild(groupHeader);

                        // 🆕 GRUP İÇERİĞİ (DERSLER)
                        var groupContent = document.createElement("div");
                        groupContent.id = groupId;

                        // 🆕 localStorage'dan durumu kontrol et
                        var savedState = localStorage.getItem('classGroup_' + groupId);
                        if (savedState === 'closed') {
                            groupContent.style.display = "none";
                            var arrowElem = document.getElementById('arrow_' + groupId);
                            if (arrowElem) {
                                arrowElem.style.transform = "rotate(-90deg)";
                            }
                        } else {
                            // Varsayılan açık
                            groupContent.style.display = "block";
                        }

                        groupContent.style.overflow = "hidden";
                        groupContent.style.transition = "all 0.3s ease";

                        // GRUP İÇİNDEKİ DERSLER
                        groupedByClass[className].forEach(function(lesson) {
                        var weekText = lesson.weeks === "all" ? "Her hafta" : "Hafta " + lesson.weeks;

                        var card = document.createElement("div");
                        card.style.border = "2px solid #10b981";
                        card.style.borderRadius = "12px";
                        card.style.padding = "15px";
                        card.style.marginBottom = "12px";
                        card.style.marginLeft = "15px";  // Grup altında girintili
                        card.style.background = "linear-gradient(135deg, #f0fdf4 0%, #d1fae5 100%)";

                        var flexDiv = document.createElement("div");
                        flexDiv.style.display = "flex";
                        flexDiv.style.justifyContent = "space-between";
                        flexDiv.style.alignItems = "start";

                        var contentDiv = document.createElement("div");
                        contentDiv.style.flex = "1";

                        // 🆕 GÜN VE SAAT (başlık olarak)
                        var timeHeader = document.createElement("div");
                        timeHeader.style.color = "#065f46";
                        timeHeader.style.marginBottom = "8px";
                        timeHeader.style.fontSize = "1.1em";
                        timeHeader.style.fontWeight = "bold";
                        timeHeader.innerHTML = '<i class="fas fa-calendar"></i> ' + lesson.day + ' ' + lesson.start_time + '-' + lesson.end_time;

                        var teacherDiv = document.createElement("div");
                        teacherDiv.style.fontSize = "0.95em";
                        teacherDiv.style.color = "#047857";
                        teacherDiv.style.marginBottom = "5px";
                        teacherDiv.innerHTML = '<i class="fas fa-chalkboard-teacher"></i> <strong>' + lesson.teacher_name + '</strong> (' + lesson.teacher_branch + ')';

                        var infoDiv = document.createElement("div");
                        infoDiv.style.fontSize = "0.9em";
                        infoDiv.style.color = "#059669";
                        infoDiv.style.marginTop = "5px";

                        // 🆕 GRUP DERSİ BİLGİSİ
                        if (lesson.is_group === 1) {
                            // Aynı gün/saat/öğretmende başka sınıfları bul
                            var groupClasses = lessons.filter(function(l) {
                                return l.teacher_id === lesson.teacher_id &&
                                       l.day === lesson.day &&
                                       l.start_time === lesson.start_time &&
                                       l.end_time === lesson.end_time &&
                                       l.class_name !== lesson.class_name;
                            });

                            var allClasses = [lesson.class_name];
                            var totalStudents = lesson.student_count;

                            groupClasses.forEach(function(gc) {
                                allClasses.push(gc.class_name);
                                totalStudents += gc.student_count;
                            });

                            infoDiv.innerHTML = '<i class="fas fa-link" style="color: #f59e0b;"></i> <strong style="color: #f59e0b;">GRUP DERSİ</strong><br>' +
                                '<span style="font-size: 0.85em;">Katılımcı Sınıflar: ' + allClasses.join(', ') + '</span><br>' +
                                '<i class="fas fa-users"></i> Toplam ' + totalStudents + ' öğrenci | <i class="fas fa-repeat"></i> ' + weekText;
                        } else {
                            infoDiv.innerHTML = '<i class="fas fa-users"></i> ' + lesson.student_count + ' öğrenci | <i class="fas fa-repeat"></i> ' + weekText;
                        }

                        // 🆕 BUTON CONTAINER
                        var buttonDiv = document.createElement("div");
                        buttonDiv.style.display = "flex";
                        buttonDiv.style.gap = "10px";
                        buttonDiv.style.flexDirection = "column";

                        // 🆕 DÜZENLE BUTONU
                        var editBtn = document.createElement("button");
                        editBtn.style.background = "#3b82f6";
                        editBtn.style.color = "white";
                        editBtn.style.border = "none";
                        editBtn.style.padding = "8px 16px";
                        editBtn.style.borderRadius = "8px";
                        editBtn.style.cursor = "pointer";
                        editBtn.style.fontWeight = "600";
                        editBtn.style.transition = "all 0.3s";
                        editBtn.innerHTML = '<i class="fas fa-edit"></i> Düzenle';
                        editBtn.onmouseover = function() { this.style.background = "#2563eb"; };
                        editBtn.onmouseout = function() { this.style.background = "#3b82f6"; };
                        editBtn.onclick = function() { editClassLesson(lesson); };

                        // SİL BUTONU
                        var deleteBtn = document.createElement("button");
                        deleteBtn.style.background = "#ef4444";
                        deleteBtn.style.color = "white";
                        deleteBtn.style.border = "none";
                        deleteBtn.style.padding = "8px 16px";
                        deleteBtn.style.borderRadius = "8px";
                        deleteBtn.style.cursor = "pointer";
                        deleteBtn.style.fontWeight = "600";
                        deleteBtn.style.transition = "all 0.3s";
                        deleteBtn.innerHTML = '<i class="fas fa-trash"></i> Sil';
                        deleteBtn.onmouseover = function() { this.style.background = "#dc2626"; };
                        deleteBtn.onmouseout = function() { this.style.background = "#ef4444"; };
                        deleteBtn.onclick = function() { deleteClassLesson(lesson.id); };

                        buttonDiv.appendChild(editBtn);
                        buttonDiv.appendChild(deleteBtn);

                        contentDiv.appendChild(timeHeader);
                        contentDiv.appendChild(teacherDiv);
                        contentDiv.appendChild(infoDiv);

                        flexDiv.appendChild(contentDiv);
                        flexDiv.appendChild(buttonDiv);

                        card.appendChild(flexDiv);
                        groupContent.appendChild(card);  // 🆕 Kartı grup içeriğine ekle
                        });  // lesson forEach sonu

                        // 🆕 Grup içeriğini listeye ekle
                        listDiv.appendChild(groupContent);

                        // 🆕 Arrow durumunu localStorage'a göre güncelle
                        var savedState = localStorage.getItem('classGroup_' + groupId);
                        var arrow = document.getElementById('arrow_' + groupId);
                        if (savedState === 'closed' && arrow) {
                            arrow.style.transform = "rotate(-90deg)";
                        }
                    });  // className forEach sonu
                }
            } catch (error) {
                showError("Sinif dersleri yuklenirken hata!");
            }
        }

        // 🆕 AKORDIYON TOGGLE FONKSİYONU
        function toggleClassGroup(groupId) {
            var content = document.getElementById(groupId);
            var arrow = document.getElementById('arrow_' + groupId);

            if (content.style.display === "none") {
                content.style.display = "block";
                arrow.style.transform = "rotate(0deg)";
                // 🆕 Açık durumu kaydet
                localStorage.setItem('classGroup_' + groupId, 'open');
            } else {
                content.style.display = "none";
                arrow.style.transform = "rotate(-90deg)";
                // 🆕 Kapalı durumu kaydet
                localStorage.setItem('classGroup_' + groupId, 'closed');
            }
        }

        // 🆕 ANA MENÜ AKORDIYON TOGGLE
        function toggleMainClassLessons() {
            var content = document.getElementById("classLessonsContent");
            var arrow = document.getElementById("mainClassLessonsArrow");

            content.classList.toggle('open');
            arrow.classList.toggle('open');
            // 🆕 localStorage kaldırıldı - her zaman kapalı başlayacak
        }

        // 🆕 GRUP DERSİ ONAY MODALI FONKSİYONLARI
        let pendingGroupLessonData = null;
        let pendingGroupLessonEndpoint = null;

        function showGroupLessonConfirm(result, data, endpoint) {
            // Veriyi sakla
            pendingGroupLessonData = data;
            pendingGroupLessonEndpoint = endpoint;

            // Modal içeriğini doldur
            document.getElementById("groupLessonMessage").innerHTML = result.error;

            // 🎯 Tüm sınıfları göster
            const allClasses = result.all_classes || [result.existing_class, data.class_name];
            document.getElementById("groupClassList").textContent = allClasses.join(", ");

            // Modalı göster
            document.getElementById("groupLessonConfirmModal").style.display = "flex";
        }

        function cancelGroupLesson() {
            // Modalı kapat
            document.getElementById("groupLessonConfirmModal").style.display = "none";
            pendingGroupLessonData = null;
            pendingGroupLessonEndpoint = null;
        }

        async function confirmGroupLesson() {
            if (!pendingGroupLessonData || !pendingGroupLessonEndpoint) {
                return;
            }

            // Modalı kapat
            document.getElementById("groupLessonConfirmModal").style.display = "none";

            // Grup dersi olarak kaydet
            pendingGroupLessonData.force_group = true;

            try {
                const groupResponse = await fetch(pendingGroupLessonEndpoint, {
                    method: "POST",
                    headers: { "Content-Type": "application/json" },
                    body: JSON.stringify(pendingGroupLessonData)
                });
                const groupResult = await groupResponse.json();

                // 🆕 ÖĞRENCİ UYARILARI KONTROLÜ
                if (groupResult.warnings && groupResult.warnings.length > 0) {
                    console.log('⚠️ Grup dersi kaydında öğrenci uyarıları bulundu, warnings modalı açılıyor');
                    // Öğrenci uyarıları modalını göster
                    showWarningsModal(groupResult.warnings, pendingGroupLessonData, pendingGroupLessonEndpoint);
                    // Temizle
                    pendingGroupLessonData = null;
                    pendingGroupLessonEndpoint = null;
                    return;
                }

                if (groupResult.success) {
                    showSuccess("Grup dersi olarak kaydedildi!");
                    closeClassLessonModal();
                    loadClassLessons();

                    // 🆕 Çakışma badge'ini güncelle
                    if (globalScheduleData) {
                        setTimeout(() => {
                            checkConflictsInBackground();
                        }, 500);
                    }
                } else {
                    showError(groupResult.error || "Kayıt başarısız!");
                }
            } catch (error) {
                showError("Kayıt sırasında hata!");
            }

            // Temizle
            pendingGroupLessonData = null;
            pendingGroupLessonEndpoint = null;
        }

        // 🆕 ÖĞRENCİ UYARILARI MODALI FONKSİYONLARI
        let pendingWarningsData = null;
        let pendingWarningsEndpoint = null;

        function showWarningsModal(warnings, data, endpoint) {
            // Veriyi sakla
            pendingWarningsData = data;
            pendingWarningsEndpoint = endpoint;

            // 🔍 DEBUG
            console.log('🔍 showWarningsModal çağrıldı');
            console.log('📋 Gelen warnings:', warnings);
            console.log('📊 Warnings sayısı:', warnings.length);

            // Uyarıları grupla
            const restrictions = warnings.filter(w => w.type === 'day_restriction' || w.type === 'time_restriction');
            const blocks = warnings.filter(w => w.type === 'teacher_blocked');

            console.log('📋 Kısıtlamalar:', restrictions.length, restrictions);
            console.log('🚫 Engellemeler:', blocks.length, blocks);

            // HTML oluştur
            let html = '';

            // Kısıtlamalar
            if (restrictions.length > 0) {
                html += '<div style="background: white; border-radius: 8px; padding: 12px; margin-bottom: 12px;">';
                html += '<h4 style="color: #3b82f6; margin: 0 0 10px 0; font-size: 1em;"><i class="fas fa-calendar-times"></i> Zaman Kısıtlamaları (' + restrictions.length + ')</h4>';
                html += '<ul style="margin: 0; padding-left: 20px;">';
                restrictions.forEach(r => {
                    html += '<li style="color: #1e40af; margin: 5px 0;"><strong>' + r.student + ':</strong> ' + r.message + '</li>';
                });
                html += '</ul></div>';
            }

            // Engellemeler
            if (blocks.length > 0) {
                html += '<div style="background: white; border-radius: 8px; padding: 12px;">';
                html += '<h4 style="color: #ef4444; margin: 0 0 10px 0; font-size: 1em;"><i class="fas fa-user-times"></i> Öğretmen Engellemeleri (' + blocks.length + ')</h4>';
                html += '<ul style="margin: 0; padding-left: 20px;">';
                blocks.forEach(b => {
                    html += '<li style="color: #991b1b; margin: 5px 0;"><strong>' + b.student + ':</strong> ' + b.message + '</li>';
                });
                html += '</ul></div>';
            }

            document.getElementById("warningsList").innerHTML = html;

            // Modalı göster
            document.getElementById("studentWarningsModal").style.display = "flex";
        }

        function cancelWarnings() {
            // Modalı kapat
            document.getElementById("studentWarningsModal").style.display = "none";
            pendingWarningsData = null;
            pendingWarningsEndpoint = null;
        }

        async function confirmWithWarnings() {
            if (!pendingWarningsData || !pendingWarningsEndpoint) {
                return;
            }

            // Modalı kapat
            document.getElementById("studentWarningsModal").style.display = "none";

            // force=true ile tekrar gönder
            pendingWarningsData.force = true;

            try {
                const response = await fetch(pendingWarningsEndpoint, {
                    method: "POST",
                    headers: { "Content-Type": "application/json" },
                    body: JSON.stringify(pendingWarningsData)
                });
                const result = await response.json();

                if (result.success) {
                    showSuccess("Ders kaydedildi! (Uyarılar göz ardı edildi)");
                    closeClassLessonModal();
                    loadClassLessons();

                    // 🆕 Çakışma badge'ini güncelle
                    if (globalScheduleData) {
                        setTimeout(() => {
                            checkConflictsInBackground();
                        }, 500);
                    }
                } else {
                    showError(result.error || "Kayıt başarısız!");
                }
            } catch (error) {
                showError("Kayıt sırasında hata!");
            }

            // Temizle
            pendingWarningsData = null;
            pendingWarningsEndpoint = null;
        }

        // 🆕 ÖĞRETMEN ÇAKIŞMASI ONAY FONKSİYONLARI
        let pendingTeacherConflictData = null;
        let pendingTeacherConflictEndpoint = null;

        function showTeacherConflictModal(message, data, endpoint) {
            // Veriyi sakla
            pendingTeacherConflictData = data;
            pendingTeacherConflictEndpoint = endpoint;

            console.log('🔍 showTeacherConflictModal çağrıldı');
            console.log('📋 Data:', data);
            console.log('📊 Endpoint:', endpoint);

            // Mesajı göster
            document.getElementById("teacherConflictMessage").textContent = message;

            // Modalı aç
            document.getElementById("teacherConflictModal").style.display = "flex";
        }

        function cancelTeacherConflict() {
            // Modalı kapat
            document.getElementById("teacherConflictModal").style.display = "none";
            pendingTeacherConflictData = null;
            pendingTeacherConflictEndpoint = null;
        }

        async function confirmTeacherConflict() {
            if (!pendingTeacherConflictData || !pendingTeacherConflictEndpoint) {
                return;
            }

            // Modalı kapat
            document.getElementById("teacherConflictModal").style.display = "none";

            // force_teacher_conflict=true ile tekrar gönder
            pendingTeacherConflictData.force_teacher_conflict = true;

            console.log('✅ Öğretmen çakışması onaylandı, tekrar gönderiliyor:', pendingTeacherConflictData);

            try {
                const response = await fetch(pendingTeacherConflictEndpoint, {
                    method: "POST",
                    headers: { "Content-Type": "application/json" },
                    body: JSON.stringify(pendingTeacherConflictData)
                });
                const result = await response.json();

                console.log('📥 İkinci yanıt:', result);

                // 🆕 ÖĞRENCİ UYARILARI VAR MI KONTROL ET
                if (result.warnings && result.warnings.length > 0) {
                    console.log('⚠️ Öğrenci uyarıları bulundu, warnings modalı açılıyor');
                    // Öğrenci uyarıları modalını göster
                    showWarningsModal(result.warnings, pendingTeacherConflictData, pendingTeacherConflictEndpoint);
                    // Temizle
                    pendingTeacherConflictData = null;
                    pendingTeacherConflictEndpoint = null;
                    return;
                }

                // 🆕 GRUP DERSİ SEÇENEĞİ VAR MI KONTROL ET
                if (result.group_option) {
                    console.log('🔗 Grup dersi seçeneği bulundu, grup modalı açılıyor');
                    showGroupLessonConfirm(result, pendingTeacherConflictData, pendingTeacherConflictEndpoint);
                    // Temizle
                    pendingTeacherConflictData = null;
                    pendingTeacherConflictEndpoint = null;
                    return;
                }

                if (result.success) {
                    showSuccess("Sınıf dersi başarıyla kaydedildi!");
                    closeClassLessonModal();
                    loadClassLessons();

                    // Çakışma badge'ini güncelle
                    if (globalScheduleData) {
                        setTimeout(() => {
                            checkConflictsInBackground();
                        }, 500);
                    }
                } else if (result.error) {
                    showError(result.error);
                }
            } catch (error) {
                console.error('❌ Hata:', error);
                showError("Kayıt sırasında hata!");
            }

            // Temizle
            pendingTeacherConflictData = null;
            pendingTeacherConflictEndpoint = null;
        }

        async function deleteClassLesson(lessonId) {
            if (!confirm("Bu sinif dersini silmek istediginizden emin misiniz?")) {
                return;
            }

            try {
                const response = await fetch("/delete_class_lesson/" + lessonId, {
                    method: "DELETE"
                });

                const result = await response.json();

                if (result.success) {
                    showSuccess("Sinif dersi silindi!");
                    loadClassLessons();

                    // 🆕 Çakışma badge'ini güncelle
                    if (globalScheduleData) {
                        setTimeout(() => {
                            checkConflictsInBackground();
                        }, 500);
                    }
                } else {
                    showError("Silme islemi basarisiz!");
                }
            } catch (error) {
                showError("Silme sirasinda hata!");
            }
        }

        // 🆕 GLOBAL DEĞİŞKEN - Düzenlenen ders ID'si
        let editingClassLessonId = null;

        // 🆕 SINIF DERSİ DÜZENLE FONKSİYONU
        async function editClassLesson(lesson) {
            console.log("Düzenlenen ders:", lesson);

            // 🆕 ÖNCELİKLE FORMU TEMİZLE
            resetClassLessonForm();

            // Düzenleme modunu aktif et
            editingClassLessonId = lesson.id;

            // Modalı aç
            document.getElementById("classLessonModal").style.display = "block";

            // 🆕 LİSTELERİ YÜKLE
            await loadClasses();
            await loadTeachersForClassLesson();

            // Modal başlığını değiştir
            document.querySelector("#classLessonModal h2").innerHTML = '✏️ Sınıf Dersini Düzenle';
            document.getElementById("saveClassLessonBtn").innerHTML = '<i class="fas fa-save"></i> Güncelle';

            // Sınıf seç
            document.getElementById("classLessonClass").value = lesson.class_name;
            await updateClassStudentCount();

            // Öğretmen seç
            document.getElementById("classLessonTeacher").value = lesson.teacher_id;
            await updateTeacherSchedule();

            // Tüm grupları göster
            document.getElementById("daySelectionGroup").style.display = "block";
            document.getElementById("timeSelectionGroup").style.display = "block";
            document.getElementById("weekSelectionGroup").style.display = "block";
            document.getElementById("saveClassLessonBtn").style.display = "block";

            // 🆕 Gün seç - DİREKT OLARAK (setTimeout yok!)
            classLessonSelectedDay = lesson.day;

            // Günleri işaretle
            const dayButtons = document.querySelectorAll("#availableDays button");
            dayButtons.forEach(btn => {
                if (btn.textContent === lesson.day) {
                    btn.style.background = "linear-gradient(135deg, #10b981 0%, #059669 100%)";
                    btn.style.color = "white";
                    btn.style.border = "2px solid #10b981";
                }
            });

            // 🆕 SAATLERİ GÖSTER - selectDay gibi davran!
            displayAvailableTimes(lesson.day);

            // 🆕 Saat seç - DİREKT OLARAK (setTimeout yok!)
            const timeSlot = lesson.start_time + "-" + lesson.end_time;
            classLessonSelectedTime = [timeSlot];

            // Saatleri işaretle
            const timeLabels = document.querySelectorAll("#availableTimes label");
            timeLabels.forEach(lbl => {
                const checkbox = lbl.querySelector("input[type='checkbox']");
                if (checkbox && checkbox.value === timeSlot) {
                    lbl.style.borderColor = "#10b981";
                    lbl.style.background = "#f0fdf4";
                    checkbox.checked = true;
                }
            });

            // Hafta seç
            if (lesson.weeks === "all") {
                document.getElementById("classLessonAllWeeks").checked = true;
                document.getElementById("individualWeeks").style.display = "none";
            } else {
                document.getElementById("classLessonAllWeeks").checked = false;
                document.getElementById("individualWeeks").style.display = "flex";

                const weeks = lesson.weeks.split(",");
                document.querySelectorAll(".week-checkbox").forEach(cb => {
                    cb.checked = weeks.includes(cb.value);
                });
            }
        }

        document.addEventListener("DOMContentLoaded", function() {
            loadClassLessons();
        });

        // ═══════════════════════════════════════════════════════════════
        // 🆕 YENİ İHLAL KONTROL PANELİ V2 FONKSİYONLARI
        // ═══════════════════════════════════════════════════════════════

        function openConflictPanelV2() {
            const modal = document.getElementById('conflictPanelV2');
            const loading = document.getElementById('panelV2Loading');
            const content = document.getElementById('panelV2Content');
            const empty = document.getElementById('panelV2Empty');
            const cards = document.getElementById('panelV2Cards');

            // Modal'ı aç
            modal.style.display = 'block';
            loading.style.display = 'block';
            content.style.display = 'none';

            // API çağrısı
            fetch('/check_conflicts_v2', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' }
            })
            .then(response => response.json())
            .then(data => {
                loading.style.display = 'none';
                content.style.display = 'block';

                // 🆕 AYKIRI SWAP'LARI SESSIONSTORAGE'DAN YÜKLE
                let aykiriSwapCards = [];
                try {
                    const stored = sessionStorage.getItem('aykiriSwapViolations');
                    if (stored) {
                        const violations = JSON.parse(stored);
                        aykiriSwapCards = violations.map((violation, index) => ({
                            type: 'aykiri_swap',
                            branch: '⚠️ Aykırı Swap',
                            teacher: '-',
                            day: violation.day,
                            time: violation.time,
                            week: currentWeekView || 1,
                            student_count: violation.swappedStudents.length,
                            class: '-',
                            students: violation.swappedStudents,
                            issues: violation.conflictingStudents.map(student => ({
                                type: 'aykiri_swap',
                                student: student,
                                detail: 'Bu öğrencinin aynı saatte başka dersi var (Manuel onaylanmış swap)'
                            })),
                            severity: 'critical',
                            approved: false,
                            borderColor: violation.borderColor,
                            timestamp: violation.timestamp
                        }));
                    }
                } catch (e) {
                    console.error('Aykırı swap ihlalleri okunamadı:', e);
                }

                // API'den gelen kartlar + Aykırı swap kartları
                const allCards = [...aykiriSwapCards, ...(data.cards || [])];

                // İstatistikleri güncelle (aykırı swap'ları da dahil et)
                const criticalCount = (data.summary.critical || 0) + aykiriSwapCards.length;
                document.getElementById('v2CriticalCount').textContent = criticalCount;
                document.getElementById('v2HighCount').textContent = data.summary.high || 0;
                document.getElementById('v2MediumCount').textContent = data.summary.medium || 0;
                document.getElementById('v2TotalCount').textContent = (data.summary.total || 0) + aykiriSwapCards.length;

                // Kartları render et
                if (allCards.length > 0) {
                    cards.style.display = 'flex';
                    empty.style.display = 'none';
                    renderConflictCardsV2(allCards);
                } else {
                    cards.style.display = 'none';
                    empty.style.display = 'block';
                }
            })
            .catch(error => {
                loading.style.display = 'none';
                console.error('İhlal kontrolü hatası:', error);
                alert('İhlaller kontrol edilirken bir hata oluştu!');
            });
        }

        function closeConflictPanelV2() {
            document.getElementById('conflictPanelV2').style.display = 'none';
        }

        function renderConflictCardsV2(cards) {
            const container = document.getElementById('panelV2Cards');
            container.innerHTML = '';

            cards.forEach((card, index) => {
                // Renk şeması
                let bgColor, borderColor, badgeColor, iconColor, icon;

                if (card.approved) {
                    // Yeşil - Onaylı
                    bgColor = 'linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%)';
                    borderColor = '#22c55e';
                    badgeColor = 'background: #22c55e; color: white;';
                    iconColor = '#16a34a';
                    icon = '✅';
                } else {
                    // Kırmızı - Onaysız
                    if (card.severity === 'critical') {
                        bgColor = 'linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%)';
                        borderColor = '#ef4444';
                        badgeColor = 'background: #ef4444; color: white;';
                        iconColor = '#dc2626';
                        icon = '🔴';
                    } else if (card.severity === 'high') {
                        bgColor = 'linear-gradient(135deg, #fffbeb 0%, #fef3c7 100%)';
                        borderColor = '#f59e0b';
                        badgeColor = 'background: #f59e0b; color: white;';
                        iconColor = '#d97706';
                        icon = '⚠️';
                    } else {
                        bgColor = 'linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%)';
                        borderColor = '#3b82f6';
                        badgeColor = 'background: #3b82f6; color: white;';
                        iconColor = '#2563eb';
                        icon = 'ℹ️';
                    }
                }

                // Kart türü başlığı
                let typeLabel = '';
                if (card.type === 'aykiri_swap') {
                    typeLabel = 'Aykırı Swap (Manuel Onaylı)';
                } else if (card.type === 'class_lesson') {
                    typeLabel = 'Sınıf Dersi';
                } else if (card.type === 'group_lesson') {
                    typeLabel = 'Grup Dersi';
                } else {
                    typeLabel = 'İhlal';
                }

                // Accordion ID
                const accordionId = `accordion_${index}`;

                // Kart HTML
                const cardHtml = `
                    <div style="background: ${bgColor}; border-left: 5px solid ${borderColor}; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1); transition: all 0.3s;">
                        <!-- Başlık -->
                        <div style="padding: 20px; cursor: pointer; user-select: none; -webkit-user-select: none; -moz-user-select: none; -ms-user-select: none;" onclick="toggleConflictCardV2('${accordionId}')">
                            <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 15px;">
                                <div style="display: flex; align-items: center; gap: 15px;">
                                    <div style="font-size: 2.5em;">${icon}</div>
                                    <div>
                                        <div style="font-size: 1.3em; font-weight: 700; color: ${iconColor}; margin-bottom: 5px;">
                                            ${card.branch}
                                        </div>
                                        <div style="display: inline-block; ${badgeColor} padding: 4px 12px; border-radius: 20px; font-size: 0.85em; font-weight: 600;">
                                            ${typeLabel}
                                        </div>
                                    </div>
                                </div>
                                <div style="font-size: 1.5em; color: ${iconColor}; transition: transform 0.3s;" id="${accordionId}_icon">▼</div>
                            </div>

                            <!-- Özet Bilgiler -->
                            <!-- 1. SATIR: Öğretmen, Gün, Saat, Hafta -->
                            <div style="display: flex; flex-wrap: wrap; gap: 15px; font-size: 0.9em; color: #4b5563; margin-bottom: 8px;">
                                <div><strong>👨‍🏫 Öğretmen:</strong> ${card.teacher}</div>
                                <div><strong>📅 Gün:</strong> ${card.day}</div>
                                <div><strong>🕐 Saat:</strong> ${card.time}</div>
                                <div><strong>📌 Hafta:</strong> ${card.week}</div>
                            </div>

                            <!-- 2. SATIR: Öğrenci + Sınıf/Katılımcı Sınıflar + İhlal Sayısı -->
                            <div style="display: flex; flex-wrap: wrap; gap: 15px; font-size: 0.9em; color: #4b5563;">
                                <div><strong>👥 Öğrenci:</strong> ${card.student_count} kişi</div>
                                <div><strong>🎓 ${card.type === 'group_lesson' ? 'Katılımcı Sınıflar' : 'Sınıf'}:</strong> ${card.class || '-'}</div>
                                <div><strong>⚠️ İhlaller:</strong> ${card.issues ? card.issues.length : 0}</div>
                            </div>
                        </div>

                        <!-- Detaylar (Accordion) -->
                        <div id="${accordionId}" style="max-height: 0; overflow: hidden; transition: max-height 0.3s ease;">
                            <div style="padding: 0 20px 20px 20px; border-top: 2px dashed ${borderColor}; margin-top: 0; padding-top: 20px;">
                                <!-- Öğrenci Listesi -->
                                <div style="margin-bottom: 20px;">
                                    <h4 style="color: ${iconColor}; margin-bottom: 10px; font-size: 1.1em;">
                                        👥 Katılımcı Öğrenciler (${card.students.length})
                                    </h4>
                                    <div style="display: flex; flex-wrap: wrap; gap: 8px;">
                                        ${card.students.map(student => `
                                            <span style="background: white; padding: 6px 12px; border-radius: 20px; font-size: 0.9em; border: 1px solid #e5e7eb; color: #374151;">
                                                ${student}
                                            </span>
                                        `).join('')}
                                    </div>
                                </div>

                                ${card.borderColor ? `
                                <!-- 🎨 Tabloda Renkli İşaretleme -->
                                <div style="margin-bottom: 20px; background: rgba(239, 68, 68, 0.05); padding: 15px; border-radius: 8px; border: 2px solid ${card.borderColor};">
                                    <h4 style="color: ${iconColor}; margin-bottom: 10px; font-size: 1.1em;">
                                        🎨 Tablodaki İşaretleme
                                    </h4>
                                    <div style="display: flex; align-items: center; gap: 12px;">
                                        <div style="width: 40px; height: 40px; border-radius: 8px; border: 4px solid ${card.borderColor}; background: white;"></div>
                                        <div style="color: #6b7280; font-size: 0.9em;">
                                            <strong>Bu ders programında bu renkle işaretlenmiştir</strong><br>
                                            Çakışan slotlar 3px, swap yapılan slotlar 4px border ile gösterilir.
                                        </div>
                                    </div>
                                </div>
                                ` : ''}

                                <!-- Sorunlar -->
                                ${card.issues && card.issues.length > 0 ? `
                                    <div>
                                        <h4 style="color: ${iconColor}; margin-bottom: 10px; font-size: 1.1em;">
                                            ⚠️ Tespit Edilen Sorunlar (${card.issues.length})
                                        </h4>
                                        <div style="display: flex; flex-direction: column; gap: 10px;">
                                            ${card.issues.map(issue => {
                                                let issueIcon, issueLabel, issueBorderColor;

                                                if (issue.type === 'aykiri_swap') {
                                                    issueIcon = '🔄';
                                                    issueLabel = 'Aykırı Swap';
                                                    issueBorderColor = '#ef4444';
                                                } else if (issue.type === 'class_lesson_creation') {
                                                    issueIcon = '📚';
                                                    issueLabel = 'Sınıf Dersi Oluşturuldu';
                                                    issueBorderColor = '#10b981';
                                                } else if (issue.type === 'group_lesson_participation') {
                                                    issueIcon = '👥';
                                                    issueLabel = 'Grup Dersine Katılım';
                                                    issueBorderColor = '#10b981';
                                                } else if (issue.type === 'restriction') {
                                                    issueIcon = '🚫';
                                                    issueLabel = 'Kısıtlama İhlali';
                                                    issueBorderColor = '#f59e0b';
                                                } else if (issue.type === 'teacher_block') {
                                                    issueIcon = '🔒';
                                                    issueLabel = 'Öğretmen Engeli';
                                                    issueBorderColor = '#f59e0b';
                                                } else {
                                                    issueIcon = '⚠️';
                                                    issueLabel = 'İhlal';
                                                    issueBorderColor = '#f59e0b';
                                                }

                                                return `
                                                    <div style="background: rgba(255, 255, 255, 0.7); padding: 12px; border-radius: 8px; border-left: 3px solid ${issueBorderColor}; display: flex; align-items: center; gap: 10px;">
                                                        <div style="font-size: 1.5em;">${issueIcon}</div>
                                                        <div style="flex: 1;">
                                                            <div style="font-weight: 600; color: #1f2937; margin-bottom: 2px;">
                                                                ${issue.student}
                                                            </div>
                                                            <div style="font-size: 0.9em; color: #6b7280;">
                                                                ${issueLabel}: ${issue.detail}
                                                            </div>
                                                        </div>
                                                    </div>
                                                `;
                                            }).join('')}
                                        </div>
                                    </div>
                                ` : `
                                    <div style="text-align: center; padding: 20px; color: #10b981; background: rgba(16, 185, 129, 0.1); border-radius: 8px;">
                                        <div style="font-size: 2em; margin-bottom: 10px;">✨</div>
                                        <div style="font-weight: 600;">Bu derste herhangi bir sorun tespit edilmedi.</div>
                                    </div>
                                `}
                            </div>
                        </div>
                    </div>
                `;

                container.innerHTML += cardHtml;
            });
        }

        function toggleConflictCardV2(accordionId) {
            const accordion = document.getElementById(accordionId);
            const icon = document.getElementById(accordionId + '_icon');

            if (accordion.style.maxHeight === '0px' || accordion.style.maxHeight === '') {
                accordion.style.maxHeight = accordion.scrollHeight + 'px';
                icon.style.transform = 'rotate(180deg)';
            } else {
                accordion.style.maxHeight = '0px';
                icon.style.transform = 'rotate(0deg)';
            }
        }

    </script>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

# Flask backend routes...
# (Kod çok uzun olduğu için sonraki mesajda Python backend kodunu gönderiyorum)
@app.route('/add_teacher', methods=['POST'])
def add_teacher():
    conn = get_db()
    cursor = conn.cursor()
    data = request.json

    # ✅ GÜNLERE GÖRE SIRALA
    day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
    schedule = sorted(data['schedule'], key=lambda x: day_order.index(x['day']))

    schedule_json = json.dumps(schedule, ensure_ascii=False)

    # 🆕 BLOKLAMALARI AL VE JSON'A ÇEVİR
    blocked_slots = data.get('blocked_slots', [])
    blocked_slots_json = json.dumps(blocked_slots, ensure_ascii=False)

    cursor.execute('''
        INSERT INTO teachers (name, surname, branch, schedule, blocked_slots)
        VALUES (?, ?, ?, ?, ?)
    ''', (data['name'], data['surname'], data['branch'], schedule_json, blocked_slots_json))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Öğretmen başarıyla eklendi!'})

@app.route('/update_teacher', methods=['POST'])
def update_teacher():
    conn = get_db()
    cursor = conn.cursor()
    data = request.json

    # ✅ GÜNLERE GÖRE SIRALA
    day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
    schedule = sorted(data['schedule'], key=lambda x: day_order.index(x['day']))

    schedule_json = json.dumps(schedule, ensure_ascii=False)

    # 🆕 BLOKLAMALARI AL VE JSON'A ÇEVİR
    blocked_slots = data.get('blocked_slots', [])
    blocked_slots_json = json.dumps(blocked_slots, ensure_ascii=False)

    cursor.execute('''
        UPDATE teachers
        SET name=?, surname=?, branch=?, schedule=?, blocked_slots=?
        WHERE id=?
    ''', (data['name'], data['surname'], data['branch'], schedule_json, blocked_slots_json, data['id']))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Öğretmen başarıyla güncellendi!'})

@app.route('/delete_teacher', methods=['POST'])
def delete_teacher():
    conn = get_db()
    cursor = conn.cursor()
    data = request.json

    cursor.execute('DELETE FROM teachers WHERE id=?', (data['id'],))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Öğretmen başarıyla silindi!'})

@app.route('/add_student', methods=['POST'])
def add_student():
    conn = get_db()
    cursor = conn.cursor()
    data = request.json

    restrictions_json = json.dumps(data.get('restrictions', []), ensure_ascii=False)
    priorities_json = json.dumps(data.get('priorities', {}), ensure_ascii=False)
    manual_lessons_json = json.dumps(data.get('manual_lessons', []), ensure_ascii=False)
    teacher_blocks_json = json.dumps(data.get('teacher_blocks', []), ensure_ascii=False)

    cursor.execute('''
        INSERT INTO students (name, surname, class, restrictions, priorities, manual_lessons, teacher_blocks)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (data['name'], data['surname'], data['class'], restrictions_json, priorities_json, manual_lessons_json, teacher_blocks_json))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Öğrenci başarıyla eklendi!'})

@app.route('/update_student', methods=['POST'])
def update_student():
    conn = get_db()
    cursor = conn.cursor()
    data = request.json

    restrictions_json = json.dumps(data.get('restrictions', []), ensure_ascii=False)
    priorities_json = json.dumps(data.get('priorities', {}), ensure_ascii=False)
    manual_lessons_json = json.dumps(data.get('manual_lessons', []), ensure_ascii=False)
    teacher_blocks_json = json.dumps(data.get('teacher_blocks', []), ensure_ascii=False)

    cursor.execute('''
        UPDATE students
        SET name=?, surname=?, class=?, restrictions=?, priorities=?, manual_lessons=?, teacher_blocks=?
        WHERE id=?
    ''', (data['name'], data['surname'], data['class'], restrictions_json, priorities_json, manual_lessons_json, teacher_blocks_json, data['id']))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Öğrenci başarıyla güncellendi!'})

@app.route('/delete_student', methods=['POST'])
def delete_student():
    conn = get_db()
    cursor = conn.cursor()
    data = request.json

    cursor.execute('DELETE FROM students WHERE id=?', (data['id'],))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Öğrenci başarıyla silindi!'})

@app.route('/get_teachers')
def get_teachers():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM teachers')
    rows = cursor.fetchall()

    teachers = []
    for row in rows:
        # 🆕 BLOKLAMALARI OKU
        blocked_slots = []
        try:
            if row['blocked_slots']:
                blocked_slots = json.loads(row['blocked_slots'])
        except:
            blocked_slots = []

        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule']),
            'blocked_slots': blocked_slots  # 🆕 EKLENDI
        })

    conn.close()
    return jsonify({'teachers': teachers})

@app.route('/get_students')
def get_students():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM students')
    rows = cursor.fetchall()

    students = []
    for row in rows:
        restrictions = json.loads(row['restrictions']) if row['restrictions'] else []

        priorities = {}
        manual_lessons = []

        try:
            if row['priorities']:
                priorities = json.loads(row['priorities'])
        except:
            priorities = {}

        try:
            if row['manual_lessons']:
                manual_lessons = json.loads(row['manual_lessons'])
        except:
            manual_lessons = []

        teacher_blocks = []
        try:
            if row['teacher_blocks']:
                teacher_blocks = json.loads(row['teacher_blocks'])
        except:
            teacher_blocks = []

        students.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'class': row['class'],
            'restrictions': restrictions,
            'priorities': priorities,
            'manual_lessons': manual_lessons,
            'teacher_blocks': teacher_blocks
        })

    conn.close()
    return jsonify({'students': students})

@app.route('/generate_schedule')
def generate_schedule():
    global schedule_data

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        # 🆕 BLOKLAMALARI OKU
        blocked_slots = []
        try:
            if row['blocked_slots']:
                blocked_slots = json.loads(row['blocked_slots'])
        except:
            blocked_slots = []

        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule']),
            'blocked_slots': blocked_slots  # ✅ EKLENDI!

        })

    cursor.execute('SELECT * FROM students')
    students = []
    for row in cursor.fetchall():
        priorities = {}
        manual_lessons = []

        try:
            if row['priorities']:
                priorities = json.loads(row['priorities'])
        except:
            priorities = {}

        try:
            if row['manual_lessons']:
                manual_lessons = json.loads(row['manual_lessons'])
        except:
            manual_lessons = []

        teacher_blocks = []
        try:
            if row['teacher_blocks']:
                teacher_blocks = json.loads(row['teacher_blocks'])
        except:
            teacher_blocks = []

        students.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'class': row['class'],
            'restrictions': json.loads(row['restrictions']) if row['restrictions'] else [],
            'priorities': priorities,
            'manual_lessons': manual_lessons,
            'teacher_blocks': teacher_blocks
        })

    conn.close()

    if len(teachers) == 0:
        return jsonify({'error': 'Lütfen önce en az bir öğretmen ekleyin!'})

    if len(students) == 0:
        return jsonify({'error': 'Lütfen önce en az bir öğrenci ekleyin!'})

    math_teachers = [t for t in teachers if t['branch'] == 'Matematik']
    if len(math_teachers) == 0:
        return jsonify({'error': 'Her öğrenci haftada en az 1 Matematik dersi almalı. Lütfen en az bir Matematik öğretmeni ekleyin!'})

    # 🆕 SINIF DERSLERİNİ ÇEK
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT
            cl.id,
            cl.class_name,
            cl.teacher_id,
            cl.day,
            cl.start_time,
            cl.end_time,
            cl.weeks,
            cl.is_group,
            t.name as teacher_name,
            t.surname as teacher_surname,
            t.branch as teacher_branch
        FROM class_lessons cl
        JOIN teachers t ON cl.teacher_id = t.id
    ''')

    class_lessons = []
    for row in cursor.fetchall():
        class_lessons.append({
            'id': row['id'],
            'class_name': row['class_name'],
            'teacher_id': row['teacher_id'],
            'teacher_name': f"{row['teacher_name']} {row['teacher_surname']}",
            'teacher_branch': row['teacher_branch'],
            'day': row['day'],
            'start_time': row['start_time'],
            'end_time': row['end_time'],
            'time': f"{row['start_time']}-{row['end_time']}",
            'weeks': row['weeks'],
            'is_group': row['is_group']
        })
    conn.close()

    schedule_data = create_four_week_schedule(teachers, students, class_lessons)
    return jsonify({'schedule': schedule_data})

@app.route('/save_current_schedule', methods=['POST'])
def save_current_schedule():
    """Mevcut programı kaydet"""
    global schedule_data

    if not schedule_data:
        return jsonify({'error': 'Kaydedilecek program bulunamadı!'}), 400

    data = request.json
    schedule_name = data.get('name', f"Program_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    start_date = data.get('start_date')  # Frontend'den gelen tarih (YYYY-MM-DD)

    # Eğer tarih gönderilmemişse, en yakın Pazartesi'yi hesapla
    if not start_date:
        today = datetime.now()
        days_until_monday = (7 - today.weekday()) % 7
        if days_until_monday == 0 and today.weekday() != 0:
            days_until_monday = 7
        next_monday = today + timedelta(days=days_until_monday if days_until_monday > 0 else 0)
        start_date = next_monday.strftime('%Y-%m-%d')

    conn = get_db()
    cursor = conn.cursor()

    # Öğretmen snapshot'ı al
    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        blocked_slots = []
        try:
            if row['blocked_slots']:
                blocked_slots = json.loads(row['blocked_slots'])
        except:
            blocked_slots = []

        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule']),
            'blocked_slots': blocked_slots
        })

    # Öğrenci snapshot'ı al
    cursor.execute('SELECT * FROM students')
    students = []
    for row in cursor.fetchall():
        students.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'class': row['class'],
            'restrictions': json.loads(row['restrictions']) if row['restrictions'] else [],
            'priorities': json.loads(row['priorities']) if row['priorities'] else {},
            'manual_lessons': json.loads(row['manual_lessons']) if row['manual_lessons'] else []
        })

    # Veritabanına kaydet
    cursor.execute('''
        INSERT INTO saved_schedules (name, schedule_data, teachers_snapshot, students_snapshot, start_date)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        schedule_name,
        json.dumps(schedule_data, ensure_ascii=False),
        json.dumps(teachers, ensure_ascii=False),
        json.dumps(students, ensure_ascii=False),
        start_date
    ))

    saved_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'message': 'Program başarıyla kaydedildi!', 'id': saved_id})


@app.route('/get_saved_schedules')
def get_saved_schedules():
    """Kaydedilmiş programları listele"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT id, name, created_at, schedule_data, start_date,
               LENGTH(schedule_data) as data_size
        FROM saved_schedules
        ORDER BY created_at DESC
    ''')

    schedules = []
    for row in cursor.fetchall():
        schedules.append({
            'id': row['id'],
            'name': row['name'],
            'created_at': row['created_at'],
            'schedule_data': row['schedule_data'],
            'start_date': row['start_date'],
            'data_size': row['data_size']
        })

    conn.close()
    return jsonify({'schedules': schedules})


@app.route('/load_schedule/<int:schedule_id>')
def load_schedule(schedule_id):
    """Kayıtlı programı yükle"""
    global schedule_data

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT schedule_data, teachers_snapshot, students_snapshot
        FROM saved_schedules
        WHERE id = ?
    ''', (schedule_id,))

    row = cursor.fetchone()
    conn.close()

    if not row:
        return jsonify({'error': 'Program bulunamadı!'}), 404

    schedule_data = json.loads(row['schedule_data'])

    return jsonify({
        'schedule': schedule_data,
        'teachers': json.loads(row['teachers_snapshot']),
        'students': json.loads(row['students_snapshot'])
    })


@app.route('/delete_schedule/<int:schedule_id>', methods=['POST'])
def delete_schedule(schedule_id):
    """Kayıtlı programı sil"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('DELETE FROM saved_schedules WHERE id = ?', (schedule_id,))

    conn.commit()
    conn.close()

    return jsonify({'message': 'Program başarıyla silindi!'})


@app.route('/rename_schedule/<int:schedule_id>', methods=['POST'])
def rename_schedule(schedule_id):
    """Kayıtlı programı yeniden adlandır"""
    data = request.json
    new_name = data.get('name', '').strip()

    if not new_name:
        return jsonify({'error': 'Program adı boş olamaz!'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('UPDATE saved_schedules SET name = ? WHERE id = ?', (new_name, schedule_id))

    conn.commit()
    conn.close()

    return jsonify({'message': 'Program adı güncellendi!'})

def is_student_available(student, week_num, day, lesson_start, lesson_end):
    """
    ✅ YENİ: Saat karşılaştırmalarını dakikaya çevirip yapıyoruz
    """
    if not student.get('restrictions'):
        return True

    # 🔧 YARDIMCI FONKSİYON: Saati dakikaya çevir
    def time_to_minutes(time_str):
        """14:00 -> 840"""
        try:
            hours, mins = map(int, time_str.split(':'))
            return hours * 60 + mins
        except:
            return 0

    lesson_start_mins = time_to_minutes(lesson_start)
    lesson_end_mins = time_to_minutes(lesson_end)

    for restriction in student['restrictions']:
        # ✅ GÜNLER KONTROLÜ (eski + yeni format)
        days = restriction.get('days', [])
        if not days and restriction.get('day'):
            days = [restriction.get('day')]

        if day not in days:
            continue

        # ✅ SAAT BİLGİLERİNİ AL
        restriction_start = restriction.get('start_time', '')
        restriction_end = restriction.get('end_time', '')

        if not restriction_start or not restriction_end:
            continue

        # 🔥 SAATLARI DAKİKAYA ÇEVİR
        rest_start_mins = time_to_minutes(restriction_start)
        rest_end_mins = time_to_minutes(restriction_end)

        # 🔥 ÇAKIŞMA KONTROLÜ (Dakika bazında)
        has_overlap = not (lesson_end_mins <= rest_start_mins or lesson_start_mins >= rest_end_mins)

        if has_overlap:
            # Haftalık kısıtlama ise her zaman blokla
            if restriction.get('type') == 'weekly':
                return False

            # Özel hafta seçimi
            if restriction.get('type') == 'custom':
                weeks = restriction.get('weeks', [])
                if not weeks and restriction.get('week'):
                    weeks = [restriction.get('week')]

                if week_num in weeks:
                    return False

    return True

def is_teacher_blocked_for_student(student, teacher, week, day, lesson_info):
    """
    Öğretmenin bu slot için öğrenci tarafından engellenip engellenmediğini kontrol eder

    Args:
        student: Öğrenci dict
        teacher: Öğretmen dict
        week: Hafta numarası (0-3)
        day: Gün adı
        lesson_info: Ders bilgisi dict (start_time, end_time içerir)

    Returns:
        bool: True ise öğretmen engellenmiş, False ise müsait
    """
    if not student.get('teacher_blocks'):
        return False

    for tb in student['teacher_blocks']:
        # Öğretmen kontrolü
        if tb['teacher_id'] != teacher['id']:
            continue

        # Hafta kontrolü
        if tb['type'] == 'custom':
            if (week + 1) not in tb.get('weeks', []):
                continue

        # Gün kontrolü
        if tb['day'] != 'all' and tb['day'] != day:
            continue

        # Slot kontrolü
        slot_to_check = f"{day}_{lesson_info['start_time']}-{lesson_info['end_time']}"
        if slot_to_check in tb.get('blocked_slots', []):
            return True

    return False

def is_slot_available_for_student(student, week_student_time_slots, day, lesson_info):
    """
    Öğrencinin bu slotta başka dersi olup olmadığını kontrol eder

    Args:
        student: Öğrenci dict
        week_student_time_slots: Haftalık slot takip dict
        day: Gün adı
        lesson_info: Ders bilgisi dict (start_time içerir)

    Returns:
        bool: True ise slot müsait, False ise dolu
    """
    slot_key = f"{day}_{lesson_info['start_time']}"
    student_slots = week_student_time_slots.get(student['id'], set())
    return slot_key not in student_slots

def create_four_week_schedule(teachers, students, class_lessons=[]):
    """
    ✅ ÇAKIŞMA SORUNU ÇÖZÜLMÜŞ VERSİYON
    - Öncelikli dersler: Haftada 2 ders
    - Normal dersler: Haftada 1 ders
    - GÜNLÜK LİMİT: Aynı gün aynı öğretmenden maksimum 1 ders
    - 🔥 SLOT KONTROLÜ: Aynı gün/saatte birden fazla ders ASLA yok!
    - 🆕 SINIF DERSLERİ: Otomatik olarak sınıftaki tüm öğrencilere eklenir
    """

    # MANUEL DERSLERİ TOPLA
    manual_assignments = []
    for student in students:
        if student.get('manual_lessons'):
            for manual in student['manual_lessons']:
                manual_assignments.append({
                    'student': student,
                    'week': manual['week'],
                    'day': manual['day'],
                    'teacher_id': manual['teacher_id'],
                    'time': manual['time']
                })

    weeks = []
    student_stats = {}

    for student in students:
        student_stats[student['id']] = {
            'student': student,
            'weeks': [{}, {}, {}, {}],
            'total_lessons': 0,
            'total_math_per_week': [0, 0, 0, 0],
            'teachers_taken': set(),
            'teacher_lesson_count': {},
            'total_branches': {}
        }

    for week in range(4):
        week_schedule = []
        week_student_lessons = {s['id']: 0 for s in students}
        week_student_teacher_lessons = {s['id']: {} for s in students}
        week_student_math_count = {s['id']: 0 for s in students}
        week_student_time_slots = {s['id']: set() for s in students}
        # 🆕 ÖĞRETMEN SLOT TAKİBİ
        week_teacher_used_slots = {}  # {teacher_id: {slot_key, slot_key, ...}}
        for teacher in teachers:
            week_teacher_used_slots[teacher['id']] = set()

        # 🆕 GÜNLÜK ÖĞRETMEN-DERS TAKİBİ
        week_student_daily_teacher_lessons = {s['id']: {} for s in students}

        # MANUEL DERSLERİ EKLE
        for manual in manual_assignments:
            if manual['week'] == week + 1:
                student = manual['student']

                teacher = next((t for t in teachers if t['id'] == manual['teacher_id']), None)
                if not teacher:
                    continue

                lesson = {
                    'day': manual['day'],
                    'time': manual['time'],
                    'teacher_name': f"{teacher['name']} {teacher['surname']}",
                    'branch': teacher['branch'],
                    'student_name': f"{student['name']} {student['surname']}",
                    'student_class': student['class'],
                    'week': week + 1
                }
                week_schedule.append(lesson)

                week_student_lessons[student['id']] += 1
                student_stats[student['id']]['total_lessons'] += 1

                if teacher['id'] not in week_student_teacher_lessons[student['id']]:
                    week_student_teacher_lessons[student['id']][teacher['id']] = 0
                week_student_teacher_lessons[student['id']][teacher['id']] += 1

                slot_key = f"{manual['day']}_{manual['time'].split('-')[0]}"
                week_student_time_slots[student['id']].add(slot_key)
                # 🆕 ÖĞRETMEN SLOTUNU İŞARETLE
                teacher_slot_key = f"{manual['day']}_{manual['time']}"
                week_teacher_used_slots[teacher['id']].add(teacher_slot_key)
                # 🆕 GÜNLÜK TAKİP EKLE (Manuel Dersler)
                day_teacher_key = f"{manual['day']}_{teacher['id']}"
                if day_teacher_key not in week_student_daily_teacher_lessons[student['id']]:
                    week_student_daily_teacher_lessons[student['id']][day_teacher_key] = 0
                week_student_daily_teacher_lessons[student['id']][day_teacher_key] += 1

                if teacher['branch'] == 'Matematik':
                    week_student_math_count[student['id']] += 1
                    student_stats[student['id']]['total_math_per_week'][week] += 1

                if teacher['branch'] not in student_stats[student['id']]['weeks'][week]:
                    student_stats[student['id']]['weeks'][week][teacher['branch']] = 0
                student_stats[student['id']]['weeks'][week][teacher['branch']] += 1

                if teacher['branch'] not in student_stats[student['id']]['total_branches']:
                    student_stats[student['id']]['total_branches'][teacher['branch']] = 0
                student_stats[student['id']]['total_branches'][teacher['branch']] += 1

        # 🆕 SINIF DERSLERİNİ EKLE
        for class_lesson in class_lessons:
            # Hafta kontrolü
            weeks_list = []
            if class_lesson['weeks'] == 'all':
                weeks_list = [1, 2, 3, 4]
            else:
                weeks_list = [int(w) for w in class_lesson['weeks'].split(',')]

            # Bu hafta için değilse atla
            if (week + 1) not in weeks_list:
                continue

            # Bu sınıftaki tüm öğrencileri bul
            class_students = [s for s in students if s['class'] == class_lesson['class_name']]

            if len(class_students) == 0:
                continue  # Sınıfta öğrenci yoksa atla

            # Öğretmeni bul
            teacher = next((t for t in teachers if t['id'] == class_lesson['teacher_id']), None)
            if not teacher:
                continue

            # 🔥 ÖNEMLİ: Sınıf dersi TOPLU atanır, slot kontrolü YAPILMAZ!
            # Çünkü tüm sınıf aynı saatte aynı dersi alacak

            # Bu sınıftaki HER öğrenci için ders ekle
            for student in class_students:
                # Dersi ekle
                lesson = {
                    'day': class_lesson['day'],
                    'time': class_lesson['time'],
                    'teacher_name': class_lesson['teacher_name'],
                    'branch': class_lesson['teacher_branch'],
                    'student_name': f"{student['name']} {student['surname']}",
                    'student_class': student['class'],
                    'week': week + 1,
                    'is_class_lesson': True,  # 🆕 Sınıf dersi işareti
                    'is_group': class_lesson.get('is_group', 0)  # 🆕 Grup dersi bilgisi
                }
                week_schedule.append(lesson)


                # İstatistikleri güncelle
                week_student_lessons[student['id']] += 1
                student_stats[student['id']]['total_lessons'] += 1

                if teacher['id'] not in week_student_teacher_lessons[student['id']]:
                    week_student_teacher_lessons[student['id']][teacher['id']] = 0
                week_student_teacher_lessons[student['id']][teacher['id']] += 1

                # 🆕 Slot'u işaretle (ÖNEMLİ: Bireysel ders atamasında çakışma kontrolü için)
                slot_key = f"{class_lesson['day']}_{class_lesson['start_time']}"
                week_student_time_slots[student['id']].add(slot_key)

                # Günlük takip
                day_teacher_key = f"{class_lesson['day']}_{teacher['id']}"
                if day_teacher_key not in week_student_daily_teacher_lessons[student['id']]:
                    week_student_daily_teacher_lessons[student['id']][day_teacher_key] = 0
                week_student_daily_teacher_lessons[student['id']][day_teacher_key] += 1

                # Matematik sayacı
                if teacher['branch'] == 'Matematik':
                    week_student_math_count[student['id']] += 1
                    student_stats[student['id']]['total_math_per_week'][week] += 1

                # Branş istatistikleri
                if teacher['branch'] not in student_stats[student['id']]['weeks'][week]:
                    student_stats[student['id']]['weeks'][week][teacher['branch']] = 0
                student_stats[student['id']]['weeks'][week][teacher['branch']] += 1

                if teacher['branch'] not in student_stats[student['id']]['total_branches']:
                    student_stats[student['id']]['total_branches'][teacher['branch']] = 0
                student_stats[student['id']]['total_branches'][teacher['branch']] += 1

            # Öğretmen slot'unu işaretle (sadece bir kere)
            teacher_slot_key = f"{class_lesson['day']}_{class_lesson['time']}"
            week_teacher_used_slots[teacher['id']].add(teacher_slot_key)

        all_slots = []
        for teacher in teachers:
            for day_schedule in teacher['schedule']:
                day = day_schedule['day']
                for lesson_info in day_schedule['lessons']:
                    all_slots.append({
                        'teacher': teacher,
                        'day': day,
                        'lesson_info': lesson_info
                    })



        for slot in all_slots:
            teacher = slot['teacher']
            day = slot['day']
            lesson_info = slot['lesson_info']

            # BLOKLAMA KONTROLÜ
            is_blocked = False
            if teacher.get('blocked_slots'):
                for block in teacher['blocked_slots']:
                    # Gün kontrolü
                    if block['day'] != day:
                        continue

                    # Slot kontrolü
                    slot_key = f"{lesson_info['start_time']}-{lesson_info['end_time']}"
                    if slot_key not in block.get('blocked_slots', []):
                        continue

                    # Hafta kontrolü
                    if block['type'] == 'weekly':
                        is_blocked = True
                        break
                    elif block['type'] == 'custom':
                        if (week + 1) in block.get('weeks', []):
                            is_blocked = True
                            break

            if is_blocked:
                continue

            # 🆕 BU SLOT MANUEL DERSLE DOLU MU KONTROL ET
            teacher_slot_key = f"{day}_{lesson_info['start_time']}-{lesson_info['end_time']}"
            if teacher_slot_key in week_teacher_used_slots.get(teacher['id'], set()):
                continue  # Bu slot zaten kullanılmış, atla!

            available_students = []
            never_taken_students = []  # 🆕 Bu branşı hiç almayan öğrenciler

            # ═══════════════════════════════════════════════════
            # SEVİYE 1: EN SIKI KURALLAR
            # ═══════════════════════════════════════════════════
            for student in students:
                if not is_student_available(student, week + 1, day, lesson_info['start_time'], lesson_info['end_time']):
                    continue

                # 🚫 ÖĞRETMEN ENGELLEME KONTROLÜ
                if is_teacher_blocked_for_student(student, teacher, week, day, lesson_info):
                    continue

                # 🔥 SLOT KONTROLÜ
                if not is_slot_available_for_student(student, week_student_time_slots, day, lesson_info):
                    continue

                teacher_lesson_count = week_student_teacher_lessons[student['id']].get(teacher['id'], 0)

                # ÖNCELİK KONTROLÜ
                is_priority = False
                if student.get('priorities'):
                    week_key = f'week{week + 1}'
                    week_priorities = student['priorities'].get(week_key, [])
                    if teacher['branch'] in week_priorities:
                        is_priority = True

                # 🆕 GÜNLÜK LİMİT KONTROLÜ (Aynı gün aynı öğretmenden maksimum 1 ders)
                day_teacher_key = f"{day}_{teacher['id']}"
                daily_lesson_count = week_student_daily_teacher_lessons[student['id']].get(day_teacher_key, 0)
                if daily_lesson_count >= 1:
                    continue

                # HAFTALIK LİMİT: Öncelikli değilse 1 ders, öncelikliyse 2 ders
                if not is_priority and teacher_lesson_count >= 1:
                    continue
                elif is_priority and teacher_lesson_count >= 2:
                    continue

                current_math_count = week_student_math_count[student['id']]

                if teacher['branch'] == 'Matematik':
                    if current_math_count >= 2:
                        continue

                total_branch_count = student_stats[student['id']]['total_branches'].get(teacher['branch'], 0)

                # 🆕 BRANŞ HİÇ ALINMADI MI KONTROLÜ
                if total_branch_count == 0 and week >= 2:  # Hafta 3-4'te zorunlu yap
                    never_taken_students.append(student)
                    continue  # Normal listeye ekleme, sadece never_taken'a ekle

                branch_limits = {
                    'Matematik': 8, 'Fizik': 8, 'Kimya': 8, 'Biyoloji': 8,
                    'Geometri': 8, 'Türkçe': 8, 'Edebiyat': 8, 'İngilizce': 8,







                    'Fen Bilgisi': 8, 'Sosyal Bilgiler': 8, 'Tarih': 8,



                    'Coğrafya': 8, 'Felsefe': 8, 'Din Kültürü': 8,
                    'Matematik-1': 8, 'Matematik-2': 8

                }
                max_for_branch = branch_limits.get(teacher['branch'], 999)
                if total_branch_count >= max_for_branch:
                    continue

                available_students.append(student)

            # 🆕 EĞER HİÇ ALMAYAN ÖĞRENCİLER VARSA, ONLARA MUTLAKA VER!
            if never_taken_students and week >= 2:
                available_students = never_taken_students

            # ═══════════════════════════════════════════════════
            # SEVİYE 2: MATEMATİK ÖNCELİĞİ KONTROLÜ
            # ═══════════════════════════════════════════════════
            if not available_students:
                for student in students:
                    if not is_student_available(student, week + 1, day, lesson_info["start_time"], lesson_info["end_time"]):
                        continue

                    # 🚫 ÖĞRETMEN ENGELLEME KONTROLÜ
                    if is_teacher_blocked_for_student(student, teacher, week, day, lesson_info):
                        continue

                    # 🔥 SLOT KONTROLÜ
                    if not is_slot_available_for_student(student, week_student_time_slots, day, lesson_info):
                        continue

                    teacher_lesson_count = week_student_teacher_lessons[student['id']].get(teacher['id'], 0)
                    if teacher_lesson_count >= 1:
                        continue

                    current_math_count = week_student_math_count[student['id']]

                    if teacher['branch'] == 'Matematik':
                        if current_math_count >= 2:
                            continue
                    else:
                        if current_math_count < 1:
                            continue

                    available_students.append(student)

            # ═══════════════════════════════════════════════════
            # SEVİYE 3: SADECE MATEMATİK KONTROLÜ
            # ═══════════════════════════════════════════════════
            if not available_students:
                for student in students:
                    if not is_student_available(student, week + 1, day, lesson_info["start_time"], lesson_info["end_time"]):
                        continue

                    # 🚫 ÖĞRETMEN ENGELLEME KONTROLÜ
                    if is_teacher_blocked_for_student(student, teacher, week, day, lesson_info):
                        continue

                    # 🔥 SLOT KONTROLÜ
                    if not is_slot_available_for_student(student, week_student_time_slots, day, lesson_info):
                        continue

                    current_math_count = week_student_math_count[student['id']]

                    if teacher['branch'] == 'Matematik':
                        if current_math_count >= 2:
                            continue
                    else:
                        if current_math_count < 1:
                            continue

                    available_students.append(student)

            # ═══════════════════════════════════════════════════
            # SEVİYE 4: EN GEVŞEK - SADECE KISITLAMA
            # ═══════════════════════════════════════════════════
            if not available_students:
                for student in students:
                    if not is_student_available(student, week + 1, day, lesson_info["start_time"], lesson_info["end_time"]):
                        continue

                    # 🚫 ÖĞRETMEN ENGELLEME KONTROLÜ
                    if is_teacher_blocked_for_student(student, teacher, week, day, lesson_info):
                        continue

                    # 🔥 SLOT KONTROLÜ
                    if not is_slot_available_for_student(student, week_student_time_slots, day, lesson_info):
                        continue

                    available_students.append(student)

            if not available_students:
                continue

            def get_priority(s):

                priority_bonus = 0
                if s.get('priorities'):
                    week_key = f'week{week + 1}'
                    week_priorities = s['priorities'].get(week_key, [])
                    if teacher['branch'] in week_priorities:
                        priority_index = week_priorities.index(teacher['branch'])
                        priority_bonus = -(len(week_priorities) - priority_index) * 100000

                branch_targets = {
                    'Matematik': 2, 'Fizik': 0.5, 'Kimya': 0.5, 'Biyoloji': 0.5,
                    'Geometri': 0.5, 'Türkçe': 0.5, 'Edebiyat': 0.5, 'İngilizce': 0.5,







                    'Fen Bilgisi': 0.5, 'Sosyal Bilgiler': 0.5, 'Tarih': 0.5,



                    'Coğrafya': 0.5, 'Felsefe': 0.5, 'Din Kültürü': 0.5,
                    'Matematik-1': 0.5, 'Matematik-2': 0.5

                }

                current_total = student_stats[s['id']]['total_branches'].get(teacher['branch'], 0)
                target_total = branch_targets.get(teacher['branch'], 0) * (week + 1)
                shortage = max(0, target_total - current_total) * 10000

                times_with_teacher = student_stats[s['id']]['teacher_lesson_count'].get(teacher['id'], 0)
                never_matched = 1 if times_with_teacher == 0 else 0

                weekly_lesson_shortage = max(0, 3 - week_student_lessons[s['id']])

                return (
                    priority_bonus,
                    -never_matched,
                    -shortage,
                    -weekly_lesson_shortage,
                    student_stats[s['id']]['total_lessons'],
                    week_student_lessons[s['id']]
                )

            available_students.sort(key=get_priority)

            # Öncelikli öğrenci varsa direkt seç, yoksa rastgele seç
            if len(available_students) > 0:
                first_priority = get_priority(available_students[0])
                if first_priority[0] < -50000:  # Öncelikli ders varsa
                    selected_student = available_students[0]
                else:
                    top_candidates = available_students[:min(5, len(available_students))]
                    selected_student = random.choice(top_candidates)
            else:
                continue

            lesson = {
                'day': day,
                'time': f"{lesson_info['start_time']}-{lesson_info['end_time']}",
                'teacher_name': f"{teacher['name']} {teacher['surname']}",
                'branch': teacher['branch'],
                'student_name': f"{selected_student['name']} {selected_student['surname']}",
                'student_class': selected_student['class'],
                'week': week + 1
            }
            week_schedule.append(lesson)

            slot_key = f"{day}_{lesson_info['start_time']}"
            week_student_time_slots[selected_student['id']].add(slot_key)
            # 🆕 ÖĞRETMEN SLOTUNU DA İŞARETLE
            teacher_slot_key = f"{day}_{lesson_info['start_time']}-{lesson_info['end_time']}"
            week_teacher_used_slots[teacher['id']].add(teacher_slot_key)

            week_student_lessons[selected_student['id']] += 1
            student_stats[selected_student['id']]['total_lessons'] += 1
            student_stats[selected_student['id']]['teachers_taken'].add(teacher['id'])

            if teacher['id'] not in student_stats[selected_student['id']]['teacher_lesson_count']:
                student_stats[selected_student['id']]['teacher_lesson_count'][teacher['id']] = 0
            student_stats[selected_student['id']]['teacher_lesson_count'][teacher['id']] += 1

            if teacher['id'] not in week_student_teacher_lessons[selected_student['id']]:
                week_student_teacher_lessons[selected_student['id']][teacher['id']] = 0
            week_student_teacher_lessons[selected_student['id']][teacher['id']] += 1


            day_teacher_key = f"{day}_{teacher['id']}"
            if day_teacher_key not in week_student_daily_teacher_lessons[selected_student['id']]:
                week_student_daily_teacher_lessons[selected_student['id']][day_teacher_key] = 0
            week_student_daily_teacher_lessons[selected_student['id']][day_teacher_key] += 1

            if teacher['branch'] == 'Matematik':
                week_student_math_count[selected_student['id']] += 1
                student_stats[selected_student['id']]['total_math_per_week'][week] += 1

            if teacher['branch'] not in student_stats[selected_student['id']]['weeks'][week]:
                student_stats[selected_student['id']]['weeks'][week][teacher['branch']] = 0
            student_stats[selected_student['id']]['weeks'][week][teacher['branch']] += 1

            if teacher['branch'] not in student_stats[selected_student['id']]['total_branches']:
                student_stats[selected_student['id']]['total_branches'][teacher['branch']] = 0
            student_stats[selected_student['id']]['total_branches'][teacher['branch']] += 1

        day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
        week_schedule.sort(key=lambda x: (day_order.index(x['day']), x['time']))
        weeks.append(week_schedule)

    stats = []
    for student_id, stat in student_stats.items():
        week_summaries = []
        for w in stat['weeks']:
            if w:
                branches = ', '.join([f"{b}({c})" for b, c in w.items()])
                week_summaries.append(branches)
            else:
                week_summaries.append('-')

        total_lessons = stat['total_lessons']

        stats.append({
            'student_name': f"{stat['student']['name']} {stat['student']['surname']}",
            'student_class': stat['student']['class'],
            'week1': week_summaries[0],
            'week2': week_summaries[1],
            'week3': week_summaries[2],
            'week4': week_summaries[3],
            'total': total_lessons
        })

    return {'weeks': weeks, 'stats': stats}

@app.route('/export_excel')
def export_excel():
    if not schedule_data:
        return "Lütfen önce program oluşturun!", 400

    wb = Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')

    day_fill = PatternFill(start_color="9575CD", end_color="9575CD", fill_type="solid")
    day_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')

    time_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    time_font = Font(bold=False, size=10, name='Calibri', color="1565C0")

    student_font = Font(size=10, name='Calibri')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule'])
        })

    # ✅ Python'da Türkçe desteğiyle alfabetik sırala
    import locale
    try:
        locale.setlocale(locale.LC_COLLATE, 'tr_TR.UTF-8')
        teachers.sort(key=lambda t: (locale.strxfrm(t['branch']), locale.strxfrm(t['name']), locale.strxfrm(t['surname'])))
    except:
        # Locale bulunamazsa normal sıralama
        teachers.sort(key=lambda t: (t['branch'], t['name'], t['surname']))

    conn.close()

    for week_num in range(1, 5):
        ws = wb.create_sheet(title=f"Hafta {week_num}")

        ws['A1'] = ''
        ws['A1'].fill = header_fill
        ws['A1'].border = thin_border

        col_idx = 2
        teacher_columns = {}

        for teacher in teachers:
            cell = ws.cell(row=1, column=col_idx)
            # ✅ BRANŞ VE İSİM AYRI SATIRLARDA
            cell.value = f"{teacher['branch'].upper()}\n({teacher['name'].upper()} {teacher['surname'].upper()})"
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            # ✅ WRAP_TEXT AKTİF + DİKEY HİZALAMA
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            teacher_columns[teacher['id']] = col_idx
            ws.column_dimensions[cell.column_letter].width = 18
            col_idx += 1

        ws.column_dimensions['A'].width = 15
        # ✅ SATIR YÜKSEKLİĞİNİ ARTIR (40 → 50)
        ws.row_dimensions[1].height = 50

        all_slots = []
        for teacher in teachers:
            for day_schedule in teacher['schedule']:
                day = day_schedule['day']
                for lesson in day_schedule['lessons']:
                    slot_key = f"{day}_{lesson['start_time']}_{lesson['end_time']}"
                    slot_info = {
                        'day': day,
                        'start_time': lesson['start_time'],
                        'end_time': lesson['end_time'],
                        'key': slot_key
                    }
                    if slot_info not in all_slots:
                        all_slots.append(slot_info)

        day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
        all_slots.sort(key=lambda x: (day_order.index(x['day']), x['start_time']))

        current_row = 2
        current_day = None

        for slot in all_slots:
            if slot['day'] != current_day:
                day_cell = ws.cell(row=current_row, column=1)
                day_cell.value = slot['day'].upper()
                day_cell.fill = day_fill
                day_cell.font = day_font
                day_cell.border = thin_border
                day_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[current_row].height = 25

                for col in range(2, col_idx):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = day_fill
                    cell.border = thin_border

                # ✅ TÜM KOLONLARI BİRLEŞTİR
                from openpyxl.utils import get_column_letter
                start_col = get_column_letter(1)
                end_col = get_column_letter(col_idx - 1)
                ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')

                current_day = slot['day']
                current_row += 1

            time_cell = ws.cell(row=current_row, column=1)
            time_cell.value = f"{slot['start_time']}-{slot['end_time']}"
            time_cell.fill = time_fill
            time_cell.font = time_font
            time_cell.border = thin_border
            time_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[current_row].height = 20

            week_data = schedule_data['weeks'][week_num - 1]

            for teacher in teachers:
                col = teacher_columns[teacher['id']]
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = student_font

                teacher_full_name = f"{teacher['name']} {teacher['surname']}"

                # 🆕 TÜM eşleşen dersleri bul
                matching_lessons = [
                    lesson for lesson in week_data
                    if (lesson['teacher_name'] == teacher_full_name and
                        lesson['day'] == slot['day'] and
                        lesson['time'] == f"{slot['start_time']}-{slot['end_time']}")
                ]

                # 🆕 Gruplama mantığı
                if len(matching_lessons) == 0:
                    cell.value = ''
                elif len(matching_lessons) == 1:
                    cell.value = matching_lessons[0]['student_name'].upper()
                else:
                    # 🆕 GRUP DERSİ - TÜM SINIFLARI TOPLA
                    unique_classes = list(set([l.get('student_class', '') for l in matching_lessons if l.get('student_class')]))
                    if unique_classes:
                        classes_str = ', '.join(sorted(unique_classes))
                        cell.value = f"{classes_str} ({len(matching_lessons)} öğr)"
                    else:
                        cell.value = f"{matching_lessons[0]['student_name'].upper()} +{len(matching_lessons)-1}"

            current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'ders_programi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@app.route('/export_html')
def export_html():
    if not schedule_data:
        return "Lütfen önce program oluşturun!", 400

    # Öğretmenleri çek (Excel gibi)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule'])
        })

    # ✅ Python'da Türkçe desteğiyle alfabetik sırala
    import locale
    try:
        locale.setlocale(locale.LC_COLLATE, 'tr_TR.UTF-8')
        teachers.sort(key=lambda t: (locale.strxfrm(t['branch']), locale.strxfrm(t['name']), locale.strxfrm(t['surname'])))
    except:
        teachers.sort(key=lambda t: (t['branch'], t['name'], t['surname']))

    conn.close()

    html_content = f'''<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>4 Haftalık Ders Programı - {datetime.now().strftime("%d.%m.%Y")}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1600px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        }}
        h1 {{
            text-align: center;
            color: #667eea;
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: bold;
            text-transform: uppercase;
        }}
        .date {{
            text-align: center;
            color: #666;
            font-size: 1.1em;
            margin-bottom: 40px;
        }}
        .week-section {{
            margin-bottom: 50px;
            page-break-after: always;
        }}
        .week-header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 15px 15px 0 0;
            font-size: 2em;
            font-weight: bold;
            text-align: center;
            margin-bottom: 0;
        }}
        .schedule-table {{
            width: 100%;
            border-collapse: collapse;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            border-radius: 0 0 15px 15px;
            overflow: hidden;
        }}
        .schedule-table thead tr {{
            background: linear-gradient(135deg, #4472C4 0%, #5B9BD5 100%);
        }}
        .schedule-table th {{
            color: white;
            padding: 12px 6px;
            text-align: center;
            font-weight: bold;
            font-size: 0.75em;
            border-right: 1px solid rgba(255,255,255,0.2);
            line-height: 1.3;
            white-space: normal;
            word-wrap: break-word;
            vertical-align: middle;
            text-transform: uppercase;
        }}
        .schedule-table th:first-child {{
            text-align: left;
            padding-left: 15px;
            min-width: 100px;
            max-width: 120px;
            width: 100px;
        }}
        .schedule-table th:last-child {{
            border-right: none;
        }}
        .day-header {{
            background: #9575CD !important;
            color: white !important;
            font-weight: bold;
            font-size: 1.1em;
            padding: 12px 20px !important;
            text-align: center !important;
        }}
        .time-row {{
            background: #E3F2FD;
        }}
        .time-row td:first-child {{
            font-weight: 600;
            color: #1565C0;
            padding: 12px 10px;
            width: 100px;
            max-width: 120px;
            white-space: nowrap;
        }}
        .schedule-table td {{
            padding: 10px 8px;
            text-align: center;
            border: 1px solid #e5e7eb;
            font-size: 0.8em;
            vertical-align: middle;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            max-width: 120px;
        }}
        .schedule-table tbody tr:hover {{
            background: #f0f4ff;
            transition: all 0.2s;
        }}
        .student-name {{
            font-weight: 600;
            color: #1f2937;
        }}
        .stats-section {{
            margin-top: 50px;
            page-break-before: always;
        }}
        .stats-header {{
            background: linear-gradient(135deg, #16a34a 0%, #15803d 100%);
            color: white;
            padding: 20px;
            border-radius: 15px;
            font-size: 2em;
            font-weight: bold;
            text-align: center;
            margin-bottom: 20px;
        }}
        .stats-table {{
            width: 100%;
            border-collapse: collapse;
            box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            border-radius: 15px;
            overflow: hidden;
        }}
        .stats-table thead {{
            background: linear-gradient(135deg, #16a34a 0%, #15803d 100%);
        }}
        .stats-table th {{
            color: white;
            padding: 18px 20px;
            text-align: center;
            font-weight: bold;
            font-size: 1.1em;
            border: 1px solid #15803d;
        }}
        .stats-table th:first-child {{
            text-align: left;
        }}
        .stats-table td {{
            padding: 15px 20px;
            border: 1px solid #e5e7eb;
            text-align: center;
        }}
        .stats-table .student-name-col {{
            text-align: left;
            font-weight: 600;
            color: #1f2937;
        }}
        .stats-table tbody tr:nth-child(even) {{
            background: #f9fafb;
        }}
        .stats-table tbody tr:hover {{
            background: #f0fdf4;
            transition: all 0.2s;
        }}
        .total-cell {{
            font-weight: bold;
            color: #15803d;
            background: #dcfce7 !important;
            font-size: 1.1em;
        }}

        @media print {{
            @page {{
                size: A4 landscape;
                margin: 10mm;
            }}

            body {{
                background: white !important;
                padding: 0 !important;
            }}

            .container {{
                padding: 10px !important;
                box-shadow: none !important;
            }}

            .week-section {{
                page-break-after: always;
                page-break-inside: avoid;
            }}

            .schedule-table {{
                font-size: 0.65em !important;
                width: 100% !important;
            }}

            .schedule-table th:first-child,
            .schedule-table td:first-child {{
                width: 80px !important;
                max-width: 80px !important;
                min-width: 80px !important;
                padding: 6px 8px !important;
            }}

            .schedule-table th,
            .schedule-table td {{
                padding: 6px 4px !important;
                font-size: 0.75em !important;
            }}

            .schedule-table th {{
                font-size: 0.7em !important;
                line-height: 1.2 !important;
            }}

            .week-header {{
                font-size: 1.5em !important;
                padding: 15px !important;
            }}

            h1 {{
                font-size: 1.8em !important;
            }}

            .stats-section {{
                page-break-before: always;
                page-break-inside: avoid !important;
            }}

            .stats-header {{
                font-size: 1.3em !important;
                padding: 12px !important;
                margin-bottom: 10px !important;
            }}

            .stats-table {{
                font-size: 0.7em !important;
            }}

            .stats-table th,
            .stats-table td {{
                padding: 8px 10px !important;
                font-size: 0.85em !important;
            }}

            .stats-table th {{
                font-size: 0.9em !important;
            }}

            * {{
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>4 HAFTALIK DERS PROGRAMI</h1>
        <div class="date">{datetime.now().strftime("%d %B %Y - %A")}</div>
'''

    # HER HAFTA İÇİN TABLO OLUŞTUR
    for week_num in range(1, 5):
        html_content += f'''
        <div class="week-section">
            <div class="week-header">HAFTA {week_num}</div>
            <table class="schedule-table">
                <thead>
                    <tr>
                        <th>GÜN / SAAT</th>
'''
        for teacher in teachers:
            html_content += f'''
                        <th>{teacher['branch'].upper()}<br>({teacher['name'].upper()} {teacher['surname'].upper()})</th>
'''
        html_content += '''
                    </tr>
                </thead>
                <tbody>
'''

        all_slots = []
        for teacher in teachers:
            for day_schedule in teacher['schedule']:
                day = day_schedule['day']
                for lesson in day_schedule['lessons']:
                    slot_key = f"{day}_{lesson['start_time']}_{lesson['end_time']}"
                    slot_info = {
                        'day': day,
                        'start_time': lesson['start_time'],
                        'end_time': lesson['end_time'],
                        'key': slot_key
                    }
                    if slot_info not in all_slots:
                        all_slots.append(slot_info)

        day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
        all_slots.sort(key=lambda x: (day_order.index(x['day']), x['start_time']))

        current_day = None
        week_data = schedule_data['weeks'][week_num - 1]

        for slot in all_slots:
            if slot['day'] != current_day:
                html_content += f'''
                    <tr>
                        <td colspan="{len(teachers) + 1}" class="day-header">{slot['day'].upper()}</td>
                    </tr>
'''
                current_day = slot['day']

            html_content += f'''
                    <tr class="time-row">
                        <td>{slot['start_time']}-{slot['end_time']}</td>
'''
            for teacher in teachers:
                teacher_full_name = f"{teacher['name']} {teacher['surname']}"

                # 🆕 Aynı slot'taki TÜM dersleri topla
                matching_lessons = []
                for lesson in week_data:
                    if (lesson['teacher_name'] == teacher_full_name and
                        lesson['day'] == slot['day'] and
                        lesson['time'] == f"{slot['start_time']}-{slot['end_time']}"):
                        matching_lessons.append(lesson)

                # 🐛 DEBUG
                if len(matching_lessons) > 1:
                    print(f"\n{'='*80}")
                    print(f"🔍 ANA TABLO DEBUG")
                    print(f"Öğretmen: {teacher_full_name}")
                    print(f"Slot: {slot['day']} {slot['start_time']}-{slot['end_time']}")
                    print(f"Toplam ders: {len(matching_lessons)}")
                    print(f"{'='*80}")
                    for idx, les in enumerate(matching_lessons):
                        print(f"  [{idx+1}] Öğrenci: {les['student_name']:30} | Sınıf: {les.get('student_class', 'YOK'):10} | is_group: {les.get('is_group', 'YOK')}")
                    print(f"{'='*80}\n")

                # 🆕 Sınıf dersiyse gruplayarak göster
                if len(matching_lessons) == 0:
                    student_name = ''
                elif len(matching_lessons) == 1:
                    student_name = matching_lessons[0]['student_name'].upper()
                else:
                    # Birden fazla öğrenci varsa sınıf dersi olarak göster
                    # 🎯 GRUP DERSİ KONTROLÜ - Farklı sınıflar var mı?
                    unique_classes = sorted(list(set([les.get('student_class', '') for les in matching_lessons if les.get('student_class', '')])))

                    print(f"  → 🎯 Benzersiz sınıflar: {unique_classes}")
                    print(f"  → 🎯 Sınıf sayısı: {len(unique_classes)}")
                    print(f"  → 🎯 is_group değerleri: {[les.get('is_group', 0) for les in matching_lessons]}")

                    if unique_classes:
                        if len(unique_classes) > 1:
                            # GRUP DERSİ: "10A, 10B (45 öğr)"
                            class_list = ', '.join(unique_classes)
                            student_name = f"{class_list} ({len(matching_lessons)} öğr)"
                            print(f"  → ✅ GRUP DERSİ tespit edildi: '{student_name}'")
                        else:
                            # TEK SINIF: "10A (45 öğrenci)"
                            student_name = f"{unique_classes[0]} ({len(matching_lessons)} öğr)"
                            print(f"  → ✅ SINIF DERSİ tespit edildi: '{student_name}'")
                    else:
                        # Fallback: Sınıf bilgisi yok, isimleri listele
                        student_name = matching_lessons[0]['student_name'].upper() + f" +{len(matching_lessons)-1}"
                        print(f"  → ⚠️ Fallback kullanıldı: '{student_name}'")

                    print(f"{'='*80}\n")

                html_content += f'''
                        <td class="student-name">{student_name}</td>
'''
            html_content += '''
                    </tr>
'''
        html_content += '''
                </tbody>
            </table>
        </div>
'''

    # ÖZET İSTATİSTİK
    html_content += '''
        <div class="stats-section">
            <div class="stats-header">ÖZET İSTATİSTİK</div>
            <table class="stats-table">
                <thead>
                    <tr>
                        <th>ÖĞRENCİ</th>
                        <th>SINIF</th>
                        <th>1. HAFTA</th>
                        <th>2. HAFTA</th>
                        <th>3. HAFTA</th>
                        <th>4. HAFTA</th>
                        <th>TOPLAM</th>
                    </tr>
                </thead>
                <tbody>
'''
    for stat in schedule_data['stats']:
        html_content += f'''
                    <tr>
                        <td class="student-name-col">{stat["student_name"]}</td>
                        <td>{stat["student_class"]}</td>
                        <td>{stat["week1"]}</td>
                        <td>{stat["week2"]}</td>
                        <td>{stat["week3"]}</td>
                        <td>{stat["week4"]}</td>
                        <td class="total-cell">{stat["total"]}</td>
                    </tr>
'''
    html_content += '''
                </tbody>
            </table>
        </div>
    </div>
</body>
</html>
'''

    response = make_response(html_content)
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=ders_programi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html'
    return response
# ═══════════════════════════════════════════════════════════════
# 🔥 SEVİYE 3: GELİŞMİŞ ÇAKIŞMA KONTROL SİSTEMİ
# ═══════════════════════════════════════════════════════════════

def time_to_minutes(time_str):
    """Saat stringini dakikaya çevir: '14:00' -> 840"""
    try:
        hours, mins = map(int, time_str.split(':'))
        return hours * 60 + mins
    except:
        return 0

def check_time_overlap(start1, end1, start2, end2):
    """İki zaman aralığının çakışıp çakışmadığını kontrol eder"""
    start1_mins = time_to_minutes(start1)
    end1_mins = time_to_minutes(end1)
    start2_mins = time_to_minutes(start2)
    end2_mins = time_to_minutes(end2)

    # Çakışma varsa True, yoksa False
    return not (end1_mins <= start2_mins or start1_mins >= end2_mins)

def detect_all_conflicts(schedule_data, teachers, students):
    """
    Tüm ihlalleri tespit et
    Returns: {
        'conflicts': [...],
        'group_lessons': [...],  # Onaylanmış grup dersleri
        'approved_violations': [...],  # 🆕 Force ile onaylanmış ihlalli dersler
        'grouped_violations': {...},  # 🆕 Öğretmen-gün-saat bazında gruplanmış ihlaller
        'summary': {...},
        'severity': 'low'/'medium'/'high'
    }
    """
    conflicts = []
    group_lessons = []  # Grup derslerini ayrı topla
    approved_violations = []  # 🆕 Force ile onaylanmış ihlalli dersleri topla
    forced_lesson_keys = set()  # 🆕 Force ile kaydedilen ders anahtarları

    if not schedule_data or not schedule_data.get('weeks'):
        return {'conflicts': [], 'group_lessons': [], 'approved_violations': [], 'grouped_violations': {}, 'summary': {}, 'severity': 'none'}

    # 🆕 FORCED DERSLERİ TESPİT ET
    for week_num, week_data in enumerate(schedule_data['weeks']):
        for lesson in week_data:
            if lesson.get('is_forced', 0) == 1:
                teacher_name = lesson.get('teacher_name', '')
                day = lesson.get('day', '')
                time = lesson.get('time', '')
                key = f"{teacher_name}_{day}_{time}_{week_num+1}"
                forced_lesson_keys.add(key)

    for week_num, week_data in enumerate(schedule_data['weeks']):
        # Her gün için öğrenci ve öğretmen slotlarını takip et
        daily_student_slots = {}
        daily_teacher_slots = {}

        for lesson in week_data:
            day = lesson['day']
            time = lesson['time']
            student_name = lesson['student_name']
            teacher_name = lesson['teacher_name']

            # Saat aralığını parse et
            try:
                start_time, end_time = time.split('-')
            except:
                continue

            # ÖĞRENCİ ÇAKIŞMASI KONTROLÜ
            student_key = f"{student_name}_{day}"
            if student_key not in daily_student_slots:
                daily_student_slots[student_key] = []

            # Bu öğrencinin bu gündeki diğer dersleriyle çakışma var mı?
            for existing_lesson in daily_student_slots[student_key]:
                if check_time_overlap(start_time, end_time, existing_lesson['start'], existing_lesson['end']):
                    conflicts.append({
                        'type': 'student',
                        'severity': 'high',
                        'week': week_num + 1,
                        'day': day,
                        'time': time,
                        'student': student_name,
                        'lesson1': f"{lesson['branch']} ({teacher_name})",
                        'lesson2': f"{existing_lesson['branch']} ({existing_lesson['teacher']})",
                        'message': f"⚠️ {student_name} - {day} {time}: {lesson['branch']} ile {existing_lesson['branch']} çakışıyor!"
                    })

            daily_student_slots[student_key].append({
                'start': start_time,
                'end': end_time,
                'branch': lesson['branch'],
                'teacher': teacher_name
            })

            # ÖĞRETMEN ÇAKIŞMASI KONTROLÜ
            teacher_key = f"{teacher_name}_{day}"
            if teacher_key not in daily_teacher_slots:
                daily_teacher_slots[teacher_key] = []

            # Bu öğretmenin bu gündeki diğer dersleriyle çakışma var mı?
            for existing_lesson in daily_teacher_slots[teacher_key]:
                if check_time_overlap(start_time, end_time, existing_lesson['start'], existing_lesson['end']):
                    # 🆕 GRUP DERSİ KONTROLÜ - Her ikisi de grup dersiyse çakışma sayılmaz!
                    current_is_group = lesson.get('is_group', 0)
                    existing_is_group = existing_lesson.get('is_group', 0)

                    if current_is_group == 1 and existing_is_group == 1:
                        # İkisi de grup dersi - onaylanmış grup dersi olarak kaydet!
                        # Tekrar eklememek için benzersiz key oluştur
                        group_key = f"{teacher_name}_{day}_{time}_week{week_num+1}"

                        # Katılımcı sınıfları topla
                        current_class = lesson.get('student_class', '')
                        existing_class = existing_lesson.get('student_class', '')

                        print(f"   Current class: '{current_class}' | Existing class: '{existing_class}'")

                        # Bu grup dersi daha önce eklenmişse güncelle, yoksa yeni ekle
                        existing_group = next((g for g in group_lessons if g.get('unique_key') == group_key), None)

                        if existing_group:

                            # Mevcut grup dersini güncelle - yeni sınıfları ekle
                            if current_class and current_class not in existing_group['classes']:
                                existing_group['classes'].append(current_class)
                                existing_group['classes'].sort()
                            if existing_class and existing_class not in existing_group['classes']:
                                existing_group['classes'].append(existing_class)
                                existing_group['classes'].sort()


                            # Mesajı güncelle
                            existing_group['message'] = f"✅ {teacher_name} - {day} {time}: Grup dersi ({', '.join(existing_group['classes'])})"
                            # ⚠️ İHLAL KONTROLÜ YAPMA - Zaten yapıldı!
                        else:
                            # Yeni grup dersi ekle
                            classes = set()
                            if current_class:
                                classes.add(current_class)
                            if existing_class:
                                classes.add(existing_class)


                            group_lessons.append({
                                'unique_key': group_key,
                                'type': 'group_lesson',
                                'week': week_num + 1,
                                'day': day,
                                'time': time,
                                'teacher': teacher_name,
                                'branch': lesson.get('branch', ''),
                                'classes': sorted(list(classes)),
                                'message': f"✅ {teacher_name} - {day} {time}: Grup dersi ({', '.join(sorted(list(classes)))})"
                            })

                            # 🆕 GRUP DERSİNDEKİ ÖĞRENCİLERİN İHLALLERİNİ KONTROL ET (SADECE İLK KERE!)

                            # 1️⃣ Bu slottaki TÜM grup dersi öğrencilerini bul
                            group_students_in_slot = []
                            for check_lesson in week_data:
                                if (check_lesson.get('teacher_name') == teacher_name and
                                    check_lesson.get('day') == day and
                                    check_lesson.get('time') == time and
                                    check_lesson.get('is_group') == 1):
                                    student_name_check = check_lesson.get('student_name', '')
                                    if student_name_check and student_name_check not in group_students_in_slot:
                                        group_students_in_slot.append(student_name_check)


                            # 3️⃣ Her öğrenci için ÖĞRENCİ ÇAKIŞMASI kontrolü
                            for student_name_to_check in group_students_in_slot:
                                # Bu öğrencinin bu gündeki diğer dersleriyle çakışma var mı?
                                student_check_key = f"{student_name_to_check}_{day}"
                                if student_check_key in daily_student_slots:
                                    for other_lesson in daily_student_slots[student_check_key]:
                                        if check_time_overlap(start_time, end_time, other_lesson['start'], other_lesson['end']):
                                            # ⚠️ AYNI GRUP DERSİ Mİ KONTROL ET!
                                            if (other_lesson.get('teacher') == teacher_name and
                                                other_lesson['start'] == start_time and
                                                other_lesson['end'] == end_time):
                                                # Bu aynı grup dersi, çakışma değil!
                                                continue

                                            # Farklı ders - çakışma var!
                                            conflicts.append({
                                                'type': 'student',
                                                'severity': 'high',
                                                'week': week_num + 1,
                                                'day': day,
                                                'time': time,
                                                'student': student_name_to_check,
                                                'teacher': teacher_name,
                                                'branch': lesson.get('branch', ''),
                                                'message': f"🔴 {student_name_to_check} - {day} {time}: Başka bir derse de kayıtlı! ({other_lesson.get('teacher', 'Bilinmeyen')} ile çakışma)"
                                            })

                            # 4️⃣ Her öğrenci için KISITLAMA ve ÖĞRETMEN ENGELİ kontrolü
                            for student_name_to_check in group_students_in_slot:
                                student_obj = next((s for s in students if f"{s['name']} {s['surname']}" == student_name_to_check), None)

                                if not student_obj:
                                    continue

                                # KISITLAMA KONTROLÜ
                                if student_obj.get('restrictions'):
                                    for restriction in student_obj['restrictions']:
                                        # Gün kontrolü
                                        days = restriction.get('days', [])
                                        if not days and restriction.get('day'):
                                            days = [restriction.get('day')]

                                        if day not in days:
                                            continue

                                        # Hafta kontrolü
                                        if restriction.get('type') == 'custom':
                                            weeks = restriction.get('weeks', [])
                                            if weeks and (week_num + 1) not in weeks:
                                                continue

                                        # Saat kontrolü
                                        rest_start = restriction.get('start_time', '')
                                        rest_end = restriction.get('end_time', '')

                                        if rest_start and rest_end:
                                            if check_time_overlap(start_time, end_time, rest_start, rest_end):
                                                conflicts.append({
                                                    'type': 'restriction',
                                                    'severity': 'medium',
                                                    'week': week_num + 1,
                                                    'day': day,
                                                    'time': time,
                                                    'student': student_name_to_check,
                                                    'teacher': teacher_name,
                                                    'branch': lesson.get('branch', ''),
                                                    'restriction': f"{rest_start}-{rest_end}",
                                                    'message': f"⚠️ {student_name_to_check} - {day} {time}: Kısıtlama saatinde ders! ({rest_start}-{rest_end})"
                                                })

                                # ÖĞRETMEN ENGELİ KONTROLÜ
                                teacher_blocks = student_obj.get('teacher_blocks', '[]')
                                if isinstance(teacher_blocks, str):
                                    try:
                                        teacher_blocks = json.loads(teacher_blocks)
                                    except:
                                        teacher_blocks = []

                                if teacher_blocks and teacher_name in teacher_blocks:
                                    conflicts.append({
                                        'type': 'teacher_block',
                                        'severity': 'medium',
                                        'week': week_num + 1,
                                        'day': day,
                                        'time': time,
                                        'student': student_name_to_check,
                                        'teacher': teacher_name,
                                        'branch': lesson.get('branch', ''),
                                        'message': f"🚫 {student_name_to_check} - {day} {time}: {teacher_name} engellenmiş öğretmen!"
                                    })

                    # 🆕 SINIF DERSİ KONTROLÜ - Aynı sınıftan mı?
                    elif lesson.get('student_class', '') and existing_lesson.get('student_class', '') and lesson.get('student_class', '') == existing_lesson.get('student_class', ''):
                        # Aynı sınıftan - sınıf dersi, çakışma sayılmaz
                        pass
                    else:
                        # Farklı sınıflar veya bireysel dersler - ÇAKIŞMA VAR!
                        conflicts.append({
                            'type': 'teacher',
                            'severity': 'critical',
                            'week': week_num + 1,
                            'day': day,
                            'time': time,
                            'teacher': teacher_name,
                            'student1': student_name,
                            'student2': existing_lesson['student'],
                            'message': f"🔴 {teacher_name} - {day} {time}: {student_name} ve {existing_lesson['student']} ile çakışma!"
                        })

            daily_teacher_slots[teacher_key].append({
                'start': start_time,
                'end': end_time,
                'student': student_name,
                'student_class': lesson.get('student_class', ''),  # 🆕 Sınıf bilgisini kaydet
                'is_group': lesson.get('is_group', 0)  # 🆕 Grup dersi bilgisini kaydet
            })

            # ÖĞRENCİ KISITLAMA KONTROLÜ
            student = next((s for s in students if f"{s['name']} {s['surname']}" == student_name), None)
            if student and student.get('restrictions'):
                for restriction in student['restrictions']:
                    # Gün kontrolü
                    days = restriction.get('days', [])
                    if not days and restriction.get('day'):
                        days = [restriction.get('day')]

                    if day not in days:
                        continue

                    # Hafta kontrolü
                    if restriction.get('type') == 'custom':
                        weeks = restriction.get('weeks', [])
                        if weeks and (week_num + 1) not in weeks:
                            continue

                    # Saat kontrolü
                    rest_start = restriction.get('start_time', '')
                    rest_end = restriction.get('end_time', '')

                    if rest_start and rest_end:
                        if check_time_overlap(start_time, end_time, rest_start, rest_end):
                            conflict_obj = {
                                'type': 'restriction',
                                'severity': 'medium',
                                'week': week_num + 1,
                                'day': day,
                                'time': time,
                                'student': student_name,
                                'branch': lesson['branch'],
                                'restriction': f"{rest_start}-{rest_end}",
                                'message': f"⚠️ {student_name} - {day} {time}: Kısıtlama saatinde ders! ({rest_start}-{rest_end})"
                            }
                            conflicts.append(conflict_obj)

                            # 🆕 Bu ders force ile onaylanmışsa approved_violations'a ekle
                            if lesson.get('is_class_lesson') and not lesson.get('is_group'):
                                key = f"{teacher_name}_{day}_{time}_{week_num+1}"
                                existing_approved = next((a for a in approved_violations if a.get('key') == key), None)

                                if not existing_approved:
                                    approved_violations.append({
                                        'key': key,
                                        'teacher': teacher_name,
                                        'day': day,
                                        'time': time,
                                        'branch': lesson['branch'],
                                        'week': week_num + 1,
                                        'class_name': lesson.get('student_class', ''),
                                        'violations': [conflict_obj]
                                    })
                                else:
                                    existing_approved['violations'].append(conflict_obj)

            # 🆕 ÖĞRETMEN ENGELİ KONTROLÜ
            student = next((s for s in students if f"{s['name']} {s['surname']}" == student_name), None)
            if student:
                # teacher_blocks kontrolü
                teacher_blocks = []
                if 'teacher_blocks' in student:
                    try:
                        teacher_blocks = json.loads(student['teacher_blocks']) if isinstance(student['teacher_blocks'], str) else student['teacher_blocks']
                    except:
                        teacher_blocks = []

                # Bu dersin öğretmenini bul
                teacher = next((t for t in teachers if f"{t['name']} {t['surname']}" == teacher_name), None)
                if teacher:
                    # Öğrenci bu öğretmeni engellemiş mi?
                    for block in teacher_blocks:
                        if block.get('teacher_id') == teacher['id']:
                            conflict_obj = {
                                'type': 'teacher_block',
                                'severity': 'medium',
                                'week': week_num + 1,
                                'day': day,
                                'time': time,
                                'student': student_name,
                                'teacher': teacher_name,
                                'branch': lesson['branch'],
                                'message': f"🚫 {student_name} - {day} {time}: {teacher_name} engellenmiş öğretmen!"
                            }
                            conflicts.append(conflict_obj)

                            # 🆕 Bu ders force ile onaylanmışsa approved_violations'a ekle
                            if lesson.get('is_class_lesson') and not lesson.get('is_group'):
                                # Sınıf dersi ama grup değil = force ile kaydedilmiş
                                key = f"{teacher_name}_{day}_{time}_{week_num+1}"
                                existing_approved = next((a for a in approved_violations if a.get('key') == key), None)

                                if not existing_approved:
                                    approved_violations.append({
                                        'key': key,
                                        'teacher': teacher_name,
                                        'day': day,
                                        'time': time,
                                        'branch': lesson['branch'],
                                        'week': week_num + 1,
                                        'class_name': lesson.get('student_class', ''),
                                        'violations': [conflict_obj]
                                    })
                                else:
                                    existing_approved['violations'].append(conflict_obj)
                            break

    # 🆕 GRUP DERSLERİ İÇİN ÖĞRETMEN ÇAKIŞMASI EKLE (tüm sınıflar toplandıktan sonra!)
    for group_lesson in group_lessons:
        conflicts.append({
            'type': 'teacher',
            'severity': 'critical',
            'week': group_lesson['week'],
            'day': group_lesson['day'],
            'time': group_lesson['time'],
            'teacher': group_lesson['teacher'],
            'branch': group_lesson.get('branch', ''),
            'message': f"🔴 ÖĞRETMEN ÇAKIŞMASI: {group_lesson['teacher']} birden fazla sınıfla aynı anda ders veriyor ({', '.join(group_lesson['classes'])})"
        })

    # 🆕 İHLALLERİ GRUPLA (öğretmen-gün-saat bazında)
    grouped_violations = {}

    # Grup derslerinin ihlallerini ekle
    for group_lesson in group_lessons:
        key = f"{group_lesson['teacher']}_{group_lesson['day']}_{group_lesson['time']}_{group_lesson['week']}"

        # Bu grup dersi için çakışmaları bul
        group_conflicts = [c for c in conflicts
                          if c.get('teacher') == group_lesson['teacher']
                          and c.get('day') == group_lesson['day']
                          and c.get('time') == group_lesson['time']
                          and c.get('week') == group_lesson['week']]

        if group_conflicts:
            grouped_violations[key] = {
                'teacher': group_lesson['teacher'],
                'day': group_lesson['day'],
                'time': group_lesson['time'],
                'branch': group_lesson['branch'],
                'week': group_lesson['week'],
                'is_group': True,
                'classes': group_lesson['classes'],
                'violations': group_conflicts
            }

    # Diğer ihlalleri grupla
    for conflict in conflicts:
        # Grup anahtarı oluştur
        if conflict['type'] in ['restriction', 'teacher_block']:
            # Öğrenci ihlalleri için dersin bilgilerini al
            teacher = conflict.get('teacher', 'Bilinmeyen')
            day = conflict.get('day', '')
            time = conflict.get('time', '')
            branch = conflict.get('branch', '')
            week = conflict.get('week', 1)

            key = f"{teacher}_{day}_{time}_{week}"

            # Zaten grup dersi olarak eklendiyse atla
            if key in grouped_violations:
                continue

            if key not in grouped_violations:
                grouped_violations[key] = {
                    'teacher': teacher,
                    'day': day,
                    'time': time,
                    'branch': branch,
                    'week': week,
                    'is_group': False,
                    'violations': []
                }

            grouped_violations[key]['violations'].append(conflict)
        elif conflict['type'] in ['student', 'teacher']:
            # Öğrenci/Öğretmen çakışmaları
            teacher = conflict.get('teacher', conflict.get('teacher_name', 'Bilinmeyen'))
            day = conflict.get('day', '')
            time = conflict.get('time', '')
            week = conflict.get('week', 1)

            key = f"{teacher}_{day}_{time}_{week}"

            if key in grouped_violations:
                continue

            if key not in grouped_violations:
                grouped_violations[key] = {
                    'teacher': teacher,
                    'day': day,
                    'time': time,
                    'branch': '',
                    'week': week,
                    'is_group': False,
                    'violations': []
                }

            grouped_violations[key]['violations'].append(conflict)

    # 🆕 FORCED DERSLERİN İHLALLERİNİ AYIR
    for key in list(grouped_violations.keys()):
        if key in forced_lesson_keys:
            # Bu ders force ile onaylanmış
            approved_violations.append(grouped_violations[key])
            # conflicts'ten çıkar
            conflicts = [c for c in conflicts if not (
                c.get('teacher', c.get('teacher_name', '')) == grouped_violations[key]['teacher'] and
                c.get('day', '') == grouped_violations[key]['day'] and
                c.get('time', '') == grouped_violations[key]['time'] and
                c.get('week', 0) == grouped_violations[key]['week']
            )]
            # grouped_violations'tan çıkar
            del grouped_violations[key]


    # 🆕 GRUP DERSLERİNE AİT İHLALLERİ conflicts'TEN ÇIKAR (çift sayılmasın!)
    group_violation_ids = set()
    for group_lesson in group_lessons:
        for c in conflicts:
            if (c.get('teacher') == group_lesson['teacher'] and
                c.get('day') == group_lesson['day'] and
                c.get('time') == group_lesson['time'] and
                c.get('week') == group_lesson['week']):
                group_violation_ids.add(id(c))

    # Grup dersi olmayan ihlaller
    non_group_conflicts = [c for c in conflicts if id(c) not in group_violation_ids]


    # 🆕 GROUPED_VIOLATIONS VE APPROVED_VIOLATIONS'TAKİ TÜM İHLALLERİ SAY
    # Grup derslerindeki ihlalleri say
    group_violation_count = 0
    group_critical = 0
    group_high = 0
    group_medium = 0
    for key, data in grouped_violations.items():
        violations = data.get('violations', [])
        group_violation_count += len(violations)
        group_critical += len([v for v in violations if v.get('severity') == 'critical'])
        group_high += len([v for v in violations if v.get('severity') == 'high'])
        group_medium += len([v for v in violations if v.get('severity') == 'medium'])

    # Onaylı ihlalleri say
    approved_violation_count = 0
    approved_critical = 0
    approved_high = 0
    approved_medium = 0
    for approved in approved_violations:
        violations = approved.get('violations', [])
        approved_violation_count += len(violations)
        approved_critical += len([v for v in violations if v.get('severity') == 'critical'])
        approved_high += len([v for v in violations if v.get('severity') == 'high'])
        approved_medium += len([v for v in violations if v.get('severity') == 'medium'])

    # ÖZET İSTATİSTİKLER (grup + onaylı ihlaller)
    total_all_violations = group_violation_count + approved_violation_count
    total_critical = group_critical + approved_critical
    total_high = group_high + approved_high
    total_medium = group_medium + approved_medium

    summary = {
        'total': total_all_violations,
        'critical': total_critical,
        'high': total_high,
        'medium': total_medium,
        'by_week': {}
    }

    # Haftalık dağılım (grup + onaylı)
    for i in range(1, 5):
        week_count = 0
        # Grup derslerinden
        for key, data in grouped_violations.items():
            if data.get('week') == i:
                week_count += len(data.get('violations', []))
        # Onaylı derslerden
        for approved in approved_violations:
            if approved.get('week') == i:
                week_count += len(approved.get('violations', []))
        summary['by_week'][f'week{i}'] = week_count


    # GENEL CİDDİYET SEVİYESİ
    if total_critical > 0:
        severity = 'critical'
    elif total_high > 2:
        severity = 'high'
    elif total_medium > 5:
        severity = 'medium'
    elif total_all_violations > 0:
        severity = 'low'
    else:
        severity = 'none'

    # 🔍 DEBUG - Grup derslerini logla
    for idx, gl in enumerate(group_lessons):
        print(f"  Grup Dersi {idx + 1}:")
        print(f"    Öğretmen: {gl['teacher']}")
        print(f"    Gün: {gl['day']} {gl['time']}")
        print(f"    Sınıflar: {gl['classes']}")
        print(f"    Sınıf Sayısı: {len(gl['classes'])}")

    return {
        'conflicts': conflicts,
        'group_lessons': group_lessons,
        'approved_violations': approved_violations,  # 🆕
        'grouped_violations': grouped_violations,  # 🆕
        'summary': summary,
        'severity': severity
    }

def detect_conflicts_v2(schedule_data, teachers, students):
    """
    🆕 YENİ İHLAL TESPİT FONKSİYONU

    Hiyerarşik yapı: 1 ders = 1 kart, içinde detaylar

    Returns: {
        'cards': [
            {
                'id': 'unique_id',
                'type': 'class_lesson' / 'group_lesson' / 'conflict',
                'approved': True/False,
                'severity': 'critical'/'high'/'medium',
                'teacher': 'Ayşe Çolak',
                'branch': 'Biyoloji',
                'class': '11A',
                'day': 'Pazartesi',
                'time': '14:50-15:30',
                'week': 1,
                'student_count': 3,
                'issues': [
                    {'student': 'Zeynep', 'type': 'restriction', 'detail': '14:00-16:20'},
                    ...
                ]
            }
        ],
        'summary': {
            'total': 1,
            'critical': 0,
            'high': 0,
            'medium': 1,
            'approved': 1,
            'unapproved': 0
        }
    }
    """

    if not schedule_data or not schedule_data.get('weeks'):
        return {'cards': [], 'summary': {'total': 0, 'critical': 0, 'high': 0, 'medium': 0, 'approved': 0, 'unapproved': 0}}

    cards = []
    processed_slots = set()  # Aynı dersi iki kere işlememek için

    # Her hafta için
    for week_num, week_data in enumerate(schedule_data['weeks'], 1):

        # Her slot için dersleri grupla
        slot_lessons = {}  # key: "teacher_day_time", value: [lessons]

        for lesson in week_data:
            teacher = lesson.get('teacher_name', '')
            day = lesson.get('day', '')
            time = lesson.get('time', '')
            key = f"{teacher}_{day}_{time}_{week_num}"

            if key not in slot_lessons:
                slot_lessons[key] = []
            slot_lessons[key].append(lesson)

        # Her slot için kart oluştur
        for slot_key, lessons in slot_lessons.items():
            if slot_key in processed_slots:
                continue
            processed_slots.add(slot_key)

            if not lessons:
                continue

            first_lesson = lessons[0]
            teacher = first_lesson.get('teacher_name', '')
            branch = first_lesson.get('branch', '')
            day = first_lesson.get('day', '')
            time = first_lesson.get('time', '')

            # Ders türünü belirle
            is_class_lesson = first_lesson.get('is_class_lesson', 0) == 1
            is_group = first_lesson.get('is_group', 0) == 1
            lesson_count = len(lessons)

            # İçindeki sorunları topla
            issues = []

            for lesson in lessons:
                student_name = lesson.get('student_name', '')
                student_class = lesson.get('student_class', '')

                # Bu öğrenciyi bul
                student = next((s for s in students if f"{s['name']} {s['surname']}" == student_name), None)

                if student:
                    # Kısıtlama kontrolü
                    if student.get('restrictions'):
                        for restriction in student['restrictions']:
                            # Gün kontrolü
                            days = restriction.get('days', [])
                            if not days and restriction.get('day'):
                                days = [restriction.get('day')]

                            if day not in days:
                                continue

                            # Hafta kontrolü
                            if restriction.get('type') == 'custom':
                                weeks = restriction.get('weeks', [])
                                if weeks and week_num not in weeks:
                                    continue

                            # Saat kontrolü
                            rest_start = restriction.get('start_time', '')
                            rest_end = restriction.get('end_time', '')

                            if rest_start and rest_end:
                                try:
                                    lesson_start, lesson_end = time.split('-')
                                    if check_time_overlap(lesson_start, lesson_end, rest_start, rest_end):
                                        issues.append({
                                            'student': student_name,
                                            'type': 'restriction',
                                            'detail': f"{rest_start}-{rest_end}"
                                        })
                                except:
                                    pass

                    # Öğretmen engeli kontrolü
                    teacher_blocks = student.get('teacher_blocks', '[]')
                    if isinstance(teacher_blocks, str):
                        try:
                            teacher_blocks = json.loads(teacher_blocks)
                        except:
                            teacher_blocks = []

                    # Bu dersin öğretmenini bul
                    teacher_obj = next((t for t in teachers if f"{t['name']} {t['surname']}" == teacher), None)
                    if teacher_obj:
                        for block in teacher_blocks:
                            if block.get('teacher_id') == teacher_obj['id']:
                                issues.append({
                                    'student': student_name,
                                    'type': 'teacher_block',
                                    'detail': f"{teacher} engellenmiş"
                                })
                                break

            # Kart türünü ve onay durumunu belirle
            if is_class_lesson and not is_group:
                # Sınıf dersi
                card_type = 'class_lesson'
                approved = True
                severity = 'medium'
                class_name = first_lesson.get('student_class', '')

                # 🆕 SINIF DERSİ KENDİSİ BİR İHLAL
                issues.insert(0, {
                    'student': f"{class_name} Sınıfı",
                    'type': 'class_lesson_creation',
                    'detail': f"{lesson_count} öğrenci ile sınıf dersi oluşturuldu"
                })

            elif is_group:
                # Grup dersi
                card_type = 'group_lesson'
                approved = True
                severity = 'medium'

                # 🆕 GRUP DERSİNDEKİ TÜM SINIFLARI BİRLEŞTİR
                unique_classes = list(set([l.get('student_class', '') for l in lessons if l.get('student_class')]))
                class_name = ', '.join(sorted(unique_classes))  # "12A, 12B" formatı


                # 🆕 GRUP DERSİNDEKİ HER SINIF BİR İHLAL
                for cls in unique_classes:
                    student_count_in_class = len([l for l in lessons if l.get('student_class') == cls])
                    issues.insert(0, {
                        'student': f"{cls} Sınıfı",
                        'type': 'group_lesson_participation',
                        'detail': f"Grup dersinde {student_count_in_class} öğrenci ile katılıyor"
                    })

            else:
                # Diğer durumlar (bireysel ders sorunları vs)
                if issues:
                    card_type = 'conflict'
                    approved = False
                    severity = 'high' if any(i['type'] == 'teacher_block' for i in issues) else 'medium'
                    class_name = first_lesson.get('student_class', '')
                else:
                    # Sorun yok, kart oluşturma
                    continue

            # Kartı oluştur
            card = {
                'id': slot_key,
                'type': card_type,
                'approved': approved,
                'severity': severity,
                'teacher': teacher,
                'branch': branch,
                'class': class_name,
                'day': day,
                'time': time,
                'week': week_num,
                'student_count': lesson_count,
                'students': [l.get('student_name', '') for l in lessons],
                'issues': issues
            }

            cards.append(card)

    # Özet istatistikler
    total = len(cards)
    critical = len([c for c in cards if c['severity'] == 'critical'])
    high = len([c for c in cards if c['severity'] == 'high'])
    medium = len([c for c in cards if c['severity'] == 'medium'])
    approved = len([c for c in cards if c['approved']])
    unapproved = len([c for c in cards if not c['approved']])

    # 🆕 TOPLAM SORUN SAYISI (tüm kartlardaki issues toplamı)
    total_issues = sum(len(c.get('issues', [])) for c in cards)

    return {
        'cards': cards,
        'summary': {
            'total': total,
            'total_issues': total_issues,  # 🆕 Yeni alan
            'critical': critical,
            'high': high,
            'medium': medium,
            'approved': approved,
            'unapproved': unapproved
        }
    }

@app.route('/check_conflicts', methods=['POST'])
def check_conflicts():
    """Mevcut program için çakışmaları kontrol et"""
    global schedule_data

    if not schedule_data:
        return jsonify({'error': 'Program bulunamadı!'}), 400

    # Öğretmen ve öğrencileri çek
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        blocked_slots = []
        try:
            if row['blocked_slots']:
                blocked_slots = json.loads(row['blocked_slots'])
        except:
            pass

        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule']),
            'blocked_slots': blocked_slots
        })

    cursor.execute('SELECT * FROM students')
    students = []
    for row in cursor.fetchall():
        students.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'class': row['class'],
            'restrictions': json.loads(row['restrictions']) if row['restrictions'] else [],
            'teacher_blocks': row['teacher_blocks'] if row['teacher_blocks'] else '[]'  # 🆕 Eklendi
        })

    conn.close()

    # Çakışmaları tespit et
    result = detect_all_conflicts(schedule_data, teachers, students)

    return jsonify(result)

@app.route('/check_conflicts_v2', methods=['POST'])
def check_conflicts_v2():
    """🆕 YENİ: Hiyerarşik ihlal kontrolü"""
    global schedule_data

    if not schedule_data:
        return jsonify({'error': 'Program bulunamadı!'}), 400

    # Öğretmen ve öğrencileri çek
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule'])
        })

    cursor.execute('SELECT * FROM students')
    students = []
    for row in cursor.fetchall():
        students.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'class': row['class'],
            'restrictions': json.loads(row['restrictions']) if row['restrictions'] else [],
            'teacher_blocks': row['teacher_blocks'] if row['teacher_blocks'] else '[]'
        })

    conn.close()

    # Yeni ihlal tespit sistemi
    result = detect_conflicts_v2(schedule_data, teachers, students)

    return jsonify(result)

@app.route('/suggest_alternative_slots', methods=['POST'])
def suggest_alternative_slots():
    """
    Bir ders için alternatif slotlar öner
    POST data: {
        'student_id': int,
        'teacher_id': int,
        'week': int,
        'current_day': str (optional),
        'current_time': str (optional)
    }
    """
    global schedule_data

    data = request.json
    student_id = data.get('student_id')
    teacher_id = data.get('teacher_id')
    week = data.get('week', 1)
    current_day = data.get('current_day')
    current_time = data.get('current_time')

    conn = get_db()
    cursor = conn.cursor()

    # Öğretmeni çek
    cursor.execute('SELECT * FROM teachers WHERE id=?', (teacher_id,))
    teacher_row = cursor.fetchone()
    if not teacher_row:
        conn.close()
        return jsonify({'error': 'Öğretmen bulunamadı!'}), 404

    teacher = {
        'id': teacher_row['id'],
        'name': teacher_row['name'],
        'surname': teacher_row['surname'],
        'branch': teacher_row['branch'],
        'schedule': json.loads(teacher_row['schedule'])
    }

    # Öğrenciyi çek
    cursor.execute('SELECT * FROM students WHERE id=?', (student_id,))
    student_row = cursor.fetchone()
    if not student_row:
        conn.close()
        return jsonify({'error': 'Öğrenci bulunamadı!'}), 404

    student = {
        'id': student_row['id'],
        'name': student_row['name'],
        'surname': student_row['surname'],
        'restrictions': json.loads(student_row['restrictions']) if student_row['restrictions'] else []
    }

    conn.close()

    # Mevcut programdaki slotları kontrol et
    occupied_slots = set()
    if schedule_data and schedule_data.get('weeks'):
        week_data = schedule_data['weeks'][week - 1]
        student_name = f"{student['name']} {student['surname']}"
        teacher_name = f"{teacher['name']} {teacher['surname']}"

        for lesson in week_data:
            if lesson['student_name'] == student_name:
                occupied_slots.add(f"{lesson['day']}_{lesson['time'].split('-')[0]}")
            if lesson['teacher_name'] == teacher_name:
                occupied_slots.add(f"{lesson['day']}_{lesson['time']}_teacher")

    # Alternatif slotlar bul
    suggestions = []

    for day_schedule in teacher['schedule']:
        day = day_schedule['day']

        for lesson_info in day_schedule['lessons']:
            start_time = lesson_info['start_time']
            end_time = lesson_info['end_time']
            slot_key = f"{day}_{start_time}"
            teacher_slot_key = f"{day}_{start_time}-{end_time}_teacher"

            # Mevcut slot ise atla
            if current_day == day and current_time == f"{start_time}-{end_time}":
                continue

            # Dolu slot ise atla
            if slot_key in occupied_slots or teacher_slot_key in occupied_slots:
                continue

            # Öğrenci kısıtlaması var mı kontrol et
            is_restricted = False
            if student.get('restrictions'):
                for restriction in student['restrictions']:
                    days = restriction.get('days', [])
                    if not days and restriction.get('day'):
                        days = [restriction.get('day')]

                    if day not in days:
                        continue

                    # Hafta kontrolü
                    if restriction.get('type') == 'custom':
                        weeks = restriction.get('weeks', [])
                        if weeks and week not in weeks:
                            continue

                    rest_start = restriction.get('start_time', '')
                    rest_end = restriction.get('end_time', '')

                    if rest_start and rest_end:
                        if check_time_overlap(start_time, end_time, rest_start, rest_end):
                            is_restricted = True
                            break

            if is_restricted:
                continue

            # Skor hesapla (basit öncelik sistemi)
            score = 100

            # Aynı gündeki dersler tercih edilir
            if current_day and day == current_day:
                score += 20

            # Sabah saatleri tercih edilir
            hour = int(start_time.split(':')[0])
            if 9 <= hour <= 12:
                score += 10

            suggestions.append({
                'day': day,
                'time': f"{start_time}-{end_time}",
                'start_time': start_time,
                'end_time': end_time,
                'score': score,
                'teacher': f"{teacher['name']} {teacher['surname']}",
                'branch': teacher['branch']
            })

    # Skora göre sırala
    suggestions.sort(key=lambda x: x['score'], reverse=True)

    return jsonify({
        'suggestions': suggestions[:10],  # En iyi 10 öneriyi gönder
        'total_available': len(suggestions)
    })

@app.route('/get_teacher_timeline/<int:teacher_id>')
def get_teacher_timeline(teacher_id):
    """Öğretmenin 4 haftalık timeline'ını döndür"""
    global schedule_data

    if not schedule_data:
        return jsonify({'error': 'Program bulunamadı!'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM teachers WHERE id=?', (teacher_id,))
    teacher_row = cursor.fetchone()
    conn.close()

    if not teacher_row:
        return jsonify({'error': 'Öğretmen bulunamadı!'}), 404

    teacher_name = f"{teacher_row['name']} {teacher_row['surname']}"

    # Timeline datasını oluştur
    timeline = {
        'teacher': teacher_name,
        'branch': teacher_row['branch'],
        'weeks': []
    }

    for week_num, week_data in enumerate(schedule_data['weeks']):
        week_timeline = {'week': week_num + 1, 'days': {}}

        for lesson in week_data:
            if lesson['teacher_name'] == teacher_name:
                day = lesson['day']
                if day not in week_timeline['days']:
                    week_timeline['days'][day] = []

                week_timeline['days'][day].append({
                    'time': lesson['time'],
                    'student': lesson['student_name'],
                    'student_class': lesson['student_class']
                })

        # Günleri sırala
        day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
        sorted_days = {}
        for day in day_order:
            if day in week_timeline['days']:
                # Saatlere göre sırala
                week_timeline['days'][day].sort(key=lambda x: x['time'])
                sorted_days[day] = week_timeline['days'][day]

        week_timeline['days'] = sorted_days
        timeline['weeks'].append(week_timeline)

    return jsonify(timeline)

@app.route('/get_student_timeline/<int:student_id>')
def get_student_timeline(student_id):
    """Öğrencinin 4 haftalık timeline'ını döndür"""
    global schedule_data

    if not schedule_data:
        return jsonify({'error': 'Program bulunamadı!'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM students WHERE id=?', (student_id,))
    student_row = cursor.fetchone()
    conn.close()

    if not student_row:
        return jsonify({'error': 'Öğrenci bulunamadı!'}), 404

    student_name = f"{student_row['name']} {student_row['surname']}"

    # Timeline datasını oluştur
    timeline = {
        'student': student_name,
        'class': student_row['class'],
        'weeks': []
    }

    for week_num, week_data in enumerate(schedule_data['weeks']):
        week_timeline = {'week': week_num + 1, 'days': {}}

        for lesson in week_data:
            if lesson['student_name'] == student_name:
                day = lesson['day']
                if day not in week_timeline['days']:
                    week_timeline['days'][day] = []

                week_timeline['days'][day].append({
                    'time': lesson['time'],
                    'teacher': lesson['teacher_name'],
                    'branch': lesson['branch']
                })

        # Günleri sırala
        day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
        sorted_days = {}
        for day in day_order:
            if day in week_timeline['days']:
                # Saatlere göre sırala
                week_timeline['days'][day].sort(key=lambda x: x['time'])
                sorted_days[day] = week_timeline['days'][day]

        week_timeline['days'] = sorted_days
        timeline['weeks'].append(week_timeline)

    return jsonify(timeline)

@app.route('/auto_fix_conflicts', methods=['POST'])
def auto_fix_conflicts():
    """
    Çakışmaları otomatik olarak düzeltmeye çalış
    NOT: Bu fonksiyon sadece basit çakışmaları düzeltir
    """
    global schedule_data

    if not schedule_data:
        return jsonify({'error': 'Program bulunamadı!'}), 400

    # Öğretmen ve öğrencileri çek
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule'])
        })

    cursor.execute('SELECT * FROM students')
    students = []
    for row in cursor.fetchall():
        students.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'restrictions': json.loads(row['restrictions']) if row['restrictions'] else []
        })

    conn.close()

    # Çakışmaları tespit et
    conflicts_result = detect_all_conflicts(schedule_data, teachers, students)
    conflicts = conflicts_result['conflicts']

    if not conflicts:
        return jsonify({'message': 'Çakışma bulunamadı!', 'fixed': 0})

    fixed_count = 0
    fixed_details = []

    # Sadece öğrenci çakışmalarını düzeltmeye çalış (öğretmen çakışmaları daha kritik)
    student_conflicts = [c for c in conflicts if c['type'] == 'student']

    for conflict in student_conflicts:
        week_idx = conflict['week'] - 1
        week_data = schedule_data['weeks'][week_idx]

        # Çakışan dersleri bul
        conflicting_lessons = [
            l for l in week_data
            if l['student_name'] == conflict['student']
            and l['day'] == conflict['day']
            and check_time_overlap(
                l['time'].split('-')[0],
                l['time'].split('-')[1],
                conflict['time'].split('-')[0],
                conflict['time'].split('-')[1]
            )
        ]

        if len(conflicting_lessons) < 2:
            continue

        # İkinci dersi başka bir slota taşımaya çalış
        lesson_to_move = conflicting_lessons[1]

        # Öğretmeni bul
        teacher = next((t for t in teachers if f"{t['name']} {t['surname']}" == lesson_to_move['teacher_name']), None)
        if not teacher:
            continue

        # Öğrenciyi bul
        student = next((s for s in students if f"{s['name']} {s['surname']}" == conflict['student']), None)
        if not student:
            continue

        # Alternatif slot bulmak karmaşık olduğu için şimdilik basitleştirilmiş versiyon:
        # Çakışan dersi kaldır (gerçek uygulamada alternatif slot bulunmalı)
        schedule_data['weeks'][week_idx] = [
            l for l in week_data
            if not (l == lesson_to_move)
        ]

        fixed_count += 1
        fixed_details.append({
            'student': conflict['student'],
            'week': conflict['week'],
            'day': conflict['day'],
            'removed_lesson': f"{lesson_to_move['branch']} ({lesson_to_move['time']})"
        })

    return jsonify({
        'message': f'{fixed_count} çakışma düzeltildi!',
        'fixed': fixed_count,
        'details': fixed_details,
        'remaining': len(conflicts) - fixed_count
    })

# ═══════════════════════════════════════════════════════════════
# 🔥 PART 1 SONU - Backend Fonksiyonları Eklendi
# ═══════════════════════════════════════════════════════════════

@app.route('/export_conflict_report')
def export_conflict_report():
    """Çakışma raporunu Excel olarak indir"""
    global schedule_data

    if not schedule_data:
        return "Program bulunamadı!", 400

    # Öğretmen ve öğrencileri çek
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        teachers.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'branch': row['branch'],
            'schedule': json.loads(row['schedule'])
        })

    cursor.execute('SELECT * FROM students')
    students = []
    for row in cursor.fetchall():
        students.append({
            'id': row['id'],
            'name': row['name'],
            'surname': row['surname'],
            'restrictions': json.loads(row['restrictions']) if row['restrictions'] else []
        })

    conn.close()

    # Çakışmaları tespit et
    conflicts_result = detect_all_conflicts(schedule_data, teachers, students)
    conflicts = conflicts_result['conflicts']

    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Çakışma Raporu"

    # Başlık satırı
    headers = ['📋 Tip', '⚠️ Ciddiyet', '📅 Hafta', '📆 Gün', '🕐 Saat', '📝 Detay']
    ws.append(headers)

    # Başlık stilini ayarla
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Çakışmaları ekle
    for conflict in conflicts:
        if conflict['type'] == 'student':
            detail = f"{conflict['student']}: {conflict['lesson1']} ↔ {conflict['lesson2']}"
        elif conflict['type'] == 'teacher':
            detail = f"{conflict['teacher']}: {conflict['student1']} ↔ {conflict['student2']}"
        else:
            detail = f"{conflict['student']}: {conflict['branch']} ({conflict['restriction']})"

        # 🆕 TÜRKÇE ÇEVİRİ
        type_tr = {
            'student': 'ÖĞRENCİ',
            'teacher': 'ÖĞRETMEN',
            'restriction': 'KISITLAMA'
        }.get(conflict['type'], conflict['type'].upper())

        severity_tr = {
            'critical': 'KRİTİK',
            'high': 'YÜKSEK',
            'medium': 'ORTA',
            'low': 'DÜŞÜK'
        }.get(conflict['severity'], conflict['severity'].upper())

        row = [
            type_tr,
            severity_tr,
            conflict['week'],
            conflict['day'],
            conflict['time'],
            detail
        ]
        ws.append(row)

        # Ciddiyet rengini ayarla
        row_idx = ws.max_row
        if conflict['severity'] == 'critical':
            fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        elif conflict['severity'] == 'high':
            fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

        for cell in ws[row_idx]:
            cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Kolon genişliklerini ayarla
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 60

    # Excel'i kaydet
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cakisma_raporu_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/export_weekly_pdf_server/<int:week_num>')
def export_weekly_pdf_server(week_num):
    if not schedule_data:
        return "Program bulunamadı!", 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        teachers.append({'name': row['name'], 'surname': row['surname'], 'branch': row['branch'], 'schedule': json.loads(row['schedule'])})
    teachers.sort(key=lambda t: (t['branch'], t['name'], t['surname']))
    conn.close()

    all_slots = []
    for teacher in teachers:
        for day_schedule in teacher['schedule']:
            for lesson in day_schedule['lessons']:
                slot_key = f"{day_schedule['day']}_{lesson['start_time']}_{lesson['end_time']}"
                if not any(s['key'] == slot_key for s in all_slots):
                    all_slots.append({'day': day_schedule['day'], 'start_time': lesson['start_time'], 'end_time': lesson['end_time'], 'key': slot_key})

    day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
    all_slots.sort(key=lambda x: (day_order.index(x['day']), x['start_time']))
    week_data = schedule_data['weeks'][week_num - 1]

    html = f'<html><head><meta charset="UTF-8"><style>@page{{size:A4 landscape;margin:5mm}}body{{font:6pt Arial}}table{{width:100%;border-collapse:collapse;table-layout:fixed}}caption{{background:#667eea;color:white;padding:10px;font-weight:bold;font-size:12pt;text-align:center}}th{{background:#4472C4;color:white;padding:6px 3px;border:1px solid rgba(255,255,255,0.2);font-size:6pt;line-height:1.2;text-align:center}}th:first-child{{width:50px}}td{{padding:5px 2px;border:1px solid #e5e7eb;font-size:6pt;text-align:center;line-height:1.2;word-wrap:break-word}}.day-header{{background:#9575CD!important;color:white!important;font-weight:bold;padding:6px!important;font-size:7pt}}.time-cell{{background:#E3F2FD;color:#1565C0;font-weight:600;text-align:center;white-space:nowrap}}</style></head><body><table><caption>HAFTA {week_num}</caption><colgroup><col style="width:50px">'
    for _ in teachers:
        html += '<col>'
    html += '</colgroup><thead><tr><th>GÜN/SAAT</th>'

    for t in teachers:
        html += f'<th>{t["branch"].upper()}<br><span style="font-size:5pt">({t["name"].upper()} {t["surname"].upper()})</span></th>'
    html += '</tr></thead><tbody>'

    current_day = None
    for slot in all_slots:
        if slot['day'] != current_day:
            html += f'<tr><td colspan="{len(teachers)+1}" class="day-header">{slot["day"].upper()}</td></tr>'
            current_day = slot['day']
        html += f'<tr><td class="time-cell">{slot["start_time"]}-{slot["end_time"]}</td>'
        for teacher in teachers:
            teacher_full = f"{teacher['name']} {teacher['surname']}"

            # 🆕 TÜM eşleşen dersleri bul
            matching_lessons = [
                lesson for lesson in week_data
                if (lesson['teacher_name'] == teacher_full and
                    lesson['day'] == slot['day'] and
                    lesson['time'] == f"{slot['start_time']}-{slot['end_time']}")
            ]

            # 🆕 Gruplama mantığı
            if len(matching_lessons) == 0:
                student = ''
            elif len(matching_lessons) == 1:
                student = matching_lessons[0]['student_name'].upper()
            else:
                # 🆕 GRUP DERSİ - TÜM SINIFLARI TOPLA
                unique_classes = list(set([l.get('student_class', '') for l in matching_lessons if l.get('student_class')]))
                if unique_classes:
                    classes_str = ', '.join(sorted(unique_classes))
                    student = f"{classes_str} ({len(matching_lessons)} öğr)"
                else:
                    student = f"{matching_lessons[0]['student_name'].upper()} +{len(matching_lessons)-1}"

            html += f'<td>{student}</td>'
        html += '</tr>'
    html += '</tbody></table></body></html>'

    pdf = HTML(string=html).write_pdf()
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=Hafta_{week_num}.pdf'
    return response

@app.route('/export_all_weeks_pdf_server')
def export_all_weeks_pdf_server():
    if not schedule_data:
        return "Program bulunamadı!", 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM teachers')
    teachers = []
    for row in cursor.fetchall():
        teachers.append({'name': row['name'], 'surname': row['surname'], 'branch': row['branch'], 'schedule': json.loads(row['schedule'])})
    teachers.sort(key=lambda t: (t['branch'], t['name'], t['surname']))
    conn.close()

    all_slots = []
    for teacher in teachers:
        for day_schedule in teacher['schedule']:
            for lesson in day_schedule['lessons']:
                slot_key = f"{day_schedule['day']}_{lesson['start_time']}_{lesson['end_time']}"
                if not any(s['key'] == slot_key for s in all_slots):
                    all_slots.append({'day': day_schedule['day'], 'start_time': lesson['start_time'], 'end_time': lesson['end_time'], 'key': slot_key})

    day_order = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
    all_slots.sort(key=lambda x: (day_order.index(x['day']), x['start_time']))

    html = '<html><head><meta charset="UTF-8"><style>@page{size:A4 landscape;margin:5mm}body{font:6pt Arial}.week-section{page-break-after:always}.week-section:last-child{page-break-after:auto}table{width:100%;border-collapse:collapse;margin-bottom:20px;table-layout:fixed}caption{background:#667eea;color:white;padding:10px;font-weight:bold;font-size:12pt;text-align:center}th{background:#4472C4;color:white;padding:6px 3px;border:1px solid rgba(255,255,255,0.2);font-size:6pt;line-height:1.2;text-align:center}th:first-child{width:50px}td{padding:5px 2px;border:1px solid #e5e7eb;font-size:6pt;text-align:center;line-height:1.2;word-wrap:break-word}.day-header{background:#9575CD!important;color:white!important;font-weight:bold;padding:6px!important;font-size:7pt}.time-cell{background:#E3F2FD;color:#1565C0;font-weight:600;text-align:center;white-space:nowrap}</style></head><body>'
    for week_num in range(1, 5):
        week_data = schedule_data['weeks'][week_num - 1]
        html += f'<div class="week-section"><table><caption>HAFTA {week_num}</caption><colgroup><col style="width:50px">'
        for _ in teachers:
            html += '<col>'
        html += '</colgroup><thead><tr><th>GÜN/SAAT</th>'
        for t in teachers:
            html += f'<th>{t["branch"].upper()}<br><span style="font-size:5pt">({t["name"].upper()} {t["surname"].upper()})</span></th>'
        html += '</tr></thead><tbody>'
        current_day = None
        for slot in all_slots:
            if slot['day'] != current_day:
                html += f'<tr><td colspan="{len(teachers)+1}" class="day-header">{slot["day"].upper()}</td></tr>'
                current_day = slot['day']
            html += f'<tr><td class="time-cell">{slot["start_time"]}-{slot["end_time"]}</td>'
            for teacher in teachers:
                teacher_full = f"{teacher['name']} {teacher['surname']}"

                # 🆕 TÜM eşleşen dersleri bul
                matching_lessons = [
                    lesson for lesson in week_data
                    if (lesson['teacher_name'] == teacher_full and
                        lesson['day'] == slot['day'] and
                        lesson['time'] == f"{slot['start_time']}-{slot['end_time']}")
                ]

                # 🆕 Gruplama mantığı
                if len(matching_lessons) == 0:
                    student = ''
                elif len(matching_lessons) == 1:
                    student = matching_lessons[0]['student_name'].upper()
                else:
                    # 🆕 GRUP DERSİ - TÜM SINIFLARI TOPLA
                    unique_classes = list(set([l.get('student_class', '') for l in matching_lessons if l.get('student_class')]))
                    if unique_classes:
                        classes_str = ', '.join(sorted(unique_classes))
                        student = f"{classes_str} ({len(matching_lessons)} öğr)"
                    else:
                        student = f"{matching_lessons[0]['student_name'].upper()} +{len(matching_lessons)-1}"

                html += f'<td>{student}</td>'
            html += '</tr>'
        html += '</tbody></table></div>'
    html += '</body></html>'

    pdf = HTML(string=html).write_pdf()
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=4_Haftalik_Program.pdf'
    return response

@app.route('/swap_lessons', methods=['POST'])
def swap_lessons():
    """Sürükle-bırak ile ders değiştirme - Sınıf dersi swap desteği"""
    global schedule_data

    if not schedule_data:
        return jsonify({'error': 'Program bulunamadı!'}), 400

    data = request.json
    week = data.get('week', 1)
    source = data.get('source', {})
    target = data.get('target', {})

    if week < 1 or week > 4:
        return jsonify({'error': 'Geçersiz hafta numarası!'}), 400

    week_idx = week - 1
    week_data = schedule_data['weeks'][week_idx]

    # KAYNAK SINIF DERSI MI?
    source_is_class = source.get('isClassLesson', False)
    source_student_names = source.get('studentNames', [source.get('student')])
    source_teacher = source.get('teacher', '')  # ✅ Kaynak öğretmen bilgisi

    # HEDEF BOSSA (target.student None ise)
    if not target.get('student'):
        # BOŞ SLOTA TAŞIMA - SINIF veya BİREYSEL
        moved_lessons = []

        for student_name in source_student_names:
            for idx, lesson in enumerate(week_data):
                # ✅ Öğretmen kontrolü eklendi
                if (lesson['day'] == source['day'] and
                    lesson['time'] == source['time'] and
                    lesson['student_name'] == student_name and
                    lesson['teacher_name'] == source_teacher):

                    lesson['day'] = target['day']
                    lesson['time'] = target['time']
                    moved_lessons.append(lesson)
                    break

        if not moved_lessons:
            return jsonify({'error': 'Kaynak ders bulunamadı!'}), 404

        return jsonify({
            'message': f'Ders başarıyla taşındı! ({len(moved_lessons)} öğrenci)' if source_is_class else 'Ders başarıyla taşındı!',
            'swapped': False,
            'moved': moved_lessons
        })

    # HEDEF DOLU - SWAP YAPILACAK
    # Hedef taraftaki TÜM dersleri bul (sınıf dersi olabilir)
    target_student_text = target.get('student')  # "11A (3 öğrenci)" veya "ZEYNEP YAVUZ"
    target_teacher = target.get('teacher', '')  # ✅ Hedef öğretmen bilgisi

    # ✅ Hedef slottaki dersleri bul - SADECE BU ÖĞRETMENİN DERSLERİ
    target_lessons = []
    for lesson in week_data:
        if (lesson['day'] == target['day'] and
            lesson['time'] == target['time'] and
            lesson['teacher_name'] == target_teacher):  # ✅ Öğretmen filtresi eklendi
            target_lessons.append(lesson)

    if not target_lessons:
        return jsonify({'error': 'Hedef ders(ler) bulunamadı!'}), 404

    # Hedef sınıf dersi mi? (Aynı slotta birden fazla ders var mı?)
    target_is_class = len(target_lessons) > 1
    target_student_names = [l['student_name'] for l in target_lessons]

    # ✅ KAYNAK DERSLERİ BUL - SADECE BU ÖĞRETMENİN DERSLERİ
    source_lessons = []
    for student_name in source_student_names:
        for lesson in week_data:
            if (lesson['day'] == source['day'] and
                lesson['time'] == source['time'] and
                lesson['student_name'] == student_name and
                lesson['teacher_name'] == source_teacher):  # ✅ Öğretmen filtresi eklendi
                source_lessons.append(lesson)
                break

    if not source_lessons:
        return jsonify({'error': 'Kaynak ders(ler) bulunamadı!'}), 404

    # YER DEĞİŞTİR (SWAP)
    # Geçici olarak kaynak bilgilerini sakla
    temp_day = source['day']
    temp_time = source['time']
    temp_teacher = source_teacher  # ✅ Kaynak öğretmeni de sakla

    # Kaynak dersleri hedef slota taşı
    for lesson in source_lessons:
        lesson['day'] = target['day']
        lesson['time'] = target['time']
        lesson['teacher_name'] = target_teacher  # ✅ Hedef öğretmene değiştir

    # Hedef dersleri kaynak slota taşı
    for lesson in target_lessons:
        lesson['day'] = temp_day
        lesson['time'] = temp_time
        lesson['teacher_name'] = temp_teacher  # ✅ Kaynak öğretmene değiştir

    swap_type = ''
    if source_is_class and target_is_class:
        swap_type = f'Sınıf dersleri yer değiştirdi! ({len(source_lessons)} ↔ {len(target_lessons)} öğrenci)'
    elif source_is_class:
        swap_type = f'Sınıf dersi ile bireysel ders yer değiştirdi! ({len(source_lessons)} öğrenci ↔ 1)'
    elif target_is_class:
        swap_type = f'Bireysel ders ile sınıf dersi yer değiştirdi! (1 ↔ {len(target_lessons)} öğrenci)'
    else:
        swap_type = 'Dersler başarıyla yer değiştirdi!'

    return jsonify({
        'message': swap_type,
        'swapped': True,
        'source_count': len(source_lessons),
        'target_count': len(target_lessons)
    })

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 📚 SINIF DERSİ YÖNETİMİ ENDPOINT'LERİ
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route('/get_unique_classes')
def get_unique_classes():
    """Veritabanındaki benzersiz sınıfları getir"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT DISTINCT class
        FROM students
        WHERE class IS NOT NULL AND class != ''
        ORDER BY class
    ''')

    classes = [row['class'] for row in cursor.fetchall()]
    conn.close()

    return jsonify(classes)

@app.route('/get_students_by_class/<class_name>')
def get_students_by_class(class_name):
    """Belirli bir sınıftaki öğrencileri getir"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT id, name, surname
        FROM students
        WHERE class = ?
        ORDER BY name, surname
    ''', (class_name,))

    students = []
    for row in cursor.fetchall():
        students.append({
            'id': row['id'],
            'name': f"{row['name']} {row['surname']}"
        })

    conn.close()
    return jsonify({'students': students, 'count': len(students)})

@app.route('/save_class_lesson', methods=['POST'])
def save_class_lesson():
    """Sınıf dersini kaydet - çakışma kontrolü ile"""
    data = request.json

    print("🔥 SAVE_CLASS_LESSON ÇAĞRILDI!")
    print(f"📥 Gelen data: {data}")

    # Gerekli alanları kontrol et
    required_fields = ['class_name', 'teacher_id', 'day', 'start_time', 'end_time', 'weeks']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'{field} alanı eksik!'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # Sınıftaki öğrencileri getir
    cursor.execute('SELECT id, name, surname, restrictions, teacher_blocks FROM students WHERE class = ?',
                   (data['class_name'],))
    students = cursor.fetchall()

    print(f"👥 {data['class_name']} sınıfında {len(students)} öğrenci bulundu")

    if not students:
        conn.close()
        return jsonify({'error': 'Bu sınıfta öğrenci bulunamadı!'}), 400

    # Öğretmeni getir
    cursor.execute('SELECT id, name, surname, branch, schedule FROM teachers WHERE id = ?',
                   (data['teacher_id'],))
    teacher = cursor.fetchone()

    if not teacher:
        conn.close()
        return jsonify({'error': 'Öğretmen bulunamadı!'}), 400

    print(f"👨‍🏫 Öğretmen: {teacher['name']} {teacher['surname']} ({teacher['branch']})")

    # 🆕 1. AŞAMA: ÖĞRETMEN ÇAKIŞMASI ONAY KONTROLÜ
    # Sınıf dersi = öğretmen çakışması demektir (aynı öğretmen birden fazla öğrenciye ders veriyor)
    force_teacher_conflict = data.get('force_teacher_conflict', False)
    print(f"🔒 Force Teacher Conflict: {force_teacher_conflict}")

    if not force_teacher_conflict:
        # Kullanıcı henüz öğretmen çakışmasını onaylamadı
        # Uyarı ver ve onay iste
        conn.close()
        return jsonify({
            'success': False,
            'teacher_conflict_warning': True,
            'message': f"⚠️ ÖĞRETMEN ÇAKIŞMASI!\n\n{data['class_name']} sınıfına {teacher['name']} {teacher['surname']} öğretmeni atanacak.\n\n{data['day']} günü {data['start_time']}-{data['end_time']} saatinde bu sınıftaki tüm öğrenciler aynı öğretmenden ders alacak.\n\nBu işlem öğretmen çakışması oluşturacaktır. Devam etmek istiyor musunuz?"
        })

    # 🆕 2. AŞAMA: ÖĞRENCİ KISITLAMALARI VE DİĞER KONTROLLER
    # Kullanıcı force=True göndermemişse uyarıları kontrol et
    force_mode = data.get('force', False)
    print(f"🔒 Force modu: {force_mode}")

    if not force_mode:
        print("✅ Çakışma kontrolü BAŞLIYOR...")
        warnings = []

        # ═══════════════════════════════════════════════════════════
        # 🆕 ÖNCELİKLE SINIF VE ÖĞRETMEN ÇAKIŞMALARINI KONTROL ET
        # ═══════════════════════════════════════════════════════════

        print(f"🔍 Çakışma kontrolü: {data['class_name']} - {data['day']} {data['start_time']}-{data['end_time']}")

        # 1️⃣ AYNI SINIFA AYNI GÜN/SAATTE BAŞKA DERS ATANMIŞ MI? (aynı öğretmen dahil!)
        cursor.execute('''
            SELECT t.name, t.surname, t.branch, cl.start_time, cl.end_time, cl.weeks, cl.teacher_id
            FROM class_lessons cl
            JOIN teachers t ON cl.teacher_id = t.id
            WHERE cl.class_name = ?
            AND cl.day = ?
        ''', (data['class_name'], data['day']))

        existing_class_lessons = cursor.fetchall()
        print(f"📋 Sınıf için mevcut dersler: {len(existing_class_lessons)}")

        for existing in existing_class_lessons:
            print(f"  Kontrol: {existing['start_time']}-{existing['end_time']} vs {data['start_time']}-{data['end_time']}")
            # Saat çakışması var mı kontrol et
            if check_time_overlap(data['start_time'], data['end_time'],
                                  existing['start_time'], existing['end_time']):
                print(f"  ⚠️ Saat çakışması bulundu!")
                # Hafta kontrolü
                requested_weeks = data['weeks'].split(',') if data['weeks'] != 'all' else ['1', '2', '3', '4']
                existing_weeks = existing['weeks'].split(',') if existing['weeks'] != 'all' else ['1', '2', '3', '4']

                common_weeks = set(requested_weeks) & set(existing_weeks)
                print(f"  Ortak haftalar: {common_weeks}")
                if common_weeks:
                    week_text = ', '.join(sorted(common_weeks)) if len(common_weeks) < 4 else 'Tüm haftalarda'

                    # AYNI ÖĞRETMEN Mİ FARKLI ÖĞRETMEN Mİ?
                    if existing['teacher_id'] == data['teacher_id']:
                        # AYNI ÖĞRETMEN - Tekrar atama!
                        error_msg = f"🚫 TEKRAR ATAMA!\n\n{data['class_name']} sınıfının {data['day']} günü {data['start_time']}-{data['end_time']} saatinde zaten bu ders kayıtlı!\n\nHafta: {week_text}\n\nAynı sınıfa aynı gün ve saatte aynı dersi tekrar atayamazsınız!"
                    else:
                        # FARKLI ÖĞRETMEN
                        error_msg = f"🚫 SINIF ÇAKIŞMASI!\n\n{data['class_name']} sınıfının {data['day']} günü {data['start_time']}-{data['end_time']} saatinde zaten {existing['name']} {existing['surname']} ({existing['branch']}) ile dersi var!\n\nHafta: {week_text}\n\nAynı sınıfa aynı gün ve saatte iki öğretmen atanamaz!"

                    conn.close()
                    print(f"🚫 HATA: {error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    })

        # 2️⃣ AYNI ÖĞRETMENE AYNI GÜN/SAATTE BAŞKA SINIF ATANMIŞ MI?
        cursor.execute('''
            SELECT cl.class_name, cl.start_time, cl.end_time, cl.weeks
            FROM class_lessons cl
            WHERE cl.teacher_id = ?
            AND cl.day = ?
            AND cl.class_name != ?
        ''', (data['teacher_id'], data['day'], data['class_name']))

        existing_teacher_lessons = cursor.fetchall()
        print(f"📋 Öğretmen için mevcut dersler: {len(existing_teacher_lessons)}")

        # 🎯 ÖNCELİKLE: Aynı slot'taki TÜM sınıfları topla
        all_conflicting_classes = []
        for lesson in existing_teacher_lessons:
            if check_time_overlap(data['start_time'], data['end_time'],
                                  lesson['start_time'], lesson['end_time']):
                # Hafta kontrolü
                req_weeks = data['weeks'].split(',') if data['weeks'] != 'all' else ['1', '2', '3', '4']
                les_weeks = lesson['weeks'].split(',') if lesson['weeks'] != 'all' else ['1', '2', '3', '4']
                if set(req_weeks) & set(les_weeks):
                    all_conflicting_classes.append(lesson['class_name'])

        # Benzersiz sınıfları al
        unique_conflicting_classes = sorted(list(set(all_conflicting_classes)))
        print(f"🔍 Çakışan sınıflar: {unique_conflicting_classes}")

        for existing in existing_teacher_lessons:
            print(f"  Kontrol: {existing['start_time']}-{existing['end_time']} vs {data['start_time']}-{data['end_time']}")
            # Saat çakışması var mı kontrol et
            if check_time_overlap(data['start_time'], data['end_time'],
                                  existing['start_time'], existing['end_time']):
                print(f"  ⚠️ Saat çakışması bulundu!")
                # Hafta kontrolü
                requested_weeks = data['weeks'].split(',') if data['weeks'] != 'all' else ['1', '2', '3', '4']
                existing_weeks = existing['weeks'].split(',') if existing['weeks'] != 'all' else ['1', '2', '3', '4']

                common_weeks = set(requested_weeks) & set(existing_weeks)
                print(f"  Ortak haftalar: {common_weeks}")
                if common_weeks:
                    week_text = ', '.join(sorted(common_weeks)) if len(common_weeks) < 4 else 'Tüm haftalarda'

                    # 🆕 GRUP DERSİ SEÇENEĞİ
                    # force_group=True gelirse grup dersi olarak kaydet
                    if not data.get('force_group', False):
                        # Önceden toplanan çakışan sınıfları kullan
                        classes_text = ', '.join(unique_conflicting_classes) if unique_conflicting_classes else existing['class_name']

                        # Yeni sınıf ile birlikte toplam liste
                        all_classes_with_new = sorted(list(set(unique_conflicting_classes + [data['class_name']])))

                        conn.close()
                        error_msg = f"⚠️ ÖĞRETMEN ÇAKIŞMASI!\n\n{teacher['name']} {teacher['surname']} öğretmenin {data['day']} günü {data['start_time']}-{data['end_time']} saatinde zaten {classes_text} sınıfları ile dersi var!\n\nHafta: {week_text}\n\nBu grup dersi olarak kaydedilsin mi?"
                        print(f"⚠️ UYARI: {error_msg}")
                        return jsonify({
                            'success': False,
                            'group_option': True,  # 🆕 Grup dersi seçeneği sun
                            'existing_classes': unique_conflicting_classes,  # 🆕 Tüm sınıflar
                            'all_classes': all_classes_with_new,  # 🆕 Yeni sınıf dahil
                            'error': error_msg
                        })
                    else:
                        print("✅ force_group=True - Grup dersi olarak kaydedilecek!")


        # ═══════════════════════════════════════════════════════════
        # 🆕 ÖĞRENCİ UYARILARI (force ile geçilebilir)
        # ═══════════════════════════════════════════════════════════

        # Öğrenci kontrolü
        for student in students:
            student_name = f"{student['name']} {student['surname']}"

            print(f"🔍 Öğrenci kontrolü: {student_name}")

            # A) Kısıtlama kontrolü
            restrictions = []
            if student['restrictions']:
                try:
                    restrictions = json.loads(student['restrictions'])
                    print(f"  📋 Kısıtlamalar: {restrictions}")
                except:
                    restrictions = []
                    print(f"  ⚠️ Kısıtlama parse hatası!")
            else:
                print(f"  ℹ️ Kısıtlama yok")

            for restriction in restrictions:
                print(f"  🔍 Kısıtlama kontrol: {restriction}")

                # days array olarak gelebilir (yeni format) veya day string olabilir (eski format)
                restriction_days = restriction.get('days', [])
                if not restriction_days and restriction.get('day'):
                    # Eski format: day string → days array'e çevir
                    restriction_days = [restriction.get('day')]

                print(f"  📅 Kısıtlama günleri: {restriction_days}")
                print(f"  🎯 Atanan gün: {data['day']}")

                # Gün kontrolü - atanan gün, kısıtlama günleri arasında mı?
                if data['day'] in restriction_days:
                    print(f"  ✅ Gün eşleşti: {data['day']} in {restriction_days}")

                    # Saat aralığı kontrolü
                    start_time = restriction.get('start_time')
                    end_time = restriction.get('end_time')

                    print(f"  ⏰ Kısıtlama saati: {start_time}-{end_time}")
                    print(f"  ⏰ Atanan saat: {data['start_time']}-{data['end_time']}")

                    if start_time and end_time:
                        # Saat çakışması var mı kontrol et
                        if check_time_overlap(data['start_time'], data['end_time'], start_time, end_time):
                            print(f"  ⚠️ UYARI: Saat kısıtlaması çakışıyor!")
                            warnings.append({
                                'student': student_name,
                                'type': 'time_restriction',
                                'message': f"{start_time}-{end_time} saatinde müsait değil"
                            })
                        else:
                            print(f"  ℹ️ Saat çakışmıyor")
                    else:
                        # Sadece gün kısıtlaması (saat yok)
                        print(f"  ⚠️ UYARI: Sadece gün kısıtlaması")
                        warnings.append({
                            'student': student_name,
                            'type': 'day_restriction',
                            'message': f"{data['day']} günü müsait değil"
                        })
                else:
                    print(f"  ℹ️ Gün eşleşmedi: {data['day']} not in {restriction_days}")

            # B) Öğretmen engeli kontrolü
            teacher_blocks = []
            if student['teacher_blocks']:
                try:
                    teacher_blocks = json.loads(student['teacher_blocks'])
                    print(f"  🚫 Engellenen öğretmenler: {teacher_blocks}")
                except:
                    teacher_blocks = []
                    print(f"  ⚠️ Engelleme parse hatası!")
            else:
                print(f"  ℹ️ Öğretmen engeli yok")

            for block in teacher_blocks:
                if block.get('teacher_id') == data['teacher_id']:
                    print(f"  ⚠️ UYARI: {student_name} bu öğretmeni engellemiş!")
                    warnings.append({
                        'student': student_name,
                        'type': 'teacher_blocked',
                        'message': f"{teacher['name']} {teacher['surname']}'dan ders alamaz"
                    })

        # Uyarılar varsa döndür
        if warnings:
            print(f"⚠️ {len(warnings)} uyarı bulundu!")
            for w in warnings:
                print(f"  - {w['student']}: [{w['type']}] {w['message']}")
            conn.close()
            return jsonify({
                'success': False,
                'warnings': warnings,
                'message': 'Çakışmalar tespit edildi! Yine de kaydetmek ister misiniz?'
            })

        print("✅ Çakışma kontrolü TAMAMLANDI - Sorun yok!")
    else:
        print("⏭️ Çakışma kontrolü ATLANDI (force=True)")

    # Kaydet
    print("💾 Veritabanına kaydediliyor...")

    # 🆕 Grup dersi mi? Force mi?
    is_group = 1 if data.get('force_group', False) else 0
    is_forced = 1 if force_mode else 0

    # 🆕 GRUP DERSİ İSE, MEVCUT AYNI GÜN/SAAT/ÖĞRETMEN KAYITLARINI DA GÜNCELLE
    if is_group == 1:
        print("🔄 Grup dersi olarak kaydediliyor - mevcut kayıtlar güncelleniyor...")
        cursor.execute('''
            UPDATE class_lessons
            SET is_group = 1
            WHERE teacher_id = ?
            AND day = ?
            AND start_time = ?
            AND end_time = ?
        ''', (data['teacher_id'], data['day'], data['start_time'], data['end_time']))
        updated_count = cursor.rowcount
        print(f"✅ {updated_count} mevcut kayıt is_group=1 olarak güncellendi")

    cursor.execute('''
        INSERT INTO class_lessons (class_name, teacher_id, day, start_time, end_time, weeks, is_group, is_forced)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (data['class_name'], data['teacher_id'], data['day'],
          data['start_time'], data['end_time'], data['weeks'], is_group, is_forced))


    conn.commit()
    lesson_id = cursor.lastrowid
    print(f"✅ Kayıt BAŞARILI - lesson_id: {lesson_id}")

    # 🆕 DİNAMİK SCHEDULE_DATA GÜNCELLEMESİ
    global schedule_data
    if schedule_data and schedule_data.get('weeks'):
        print("🔄 Schedule_data dinamik olarak güncelleniyor...")

        # Öğretmen bilgilerini çek
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM teachers WHERE id = ?', (data['teacher_id'],))
        teacher_row = cursor.fetchone()

        if teacher_row:
            teacher_name = f"{teacher_row['name']} {teacher_row['surname']}"
            teacher_branch = teacher_row['branch']

            # Bu sınıftaki öğrencileri çek
            cursor.execute('SELECT id, name, surname, class FROM students WHERE class = ?', (data['class_name'],))
            students = cursor.fetchall()

            # Hangi haftalara eklenecek?
            weeks_list = []
            if data['weeks'] == 'all':
                weeks_list = [1, 2, 3, 4]
            else:
                weeks_list = [int(w) for w in data['weeks'].split(',')]

            print(f"📅 {len(students)} öğrenci için {len(weeks_list)} haftaya ders ekleniyor...")

            # Her hafta için
            for week_num in weeks_list:
                if week_num <= len(schedule_data['weeks']):
                    # Her öğrenci için ders objesi oluştur
                    for student in students:
                        student_name = f"{student['name']} {student['surname']}"

                        lesson = {
                            'day': data['day'],
                            'time': f"{data['start_time']}-{data['end_time']}",
                            'teacher_name': teacher_name,
                            'branch': teacher_branch,
                            'student_name': student_name,
                            'student_class': data['class_name'],
                            'week': week_num,
                            'is_class_lesson': True,
                            'is_group': is_group,
                            'is_forced': is_forced
                        }

                        # Schedule_data'ya ekle
                        schedule_data['weeks'][week_num - 1].append(lesson)

            print(f"✅ Schedule_data güncellendi! Toplam {len(students) * len(weeks_list)} ders eklendi")

        cursor.close()

    conn.close()

    return jsonify({
        'success': True,
        'message': 'Sınıf dersi başarıyla kaydedildi!',
        'lesson_id': lesson_id,
        'debug_info': f"Force: {force_mode}, Checked conflicts: {not force_mode}"
    })

@app.route('/get_class_lessons')
def get_class_lessons():
    """Tüm sınıf derslerini getir"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            cl.id,
            cl.class_name,
            cl.teacher_id,
            cl.day,
            cl.start_time,
            cl.end_time,
            cl.weeks,
            cl.is_group,
            cl.created_at,
            t.name as teacher_name,
            t.surname as teacher_surname,
            t.branch as teacher_branch
        FROM class_lessons cl
        JOIN teachers t ON cl.teacher_id = t.id
        ORDER BY cl.day, cl.start_time
    ''')

    lessons = []
    for row in cursor.fetchall():
        # Bu sınıftaki öğrenci sayısını al
        cursor.execute('SELECT COUNT(*) as count FROM students WHERE class = ?',
                      (row['class_name'],))
        student_count = cursor.fetchone()['count']

        lessons.append({
            'id': row['id'],
            'class_name': row['class_name'],
            'teacher_id': row['teacher_id'],
            'teacher_name': f"{row['teacher_name']} {row['teacher_surname']}",
            'teacher_branch': row['teacher_branch'],
            'day': row['day'],
            'start_time': row['start_time'],
            'end_time': row['end_time'],
            'weeks': row['weeks'],
            'is_group': row['is_group'],
            'student_count': student_count,
            'created_at': row['created_at']
        })

    conn.close()
    return jsonify(lessons)

@app.route('/delete_class_lesson/<int:lesson_id>', methods=['DELETE'])
def delete_class_lesson(lesson_id):
    """Sınıf dersini sil"""
    conn = get_db()
    cursor = conn.cursor()

    # Silmeden önce ders bilgilerini al (schedule_data'dan silmek için)
    cursor.execute('SELECT * FROM class_lessons WHERE id = ?', (lesson_id,))
    lesson_row = cursor.fetchone()

    if lesson_row:
        cursor.execute('DELETE FROM class_lessons WHERE id = ?', (lesson_id,))
        conn.commit()

        # 🆕 DİNAMİK SCHEDULE_DATA GÜNCELLEMESİ
        global schedule_data
        if schedule_data and schedule_data.get('weeks'):
            print("🔄 Schedule_data'dan ders siliniyor...")

            # Öğretmen bilgisini al
            cursor.execute('SELECT name, surname FROM teachers WHERE id = ?', (lesson_row['teacher_id'],))
            teacher_row = cursor.fetchone()
            teacher_name = f"{teacher_row['name']} {teacher_row['surname']}" if teacher_row else ""

            # Sınıftaki öğrencileri al
            cursor.execute('SELECT name, surname FROM students WHERE class = ?', (lesson_row['class_name'],))
            students = cursor.fetchall()

            time_str = f"{lesson_row['start_time']}-{lesson_row['end_time']}"

            # Hangi haftalardaki dersler silinecek?
            weeks_list = []
            if lesson_row['weeks'] == 'all':
                weeks_list = [1, 2, 3, 4]
            else:
                weeks_list = [int(w) for w in lesson_row['weeks'].split(',')]

            deleted_count = 0
            for week_num in weeks_list:
                if week_num <= len(schedule_data['weeks']):
                    # Bu haftadaki dersleri filtrele - öğrencilerin derslerini sil
                    original_len = len(schedule_data['weeks'][week_num - 1])

                    schedule_data['weeks'][week_num - 1] = [
                        lesson for lesson in schedule_data['weeks'][week_num - 1]
                        if not (
                            lesson.get('day') == lesson_row['day'] and
                            lesson.get('time') == time_str and
                            lesson.get('teacher_name') == teacher_name and
                            lesson.get('student_class') == lesson_row['class_name']
                        )
                    ]

                    deleted_count += original_len - len(schedule_data['weeks'][week_num - 1])

            print(f"✅ Schedule_data'dan {deleted_count} ders silindi")

            # 🆕 GRUP DERSİ DURUMU KONTROLÜ
            # Aynı slot'ta başka sınıf kaldı mı?
            cursor.execute('''
                SELECT DISTINCT class_name
                FROM class_lessons
                WHERE teacher_id = ?
                AND day = ?
                AND start_time = ?
                AND end_time = ?
            ''', (lesson_row['teacher_id'], lesson_row['day'],
                  lesson_row['start_time'], lesson_row['end_time']))

            remaining_classes = [row['class_name'] for row in cursor.fetchall()]
            print(f"🔍 Aynı slot'ta kalan sınıflar: {remaining_classes}")

            if len(remaining_classes) <= 1:
                # Artık grup dersi değil! is_group=0 yap
                print(f"⚠️ {len(remaining_classes)} sınıf kaldı - GRUP DERSİ DEĞİL!")

                cursor.execute('''
                    UPDATE class_lessons
                    SET is_group = 0
                    WHERE teacher_id = ?
                    AND day = ?
                    AND start_time = ?
                    AND end_time = ?
                ''', (lesson_row['teacher_id'], lesson_row['day'],
                      lesson_row['start_time'], lesson_row['end_time']))
                conn.commit()
                print(f"✅ is_group=0 güncellendi")

                # Schedule_data'da da güncelle
                for week_num in weeks_list:
                    if week_num <= len(schedule_data['weeks']):
                        for lesson in schedule_data['weeks'][week_num - 1]:
                            if (lesson.get('day') == lesson_row['day'] and
                                lesson.get('time') == time_str and
                                lesson.get('teacher_name') == teacher_name):
                                lesson['is_group'] = 0
                print(f"✅ Schedule_data'da is_group güncellendi")

        conn.close()
        return jsonify({'success': True, 'message': 'Sınıf dersi silindi!'})
    else:
        conn.close()
        return jsonify({'error': 'Ders bulunamadı!'}), 404

@app.route('/update_class_lesson', methods=['POST'])
def update_class_lesson():
    """Sınıf dersini güncelle - çakışma kontrolü ile"""
    data = request.json

    # Gerekli alanları kontrol et
    required_fields = ['lesson_id', 'class_name', 'teacher_id', 'day', 'start_time', 'end_time', 'weeks']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'{field} alanı eksik!'}), 400

    lesson_id = data['lesson_id']

    conn = get_db()
    cursor = conn.cursor()

    # Sınıftaki öğrencileri getir
    cursor.execute('SELECT id, name, surname, restrictions, teacher_blocks FROM students WHERE class = ?',
                   (data['class_name'],))
    students = cursor.fetchall()

    if not students:
        conn.close()
        return jsonify({'error': 'Bu sınıfta öğrenci bulunamadı!'}), 400

    # Öğretmeni getir
    cursor.execute('SELECT id, name, surname, branch, schedule FROM teachers WHERE id = ?',
                   (data['teacher_id'],))
    teacher = cursor.fetchone()

    if not teacher:
        conn.close()
        return jsonify({'error': 'Öğretmen bulunamadı!'}), 400

    # 🆕 1. AŞAMA: ÖĞRETMEN ÇAKIŞMASI ONAY KONTROLÜ
    # Sınıf dersi = öğretmen çakışması demektir (aynı öğretmen birden fazla öğrenciye ders veriyor)
    force_teacher_conflict = data.get('force_teacher_conflict', False)
    print(f"🔒 Force Teacher Conflict: {force_teacher_conflict}")

    if not force_teacher_conflict:
        # Kullanıcı henüz öğretmen çakışmasını onaylamadı
        # Uyarı ver ve onay iste
        conn.close()
        return jsonify({
            'success': False,
            'teacher_conflict_warning': True,
            'message': f"⚠️ ÖĞRETMEN ÇAKIŞMASI!\n\n{data['class_name']} sınıfına {teacher['name']} {teacher['surname']} öğretmeni atanacak.\n\n{data['day']} günü {data['start_time']}-{data['end_time']} saatinde bu sınıftaki tüm öğrenciler aynı öğretmenden ders alacak.\n\nBu işlem öğretmen çakışması oluşturacaktır. Devam etmek istiyor musunuz?"
        })

    # 🆕 2. AŞAMA: ÖĞRENCİ KISITLAMALARI VE DİĞER KONTROLLER
    # Kullanıcı force=True göndermemişse uyarıları kontrol et
    if not data.get('force', False):
        warnings = []

        print(f"🔍 Güncelleme çakışma kontrolü: {data['class_name']} - {data['day']} {data['start_time']}-{data['end_time']}")

        # 1️⃣ AYNI SINIFA AYNI GÜN/SAATTE BAŞKA DERS ATANMIŞ MI? (kendisi hariç!)
        cursor.execute('''
            SELECT t.name, t.surname, t.branch, cl.start_time, cl.end_time, cl.weeks, cl.teacher_id
            FROM class_lessons cl
            JOIN teachers t ON cl.teacher_id = t.id
            WHERE cl.class_name = ?
            AND cl.day = ?
            AND cl.id != ?
        ''', (data['class_name'], data['day'], lesson_id))

        existing_class_lessons = cursor.fetchall()
        print(f"📋 Sınıf için mevcut dersler: {len(existing_class_lessons)}")

        for existing in existing_class_lessons:
            print(f"  Kontrol: {existing['start_time']}-{existing['end_time']} vs {data['start_time']}-{data['end_time']}")
            if check_time_overlap(data['start_time'], data['end_time'],
                                  existing['start_time'], existing['end_time']):
                print(f"  ⚠️ Saat çakışması bulundu!")
                requested_weeks = data['weeks'].split(',') if data['weeks'] != 'all' else ['1', '2', '3', '4']
                existing_weeks = existing['weeks'].split(',') if existing['weeks'] != 'all' else ['1', '2', '3', '4']

                common_weeks = set(requested_weeks) & set(existing_weeks)
                print(f"  Ortak haftalar: {common_weeks}")
                if common_weeks:
                    week_text = ', '.join(sorted(common_weeks)) if len(common_weeks) < 4 else 'Tüm haftalarda'

                    if existing['teacher_id'] == data['teacher_id']:
                        error_msg = f"🚫 TEKRAR ATAMA!\n\n{data['class_name']} sınıfının {data['day']} günü {data['start_time']}-{data['end_time']} saatinde zaten bu ders kayıtlı!\n\nHafta: {week_text}\n\nAynı sınıfa aynı gün ve saatte aynı dersi tekrar atayamazsınız!"
                    else:
                        error_msg = f"🚫 SINIF ÇAKIŞMASI!\n\n{data['class_name']} sınıfının {data['day']} günü {data['start_time']}-{data['end_time']} saatinde zaten {existing['name']} {existing['surname']} ({existing['branch']}) ile dersi var!\n\nHafta: {week_text}\n\nAynı sınıfa aynı gün ve saatte iki öğretmen atanamaz!"

                    conn.close()
                    print(f"🚫 HATA: {error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    })

        # 2️⃣ AYNI ÖĞRETMENE AYNI GÜN/SAATTE BAŞKA SINIF ATANMIŞ MI? (kendisi hariç!)
        cursor.execute('''
            SELECT cl.class_name, cl.start_time, cl.end_time, cl.weeks
            FROM class_lessons cl
            WHERE cl.teacher_id = ?
            AND cl.day = ?
            AND cl.class_name != ?
            AND cl.id != ?
        ''', (data['teacher_id'], data['day'], data['class_name'], lesson_id))

        existing_teacher_lessons = cursor.fetchall()
        print(f"📋 Öğretmen için mevcut dersler: {len(existing_teacher_lessons)}")

        for existing in existing_teacher_lessons:
            print(f"  Kontrol: {existing['start_time']}-{existing['end_time']} vs {data['start_time']}-{data['end_time']}")
            if check_time_overlap(data['start_time'], data['end_time'],
                                  existing['start_time'], existing['end_time']):
                print(f"  ⚠️ Saat çakışması bulundu!")
                requested_weeks = data['weeks'].split(',') if data['weeks'] != 'all' else ['1', '2', '3', '4']
                existing_weeks = existing['weeks'].split(',') if existing['weeks'] != 'all' else ['1', '2', '3', '4']

                common_weeks = set(requested_weeks) & set(existing_weeks)
                print(f"  Ortak haftalar: {common_weeks}")
                if common_weeks:
                    week_text = ', '.join(sorted(common_weeks)) if len(common_weeks) < 4 else 'Tüm haftalarda'
                    conn.close()
                    error_msg = f"🚫 ÖĞRETMEN ÇAKIŞMASI!\n\n{teacher['name']} {teacher['surname']} öğretmenin {data['day']} günü {data['start_time']}-{data['end_time']} saatinde zaten {existing['class_name']} sınıfı ile dersi var!\n\nHafta: {week_text}\n\nAynı öğretmene aynı gün ve saatte iki sınıf atanamaz!"
                    print(f"🚫 HATA: {error_msg}")
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    })

        # Öğrenci uyarı kontrolleri (opsiyonel - force ile geçilebilir)
        # Şimdilik atlıyoruz, gerekirse ekleriz

    # Güncelle
    cursor.execute('''
        UPDATE class_lessons
        SET class_name = ?, teacher_id = ?, day = ?, start_time = ?, end_time = ?, weeks = ?
        WHERE id = ?
    ''', (data['class_name'], data['teacher_id'], data['day'],
          data['start_time'], data['end_time'], data['weeks'], lesson_id))

    conn.commit()

    # 🆕 DİNAMİK SCHEDULE_DATA GÜNCELLEMESİ
    global schedule_data
    if schedule_data and schedule_data.get('weeks'):
        print("🔄 Schedule_data güncelleniyor (update)...")

        # Önce eski dersleri sil
        cursor.execute('SELECT * FROM class_lessons WHERE id = ?', (lesson_id,))
        old_lesson = cursor.fetchone()

        if old_lesson:
            # Eski öğretmen bilgisi
            cursor.execute('SELECT name, surname FROM teachers WHERE id = ?', (old_lesson['teacher_id'],))
            old_teacher_row = cursor.fetchone()
            old_teacher_name = f"{old_teacher_row['name']} {old_teacher_row['surname']}" if old_teacher_row else ""
            old_time_str = f"{old_lesson['start_time']}-{old_lesson['end_time']}"

            # Eski hafta listesi
            old_weeks = []
            if old_lesson['weeks'] == 'all':
                old_weeks = [1, 2, 3, 4]
            else:
                old_weeks = [int(w) for w in old_lesson['weeks'].split(',')]

            # Eski dersleri sil
            for week_num in old_weeks:
                if week_num <= len(schedule_data['weeks']):
                    schedule_data['weeks'][week_num - 1] = [
                        lesson for lesson in schedule_data['weeks'][week_num - 1]
                        if not (
                            lesson.get('day') == old_lesson['day'] and
                            lesson.get('time') == old_time_str and
                            lesson.get('teacher_name') == old_teacher_name and
                            lesson.get('student_class') == old_lesson['class_name']
                        )
                    ]

            print(f"✅ Eski dersler schedule_data'dan silindi")

        # Şimdi yeni dersleri ekle
        cursor.execute('SELECT name, surname, branch FROM teachers WHERE id = ?', (data['teacher_id'],))
        new_teacher_row = cursor.fetchone()

        if new_teacher_row:
            new_teacher_name = f"{new_teacher_row['name']} {new_teacher_row['surname']}"
            new_teacher_branch = new_teacher_row['branch']

            # Sınıftaki öğrencileri çek
            cursor.execute('SELECT name, surname FROM students WHERE class = ?', (data['class_name'],))
            students = cursor.fetchall()

            # Yeni hafta listesi
            new_weeks = []
            if data['weeks'] == 'all':
                new_weeks = [1, 2, 3, 4]
            else:
                new_weeks = [int(w) for w in data['weeks'].split(',')]

            # Yeni dersleri ekle
            for week_num in new_weeks:
                if week_num <= len(schedule_data['weeks']):
                    for student in students:
                        student_name = f"{student['name']} {student['surname']}"

                        lesson = {
                            'day': data['day'],
                            'time': f"{data['start_time']}-{data['end_time']}",
                            'teacher_name': new_teacher_name,
                            'branch': new_teacher_branch,
                            'student_name': student_name,
                            'student_class': data['class_name'],
                            'week': week_num,
                            'is_class_lesson': True,
                            'is_group': 0  # Update'te grup bilgisi yok, varsayılan 0
                        }

                        schedule_data['weeks'][week_num - 1].append(lesson)

            print(f"✅ Yeni dersler schedule_data'ya eklendi")

    conn.close()

    return jsonify({
        'success': True,
        'message': 'Sınıf dersi başarıyla güncellendi!'
    })

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)